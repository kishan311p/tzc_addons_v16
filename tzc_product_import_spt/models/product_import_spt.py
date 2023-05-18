# -*- coding: utf-8 -*-
# Part of SnepTech. See LICENSE file for full copyright and licensing details.##
###############################################################################

from odoo import api, fields, models, _
from odoo.exceptions import UserError
import requests
import json
import base64
import math
import os
from psycopg2 import IntegrityError
import xlrd
import xlsxwriter
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import pandas as pd

folder_path = '/home/sneptech/workspace/test/'
fields_dict = {
    # 'Pos categories' : 'pos_categ_id','Available In POS': 'available_in_pos','Qty': 'qty','ETO Sale Method' : 'eto_sale_method','Shape': 'shape_ids,
    'Rim Type' : 'rim_type','Product categories': 'categ_id','Gender': 'gender','Aging': 'aging','Shape': 'shape_id','Id' : 'default_code', 'Name' : 'variant_name', 'Active' : 'active', 'List Price': 'list_price', 'Standard Price' : 'standard_price','Wholesale Price': 'price_wholesale' ,
    'MSRP Price' : 'price_msrp', 'Barcode' : 'barcode', 'Brand' : 'brand', 'Model': 'model','Flex Hinges':'flex_hinges','Image 1 URL' : 'image_url', 'Image 2 URL': 'image_secondary_url',
    'HS Code': 'hs_code','Material': 'material_id','Volume' : 'volume','Weight':'weight', 'County of Origin':'country_of_origin','Bridge Size' : 'bridge_size','Temple Size': 'temple_size','Geo Restriction': 'geo_restriction','Lense Color Name':'lense_color_name','Type': 'type','Website Description':'website_description','Website Published':'website_published','Custom Message': 'custom_message',
    'Website Url Keyword':'product_seo_keyword', 'Weight': 'weight',
    'Color Name': 'product_color_name','Color': 'color', 'HTML Color Code': 'html_color',
    # 'eCommerce Category': 'kits_ecom_categ_id', ,'Material': 'material_ids'
    'Replenishable' : 'replenishable','Secondary Color Name' : 'secondary_color_name','Secondary HTML Color Code': 'secondary_html_color',
    'Wholesale Price':'price_wholesale','MSRP Price':'price_msrp','Is Select For Lenses':'is_select_for_lenses',
    'Temporary Out Of Stock': 'temporary_out_of_stock',
    #'MSRP Price In USD':'price_msrp_usd', 'Wholesale Price In USD':'price_wholesale_usd','List Price In USD': 'lst_price_usd', 'On sale' : 'on_sale','On sale price' : 'on_sale_usd','On sale discount type': 'on_sale_usd_in_percentage'
    'Sale price': 'on_sale_usd','Sale type': 'sale_type','Sale discount type': 'on_sale_usd_in_percentage','New Arrivals' :'new_arrivals','Force Unpublished': 'is_forcefully_unpublished','B2C Published': 'is_b2c_published',
    'Length': 'length','Width': 'width','Height': 'height','Product Brand Commission':'product_brand_commission','Case Image URL':'case_image_url','Case Type':'case_type','Minimum Qty':'minimum_qty','On Consignment':'on_consignment','New Price':'is_new_price','Eye Size':'eye_size','Qty': 'qty','Customer Taxes' : 'taxes_id','Application Type': 'application_type','Is 3D Model' : 'is_3d_model','Meta Keyword':'meta_keyword',
    'Meta Title':'meta_title','Meta Description':'meta_description'
    }
class product_import_spt(models.Model):
    _name = 'product.import.spt'
    _description = 'Import Product'
    _order = 'id desc'


    name = fields.Char('Name', default='New')
    create_date = fields.Datetime('Date',readonly="1")

    attach_file = fields.Binary("Attached File",readonly=True, states={'draft': [('readonly', False)]})
    attach_file_name = fields.Char("Attached File Name")
    report_file = fields.Binary("Products Report",default='',copy=False)
    read_time_file_name = fields.Char("Read Time File Name")
    read_time = fields.Binary("Read Time Products",default='',copy=False)
    
    run_time_file_name = fields.Char("Run Time File Name")
    run_time = fields.Binary("Run Time Products",default='',copy=False)

    qty_add_in_pro_qty = fields.Boolean('Add On Product Qty')
    add_reserved_qty = fields.Boolean('Add Reserved Qty',default=True)

    import_line_ids = fields.One2many('product.import.line.spt', 'import_id', 'Product Lines', domain="['|',('active','=',True),('active','=',False)]",context={'active_test': False})
    number_of_product_variant = fields.Integer('Number Of Product variant', compute='_get_number_of_product') 
    number_of_product = fields.Integer('Number Of Product', compute='_get_number_of_product')
    old_product_ids = fields.Many2many('product.product', 'product_import_old_product_rel_spt', 'product_import_id', 'old_product_tmpl_id', string='Products',context={'active_test': False}, copy=False, domain="['|',('active','=',True),('active','=',False)]")
    product_ids = fields.Many2many('product.template', 'product_import_product_tmpl_rel_spt', 'product_import_id', 'product_tmpl_id', string=' Products',context={'active_test': False}, copy=False, domain="['|',('active','=',True),('active','=',False)]")
    product_pro_ids = fields.Many2many('product.product', 'product_import_product_product_pro_rel_spt', 'product_import_id', 'product_pro_id', string='Product Variants',context={'active_test': False}, copy=False, domain="['|',('active','=',True),('active','=',False)]")
    image_path = fields.Char('Image Path')
    state = fields.Selection([
        ('draft','Draft'),
        ('process','In Process'),
        ('done','Done'),
    ], string='State', default='draft')

    based_on_categories = fields.Selection([
        ('e_or_s','Eyeglasses or Sunglasses'),
        ('case','Cases'),
    ], string='Based On Categories', default='e_or_s')

    data_on = fields.Selection([
        ('create','Create'),
        ('update','Update'),  
        ('delete','Delete'),
        ('image_change','Rename Image'),  
    ],readonly=True, states={'draft': [('readonly', False)]}, string='Operation',required=True,default="create")

    categ_id = fields.Many2one('product.category', 'Default Category')
    delete_product_result = fields.Binary("Delete Operation Result",default='',copy=False)
    delete_product_name = fields.Char("Delete Operation Result name",default='')
    case_type = fields.Selection([('original', 'Original'),('generic', 'Generic')],"Case Type")
    case_image_url = fields.Char('Case Image Path')
    column_name = fields.Char('Archive Keyword')


    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('product.import.spt') or 'New'
        res = super(product_import_spt, self).create(vals)
        for record in res:
            if record.qty_add_in_pro_qty:
                    record.add_reserved_qty = False
        return res
    
    def write(self,vals):
        for record in self:
                if (record.qty_add_in_pro_qty and vals.get('qty_add_in_pro_qty') and vals['qty_add_in_pro_qty']) or (vals.get('qty_add_in_pro_qty') and vals['qty_add_in_pro_qty'] and not record.qty_add_in_pro_qty) :
                    vals['add_reserved_qty'] = False
        return super(product_import_spt, self).write(vals)

    @api.onchange('qty_add_in_pro_qty')
    def onchange_qty_add_in_pro_qty(self):
        for record in self:
            if record.qty_add_in_pro_qty and record.add_reserved_qty:
                record.add_reserved_qty = False

    def action_view_products(self):
        self.ensure_one()
        try:
            list_view = self.env.ref('product.product_template_tree_view')
            form_view = self.env.ref('product.product_template_only_form_view')
        except ValueError:
            list_view = False
            form_view = False
        return {
            'name': 'Products',
            'view_type': 'form',
            'view_mode': 'tree,form',
            'res_model': 'product.template',
            'view_id': False,
            'views': [(list_view.id, 'tree'),(form_view.id, 'form')],
            'type': 'ir.actions.act_window',
            'target':'current',
            'domain':['|',('active','=',True),('active','=',False),('id','in',self.product_ids.ids)],
            'context':{'active_test': False, }
        }   

    def action_view_products_variants(self):
        self.ensure_one()
        try:
            list_view = self.env.ref('product.product_product_tree_view')
            form_view = self.env.ref('product.product_normal_form_view')
        except ValueError:
            list_view = False
            form_view = False
        return {
            'name': 'Products',
            'view_type': 'form',
            'view_mode': 'tree,form',
            'res_model': 'product.product',
            'view_id': False,
            'views': [(list_view.id, 'tree'),(form_view.id, 'form')],
            'type': 'ir.actions.act_window',
            'target':'current',
            'domain':[('id','in',self.product_pro_ids.ids)],
            'context':{'active_test': False, 'pending_price' :True}
        }   
    
    def _get_number_of_product(self):
        for record in self:
            record.number_of_product = len(record.product_ids)
            record.number_of_product_variant = len(record.product_pro_ids)
    
    def action_set_to_done(self):
        for record in self:
            record.state = 'done'
    
    def action_reset_to_draft(self):
        for record in self:
            record.state = 'draft'

    def action_update_product_check_error_file(self):
        self.ensure_one()
        if self.read_time:
            return {
                'name': 'Error',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'kits.read.file.error.message',
                'views': [(self.env.ref('tzc_product_import_spt.kits_read_file_error_message_form_view').id, 'form')],
                'type': 'ir.actions.act_window',
                'target':'new',
                'context':{'default_product_import_id': self.id, 'action_process':'action_update_product'}
            }   
        else:
            self.action_update_product()

    def action_update_product(self,wrong_lines=[],product_ids=[]):
        if not wrong_lines:
            wrong_lines = [['Id','Error']]
        self.ensure_one()
        pricelist = self.env['product.pricelist'].search([]).mapped('name')
        product_obj = self.env['product.product']
        # inventory_obj = self.env['stock.inventory']
        # inventory_line_obj = self.env['stock.inventory.line']
        quant_obj = self.env['stock.quant']
        import_line_obj = self.env['product.import.line.spt']
        product_pro_dict = {}
        inventory_line_list = []
        inventory_product_list = []
        product_pro_ids_list = []
        fields_dict,heading,file_data=self.get_header_data()
        heading.append('Error')
        # if 'Qty' in heading:
        #     inventory_id = inventory_obj.create({
        #                             'name' : self.name or self.attach_file_name,
        #                             })
        error_default_code_list = list(map(lambda x: x[0],wrong_lines))
        for field in tuple(fields_dict.keys()):
            if field in pricelist :
                del fields_dict[field] 
        search_filed_list = list(fields_dict.keys())
        import_line_ids =  self.import_line_ids.search([('import_id','=',self.id),'|',('active','=',False),('active','=',True)])
        for import_line in range(0,len(import_line_ids)):
            print(str(import_line))
            import_line = import_line_ids[import_line]
            product_pro_id = product_obj.with_context(pending_price=True).sudo().search([('default_code','=',import_line.default_code),'|',('active','=',False),('active','=',True)],limit=1)
            try:
                if len(product_pro_id)>1:
                    if import_line.default_code not in error_default_code_list:
                        wrong_lines.append([import_line.default_code,'Error:'+str("Duplicate sku found, so can't update a product.")])
                    continue
                if product_pro_id and product_pro_id.default_code not in error_default_code_list:
                    if self.data_on == 'create':
                        search_filed_list.extend(['meta_keyword','meta_title','meta_description'])
                    line_data = import_line_obj.search_read([('id','=',import_line.id),'|',('active','=',False),('active','=',True)],search_filed_list)[0]
                    product_pro_dict = self.convert_dict(fields_dict,line_data)
                    if 'Website Published' in heading:
                        product_pro_dict['is_published_spt'] = product_pro_dict.get('website_published')
                    if 'website_published' in product_pro_dict.keys():
                        del product_pro_dict['website_published']

                    if product_pro_dict.get('sale_type'):
                        if product_pro_dict['sale_type'] != 'on_sale':
                            product_pro_dict['clearance_usd_in_percentage'] =  product_pro_dict['on_sale_usd_in_percentage'] 
                            product_pro_dict['clearance_usd'] =  product_pro_dict['on_sale_usd'] 
                            del product_pro_dict['on_sale_usd']
                            del product_pro_dict['on_sale_usd_in_percentage']
                    if 'Qty' in heading: 
                        if ('New Arrivals' not in heading and self.qty_add_in_pro_qty) or (self.data_on == 'create' and 'New Arrivals' not in heading):
                            product_pro_dict['new_arrivals'] = True
                    if 'B2C Published'  in heading:
                        self._cr.execute("select id from ir_model where model='kits.b2c.website'")
                        model = self._cr.fetchall()
                        website_id = False
                        if model and model[0] and model[0][0]:
                            website_ids = self.env['kits.b2c.website'].search([])
                        if website_ids:
                            if product_pro_dict.get('is_b2c_published'):
                                product_pro_id.website_ids = [(6,0,website_ids.ids)]
                            else:
                                product_pro_id.website_ids = False
                        else:
                            wrong_lines.append([import_line.default_code,'Error:'+str("Website id not found.")])
                    if 'is_b2c_published' in product_pro_dict.keys():
                        del product_pro_dict['is_b2c_published']
                    product_pro_id.with_context(from_product_import=True).write(product_pro_dict)
                    product_pro_ids_list.append(product_pro_id.id)

                    if 'Qty' in heading and product_pro_id:
                        if self.data_on == 'create' or self.qty_add_in_pro_qty:
                            product_pro_id.last_qty_update = fields.Datetime.now()
                        product_qty = product_pro_id.qty_available + import_line.qty if self.qty_add_in_pro_qty else import_line.qty
                        if product_pro_id.qty_available < 0:
                            product_qty = product_qty + abs( product_pro_id.qty_available)
                        if import_line.qty == 0.0 or (not self.qty_add_in_pro_qty and product_qty < 0) or product_qty < 0:
                            product_qty = 0
                        warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.user.company_id.id)], limit=1)
                        if warehouse and warehouse.lot_stock_id:
                           inventory_id = quant_obj.with_context(inventory_mode=True,inventory_name=self.name).create({
                                'location_id': warehouse.lot_stock_id.id,
                                'product_id': product_pro_id.id,
                                'inventory_quantity': product_qty
                            }).with_context(default_product_import_id=self.id).action_apply_inventory()
                        else:
                            if import_line.default_code not in error_default_code_list:
                                wrong_lines.append([import_line.default_code,'Error:'+str('stock location not found, stock updation operation can not preformed')])
                        
                        # inventory_line_dict  = {
                        #     'location_id' : location_obj.search([('name','=','Stock'),('usage','=','internal')],limit=1).id,
                        #     'product_id' :  product_pro_id.id,
                        #     'product_uom_id' : product_pro_id.uom_po_id.id,
                        #     'product_qty' :  product_qty + product_pro_id.reversed_qty_spt if self.add_reserved_qty and not self.qty_add_in_pro_qty else product_qty,
                        #     'inventory_id' : inventory_id.id,
                        # }
                        # inventory_line_id = inventory_line_obj.create(inventory_line_dict)
                        # inventory_line_list.append(inventory_line_id.id)
                        # inventory_product_list.append(product_pro_id.id)

                else:
                    if import_line.default_code not in error_default_code_list:
                        wrong_lines.append([import_line.default_code,'Error:'+str('Product Not In system.')])
            except Exception as e:
                wrong_lines.append([product_pro_id.default_code,'Error:'+str(e)])
            
            if self._context.get('from_create'):
                if import_line.default_code in error_default_code_list and product_pro_id:
                    product_pro_id.unlink()
                else:
                    product_ids = product_obj.with_context(pending_price=True).search([('brand','=',import_line.brand.id),('model','=',import_line.model.id),('categ_id','=',import_line.categ_id.id)])
                    for product in product_ids:
                        if not product.default_code:
                            self._cr.execute('delete from stock_valuation_layer where product_id = %s'%(product.id))
                            self._cr.execute('delete from stock_quant where product_id = %s'%(product.id))
                            self._cr.execute('delete from stock_move where product_id = %s'%(product.id))
                            self._cr.execute('delete from product_product where id = %s'%(product.id))
        print('\n\n')

        # try:
        #     if inventory_line_list:
        #         inventory_id.action_start()
        #         inventory_id.action_validate()
        #         inventory_id.product_ids = [(6,0,inventory_product_list)]
        # except Exception as error:
        #     wrong_lines.append(['Inventory Creation Error','Error:'+str(error)])
        self._cr.commit()
        self.product_pro_ids = [(6,0,list(set(product_pro_ids_list)))]
        self.state = 'done' 
        if fields_dict.get('sale_type'):
            self.product_pro_ids.calculate_clearance_price_for_product_import()
            self.product_pro_ids.calculate_onsale_price_for_product_import()
        fields_dict=self.get_header_data()[0]
        if any(list(map(lambda fld : True if fld in pricelist else False,fields_dict))):
            self.update_pricelist(wrong_lines)
            
        print('\n\n\n\n')
        product_id_not_set_ids = product_obj.with_context(pending_price=True).search([('default_code','in',[False,'',' ',None])])

        if product_id_not_set_ids:
            for product_id in product_id_not_set_ids:
                if not product_id.sales_count:
                    product_id.with_context(from_product_import=True).write({'active':False})
                    # product_id.active = False
                    for attribute in product_id.product_template_attribute_value_ids:
                        product_id.product_template_attribute_value_ids = [(3,attribute.id)]
                    product_id.unlink()

        self.run_time_file_name = ''
        self.run_time = False
        if len(wrong_lines) > 1:
            try:
                if self._context.get('from_create'):
                    if 'ID' in wrong_lines[0] or 'Id' in wrong_lines[0]:
                        try:
                            id_index = wrong_lines[0].index('Id') 
                        except:
                            id_index = wrong_lines[0].index('ID')
                        str_defalt_code = ','.join(list(map(lambda col : "'"+str(col[id_index])+"'" if "'" not in col[id_index] else "'"+str(col[id_index]).replace("'","''")+"'",wrong_lines[1:])))
                        
                        self._cr.execute(""" delete from product_template where default_code in (%s);
                                            delete from product_product where default_code in (%s);
                        """%(str_defalt_code,str_defalt_code))
            except:
                pass
            out = BytesIO()
            workbook = xlsxwriter.Workbook(out) 
            worksheet = workbook.add_worksheet('wrong')
            heading_line = wrong_lines.pop(0)
            col = row =0
            for heading_data in heading_line:
                worksheet.write(row,col,heading_data)
                worksheet.set_column(row,col, len(heading_data)*10)

                col+=1
            for worksheet_line in wrong_lines:
                row +=1
                col = 0
                for worksheet_line_data in worksheet_line:
                    worksheet.write(row,col,worksheet_line_data)
                    col +=1

            workbook.close()
            out.seek(0)
            data = out.read()
            out.close()
            self.run_time = base64.b64encode(data)
            self.run_time_file_name = 'wrong_product.xlsx'
        self.product_pro_ids._compute_pending_price()
        if 'active' in search_filed_list or 'qty' in search_filed_list:
            self.product_pro_ids.product_import_product_published()


    def update_pricelist(self,wrong_line=[]):
        product_obj = self.env['product.product']
        pricelist_obj = self.env['product.pricelist']
        pricelist_item_obj = self.env['product.pricelist.item']
        wrng_default_code_list = []
        if wrong_line and ('ID' in wrong_line[0] or 'Id' in wrong_line[0]):
            try:
                id_index = wrong_line[0].index('Id') 
            except:
                id_index = wrong_line[0].index('ID')
            wrng_default_code_list = [col[id_index] for col in wrong_line]
        for record in self:
            header_dict,header_data,file_data = self.get_header_data()
            pricelist_name_list = pricelist_obj.search([]).mapped('name')
            
            # if not header_dict.get('list_price'):
            #     pricelist_name_list.pop(self.env.ref('product.list0').name)

            # if not header_dict.get('lst_price_usd'):
            #     pricelist_name_list.pop(pricelist_name_list.index(self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt').name))

            for line in file_data:
                try:
                    print(line[header_dict.get('default_code')])
                    product_id =  product_obj.with_context(pending_price=True).search([('default_code','=',line[header_dict.get('default_code')].strip()),'|',('active','=',True),('active','=',False)],limit=1)
                    if product_id:
                        for pricelist in pricelist_name_list:
                            # Fenil
                            pricelist_name = ''
                            # if header_dict.get('list_price') and pricelist == self.env.ref('product.list0').name:
                            #     pricelist_id  = self.env.ref('product.list0')
                            #     pricelist_name = 'list_price'
                            # elif header_dict.get('lst_price_usd') and pricelist == self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt').name:
                            if header_dict.get('list_price') and pricelist == self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt').name:
                                pricelist_id = self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt')
                                pricelist_name = 'list_price'
                            else:
                                pricelist_id = pricelist_obj.search([('name','=',pricelist.strip())])
                                pricelist_name = pricelist_id.name
                            pricelist_item_id =  pricelist_item_obj.search([('pricelist_id','=',pricelist_id.id),('product_id.default_code','=',line[header_dict.get('default_code')].strip())])
                            if pricelist_item_id.product_id:
                                if pricelist_name in header_dict:# Fenil
                                    pricelist_item_id.write({
                                        'base' : 'list_price',
                                        'compute_price' : 'fixed',
                                        'pricelist_id' : pricelist_id.id,
                                        'product_id' : product_id.id,
                                        'fixed_price' : line[header_dict[pricelist_name]]
                                    })
                            else:
                                if product_id:
                                    if pricelist_name in header_dict:# Fenil
                                        pricelist_item_id = pricelist_item_obj.create({
                                            'applied_on' : '0_product_variant',
                                            'base' : 'list_price',
                                            'compute_price' : 'fixed',
                                            'pricelist_id' : pricelist_id.id,
                                            'product_id' : product_id.id,
                                            'fixed_price' : line[header_dict[pricelist_name]]
                                        })  
                            # print(pricelist_name,pricelist_item_id.fixed_price)
                except Exception as e:
                        if line[header_dict.get('default_code')] in wrng_default_code_list:
                            line = [col for col in wrong_line if line[header_dict.get('default_code')] == col[id_index]]
                            if line:
                                line[0][-1] = line[0][-1]+(str(e))                            
                            else:
                                wrong_line.append([line[header_dict.get('default_code')],str(e)])

                            
                        else:
                            wrong_line.append([line[header_dict.get('default_code')],str(e)])


    def action_genrate_product_report(self):
        active_id = self.ids[0]
        f_name = self.mapped('name')[0]
        out = BytesIO()
        workbook = xlsxwriter.Workbook(out) 
        worksheet = workbook.add_worksheet('Report')

        cell_format = workbook.add_format({'bold': True,'border':1})
        worksheet.write(0,0,'Id',cell_format)
        worksheet.write(0,1,"Name",cell_format)
        worksheet.write(0,3,"Product categories",cell_format)
        worksheet.write(0,2, "Barcode",cell_format)
        worksheet.write(0,4, "Brand",cell_format)
        worksheet.write(0,5, "Model",cell_format)
        worksheet.write(0,6, "Org Color Code",cell_format)
        worksheet.write(0,7, "Raw Size",cell_format)
        worksheet.write(0,8, "Qty",cell_format)
        worksheet.write(0,9, "My Color Name",cell_format)
        worksheet.write(0,10,"Secondary Color Name",cell_format)
        worksheet.write(0,11, "Lence Color Name",cell_format)
        worksheet.write(0,12, "Rim Type",cell_format)
        worksheet.write(0,13,"Shape",cell_format)
        worksheet.write(0,14, "Material",cell_format)
        worksheet.write(0,15, "Weight",cell_format)
        worksheet.write(0,16, "Gender",cell_format)
        row = 1
        for record in self:
            product_dict = {}
            for product_id in record.product_pro_ids:
                product_dict[product_id.variant_name] = product_id
            product_ids = list(product_dict.keys())
            product_ids.sort()
            for product in product_ids:
                product = product_dict[product]
                worksheet.write(row,0, product.default_code or '')
                worksheet.write(row,1, product.variant_name or '')
                worksheet.write(row,2, product.barcode or '')
                worksheet.write(row,3, product.categ_id.name or '')
                worksheet.write(row,4, product.brand.name or '')
                worksheet.write(row,5, product.model.name or '')
                # color = product.product_template_attribute_value_ids.filtered(lambda x: x.name if x.attribute_id.name == 'Color' else None)
                worksheet.write(row,6, product.color_code.name if product.color_code and product.color_code.name else '')
                size = product.size.name or '00'
                bridge_size = product.bridge_size.name or '00' 
                temple_sze = product.temple_size.name or '000'
                worksheet.write(row,7, size+'-'+bridge_size+'-'+temple_sze)
                worksheet.write(row,8, product.qty_available)
                worksheet.write(row,9, product.product_color_name.name or '')
                worksheet.write(row,10, product.secondary_color_name.name or '')
                worksheet.write(row,11, product.lense_color_name.name or '')
                worksheet.write(row,12, product.rim_type.name or '')
                worksheet.write(row,13, product.shape_id.name or '')
                worksheet.write(row,14, product.material_id.name or '')
                worksheet.write(row,15, product.weight)
                worksheet.write(row,16, 'M/F' if product.gender == 'm/f' else 'F' if product.gender == 'female' else 'M' if product.gender == 'male' else '')
                row +=1
        workbook.close()
        out.seek(0)
        data = out.read()
        out.close()
        record.report_file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=product.import.spt&download=true&field=report_file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }
            
    def read_opration_process_check_error_file(self):
        query  = """
                        DROP PROCEDURE IF EXISTS create_product_template;
                        CREATE OR REPLACE PROCEDURE public.create_product_template(
                            default_code character,
                            pname json,
                            active boolean,
                            categ_id integer,
                            userid int,
                            INOUT result json)
                        LANGUAGE 'plpgsql'
                        AS $BODY$

                        begin
                            
                            insert into product_template (default_code,name,detailed_type,active,categ_id,uom_id,uom_po_id,tracking,sale_line_warn,sale_ok,purchase_ok,invoice_policy) values (default_code,pname,'product',active,categ_id,1,1,'none','no-message','t','t','delivery');
                            
                            insert into product_product (product_tmpl_id,active,categ_id,default_code,updated_by,updated_on) values (currval('product_template_id_seq'),active,categ_id,default_code,UserId,Now());
                            
                            Select json_agg(variable) from(
                            SELECT currval('product_product_id_seq')
                            )
                            variable into result;
                        end;
                        $BODY$;

                    """
        self._cr.execute( query )
        self.ensure_one()
        if self.read_time:
            return {
                'name': 'Error',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'kits.read.file.error.message',
                'views': [(self.env.ref('tzc_product_import_spt.kits_read_file_error_message_form_view').id, 'form')],
                'type': 'ir.actions.act_window',
                'target':'new',
                'context':{'default_product_import_id': self.id, 'action_process':'read_opration_process'}
            }   
        else:
            self.read_opration_process()
            
    def action_delete_product(self,file_data,wrong_lines=[],heading=[]):
        delete_file_list =[['Id','Delete','Archive','Error']]
        self.import_line_ids.unlink()
        if not wrong_lines:
            wrong_lines.append(['Id','Error'])
        import_line_obj = self.env['product.import.line.spt']
        product_obj = self.env['product.product']
        if 'Id' in heading:
            if len(heading)==1:
                id_col = heading.index('Id')
                for line in file_data:
                    if any(line):
                        product_id = product_obj.with_context(pending_price=True).search(['|',('active','=',False),('active','=',True),('default_code','=',line[0].strip())],limit=1)
                        if not product_id:
                            delete_file_list.append([line[id_col],0,0,'Product not found.'])
                            wrong_lines.append([line[id_col],'Product not found.'])
                        else:
                            if product_id.qty_available< 1:
                                import_line_obj.create({'default_code':str(line[id_col]),'import_id': self.id})
                            else:
                                delete_file_list.append([line[id_col],0,0,"Inventory is available for this product, so you can not delete or archive."])
                                wrong_lines.append([line[id_col],"Inventory is available for this product, so you can not delete or archive."]) 
                                product_id.in_future_archive = True

                    else:
                        delete_file_list.append([line[id_col],0,0,'In line any data not found.'])
                        wrong_lines.append([line[id_col],'In line any data not found.'])

            else:
                heading.pop(heading.index('Id'))
                raise UserError(_('Please remove "%s" columns in excel.'%(','.join(heading))))        
        else:
            raise UserError(_('ID column not found in excel.'))

        if len(delete_file_list) >1 :
                out = BytesIO()
                workbook = xlsxwriter.Workbook(out) 
                worksheet = workbook.add_worksheet('wrong')
                heading_line = delete_file_list.pop(0)
                col = row =0
                for heading_data in heading_line:
                    worksheet.write(row,col,heading_data)
                    worksheet.set_column(row,col, len(heading_data)*5)

                    col+=1
                for worksheet_line in delete_file_list:
                    row +=1
                    col = 0
                    for worksheet_line_data in worksheet_line:
                        worksheet.write(row,col,worksheet_line_data)
                        col +=1

                workbook.close()
                out.seek(0)
                data = out.read()
                out.close()
                self.delete_product_name = 'delete_opration_result.xlsx'
                self.delete_product_result = base64.b64encode(data)

    def action_delete_product_process_check_error_file(self):
        self.ensure_one()
        if self.read_time:
            return {
                'name': 'Error',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'kits.read.file.error.message',
                'views': [(self.env.ref('tzc_product_import_spt.kits_read_file_error_message_form_view').id, 'form')],
                'type': 'ir.actions.act_window',
                'target':'new',
                'context':{'default_product_import_id': self.id, 'action_process':'action_delete_product_process'}
            }   
        else:
            self.action_delete_product_process()

    def action_delete_product_process(self):
        wrong_lines = [['Id','Error']]
        import_line_obj = self.env['product.import.line.spt']
        product_obj = self.env['product.product']
        sol_obj = self.env['sale.order.line']
        for record in self:
            file_product_id = {}
            delete_file_list = [['Id','Delete','Archive','Error']]
            if record.delete_product_result:
                file_datas = base64.b64decode(record.delete_product_result)
                workbook = xlrd.open_workbook(file_contents =file_datas) 
                sheet = workbook.sheet_by_index(0)
                delete_file_list = [[sheet.cell_value(r, c) for c in range(
                sheet.ncols)] for r in range(sheet.nrows)]
            import_line_ids = import_line_obj.search(['|',('active','=',False),('active','=',True),('import_id','=',record.id)])
            list(map(lambda data : file_product_id.update({data[0]:delete_file_list.index(data)}),delete_file_list))
            for import_line in import_line_ids:
                try:
                    product_id = product_obj.with_context(pending_price=True).search(['|',('active','=',False),('active','=',True),('default_code','=',import_line.default_code.strip())],limit=1)
                    if not product_id:
                        wrong_lines.append([import_line.default_code,'Product not found.'])
                    else:
                        sale_order_ids = sol_obj.search([('product_id','=',product_id.id),('order_id.state','not in',['paid','merged','open_inv','cancel'])])
                        sale_order_ids -=  sale_order_ids.filtered(lambda line: line.order_id.source_spt == 'Website' and line.order_id.state == 'draft')
                        if not sale_order_ids:
                            sale_order_ids = sol_obj.search([('product_id','=',product_id.id)])
                            delete_sale_order_ids =  sale_order_ids.filtered(lambda line: line.order_id.source_spt == 'Website' and line.order_id.state == 'draft')
                            sale_order_ids -= delete_sale_order_ids
                            if not sale_order_ids:
                                if delete_sale_order_ids:
                                    self._cr.execute('delete from sale_order_line where id in (%s)'%(','.join(list(map(lambda line: str(line),delete_sale_order_ids.ids)))))
                                self._cr.execute('delete from stock_valuation_layer where product_id = %s'%(product_id.id))
                                self._cr.execute('delete from stock_quant where product_id = %s'%(product_id.id))
                                self._cr.execute('delete from stock_move where product_id = %s'%(product_id.id))
                                self._cr.execute('delete from product_product where id = %s'%(product_id.id))
                                delete_file_list.append([import_line.default_code,1,0])
                            else:
                                print( all(list(map(lambda state: state in ['paid','merged','open_inv','draft_inv','cancel'] ,sale_order_ids.mapped('order_id.state')))))
                                if  all(list(map(lambda state: state in ['paid','merged','open_inv','draft_inv','cancel'] ,sale_order_ids.mapped('order_id.state')))):
                                    product_id.with_context(from_product_import=True).write({'active':False})
                                    # product_id.active = False
                                    new_barcode = product_id.barcode+self.column_name if self.column_name else product_id.barcode+'_Archive'
                                    new_product_seo_keyword = product_id.product_seo_keyword+self.column_name if self.column_name else  product_id.product_seo_keyword+'_Archive'
                                    new_default_code = product_id.default_code+self.column_name if self.column_name else product_id.default_code+'_Archive'
                                    du_product_id = product_obj.with_context(pending_price=True).search(['|','|',('barcode','=',new_barcode),('product_seo_keyword','=',new_product_seo_keyword),('default_code','=',new_default_code),'|',('active','=',True),('active','=',False)])
                                    if du_product_id:
                                        new_barcode = du_product_id.barcode+self.column_name if self.column_name else du_product_id.barcode+'_Archive'
                                        new_product_seo_keyword = du_product_id.product_seo_keyword+self.column_name if self.column_name else du_product_id.product_seo_keyword+'_Archive'
                                        new_default_code = du_product_id.default_code + self.column_name if self.column_name else du_product_id.default_code + '_Archive'
                                    variant_name = product_id.variant_name + self.column_name if self.column_name else product_id.variant_name+'_Archive'
                            
                                    product_id.write({'barcode':new_barcode, 'product_seo_keyword':new_product_seo_keyword,'default_code': new_default_code,'variant_name': variant_name})
                                    if import_line.default_code in file_product_id.keys():
                                        delete_file_list.pop(file_product_id[import_line.default_code])
                                        delete_file_list.insert(file_product_id[import_line.default_code],[import_line.default_code,0,1])

                                    else:
                                        delete_file_list.append([import_line.default_code,0,1])
                                else:
                                    sale_order_ids = sale_order_ids.mapped('order_id.name')
                                    delete_file_list.append([import_line.default_code,0,0,'Open order %s found so,this product can not archived.'%(','.join(sale_order_ids))])
                                    product_id.in_future_archive = True
                                    wrong_lines.append([import_line.default_code,'Open order %s found so,this product can not archived.'%(','.join(sale_order_ids))]) 

                        else:
                            print( all(list(map(lambda state: state in ['paid','merged','open_inv','draft_inv','cancel'] ,sale_order_ids.mapped('order_id.state')))))
                            if  all(list(map(lambda state: state in ['paid','merged','open_inv','draft_inv','cancel'] ,sale_order_ids.mapped('order_id.state')))):
                                product_id.with_context(from_product_import=True).write({'active':False})
                                # product_id.active = False
                                new_barcode = product_id.barcode+self.column_name if self.column_name else product_id.barcode+'_Archive'
                                new_product_seo_keyword = product_id.product_seo_keyword+self.column_name if self.column_name else  product_id.product_seo_keyword+'_Archive'
                                new_default_code = product_id.default_code+self.column_name if self.column_name else product_id.default_code+'_Archive'
                                du_product_id = product_obj.with_context(pending_price=True).search(['|','|','|',('barcode','=',new_barcode),('product_seo_keyword','=',new_product_seo_keyword),('default_code','=',new_default_code),('active','=',True),('active','=',False),('id','!=',product_id.id)])
                                if du_product_id:
                                    new_barcode = du_product_id.barcode+self.column_name if self.column_name else du_product_id.barcode+'_Archive'
                                    new_product_seo_keyword = du_product_id.product_seo_keyword+self.column_name if self.column_name else du_product_id.product_seo_keyword+'_Archive'
                                    new_default_code = du_product_id.default_code + self.column_name if self.column_name else du_product_id.default_code + '_Archive'
                                variant_name = product_id.variant_name + self.column_name if self.column_name else product_id.variant_name+'_Archive'
                                product_id.write({'barcode':new_barcode, 'product_seo_keyword':new_product_seo_keyword,'default_code': new_default_code,'variant_name': variant_name})

                                if import_line.default_code in file_product_id.keys():
                                    delete_file_list.pop(file_product_id[import_line.default_code])
                                    delete_file_list.insert(file_product_id[import_line.default_code],[import_line.default_code,0,1])

                                else:
                                    delete_file_list.append([import_line.default_code,0,1])
                            else:
                                sale_order_ids = sale_order_ids.mapped('order_id.name')
                                delete_file_list.append([import_line.default_code,0,0,'Open order %s found so,this product can not archived.'%(','.join(sale_order_ids))])
                                product_id.in_future_archive = True
                                wrong_lines.append([import_line.default_code,'Open order %s found so,this product can not archived.'%(','.join(sale_order_ids))])  

                except Exception as error:
                    wrong_lines.append([import_line.default_code,error.args[0]])
            
            record.run_time_file_name = ''
            record.run_time = False
            if len(wrong_lines) > 1:
                out = BytesIO()
                workbook = xlsxwriter.Workbook(out) 
                worksheet = workbook.add_worksheet('wrong')
                heading_line = wrong_lines.pop(0)
                col = row =0
                for heading_data in heading_line:
                    worksheet.write(row,col,heading_data)
                    worksheet.set_column(row,col, len(heading_data)*10)

                    col+=1
                for worksheet_line in wrong_lines:
                    row +=1
                    col = 0
                    for worksheet_line_data in worksheet_line:
                        worksheet.write(row,col,worksheet_line_data)
                        col +=1

                workbook.close()
                out.seek(0)
                data = out.read()
                out.close()
                record.run_time_file_name = 'Wrong_product.xlsx'
                record.run_time = base64.b64encode(data)
            record.state = 'done'

            if len(delete_file_list) >1 :
                out = BytesIO()
                workbook = xlsxwriter.Workbook(out) 
                worksheet = workbook.add_worksheet('wrong')
                heading_line = delete_file_list.pop(0)
                col = row =0
                for heading_data in heading_line:
                    worksheet.write(row,col,heading_data)
                    worksheet.set_column(row,col, len(heading_data)*5)

                    col+=1
                for worksheet_line in delete_file_list:
                    row +=1
                    col = 0
                    for worksheet_line_data in worksheet_line:
                        worksheet.write(row,col,worksheet_line_data)
                        col +=1

                workbook.close()
                out.seek(0)
                data = out.read()
                out.close()
                record.delete_product_name = 'delete_opration_result.xlsx'
                record.delete_product_result = base64.b64encode(data)

    def action_import_product_report(self):
        product_import_ids = self.env['product.import.spt'].search([])
        wizard_obj = self.env['product.info.wizard.spt']

        f_name = 'Product Import Report'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Product Import", index=0)
        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)

        table_header = 1

        sheet.cell(row=table_header, column=1).value = 'Name'
        sheet.cell(row=table_header, column=2).value = 'Attached File Name'
        sheet.cell(row=table_header, column=3).value = 'Date'
        sheet.cell(row=table_header, column=4).value = 'State'
        sheet.cell(row=table_header, column=5).value = 'Based On Categories'
        # sheet.cell(row=table_header, column=6).value = 'Object Type'
        sheet.cell(row=table_header, column=6).value = 'Operation'
        
        sheet.cell(row=table_header, column=1).font = header_font
        sheet.cell(row=table_header, column=1).border = top_bottom_border
        sheet.cell(row=table_header, column=2).font = header_font
        sheet.cell(row=table_header, column=2).border = top_bottom_border
        sheet.cell(row=table_header, column=3).font = header_font
        sheet.cell(row=table_header, column=3).border = top_bottom_border
        sheet.cell(row=table_header, column=4).font = header_font
        sheet.cell(row=table_header, column=4).border = top_bottom_border
        sheet.cell(row=table_header, column=5).font = header_font
        sheet.cell(row=table_header, column=5).border = top_bottom_border
        sheet.cell(row=table_header, column=6).font = header_font
        sheet.cell(row=table_header, column=6).border = top_bottom_border
        # sheet.cell(row=table_header, column=7).font = header_font
        # sheet.cell(row=table_header, column=7).border = top_bottom_border

        row_index=table_header+1

        query = '''SELECT COALESCE(NAME,'') AS NAME,
                    COALESCE(ATTACH_FILE_NAME,'') AS ATTACH_FILE_NAME,
                    CREATE_DATE,
                    COALESCE(STATE,'') AS STATE,
                    COALESCE(BASED_ON_CATEGORIES,'') AS BASED_ON_CATEGORY,
                    COALESCE(DATA_ON,'') AS DATA_ON
                FROM PRODUCT_IMPORT_SPT ORDER BY NAME'''
        self.env.cr.execute(query)
        report_data = self.env.cr.fetchall()

        for data in report_data:
            sheet.cell(row=row_index, column=1).value = data[0] or ''
            sheet.cell(row=row_index, column=2).value = data[1] or ''
            sheet.cell(row=row_index, column=3).value = data[2].strftime("%d-%m-%Y %H:%M:%S") or ''
            sheet.cell(row=row_index, column=4).value = dict(self._fields['state'].selection).get(data[3]) or ''
            sheet.cell(row=row_index, column=5).value = dict(self._fields['based_on_categories'].selection).get(data[4]) or ''
            # sheet.cell(row=row_index, column=6).value = "All In One" if product_import_id.all_in_one else "Product Import"
            sheet.cell(row=row_index, column=6).value = dict(self._fields['data_on'].selection).get(data[5]) or ''
            row_index += 1

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 15

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wizard_id = wizard_obj.create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=product.info.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (wizard_id.id, f_name),
            'target': 'self',
        }

    def read_opration_process(self):
        product_pro_obj = self.env['product.product']
        wrong_lines = [['ID','Error']]
        for record in self:
            fields_dict = record.get_header_data()[0]
            import_line_ids = record.import_line_ids
            for import_line_id in range(0,len(import_line_ids)):
                print(str(import_line_id)+' out of '+str(len(import_line_ids)))
                import_line_id = import_line_ids[import_line_id]
                # product_id = product_pro_obj.search([('brand','=',import_line_id.brand.id),('categ_id','=',import_line_id.categ_id.id),('model','=',import_line_id.model.id),('color_code','=',import_line_id.color.id),('eye_size','=',import_line_id.eye_size.id)])
                product_id= product_pro_obj.search(['|',('active','=',True),('active','=',False),('default_code','=',import_line_id.default_code)])
                if not product_id:
                    product_data_dict = self.convert_dict(fields_dict,import_line_id.read()[0])
                    self._cr.execute("""call create_product_template(%s,%s,%s,%s,%s,null);
                    """,[product_data_dict['default_code'],json.dumps({"name" :product_data_dict['name'].replace("'","\'")}),product_data_dict['active'],product_data_dict['categ_id'],self.env.uid])
            print('\n\n\n\n')
            self._cr.commit() 
            record.with_context(from_create=True).action_update_product(wrong_lines,product_ids = []) 
            print('\n\n\n\n')
            product_id_not_set_ids = product_pro_obj.search([('default_code','in',[False,'',' ',None])])
            if product_id_not_set_ids:
                try:
                    for product_id in product_id_not_set_ids:
                        if not product_id.sales_count:
                            product_id.with_context(from_product_import=True).write({'active':False})
                            # product_id.active = False
                            product_id.unlink()
                except Exception as error:
                    raise UserError(error.args)

    def get_header_data(self):
        self.ensure_one()
        new_header_dict = {}
        not_heading_list=[]
        file_data = []
        if self.attach_file:
            file_datas = base64.b64decode(self.attach_file)
            workbook = xlrd.open_workbook(file_contents =file_datas)
            sheet = workbook.sheet_by_index(0)
            file_data = [[sheet.cell_value(r, c) for c in range(
            sheet.ncols)] for r in range(sheet.nrows)]
            heading_list = list(fields_dict.keys())
            pricelist = self.env['product.pricelist'].search([]).mapped('name')
            heading_list.extend(pricelist)
            for col in file_data[0]:
                if col in heading_list:
                    if col in pricelist:
                        new_header_dict[col.strip()] = file_data[0].index(col)
                        continue
                    new_header_dict[fields_dict[col.strip()]]= file_data[0].index(col)
                else:
                    if col:
                        not_heading_list.append(str(col))
        if not_heading_list:
            raise UserError(_('Incorrect %s headers found,first remove headers then process will be done.'%(','.join(not_heading_list))))
        heading = file_data[0] if file_data else []
        return new_header_dict,heading,file_data[1:]
    
    def convert_dict(self,field_dict,field_dict_value_dict):
        list_to_delte_key = ['id','html_color','secondary_html_color','qty']
        new_dict = {}
        price_list = self.env['product.pricelist'].search([]).mapped('name')
        for field_name in field_dict:
            if field_name not in price_list:
                if field_name not in list_to_delte_key:
                    if isinstance(field_dict_value_dict[field_name],tuple):
                        new_dict[field_name] = field_dict_value_dict[field_name][0]
                    elif isinstance(field_dict_value_dict[field_name],list):
                        new_dict[field_name] = [(6,0,field_dict_value_dict[field_name])]
                    else:
                        new_dict[field_name] = field_dict_value_dict[field_name]
        if new_dict.get('color'):
            new_dict['color_code'] = new_dict.get('color')
            del new_dict['color']
        
        if new_dict.get('variant_name'):
            new_dict['name'] = new_dict.get('variant_name')
        if field_dict.get('brand') and field_dict.get('model') and field_dict.get('color') and field_dict.get('categ_id') and field_dict.get('eye_size'):
            new_dict['type'] = 'product'
        new_dict.update({'updated_on':fields.Datetime.now(),'updated_by':self.env.uid })
        if self.data_on == 'create':
            if 'Meta Keyword' not in field_dict and field_dict_value_dict.get('meta_keyword'):
                new_dict['meta_keyword'] = field_dict_value_dict.get('meta_keyword')
            if 'Meta Title' not in field_dict and field_dict_value_dict.get('meta_title'):
                new_dict['meta_title'] = field_dict_value_dict.get('meta_title')
            if 'Meta Description' not in field_dict and field_dict_value_dict.get('meta_description'):
                new_dict['meta_description'] = field_dict_value_dict.get('meta_description')
        return new_dict

    def file_read_opration(self):
        for record in self:
            wrong_lines =[]
            if record.attach_file:
                field_dict,heading,file_data=self.get_header_data()
                if record.data_on == 'delete':
                    record.action_delete_product(file_data,wrong_lines,heading)
                else:
                    heading.append('Error')
                    record.import_line_ids.unlink()
                    wrong_lines = [heading]
                    if record.data_on == 'create':
                        raise_data = []
                        if 'Active' not in heading:
                            raise_data.append('Please insert column Active in excel sheet')

                        if 'Color Name' not in heading:
                            raise_data.append('Please insert column Color Name in excel sheet')
                    
                        if 'Eye Size' not in heading:
                                raise_data.append('Please insert column Eye Size in excel sheet')
                        if 'HTML Color Code' not in heading:
                            raise_data.append('Please insert column HTML Color Code in excel sheet')
                        if 'Brand' not in heading:
                                raise_data.append('Please insert column Brand in excel sheet')
                        if 'Model' not in heading:
                                raise_data.append('Please insert column Model in excel sheet')
                        if 'Product categories' not in heading:
                                raise_data.append('Please insert column Product Categories in excel sheet')
                        if 'Application Type' in heading:
                            b2b_flag =False
                            b2c_flag =False
                            both_flag =False
                            application_type_error =[]
                            for line in file_data:
                                if  not line[field_dict.get('application_type')] or line[field_dict.get('application_type')] in ('0',0):
                                    b2b_flag = True
                                elif  line[field_dict.get('application_type')] in (1,'1',1.0,'2.0'):
                                    both_flag = True
                                elif  line[field_dict.get('application_type')] in (2,'2',2.0,'2.0'):
                                    b2c_flag = True
                                if b2b_flag and b2c_flag and both_flag:
                                    break
                            pricelist = []
                            if b2b_flag:
                                pricelist.extend(self.env['product.pricelist'].search([('is_for_b2c','=',False)]).mapped('name'))
                            if b2c_flag:
                                pricelist.extend(self.env['product.pricelist'].search([('is_for_b2c','=',True)]).mapped('name'))
                                if 'Is Select For Lenses' not in heading:
                                    application_type_error.append('Is Select For Lenses')
                                if 'B2C Published' not in heading:
                                    application_type_error.append('B2C Published')
                            if both_flag:
                                pricelist.extend(self.env['product.pricelist'].search([]).mapped('name'))
                                if not b2b_flag:
                                    if 'Is Select For Lenses' not in heading:
                                        application_type_error.append('Is Select For Lenses')
                                    if 'B2C Published' not in heading:
                                        application_type_error.append('B2C Published')
                            pricelist = list(set(pricelist))
                            for name in pricelist:
                                if name == 'CAD Price List':
                                    if 'List Price' in heading:
                                        continue
                                    else:
                                        application_type_error.append('List Price')

                                # if name == 'USD Price List': 
                                #     if 'List Price In USD' in heading:
                                #         continue
                                #     else:
                                #         application_type_error.append('List Price In USD')

                                if name not in heading:
                                    application_type_error.append(name)
                            if application_type_error:
                                raise_data.append((' column not found.\n'.join(application_type_error)+'  column not found.'))

                        
                        if record.based_on_categories == 'case':
                            if 'Length' not in heading:
                                raise_data.append('Please insert Length in excel sheet')
                            if 'Height' not in heading:
                                raise_data.append('Please insert Height in excel sheet')
                            if 'Width' not in heading:
                                raise_data.append('Please insert Width in excel sheet')                            
                            if 'Volume' not in heading:
                                raise_data.append('Please insert Volume in excel sheet')                            
                        if raise_data:
                            raise UserError(_('\n'.join(raise_data)))

                    record.file_read_line(file_data,heading,wrong_lines,field_dict)

            record.read_time_file_name = ''
            record.read_time = False
            record.run_time_file_name = ''
            record.run_time = False
            if len(wrong_lines) > 1:
                out = BytesIO()
                workbook = xlsxwriter.Workbook(out) 
                worksheet = workbook.add_worksheet('wrong')
                heading_line = wrong_lines.pop(0)
                col = row =0
                for heading_data in heading_line:
                    worksheet.write(row,col,heading_data)
                    worksheet.set_column(row,col, len(heading_data)*10)

                    col+=1
                for worksheet_line in wrong_lines:
                    row +=1
                    col = 0
                    for worksheet_line_data in worksheet_line:
                        worksheet.write(row,col,worksheet_line_data)
                        col +=1

                workbook.close()
                out.seek(0)
                data = out.read()
                out.close()
                record.read_time_file_name = 'Wrong_product.xlsx'
                record.read_time = base64.b64encode(data)
            record.state = 'process'               

    def file_read_line(self,file_data,fields_list,wrong_lines,fields_dict):
        color_code_obj = self.env['kits.product.color.code']
        rim_type_obj = self.env['product.rim.type.spt']
        shape_obj = self.env['product.shape.spt']
        aging_obj = self.env['product.aging.spt']
        # ecom_categ_obj = self.env['product.public.category']
        product_obj = self.env['product.product']
        import_line_obj = self.env['product.import.line.spt']
        product_categ_obj = self.env['product.category']
        brand_obj = self.env['product.brand.spt']
        model_obj = self.env['product.model.spt']
        color_obj = self.env['product.color.spt']
        size_obj = self.env['product.size.spt']
        country_of_origin_obj = self.env['res.country']
        material_obj = self.env['product.material.spt']
        bridge_size_obj = self.env['product.bridge.size.spt']
        temple_size_obj = self.env['product.temple.size.spt']
        taxes_obj= self.env['account.tax']
        record = self
        barcode_list = []
        seo_keyword_list=[]
        if fields_dict.get('barcode') and fields_dict.get('product_seo_keyword'): 
            list(map(lambda x : (barcode_list.append(x[fields_dict.get('barcode')]),seo_keyword_list.append(x[fields_dict.get('product_seo_keyword')])),file_data))
        
        elif fields_dict.get('product_seo_keyword'):
            seo_keyword_list = list(map(lambda x : x[fields_dict.get('product_seo_keyword')],file_data))
        
        elif fields_dict.get('barcode'):
            barcode_list = list(map(lambda x : x[fields_dict.get('barcode')],file_data))
        
        for line in file_data:
            if any(line):
                print(str(file_data.index(line))+':'+str(line[fields_dict.get('default_code')].strip() if fields_dict.get('default_code') or fields_dict.get('default_code') == 0 else 'Not Found'))
                brand_name = False
                model_name = False
                file_column = False
                try:
                    active = True
                    file_column = wrong_lines[0][fields_dict.get('active')] if fields_dict.get('active') else None 
                    is_active = fields_dict.get('active') if fields_dict.get('active') and int(fields_dict.get('active')) != 0 else False
                    if is_active and line[is_active] not in ('','N/A','n/a',' ','#N/A  '):
                        is_active = True if line[is_active] in (1,1.0,True,'true','t','1') else False
                    else:
                        is_active = False
                    if 'Active' in fields_list and not is_active: 
                        active = False
                    
                    file_column = wrong_lines[0][fields_dict.get('replenishable')] if fields_dict.get('replenishable') else None 
                    replenishable = fields_dict.get('replenishable') if fields_dict.get('replenishable') and int(fields_dict.get('replenishable')) != 0 else False
                    if replenishable and line[replenishable] not in ('','N/A','n/a',' ','#N/A  '):
                        replenishable = True if line[replenishable] in ('Y','y','yes','Yes','YES',1,1.0,True,'true','t','1','True','TRUE')  else False

                    file_column = wrong_lines[0][fields_dict.get('website_published')] if fields_dict.get('website_published') else None 
                    website_published = fields_dict.get('website_published') if fields_dict.get('website_published') and  int(fields_dict.get('website_published')) != 0 else False
                    if website_published and line[website_published] not in ('','N/A','n/a',' ','#N/A  '):
                        website_published = True if line[website_published] in ('Y','y','yes','Yes','YES',1,1.0,True,'true','t','1','True','TRUE') else False
                    else:
                        website_published = False

                    gender = ''
                    file_column = wrong_lines[0][fields_dict.get('gender')] if fields_dict.get('gender') else None 
                    get_gender_index = fields_dict.get('gender') if fields_dict.get('gender') and int(fields_dict.get('gender')) else False
                    if self.based_on_categories != 'case' and  int(get_gender_index) and line[get_gender_index] not in ('','N/A','n/a',' ','#N/A   '):
                        if line[get_gender_index] == 'M' or line[get_gender_index] == 'm':
                            gender = 'male' 
                        if line[get_gender_index] == 'F' or line[get_gender_index] == 'f':
                            gender = 'female'
                        if line[get_gender_index] == 'M/F' or line[get_gender_index] == 'm/f':
                            gender = 'm/f'

                    file_column = wrong_lines[0][fields_dict.get('aging')] if fields_dict.get('aging') else None 
                    aging = 0
                    get_aging_index = fields_dict.get('aging') if fields_dict.get('aging') and int(fields_dict.get('aging')) else False
                    if get_aging_index and line[get_aging_index] not in ('','N/A','n/a',' ','#N/A  '):
                        aging_id = aging_obj.search([('name','=', line[get_aging_index].strip())])
                        if not aging_id:
                            aging_id = aging_obj.create({'name': line[get_aging_index].strip() })
                        aging = aging_id.id
                    
                    file_column = wrong_lines[0][fields_dict.get('taxes_id')] if fields_dict.get('taxes_id') else None 
                    
                    
                    taxes_id = []
                    get_taxes_id_index = fields_dict.get('taxes_id') if fields_dict.get('taxes_id') and int(fields_dict.get('taxes_id')) else False
                    if get_taxes_id_index and line[get_taxes_id_index] not in ('','N/A','n/a',' ','#N/A  '):
                        get_taxes = line[get_taxes_id_index].strip().split('|')
                        for tax in get_taxes:
                            taxes_id_id = taxes_obj.search([('name','=',tax.strip())])
                            if taxes_id_id:
                                taxes_id.append(taxes_id_id.id)
                            else:
                                line.append('Tax Is Not Found.')
                                wrong_lines.append(line)
                        if len(line) == len(wrong_lines[0]):
                            continue
                    shape = []
                    # file_column = wrong_lines[0][fields_dict.get('shape_ids')] if fields_dict.get('shape_ids') else None
                    # get_shape_index = fields_dict.get('shape_ids') if fields_dict.get('shape_ids') and  int(fields_dict.get('shape_ids')) else False
                    file_column = wrong_lines[0][fields_dict.get('shape_id')] if fields_dict.get('shape_id') else None
                    get_shape_index = fields_dict.get('shape_id') if fields_dict.get('shape_id') and  int(fields_dict.get('shape_id')) else False
                    if get_shape_index and line[get_shape_index] not in ('','N/A','n/a',' ','#N/A  '):
                        shapes = line[get_shape_index].split('|')
                        for shaep_name in shapes:
                            shape_id = shape_obj.search([('name','=',shaep_name.strip())])
                            if not shape_id:
                                shape_id = shape_obj.create({'name': shaep_name.strip() })
                            shape.append(shape_id.id)

                    file_column = wrong_lines[0][fields_dict.get('rim_type')] if fields_dict.get('rim_type') else None 
                    rim_type = False
                    rim_type_id = rim_type_obj
                    get_rim_type_index = fields_dict.get('rim_type') if fields_dict.get('rim_type') and int(fields_dict.get('rim_type')) else False
                    if self.based_on_categories != 'case' and  get_rim_type_index and line[get_rim_type_index] not in ('','N/A','n/a',' ','#N/A'):
                        rim_type_id = rim_type_obj.search([('name','=',line[get_rim_type_index].strip())])
                        if not rim_type_id:
                            rim_type_id = rim_type_obj.create({'name': line[get_rim_type_index].strip() })
                        rim_type = rim_type_id.id

                    # file_column = wrong_lines[0][fields_dict.get('eto_sale_method')] if fields_dict.get('eto_sale_method') else None
                    # eto_method = ""
                    # get_eto_sale_method_index = fields_dict.get('eto_sale_method') if fields_dict.get('eto_sale_method') else False
                    # if get_eto_sale_method_index and  line[get_eto_sale_method_index] not in ('','N/A','n/a',' ','#N/A '):
                    #     eto_method = line[get_eto_sale_method_index].strip()
                    #     if eto_method not in ('fs'):
                    #         eto_method = 'regular'


                    # file_column = wrong_lines[0][fields_dict.get('kits_ecom_categ_id')] if fields_dict.get('kits_ecom_categ_id') else None 
                    # kits_ecom_categ_id = []
                    # get_kits_ecom_categ_id_index = fields_dict.get('kits_ecom_categ_id') if fields_dict.get('kits_ecom_categ_id') and int(fields_dict.get('kits_ecom_categ_id')) else False
                    # if get_kits_ecom_categ_id_index and line[get_kits_ecom_categ_id_index] not in ('','N/A','n/a',' ','#N/A'):
                    #     for public_categ in line[get_kits_ecom_categ_id_index].split('/'):
                    #         public_categ_id = ecom_categ_obj.search([('name','=',public_categ.strip())])
                    #         if not public_categ_id:
                    #             public_categ_id = ecom_categ_obj.create({'name': public_categ.strip()})
                    #         kits_ecom_categ_id.append(public_categ_id.id)
                    
                    file_column = wrong_lines[0][fields_dict.get('categ_id')] if fields_dict.get('categ_id') else None 
                    categ_id = 0
                    get_categ_id_index  = fields_dict.get('categ_id') if fields_dict.get('categ_id') and int(fields_dict.get('categ_id')) else False
                    if get_categ_id_index and line[get_categ_id_index] not in ('','N/A','n/a',' ','#N/A'):
                        category_id = product_categ_obj.search([('name','=',line[get_categ_id_index].strip())])
                        if self.based_on_categories == 'case':
                                category_id = self.env.ref('tzc_sales_customization_spt.tzc_case_product_category')
                        if not category_id:
                            category_id = product_categ_obj.create({'name':line[get_categ_id_index].strip()})
                        categ_id = category_id.id

                    # file_column = wrong_lines[0][fields_dict.get('lst_price_usd')] if fields_dict.get('lst_price_usd') else None                         
                    # lst_price_usd = 0.0
                    # if line[fields_dict.get('lst_price_usd')] != str if fields_dict.get('lst_price_usd') else False:
                    #     if line[fields_dict.get('lst_price_usd')] not in ('','N/A','n/a',' ','#N/A'):
                    #         lst_price_usd = float(line[fields_dict.get('lst_price_usd')]) if fields_dict.get('lst_price_usd') else '0.0'

                    file_column = wrong_lines[0][fields_dict.get('list_price')] if fields_dict.get('list_price') else None                         
                    list_price = 0.0
                    if line[fields_dict.get('list_price')] != str if fields_dict.get('list_price') else False:
                        if line[fields_dict.get('list_price')] not in ('','N/A','n/a',' ','#N/A'):
                            list_price = float(line[fields_dict.get('list_price')]) if fields_dict.get('list_price') else '0.0'

                    file_column = wrong_lines[0][fields_dict.get('standard_price')] if fields_dict.get('standard_price') else None 
                    standard_price = 0.0
                    if line[fields_dict.get('standard_price')] != str if fields_dict.get('standard_price') else False:
                        if line[fields_dict.get('standard_price')] not in ('','N/A','n/a',' ','#N/A'):
                            standard_price = float(line[fields_dict.get('standard_price')]) if fields_dict.get('standard_price') else '0.0'

                    file_column = wrong_lines[0][fields_dict.get('price_msrp')] if fields_dict.get('price_msrp') else None 
                    price_msrp = 0.0
                    if line[fields_dict.get('price_msrp')] != str if fields_dict.get('price_msrp') else False:
                        if line[fields_dict.get('price_msrp')] not in ('','N/A','n/a',' ','#N/A'):
                            price_msrp = float(line[fields_dict.get('price_msrp')]) if fields_dict.get('price_msrp') else '0.0'
                    
                    file_column = wrong_lines[0][fields_dict.get('price_wholesale')] if fields_dict.get('price_wholesale') else None                         
                    price_wholesale = 0.0
                    if line[fields_dict.get('price_wholesale')] != str if fields_dict.get('price_wholesale') else False:
                        if line[fields_dict.get('price_wholesale')] not in ('','N/A','n/a',' ','#N/A   '):
                            price_wholesale = float(line[fields_dict.get('price_wholesale')]) if fields_dict.get('price_wholesale') else '0.0'
                    
                    # file_column = wrong_lines[0][fields_dict.get('price_wholesale_usd')] if fields_dict.get('price_wholesale_usd') else None                         
                    # price_wholesale_in_usd = 0.0
                    # if line[fields_dict.get('price_wholesale_usd')] != str if fields_dict.get('price_wholesale_usd') else False:
                    #     if line[fields_dict.get('price_wholesale_usd')] not in ('','N/A','n/a',' ','#N/A   '):
                    #         price_wholesale_in_usd = float(line[fields_dict.get('price_wholesale_usd')]) if fields_dict.get('price_wholesale_usd') else '0.0'
                    
                    file_column = wrong_lines[0][fields_dict.get('brand')] if fields_dict.get('brand') else None 
                    brand = 0
                    brand_id = brand_obj
                    get_brand_index = fields_dict.get('brand') if fields_dict.get('brand') and int(fields_dict.get('brand')) else False
                    if get_brand_index and line[get_brand_index] not in ('','N/A','n/a',' ','#N/A  '):
                        brand_id = brand_obj.search([('name','=',str(int(line[get_brand_index])).strip() if isinstance( line[get_brand_index],float) else  line[get_brand_index].strip())])
                        if not brand_id:
                            brand_id = brand_obj.create({'name': str(int(line[get_brand_index])).strip() if isinstance( line[get_brand_index],float) else  line[get_brand_index].strip()})
                        brand_name = brand_id.name
                        brand= brand_id.id

                    file_column = wrong_lines[0][fields_dict.get('model')] if fields_dict.get('model') else None                         
                    model = 0
                    model_id = model_obj
                    get_model_index = fields_dict.get('model') if fields_dict.get('model') and int(fields_dict.get('model')) else False
                    if get_model_index and line[get_model_index] not in ('','N/A','n/a',' ','#N/A  '):
                        model_id = model_obj.search([('brand_id','=',brand_id.id),('name','=',str(int(line[get_model_index])).strip() if isinstance( line[get_model_index],float) else  line[get_model_index].strip())])
                        if not model_id:
                            model_id = model_obj.create({'name': str(int(line[get_model_index])).strip() if isinstance( line[get_model_index],float) else  line[get_model_index].strip()})
                        if brand_id:
                            model_id.brand_id = brand_id.id
                        model_name = model_id.name
                        model= model_id.id

                    file_column = wrong_lines[0][fields_dict.get('html_color')] if fields_dict.get('html_color') else None 
                    html_color = ""
                    get_html_color_index = fields_dict.get('html_color') if fields_dict.get('html_color') and int(fields_dict.get('html_color')) else False
                    if get_html_color_index and line[get_html_color_index] not in ('','N/A','n/a',' ','#N/A'):
                        html_color = line[get_html_color_index].strip()

                    file_column = wrong_lines[0][fields_dict.get('color')] if fields_dict.get('color') else None                             
                    color_code = color_code_obj 
                    if fields_dict.get('color') and line[fields_dict.get('color')] not in ['',' ','0','N/A','n/a']:
                        kits_color_code = str(int(line[fields_dict.get('color')])) if isinstance(line[fields_dict.get('color')],int) or isinstance(line[fields_dict.get('color')],float) else str(line[fields_dict.get('color')])
                        kits_color_code =kits_color_code.split('-')[0] if '-' in kits_color_code else kits_color_code
                        color_code = color_code_obj.search([('name','=',kits_color_code),('model_id','=',model_id.id)])
                        if not color_code:
                            color_code = color_code_obj.create({
                                'name':  kits_color_code,
                                'color': html_color
                            })
                        if model_id:
                            color_code.model_id = model_id.id
                        color_code = color_code.id
                    if not color_code:
                        color_code = 0

                    material = []
                    # file_column = wrong_lines[0][fields_dict.get('material_ids')] if fields_dict.get('material_ids') else None 
                    # get_material_index = fields_dict.get('material_ids') if fields_dict.get('material_ids') and int(fields_dict.get('material_ids')) else False
                    file_column = wrong_lines[0][fields_dict.get('material_id')] if fields_dict.get('material_id') else None 
                    get_material_index = fields_dict.get('material_id') if fields_dict.get('material_id') and int(fields_dict.get('material_id')) else False
                    if get_material_index and line[get_material_index] not in ('','N/A','n/a',' ','#N/A'):
                        materials = line[get_material_index].split('|')
                        for material_name in materials:
                            material_id = material_obj.search([('name','=',material_name.strip())])
                            if not material_id:
                                material_id = material_obj.create({'name': material_name.strip()})
                            material.append(material_id.id)

                    # file_column = wrong_lines[0][fields_dict.get('price_msrp_usd')] if fields_dict.get('price_msrp_usd') else None                         
                    # price_msrp_in_usd = 0
                    # if line[fields_dict.get('price_msrp_usd')] != str if fields_dict.get('price_msrp_usd') else False:
                    #     if line[fields_dict.get('price_msrp_usd')] not in ('','N/A','n/a',' ','#N/A   '):
                    #         price_msrp_in_usd = float(line[fields_dict.get('price_msrp_usd')]) if fields_dict.get('price_msrp_usd') else '0.0'
                    
                    file_column = wrong_lines[0][fields_dict.get('country_of_origin')] if fields_dict.get('country_of_origin') else None 
                    country_of_origin = 0
                    get_country_of_origin_index = fields_dict.get('country_of_origin') if fields_dict.get('country_of_origin') and int(fields_dict.get('country_of_origin')) else False
                    if get_country_of_origin_index and line[get_country_of_origin_index] not in ('','N/A','n/a',' ','#N/A  '):
                        country_of_origin_id = country_of_origin_obj.search([('name','=',line[get_country_of_origin_index].strip())])
                        if country_of_origin_id:
                            country_of_origin= country_of_origin_id.id
                    
                    file_column = wrong_lines[0][fields_dict.get('bridge_size')] if fields_dict.get('bridge_size') else None                         
                    bridge_size = 0
                    get_bridge_size_index = fields_dict.get('bridge_size') if fields_dict.get('bridge_size') and int(fields_dict.get('bridge_size')) else False
                    if self.based_on_categories != 'case' and  get_bridge_size_index and line[get_bridge_size_index] not in ('','N/A','n/a',' ','#N/A  '):
                        bridge_size_id = bridge_size_obj.search([('bridgesize_id','=',color_code),('name','=',str(int(line[get_bridge_size_index])).strip() if isinstance(line[get_bridge_size_index],float) else line[get_bridge_size_index].strip())])
                        if not bridge_size_id:
                            bridge_size_id = bridge_size_obj.create({'name': str(int(line[get_bridge_size_index])).strip() if isinstance(line[get_bridge_size_index],float) else line[get_bridge_size_index].strip()})
                        if color_code and bridge_size_id.bridgesize_id.id != color_code:
                            bridge_size_id.bridgesize_id = color_code
                        bridge_size= bridge_size_id.id
                    

                    file_column = wrong_lines[0][fields_dict.get('temple_size')] if fields_dict.get('temple_size') else None 
                    temple_size = 0
                    get_temple_size_index = fields_dict.get('temple_size') if fields_dict.get('temple_size') and int(fields_dict.get('temple_size')) else False
                    if self.based_on_categories != 'case' and  get_temple_size_index and line[get_temple_size_index] not in ('','N/A','n/a',' ','#N/A  '):
                        temple_size_id = temple_size_obj.search([('templesize_id','=',color_code),('name','=',str(int(line[get_temple_size_index])).strip() if isinstance(line[get_temple_size_index],float) else line[get_temple_size_index].strip())])
                        if not temple_size_id:
                            temple_size_id = temple_size_obj.create({'name': str(int(line[get_temple_size_index])).strip() if isinstance(line[get_temple_size_index],float) else line[get_temple_size_index].strip()})
                        if temple_size_id and temple_size_id.templesize_id.id != color_code:
                            temple_size_id.templesize_id=color_code
                        temple_size= temple_size_id.id
                    
                    file_column = wrong_lines[0][fields_dict.get('geo_restriction')] if fields_dict.get('geo_restriction') else None                         
                    geo_restriction = []
                    file_column = wrong_lines[0][fields_dict.get('geo_restriction')] if fields_dict.get('geo_restriction') else None 
                    get_geo_restriction_index = fields_dict.get('geo_restriction') if fields_dict.get('geo_restriction') and int(fields_dict.get('geo_restriction')) else False
                    if get_geo_restriction_index and line[get_geo_restriction_index] not in ('','N/A','n/a',' ','#N/A  '):
                        for geo_restriction_spt in line[get_geo_restriction_index].strip().split('|'):
                            geo_restriction_id = country_of_origin_obj.search([('name','=',geo_restriction_spt.strip())])
                            if geo_restriction_id:
                                geo_restriction.append(geo_restriction_id.id)

                    file_column = wrong_lines[0][fields_dict.get('lense_color_name')] if fields_dict.get('lense_color_name') else None 
                    lense_color_name = 0
                    get_lense_color_name_index = fields_dict.get('lense_color_name') if fields_dict.get('lense_color_name') and int(fields_dict.get('lense_color_name')) else False
                    if get_lense_color_name_index and line[get_lense_color_name_index] not in ('','N/A','n/a',' ','#N/A'):
                        lense_color_name_id = color_obj.search([('name','=',line[get_lense_color_name_index].strip())])
                        if not lense_color_name_id:
                            lense_color_name_id = color_obj.create({'name': line[get_lense_color_name_index].strip(),'color': html_color.strip()})
                        lense_color_name= lense_color_name_id.id
                    

                    file_column = wrong_lines[0][fields_dict.get('product_color_name')] if fields_dict.get('product_color_name') else None                         
                    product_color_name = 0
                    get_product_color_name_index = fields_dict.get('product_color_name') if fields_dict.get('product_color_name') and int(fields_dict.get('product_color_name')) else False
                    if get_product_color_name_index and line[get_product_color_name_index] not in ('','N/A','n/a',' ','#N/A'):
                        product_color_name_id = color_obj.search([('name','=',line[get_product_color_name_index].strip())])
                        if not product_color_name_id:
                            product_color_name_id = color_obj.create({'name': line[get_product_color_name_index].strip(),'color': html_color.strip()})
                        product_color_name= product_color_name_id.id

                    color_name = ""
                    file_column = wrong_lines[0][fields_dict.get('product_color_name')] if fields_dict.get('product_color_name') else None 
                    get_color_name_index = fields_dict.get('product_color_name') if fields_dict.get('product_color_name') and int(fields_dict.get('product_color_name')) else False
                    if get_color_name_index and line[get_color_name_index] not in ('','N/A','n/a',' ','#N/A'):
                        color_name = line[get_color_name_index].strip()
                    

                    file_column = wrong_lines[0][fields_dict.get('product_seo_keyword')] if fields_dict.get('product_seo_keyword') else None                          
                    product_seo_keyword = None
                    get_product_seo_keyword_index =  fields_dict.get('product_seo_keyword') if fields_dict.get('product_seo_keyword') and int( fields_dict.get('product_seo_keyword')) else False
                    if get_product_seo_keyword_index and line[get_product_seo_keyword_index] not in ('','N/A','n/a',' ','#N/A  '):
                        product_seo_keyword = line[get_product_seo_keyword_index].strip()
                    
                    file_column = wrong_lines[0][fields_dict.get('custom_message')] if fields_dict.get('custom_message') else None                     
                    custom_message = ''
                    get_custom_message_index =  fields_dict.get('custom_message') if fields_dict.get('custom_message') and int( fields_dict.get('custom_message')) else False
                    if get_custom_message_index and line[get_custom_message_index] not in ('','N/A','n/a',' ','#N/A'):
                        custom_message = line[get_custom_message_index]

                    file_column = wrong_lines[0][fields_dict.get('secondary_html_color')] if fields_dict.get('secondary_html_color') else None 
                    secondary_color_code = ''
                    get_secondary_color_code_index = fields_dict.get('secondary_html_color') if fields_dict.get('secondary_html_color') and int(fields_dict.get('secondary_html_color')) else False
                    if get_secondary_color_code_index and isinstance(line[get_secondary_color_code_index],str) and line[get_secondary_color_code_index] not in ('','N/A','n/a',' ','#N/A'):
                        secondary_color_code = str(line[get_secondary_color_code_index]).strip()

                    file_column = wrong_lines[0][fields_dict.get('secondary_color_name')] if fields_dict.get('secondary_color_name') else None 
                    secondary_color_name = 0
                    get_secondary_color_name_index = fields_dict.get('secondary_color_name') if fields_dict.get('secondary_color_name') and int(fields_dict.get('secondary_color_name')) else False
                    if get_secondary_color_name_index and line[get_secondary_color_name_index] not in ('','N/A','n/a',' ','#N/A'):
                        secondary_color_name_id = color_obj.search([('name','=',line[get_secondary_color_name_index].strip())])
                        if not secondary_color_name_id:
                            secondary_color_name_id = color_obj.create({'name': line[get_secondary_color_name_index].strip(),'color': secondary_color_code.strip()})
                        secondary_color_name= secondary_color_name_id.id

                    file_column = wrong_lines[0][fields_dict.get('eye_size')] if fields_dict.get('eye_size') else None                         
                    eye_size = False
                    get_seye_size_index = fields_dict.get('eye_size') if fields_dict.get('eye_size') and int(fields_dict.get('eye_size')) else False
                    if self.based_on_categories != 'case' and  get_seye_size_index and line[get_seye_size_index] not in ('','N/A','n/a',' ','#N/A'):
                        eye_size = size_obj.search([('eyesize_id','=',color_code),('name','=', str(int(line[get_seye_size_index])).strip() if isinstance( line[get_seye_size_index],float) else  line[get_seye_size_index].strip())])
                        if not eye_size:
                            eye_size = size_obj.create({'name':  str(int(line[get_seye_size_index])).strip() if isinstance( line[get_seye_size_index],float) else  line[get_seye_size_index].strip()})
                        if color_code and eye_size.eyesize_id.id != color_code:
                            eye_size.eyesize_id = color_code
                        eye_size= eye_size.id
                    
                    file_column = wrong_lines[0][fields_dict.get('barcode')] if fields_dict.get('barcode') else None                         
                    barcode = ''
                    get_barcode_index = fields_dict.get('barcode') if fields_dict.get('barcode') else False
                    if get_barcode_index and line[get_barcode_index] not in ('','N/A','n/a',' ','#N/A  '):
                        if isinstance(line[get_barcode_index],float):
                            barcode = str(int(line[get_barcode_index])).strip()
                        elif isinstance(line[get_barcode_index],int):
                            barcode = str(line[get_barcode_index]).strip()
                        elif isinstance(line[get_barcode_index],str):
                            barcode = line[get_barcode_index].strip()
                        else:
                            barcode =  line[fields_dict.get('default_code')].strip()

                    sale_type = None
                    file_column = wrong_lines[0][fields_dict.get('sale_type')] if fields_dict.get('sale_type') else None 
                    sale_type_spt = fields_dict.get('sale_type') if fields_dict.get('sale_type') and int(fields_dict.get('sale_type')) != 0 else False
                    if sale_type_spt and line[sale_type_spt] not in ('','N/A','n/a',' ','#N/A  ',0):
                        if line[sale_type_spt] in ['on sale','On sale','On Sale','on Sale','onsale']:
                            sale_type = 'on_sale'
                        if line[sale_type_spt] in ['Clearance','clearance']:
                                sale_type = 'clearance'
                                
                    on_sale_usd = None
                    file_column = wrong_lines[0][fields_dict.get('on_sale_usd')] if fields_dict.get('on_sale_usd') else None 
                    if line[fields_dict.get('on_sale_usd')] != str if fields_dict.get('on_sale_usd') else False:
                        if line[fields_dict.get('on_sale_usd')] not in ('','N/A','n/a',' ','#N/A',0,0.0):
                                on_sale_usd = float(line[fields_dict.get('on_sale_usd')]) if fields_dict.get('on_sale_usd') else None
                            
                    on_sale_usd_in_percentage = None
                    file_column = wrong_lines[0][fields_dict.get('on_sale_usd_in_percentage')] if fields_dict.get('on_sale_usd_in_percentage') else None 
                    if on_sale_usd  and line[fields_dict.get('on_sale_usd_in_percentage')] != str if fields_dict.get('on_sale_usd_in_percentage') else False:
                        if line[fields_dict.get('on_sale_usd_in_percentage')] not in ('','N/A','n/a',' ','#N/A',0,0.0):
                            if line[fields_dict.get('on_sale_usd_in_percentage')] in ['percentage','Percentage']:
                                on_sale_usd_in_percentage = line[fields_dict.get('on_sale_usd')] if fields_dict.get('on_sale_usd') else None
                                on_sale_usd = None

                    temporary_out_of_stock = None
                    file_column = wrong_lines[0][fields_dict.get('temporary_out_of_stock')] if fields_dict.get('temporary_out_of_stock') else None 
                    is_temporary_out_of_stock = fields_dict.get('temporary_out_of_stock') if fields_dict.get('temporary_out_of_stock') and int(fields_dict.get('temporary_out_of_stock')) != 0 else False
                    if is_temporary_out_of_stock and line[is_temporary_out_of_stock]  not in ('','N/A','n/a',' ','#N/A  '):
                        if line[is_temporary_out_of_stock] in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t']:
                            temporary_out_of_stock = True
                        else:
                            temporary_out_of_stock = False
                    
                    new_arrivals = None
                    file_column = wrong_lines[0][fields_dict.get('new_arrivals')] if fields_dict.get('new_arrivals') else None 
                    is_new_arrivals = fields_dict.get('new_arrivals') if fields_dict.get('new_arrivals') and int(fields_dict.get('new_arrivals')) != 0 else False
                    if is_new_arrivals and line[is_new_arrivals]  not in ('','N/A','n/a',' ','#N/A  '):
                        if line[is_new_arrivals] in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t']:
                            new_arrivals = True
                        else:
                            new_arrivals = False
                    
                    length = 0.0
                    file_column = wrong_lines[0][fields_dict.get('length')] if fields_dict.get('length') else None 
                    if line[fields_dict.get('length')] != str if fields_dict.get('length') else False:
                        if line[fields_dict.get('length')] not in ('','N/A','n/a',' ','#N/A   '):
                            length = float(line[fields_dict.get('length')]) if fields_dict.get('length') else 0.0
                    
                    width = 0.0
                    file_column = wrong_lines[0][fields_dict.get('width')] if fields_dict.get('width') else None 
                    if line[fields_dict.get('width')] != str if fields_dict.get('width') else False:
                        if line[fields_dict.get('width')] not in ('','N/A','n/a',' ','#N/A   '):
                            width = float(line[fields_dict.get('width')]) if fields_dict.get('width') else 0.0
                    is_new_price = False
                    file_column = wrong_lines[0][fields_dict.get('is_new_price')] if fields_dict.get('is_new_price') else None 
                    new_price = fields_dict.get('is_new_price') if fields_dict.get('is_new_price') else False
                    if new_price and line[new_price] not in (0.0,0,'','N/A','n/a',' ','#N/A  ','False','0'):
                        is_new_price = True if line[new_price] in (1,1.0,True) else False

                    height = 0.0
                    file_column = wrong_lines[0][fields_dict.get('height')] if fields_dict.get('height') else None 
                    if line[fields_dict.get('height')] != str if fields_dict.get('height') else False:
                        if line[fields_dict.get('height')] not in ('','N/A','n/a',' ','#N/A   '):
                            height = float(line[fields_dict.get('height')]) if fields_dict.get('height') else 0.0
        
                    qty = 0.0
                    file_column = wrong_lines[0][fields_dict.get('qty')] if fields_dict.get('qty') else None 
                    if fields_dict.get('qty'):
                        if line[fields_dict.get('qty')] not in ('','N/A','n/a',' ','#N/A   '):
                            qty = float(line[fields_dict.get('qty')]) if fields_dict.get('qty') else 0.0

                    product_brand_commission = 0.0
                    file_column = wrong_lines[0][fields_dict.get('product_brand_commission')] if fields_dict.get('product_brand_commission') else None 
                    if line[fields_dict.get('product_brand_commission')] != str if fields_dict.get('product_brand_commission') else False:
                        if line[fields_dict.get('product_brand_commission')] not in ('','N/A','n/a',' ','#N/A   '):
                            product_brand_commission = float(line[fields_dict.get('product_brand_commission')]) if fields_dict.get('product_brand_commission') else 0.0

                    case_type = ''
                    file_column = wrong_lines[0][fields_dict.get('case_type')] if fields_dict.get('case_type') else None 
                    get_case_type_index = fields_dict.get('case_type') if fields_dict.get('case_type') and int(fields_dict.get('case_type')) else False
                    if int(get_case_type_index) and line[get_case_type_index] not in ('','N/A','n/a',' ','#N/A   '):
                        if line[get_case_type_index].lower().strip() == 'original' :
                            case_type = 'original' 
                        if line[get_case_type_index].lower().strip() == 'generic' :
                            case_type = 'generic'
                        
                    on_consignment = None
                    file_column = wrong_lines[0][fields_dict.get('on_consignment')] if fields_dict.get('on_consignment') else None 
                    is_on_consignment = fields_dict.get('on_consignment') if fields_dict.get('on_consignment') and int(fields_dict.get('on_consignment')) != 0 else False
                    if is_on_consignment and line[is_on_consignment]  not in ('','N/A','n/a',' ','#N/A  '):
                        if line[is_on_consignment] in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','True','TRUE']:
                            on_consignment = True
                        else:
                            on_consignment = False

                    minimum_qty = 0.0
                    file_column = wrong_lines[0][fields_dict.get('minimum_qty')] if fields_dict.get('minimum_qty') else None 
                    if line[fields_dict.get('minimum_qty')] != str if fields_dict.get('minimum_qty') else False:
                        if line[fields_dict.get('minimum_qty')] not in ('','N/A','n/a',' ','#N/A   '):
                            minimum_qty = float(line[fields_dict.get('minimum_qty')]) if fields_dict.get('minimum_qty') else 0.0
        
                    flex_hinges = "no"
                    file_column = wrong_lines[0][fields_dict.get('flex_hinges')] if fields_dict.get('flex_hinges') else None 
                    get_flex_hinges_index = fields_dict.get('flex_hinges') if fields_dict.get('flex_hinges') and int(fields_dict.get('flex_hinges')) else False
                    if int(get_flex_hinges_index) and line[get_flex_hinges_index] not in ('','N/A','n/a',' ','#N/A   '):
                        if line[get_flex_hinges_index].strip() in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','TRUE']:
                            flex_hinges = 'yes'
                        if line[get_flex_hinges_index].strip() in ['n','no','NO','No',0.0,0,False]:
                            flex_hinges = 'no'


                    file_column = wrong_lines[0][fields_dict.get('is_select_for_lenses')] if fields_dict.get('is_select_for_lenses') else None 
                    is_select_for_lenses = None
                    get_is_select_for_lenses_index = fields_dict.get('is_select_for_lenses') if fields_dict.get('is_select_for_lenses') and int(fields_dict.get('is_select_for_lenses')) else False
                    if get_is_select_for_lenses_index and line[get_is_select_for_lenses_index] not in ('','N/A','n/a',' ','#N/A  '):
                        if str(line[get_is_select_for_lenses_index]).strip() in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','TRUE','1.0','1']:
                            is_select_for_lenses = True

                        if str(line[get_is_select_for_lenses_index]).strip() in ['0.0','0','n','no','NO','No',0.0,0,False]:
                            is_select_for_lenses = False
                    
                    file_column = wrong_lines[0][fields_dict.get('is_forcefully_unpublished')] if fields_dict.get('is_forcefully_unpublished') else None 
                    is_forcefully_unpublished = None
                    get_is_forcefully_unpublished_index = fields_dict.get('is_forcefully_unpublished') if fields_dict.get('is_forcefully_unpublished') and int(fields_dict.get('is_forcefully_unpublished')) else False
                    if get_is_forcefully_unpublished_index and line[get_is_forcefully_unpublished_index] not in ('','N/A','n/a',' ','#N/A  '):
                        if str(line[get_is_forcefully_unpublished_index]).strip() in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','TRUE','1.0','1']:
                            is_forcefully_unpublished = True
                        if str(line[get_is_forcefully_unpublished_index]).strip() in ['0.0','0','n','no','NO','No',0.0,0,False]:
                            is_forcefully_unpublished = False
                    

                    file_column = wrong_lines[0][fields_dict.get('is_b2c_published')] if fields_dict.get('is_b2c_published') else None 
                    is_b2c_published = None
                    get_is_b2c_published_index = fields_dict.get('is_b2c_published') if fields_dict.get('is_b2c_published') and int(fields_dict.get('is_b2c_published')) else False
                    if get_is_b2c_published_index and line[get_is_b2c_published_index] not in ('','N/A','n/a',' ','#N/A  '):
                        if str(line[get_is_b2c_published_index]).strip() in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','TRUE','1.0','1']:
                            is_b2c_published = True

                        if str(line[get_is_b2c_published_index]).strip() in ['0.0','0','n','no','NO','No',0.0,0,False]:
                            is_b2c_published = False


                    file_column = wrong_lines[0][fields_dict.get('is_3d_model')] if fields_dict.get('is_3d_model') else None 
                    is_3d_model = None
                    get_is_3d_model_index = fields_dict.get('is_3d_model') if fields_dict.get('is_3d_model') and int(fields_dict.get('is_3d_model')) else False
                    if get_is_3d_model_index and line[get_is_3d_model_index] not in ('','N/A','n/a',' ','#N/A  '):
                        if str(line[get_is_3d_model_index]).strip() in ['Y','y','yes','Yes','YES','1',1,1.0,True,'true','t','TRUE','1.0','1']:
                            is_3d_model = True

                        if str(line[get_is_3d_model_index]).strip() in ['0.0','0','n','no','NO','No',0.0,0,False]:
                            is_3d_model = False

                    # Application type is mandatory: 0 or blank means B2B 1 Both 2. Only B2C

                    application_type = None
                    file_column = wrong_lines[0][fields_dict.get('application_type')] if fields_dict.get('application_type') else None 
                    if line[fields_dict.get('application_type')] != str if fields_dict.get('application_type') else False:
                        if (isinstance(line[fields_dict.get('application_type')], float) and int(line[fields_dict.get('application_type')]) == 0) or line[fields_dict.get('application_type')] == '':
                            application_type = '0'
                        elif isinstance(line[fields_dict.get('application_type')], float) and int(line[fields_dict.get('application_type')]) == 2 or line[fields_dict.get('application_type')] == '2':
                            application_type = '2'
                        elif isinstance(line[fields_dict.get('application_type')], float) and int(line[fields_dict.get('application_type')]) == 1 or line[fields_dict.get('application_type')] == '1':
                            application_type = '1'
                        else:
                            line.append('Application type not found.')
                            wrong_lines.append(line)

                        if len(line) >= len(wrong_lines[0]):
                            wrong_lines.append(line)
                            continue


                    meta_keyword = None
                    file_column = wrong_lines[0][fields_dict.get('meta_keyword')] if fields_dict.get('meta_keyword') else None 
                    get_meta_keyword_index = fields_dict.get('meta_keyword') if fields_dict.get('meta_keyword') and int(fields_dict.get('meta_keyword')) else False
                    if get_meta_keyword_index and line[get_meta_keyword_index] not in ('','N/A','n/a',' ','#N/A'):
                        meta_keyword = str(line[get_meta_keyword_index]).strip()


                    meta_title = None
                    file_column = wrong_lines[0][fields_dict.get('meta_title')] if fields_dict.get('meta_title') else None 
                    get_meta_title_index = fields_dict.get('meta_title') if fields_dict.get('meta_title') and int(fields_dict.get('meta_title')) else False
                    if get_meta_title_index and line[get_meta_title_index] not in ('','N/A','n/a',' ','#N/A'):
                        meta_title = str(line[get_meta_title_index]).strip()


                    meta_description = None
                    file_column = wrong_lines[0][fields_dict.get('meta_description')] if fields_dict.get('meta_description') else None 
                    get_meta_description_index = fields_dict.get('meta_description') if fields_dict.get('meta_description') and int(fields_dict.get('meta_description')) else False
                    if get_meta_description_index and line[get_meta_description_index] not in ('','N/A','n/a',' ','#N/A'):
                        meta_description = str(line[get_meta_description_index]).strip()    

                    file_column = 'product creating'                                    
                    product_values_dict = {
                            'meta_description':meta_description,
                            'meta_keyword':meta_keyword,
                            'meta_title':meta_title,
                            'application_type':application_type,
                            'is_3d_model': is_3d_model,
                            'is_b2c_published': is_b2c_published,
                            'is_forcefully_unpublished': is_forcefully_unpublished,
                            'is_select_for_lenses':is_select_for_lenses,
                            'flex_hinges':flex_hinges,
                            'on_consignment': on_consignment,
                            'minimum_qty': minimum_qty,
                            'old_id' : None,
                            'case_image_url':line[fields_dict.get('case_image_url')].replace(' ','').strip() if fields_dict.get('case_image_url') and line[fields_dict.get('case_image_url')] else None,
                            'case_type' : case_type,
                            'default_code' : line[fields_dict.get('default_code')].strip(),
                            'name':str(brand_name+' '+model_name ) if fields_dict.get('variant_name') and brand_name and model_name else None,
                            'variant_name':str(line[fields_dict.get('variant_name')]).strip() if fields_dict.get('variant_name') else None,
                            'active': is_active if is_active else False,
                            'list_price':   list_price if list_price else None,
                            'price_msrp': price_msrp if price_msrp else None,
                            'price_wholesale': price_wholesale if price_wholesale else None,
                            'detailed_type': 'product' if fields_dict.get('type') else None,
                            'brand': brand if brand else  None,
                            'model': model if model else  None,
                            'color': color_code,
                            'categ_id':categ_id if categ_id else None,
                            'image_secondary_url':line[fields_dict.get('image_url')].replace(' ','').strip() if fields_dict.get('image_url') and line[fields_dict.get('image_url')] else None,
                            'image_url':line[fields_dict.get('image_secondary_url')].replace(' ','').strip() if fields_dict.get('image_secondary_url') and line[fields_dict.get('image_secondary_url')] else None,
                            'hs_code' : line[fields_dict.get('hs_code')] if fields_dict.get('hs_code') and line[fields_dict.get('hs_code')]  else None,
                            'color_name' : color_name if color_name else  None,
                            'material_id' : material[-1] if material else  None,
                            # 'material_ids' : [(6,0,material)] if material else  None,
                            'barcode' : barcode if barcode else None,
                            'volume' : line[fields_dict.get('volume')] if fields_dict.get('volume') and line[fields_dict.get('volume')] and line[fields_dict.get('volume')] not in ['',' ','N/A','n/a','0',0] else None,
                            'weight' : round((line[fields_dict.get('weight')]*0.001),2) if fields_dict.get('weight') and line[fields_dict.get('weight')] and line[fields_dict.get('weight')] not in ['',' ','N/A','n/a','0',0] else None,
                            'gender' : gender if gender else   None,
                            'eye_size' : eye_size if eye_size else None,
                            'temple_size' : temple_size if temple_size else  None,
                            'geo_restriction' : [(6,0,geo_restriction)] if geo_restriction else  None,
                            'lense_color_name' : lense_color_name if lense_color_name else  None,
                            'aging' : aging if aging else  None,
                            'shape_id' : shape[-1] if shape else  None,
                            # 'shape_ids' : [(6,0,shape)] if shape else  None,
                            'taxes_id' : [(6,0,taxes_id)] if taxes_id else  None,
                            'rim_type' : rim_type if rim_type else None,
                            'country_of_origin' : country_of_origin if country_of_origin else None,
                            'standard_price': standard_price if standard_price else None,
                            # 'website_published' : website_published if website_published else None,
                            'is_published_spt' : website_published if website_published else None,
                            'custom_message' : custom_message if custom_message else  None ,
                            # 'kits_ecom_categ_id' : kits_ecom_categ_id[0] if kits_ecom_categ_id else None,
                            'qty' : qty if qty else None,
                            'website_description' : line[fields_dict.get('website_description')] if fields_dict.get('website_description') and line[fields_dict.get('website_description')] else None,
                            'import_id' : record.id,
                            'bridge_size' : bridge_size if bridge_size else  None,
                            'replenishable' : replenishable if replenishable else  None,
                            'html_color' : html_color if html_color else  '',
                            'product_seo_keyword' : product_seo_keyword if product_seo_keyword else  None,
                            'product_color_name' : product_color_name if product_color_name else  None,
                            'secondary_color_name' : secondary_color_name if secondary_color_name else  None,
                            'secondary_html_color' : secondary_color_code if secondary_color_code else  None,
                            # 'price_wholesale_usd' : price_wholesale_in_usd if price_wholesale_in_usd else None,
                            # 'price_msrp_usd' : price_msrp_in_usd if price_msrp_in_usd else None,
                            # 'lst_price_usd' : lst_price_usd if lst_price_usd else None,
                            'temporary_out_of_stock' : temporary_out_of_stock if temporary_out_of_stock else None,
                            'sale_type' : sale_type if sale_type else None,                                
                            'on_sale_usd' : on_sale_usd if on_sale_usd else None,
                            'on_sale_usd_in_percentage' : on_sale_usd_in_percentage if on_sale_usd_in_percentage else None,
                            'new_arrivals' : new_arrivals,
                            'length' : length if length else None,
                            'height' : height if height else None,
                            'is_new_price' : is_new_price if is_new_price else None,
                            'width' : width if width else None,           
                            'product_brand_commission' : product_brand_commission if product_brand_commission else None                     
                        }
                    if self.data_on == 'create':
                        if not meta_title:
                           product_values_dict['meta_title'] = product_values_dict.get('variant_name') 

                        if not meta_description:
                            product_values_dict['meta_description'] = product_values_dict.get('custom_message')

                        if not meta_keyword:
                            product_values_dict['meta_keyword'] = self.genrate_keyword(product_values_dict['meta_title'],brand_id.name or '',model_id.name or '',category_id.name or '',color_name or '',shape_id.name or '',material_id.name or '',rim_type_id.name or '')                            

                    if 'Product categories' in fields_list or 'Brand' in fields_list or 'Model' in fields_list or'Color' in fields_list or'Eye Size' in fields_list:
                        msg = ''
                        if not product_values_dict.get('brand') and 'Brand' in fields_list :
                            msg = 'Brand' if not msg else msg+',Brand'
                        if not product_values_dict.get('model') and  'Model' in fields_list :
                            msg = 'Model'if not msg else msg+',Model'
                        if not product_values_dict.get('color') and 'Color' in fields_list :
                            msg = 'Color'if not msg else msg+',Color'
                        if not product_values_dict.get('categ_id') and  'Product categories' in fields_list:
                            msg = 'Product categories'if not msg else msg+',Product categories'
                        if not product_values_dict.get('eye_size') and 'Eye Size' in fields_list and self.based_on_categories != 'case':
                            msg = 'Eye Size'if not msg else msg+',Eye Size'
                        if msg:
                            msg=msg+' column is empty so operation cannot perform on this product'
                            line.append(msg)
                            wrong_lines.append(line)
                            continue

                    barcode = False
                    products_seo_id = False
                    import_line_seo_id = False
                    product_id = product_obj.with_context(pending_price=True).sudo().search([('default_code','=', line[fields_dict.get('default_code')].strip()),'|',('active','=',True),('active','=',False)])
                    if len(product_id)>1:
                        line.append("Duplicate sku found, so can't %s a product."%(self.data_on))
                        wrong_lines.append(line)
                        return 
                    if self.data_on == 'update':
                        if product_id and not product_id.active and not product_values_dict['active']:
                            active = False
                        else:
                            active = True
                    
                    
                    if product_values_dict['barcode'] not in ('',None,False,'N/A','#N/A'):
                        barcode_spt = product_values_dict['barcode']
                        barcode = product_obj.with_context(pending_price=True).sudo().search([('barcode','=',barcode_spt),'|',('active','=',True),('active','=',False),'|',('active','=',True),('active','=',False)])
                        if not barcode:
                            barcode =  import_line_obj.search([('import_id','=',self.id),('barcode','=',barcode_spt),'|',('active','=',True),('active','=',False)])
                        if barcode.default_code == line[fields_dict.get('default_code')].strip():
                            barcode = False
                        barcode = barcode_list.count(barcode_spt) if  barcode_list.count(barcode_spt) >1 else barcode if barcode else False
                    if product_values_dict['product_seo_keyword'] not in ('',None,False,'N/A','#N/A'):
                        products_seo_id =  product_obj.with_context(pending_price=True).search([('product_seo_keyword','=',product_values_dict['product_seo_keyword']),'|',('active','=',True),('active','=',False)])
                        import_line_seo_id = import_line_obj.search([('product_seo_keyword','=',product_values_dict['product_seo_keyword']),('import_id','=',self.id),'|',('active','=',True),('active','=',False)])
                        if products_seo_id.default_code == line[fields_dict.get('default_code')].strip() or import_line_seo_id.default_code == line[fields_dict.get('default_code')].strip():
                            products_seo_id = import_line_seo_id = False
                        import_line_seo_id = seo_keyword_list.count(line[fields_dict.get('product_seo_keyword')]) if  seo_keyword_list.count(line[fields_dict.get('product_seo_keyword')]) >1 else False
                    if self.data_on in ('update')and not barcode and product_id and not import_line_seo_id and not products_seo_id:
                            import_line_id = import_line_obj.create(product_values_dict)
                    elif self.data_on in ('create') and not barcode and not product_id and not import_line_seo_id and not products_seo_id:   
                            if product_values_dict['color'] and product_values_dict['eye_size']and product_values_dict['html_color']:
                                    import_line_id = import_line_obj.create(product_values_dict)
                            elif self.based_on_categories == 'case' and product_values_dict['color'] and product_values_dict['html_color']:
                                import_line_id = import_line_obj.create(product_values_dict)

                            else:
                                if not product_values_dict['color']:
                                    if len(wrong_lines[0]) - 1 < len(line):
                                        line[-1] = line[-1]+'/In Line Color Not Found.'
                                    else:
                                        line.append('In Line Color Not Found')
                                if self.based_on_categories != 'case' and not product_values_dict['eye_size']:
                                    if len(wrong_lines[0]) - 1 < len(line):
                                        line[-1] = line[-1]+'/In Line Eye Size Not Found.'
                                    else:
                                        line.append('In Line Eye Size Not Found')
                                if not product_values_dict['html_color']:
                                    if len(wrong_lines[0]) - 1 < len(line):
                                        line[-1] = line[-1]+'/In Line HTML Color Not Found.'
                                    else:
                                        line.append('In Line HTML Color Not Found')

                                wrong_lines.append(line)

                    else:
                        if import_line_seo_id or products_seo_id:
                            line.append('Product SEO Keyword Already Exist (%s)'%(','.join(products_seo_id.mapped('default_code'))))
                        elif barcode:
                            line.append('Product Barcode Already Exist (%s)'%(','.join(barcode.mapped('default_code'))))
                        elif not  product_values_dict['active'] and 'Active' in fields_list:
                            line.append('Product is Archive.')
                        elif product_id:
                            line.append('Product Already Exist')
                        else:
                            line.append('Product Not Found')
                        wrong_lines.append(line)
                        
                except Exception as e:
                    line.append(str(e)+'in column '+ file_column if file_column else str(e))
                    wrong_lines.append(line)

    def genrate_keyword(self,name,brand,model,categ,color,shape,material,rim_type):    
        keyword = [name, brand, categ, brand+' '+categ, brand+' '+model, brand+' '+color, brand+' '+model+' '+color, brand+' '+model+' '+color, brand+' '+rim_type+' '+color, brand+' '+model+' '+rim_type, brand+' '+shape+' '+rim_type, brand+' '+shape+' '+color, brand+' '+material+' '+color, brand+' '+material+' '+shape, brand+' '+material+' '+rim_type,brand+' '+categ+' '+shape, brand+' '+categ+' '+rim_type, brand+' '+categ+' '+color, brand+' '+categ+' '+shape+' '+color, brand+' '+categ+' '+material, brand+' '+rim_type]
        keyword = ','.join(keyword)
        return keyword

    def action_product_image_path_change(self):
        product_list = []
        try:
            product_obj = self.env['product.product']
            self.ensure_one()
            fields_dict = {'new_image_file_name':'New Image File Name','old_image_file_name':'Old Image File Name','id':'Id'}
            if not folder_path:
                raise UserError(_('Please add path in import file.'))
            else:
                if not os.path.exists(folder_path):
                    raise UserError(_('Path is not correct.'))

            if self.attach_file:
                file_datas = base64.b64decode(self.attach_file)
                workbook = xlrd.open_workbook(file_contents =file_datas)
                sheet = workbook.sheet_by_index(0)
                file_data = [[sheet.cell_value(r, c) for c in range(
                sheet.ncols)] for r in range(sheet.nrows)]
                heading_list = list(fields_dict.values())
                error_list = []
                for heading in file_data[0]:
                    if heading.strip() not in heading_list:
                        error_list.append(heading)
                    else:
                        fields_dict[heading.lower().strip().replace(' ','_')] = file_data[0].index(heading)
                if error_list:
                    raise UserError(_('Please remove %s column in file')%(','.join(error_list)))
                else:
                    for heading in fields_dict.keys():
                        if isinstance(fields_dict.get(heading,'0'), str):
                            error_list.append(fields_dict.get(heading,' '))
                    if error_list:
                        raise UserError(_('Please insert %s column in file')%(','.join(error_list)))

                file_data[0].append('Error')
                error_list = [file_data[0]]
                for line in file_data[1:]:
                    e = []
                    try:
                        if line[fields_dict.get('old_image_file_name',False)] and line[fields_dict.get('new_image_file_name',False)]:
                            product_id = product_obj.with_context(pending_price=True).search([('default_code','=',line[fields_dict.get('id',False)]),'|',('active','=',False),('active','=',True)])
                            if product_id:
                                if isinstance(fields_dict.get('old_primary_image'),int) or line[fields_dict.get('old_image_file_name')]:
                                    os.chdir(folder_path)
                                    if os.path.exists(line[fields_dict.get('old_image_file_name')]+'_01.jpg'):
                                        if isinstance(fields_dict.get('new_image_file_name',False),int) or line[fields_dict.get('new_image_file_name')]:
                                            os.rename(line[fields_dict.get('old_image_file_name')]+'_01.jpg',line[fields_dict.get('new_image_file_name')]+'_01.jpg')
                                        else:
                                            e.append('New primary image not found in folder path')
                                    else:
                                        e.append('Old primary image file not found in folder path')
                                    if os.path.exists(line[fields_dict.get('old_image_file_name')]+'_00.jpg'):
                                        if isinstance(fields_dict.get('new_image_file_name',False),int) or line[fields_dict.get('new_image_file_name')]:
                                            os.rename(line[fields_dict.get('old_image_file_name')]+'_00.jpg',line[fields_dict.get('new_image_file_name')]+'_00.jpg')
                                        else:
                                            e.append('New secondary image not found in folder path')
                                    else:
                                        e.append('Old secondary image file not found in folder path')
                                else:
                                    e.append('Old secondary image data are not correct.')
                            else:
                                e.append('Product not found')
                        else:
                            e.append('All columns are required.')
                    except Exception as error:
                        line.append(','.join(error.args))
                        error_list.append(line)
                    if e:
                        if os.path.exists(line[fields_dict.get('new_image_file_name')]+'_01.jpg'):
                            os.rename(line[fields_dict.get('new_image_file_name')]+'_01.jpg',line[fields_dict.get('old_image_file_name')]+'_01.jpg',)
                        if os.path.exists(line[fields_dict.get('new_image_file_name')]+'_00.jpg'):
                            os.rename(line[fields_dict.get('new_image_file_name')]+'_00.jpg',line[fields_dict.get('old_image_file_name')]+'_00.jpg',)
                        line.append(','.join(e))
                        error_list.append(line)
                    else:
                        product_list.append(product_id.id)
            self.product_pro_ids = [(6,0,list(set(product_list)))]
            self.run_time_file_name = ''
            self.run_time = False
            if len(error_list) > 1:
                out = BytesIO()
                workbook = xlsxwriter.Workbook(out) 
                worksheet = workbook.add_worksheet('wrong')
                heading_line = error_list.pop(0)
                col = row =0
                for heading_data in heading_line:
                    worksheet.write(row,col,heading_data)
                    worksheet.set_column(row,col, len(heading_data)*10)

                    col+=1
                for worksheet_line in error_list:
                    row +=1
                    col = 0
                    for worksheet_line_data in worksheet_line:
                        worksheet.write(row,col,worksheet_line_data)
                        col +=1

                workbook.close()
                out.seek(0)
                data = out.read()
                out.close()
                self.run_time_file_name = 'Wrong_product.xlsx'
                self.run_time = base64.b64encode(data)
            self.state = 'done'
        except Exception as e:
            raise UserError(_(','.join(e.args)))

    def action_product_data(self):
        wizard_obj = self.env['product.info.wizard.spt']
        query = '''select
            pp.default_code as "SKU",
            pp.variant_name as "Name",
            pbs.name as "Brand",
            pms.name as "Model",
            pcc.name as "Manufacturing Color Code",
            pp.eye_size_compute as "Eye Size",
            pc.name as "Category",
            (pp.available_qty_spt+pp.reversed_qty_spt) as "Total Qty",
            pp.available_qty_spt as "Available Qty",
            pp.reversed_qty_spt as "Reserved Qty",
            pp.lst_price_usd as "Price",
            pp.barcode as "Barcode",
            pp.image_url as "Image",
            pp.image_secondary_url as "Image 2",
            (select case when pp.is_published_spt = True then 'Yes' else 'No' end) as "Is Published",
            (SELECT CASE WHEN pp.is_image_missing = True then 'false' else 'true' end) as "Image Set",
            pcs.name as "Color Name",
            (select name from product_color_spt where id = pp.secondary_color_name) as "Secondary Color Name",
            COALESCE (pp.temporary_out_of_stock,false)::varchar as "Temporary Out Of Stock",
            pp.manufacture_color_code as "Manufacturer Color Code",
            pbss.name as "Bridge Size",
            COALESCE (ptss.name,'')::varchar as "Temple Size",
            (select name from product_color_spt where id = pp.lense_color_name) as "Lence Color Name",
            prts.name as "Rim Type",
            shape.name as "Shape",
            material.name as "Material",
            pp.flex_hinges as "Flex Hinges",
            COALESCE (pp.weight,0.0)::float as "Weight",
            (select case when pp.gender = 'male' then 'M' when pp.gender = 'female' then 'F' when pp.gender = 'm/f' then 'M/F' else '' end ) as "Gender",
            pp.create_date::Date as "Create Date",
            pp.write_Date::Date as "Modify Date",
            case when country.name is NULL then 'N/A' else country.name->>'en_US' end as "Country of Origin",
            pp.order_not_invoice as "#Open Orders",
            (select count(oder.id) from sale_order_line sol 
                        left join sale_order oder on sol.order_id = oder.id
                        where sol.product_id = pp.id and oder.state not in ('cancel','merged')) as "#Order",
            (SELECT CASE WHEN pp.is_forcefully_unpublished = True then 'Yes' else 'No' end) as "Is Forcefully Unpublished"
        from product_product pp
            left join product_template pt on pp.product_tmpl_id = pt.id
            left join product_brand_spt pbs on pp.brand=pbs.id
            left join product_model_spt pms on pp.model=pms.id
            left join product_category pc on pp.categ_id=pc.id
            left join product_color_spt pcs on pp.product_color_name=pcs.id
            left join product_size_spt pss on pp.size=pss.id 
            left join product_bridge_size_spt pbss on pp.bridge_size = pbss.id
            left join product_temple_size_spt ptss on pp.temple_size = ptss.id
            left join product_rim_type_spt prts on pp.rim_type = prts.id
            FULL OUTER join product_with_material_real material_real on pp.id = material_real.product_id
            left join product_material_spt material on material_real.material_id = material.id
            FULL OUTER join product_with_shape_real shape_real on pp.id = shape_real.product_id
            left join product_shape_spt shape on shape_real.shape_id = shape.id 
            left join res_country country on pp.country_of_origin = country.id
            left join kits_product_color_code pcc on pp.color_code = pcc.id
            where pp.active = 'True' order by pp.default_code;'''

        print('\n')
        print(query)
        print('\n')
        self.env.cr.execute(query)
        product_ids = self._cr.fetchall()
        columns = [desc[0] for desc in self.env.cr.description]
        df = pd.DataFrame(product_ids,columns=columns)
        writer = pd.ExcelWriter('/tmp/All_Prodcts_Export.xlsx')
        df.to_excel(writer,index=False,sheet_name="Products")
        writer.save()
        message=  f"From {self.env['product.product'].search_count([('active','=',True)])} products {product_ids.__len__()} products are exported."
        context = {"default_message":message,'all_product':True}
        return {
            "name": _("Exported Products"),
            "type":"ir.actions.act_window",
            "res_model":"warning.spt.wizard",
            "view_mode":"form",
            "view_id":self.env.ref('tzc_sales_customization_spt.warning_wizard_spt_form_view').id,
            "context":context,
            'target':"new",
        }


    def action_cancel_qty(self):
        for record in self:
            move_ids = self.env['stock.move'].search([('product_import_id','=',record.id)])
            for move in move_ids:
                move.quantity_done = 0
                move.state = 'cancel'
            title = _("Product Qty reset")
            message = _("Product Qty reset")
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Qty Reset'),
                'message': 'Product qty reset done.',
                'sticky': (True),
            }
        }
