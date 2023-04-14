from odoo import models,fields,api,_
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from io import BytesIO
import base64
import openpyxl
from odoo.exceptions import UserError
from urllib.request import urlopen
import xlsxwriter
from PIL import Image
import os

class kits_wizard_download_catalog_excel(models.TransientModel):
    _name = 'kits.wizard.download.catalog.excel'
    _description = 'Download Catalog Excel Report'
    
    catalog_id = fields.Many2one('sale.catalog','Catalog')

    @api.depends_context('default_catalog_id')
    def _get_domain_partner_id(self):
        catalog = self.env['sale.catalog'].browse(self._context.get('default_catalog_id'))
        return [('id','in',catalog.partner_ids.ids)]
    
    partner_id = fields.Many2one('res.partner','Customer')
    partner_ids = fields.Many2many('res.partner',string="Contacts")
    file = fields.Binary()
    currency = fields.Selection([('usd','USD'),('cad','CAD')],string='Currency',default='usd')

    def action_download_report(self):
        alignment_left = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        alignment_center = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
        main_font = Font(name='Calibri',size='16',bold=True)
        data_font = Font(name='Calibri',size='12')
        all_border = Border(left=Side(style='thin', color="000000"), 
                            right=Side(style='thin', color="000000"), 
                            top=Side(style='thin', color="000000"), 
                            bottom=Side(style='thin', color="000000"))

        if self.partner_id:
            currency_name = self.partner_id.property_product_pricelist.currency_id.name

            base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/_base_template_catalog.xlsm'
            wb = load_workbook(base_sample_file,read_only=False, keep_vba=True)
            wrksht = wb.active

            partner_line = 7
            wrksht.cell(row=partner_line, column=1).value = self.env['sale.catalog'].create_address_line_for_sale(self.partner_id, take_name=True)
            
            wrksht.cell(row=partner_line+2, column=1).value = self.catalog_id.name
            wrksht.cell(row=partner_line+2, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+2, column=1).font = main_font

            wrksht.cell(row=partner_line+3, column=1).value = 'Date: ' + self.catalog_id.create_date.strftime('%d/%m/%Y')
            wrksht.cell(row=partner_line+3, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+3, column=1).font = data_font

            wrksht.cell(row=partner_line+4, column=1).value = 'Salesperson: ' + self.catalog_id.user_id.name
            wrksht.cell(row=partner_line+4, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+4, column=1).font = data_font

            wrksht.cell(row=16, column=5).value = 'Price (%s)'%(currency_name or '')
            
            data_row = 17

            total_discount = 0.00
            total_subtotal = 0.00
            total_products = 0
            # restricted_products = self.env['product.product'].search([('geo_restriction','in',self.partner_id.country_id.ids)])
            lines = self.catalog_id.get_catalog_line()
            for line in lines:
                line_dict = lines.get(line,{})
                product_id= line_dict.get('product_id')
                # if line.product_pro_id not in restricted_products:
                if product_id.type == 'product':
                    wrksht.row_dimensions[data_row].height = 240
                    wrksht.cell(row=data_row, column=1).border = all_border
                    wrksht.cell(row=data_row, column=2).border = all_border
                    wrksht.cell(row=data_row, column=3).value = product_id.catalog_report_product_name() or ''
                    wrksht.cell(row=data_row, column=3).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=data_row, column=3).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,wrap_text=True)
                    wrksht.cell(row=data_row, column=3).border = all_border
                    wrksht.cell(row=data_row, column=4).value = int(line_dict.get('qty',0.0)) or 0
                    # total_products += line.product_qty
                    total_products += int(line_dict.get('qty',0.0)) or 0
                    wrksht.cell(row=data_row, column=4).alignment = alignment_center
                    wrksht.cell(row=data_row, column=4).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=data_row, column=4).border = all_border
                    # price = self.partner_id.property_product_pricelist._get_product_price(line.product_pro_id,line.product_qty,uom=line.product_pro_id.uom_id)
                    # if line.sale_type == 'on_sale':
                    #     # if currency_name == 'USD':
                    #     price = line.product_pro_id.on_sale_usd
                    #     # else:
                    #     #     price = line.product_pro_id.on_sale_cad
                    # if line.sale_type == 'clearance':
                    #     # if currency_name == 'CAD':
                    #     #     price = line.product_pro_id.clearance_cad
                    #     # else:
                    #     price = line.product_pro_id.clearance_usd
                    # if not price:
                    #     # if currency_name == 'USD':
                    #     #     price = line.product_pro_id.lst_price_usd
                    #     # else:
                    #     price = line.product_pro_id.lst_price
                    
                    # product_price = 0.0
                    # if 'eto dubai' in self.partner_id.property_product_pricelist.name.lower() or 'other eto' in self.partner_id.property_product_pricelist.name.lower():
                    #     product_price = self.env['product.pricelist.item'].search([('product_id','=',line.product_pro_id.id),('pricelist_id','=',self.partner_id.property_product_pricelist.id)],limit=1).fixed_price
                    # else:
                    #     # if currency_name == 'USD':
                    #     #     product_price = line.product_pro_id.lst_price_usd
                    #     # else:
                    #     product_price = line.product_pro_id.lst_price

                    # if 'eto dubai' in self.partner_id.property_product_pricelist.name.lower() or 'other eto' in self.partner_id.property_product_pricelist.name.lower():
                    #     discount_amount = (line.product_price - product_price) * line.product_qty
                    #     subtotal = product_price * line.product_qty
                    #     total_discount += round(discount_amount,2)
                    #     total_subtotal += round(subtotal,2)
                    # else:
                    #     discount_amount = product_price * line.discount * 0.01 * line.product_qty
                    #     subtotal = (product_price * line.product_qty) - discount_amount
                    #     total_discount += round(discount_amount,2)
                    #     total_subtotal += round(subtotal,2)
                    
                    # if 'eto dubai' in self.partner_id.property_product_pricelist.name.lower() or 'other eto' in self.partner_id.property_product_pricelist.name.lower():
                    #     wrksht.cell(row=data_row, column=5).value = round(product_price,2) or 0.00
                    #     wrksht.cell(row=data_row, column=5).alignment = alignment_center
                    #     wrksht.cell(row=data_row, column=5).font = Font(name='Calibri',size='16',bold=False)
                    #     wrksht.cell(row=data_row, column=5).number_format = '"$"#,##0.00'
                    #     wrksht.cell(row=data_row, column=5).border = all_border
                    # else:
                    #     wrksht.cell(row=data_row, column=5).value = round(product_price - (product_price * line.discount / 100),2) or 0.00
                    #     wrksht.cell(row=data_row, column=5).alignment = alignment_center
                    #     wrksht.cell(row=data_row, column=5).font = Font(name='Calibri',size='16',bold=False)
                    #     wrksht.cell(row=data_row, column=5).number_format = '"$"#,##0.00'
                    #     wrksht.cell(row=data_row, column=5).border = all_border
                    
                    wrksht.cell(row=data_row, column=5).value = round(line_dict.get('our_price',0),2) or 0.00
                    wrksht.cell(row=data_row, column=5).alignment = alignment_center
                    wrksht.cell(row=data_row, column=5).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=data_row, column=5).number_format = '"$"#,##0.00'
                    wrksht.cell(row=data_row, column=5).border = all_border
                    
                    total_discount = total_discount + (line_dict.get('price_unit',0) - line_dict.get('our_price',0))
                    total_subtotal = total_subtotal + line_dict.get('our_price',0)

                    wrksht.cell(row=data_row, column=6).value = '=D%s*E%s'%(str(data_row),str(data_row))
                    wrksht.cell(row=data_row, column=6).alignment = alignment_center
                    wrksht.cell(row=data_row, column=6).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=data_row, column=6).number_format = '"$"#,##0.00'
                    wrksht.cell(row=data_row, column=6).border = all_border

                    # wrksht.cell(row=data_row, column=7).hyperlink = line.product_pro_id.image_url
                    wrksht.cell(row=data_row, column=7).hyperlink = product_id.image_url if product_id.image_url else 'False' 
                    wrksht.cell(row=data_row, column=7).alignment = alignment_center
                    wrksht.cell(row=data_row, column=7).font = Font(color="FF0000FF",underline='single',name='Calibri',size='16')
                    wrksht.cell(row=data_row, column=7).border = all_border

                    # wrksht.cell(row=data_row, column=8).hyperlink = line.product_pro_id.image_secondary_url
                    wrksht.cell(row=data_row, column=8).hyperlink = product_id.image_secondary_url if product_id.image_secondary_url else 'False' 
                    wrksht.cell(row=data_row, column=8).alignment = alignment_center
                    wrksht.cell(row=data_row, column=8).font = Font(color="FF0000FF",underline='single',name='Calibri',size='16')
                    wrksht.cell(row=data_row, column=8).border = all_border

                    data_row += 1

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            self.file = base64.b64encode(data)
            return {
                "type": "ir.actions.act_url",
                "url": 'web/content/?model=kits.wizard.download.catalog.excel&download=True&field=file&id=%s&filename=%s.xlsm' % (self.id,'%s Catalog %s'%(self.partner_id.name,self.catalog_id.name)),
                "target": "new",
                }
        else:
            raise UserError(_('No Customer Selected.'))
    
    def action_download_report_without_customer(self):
        workbook = Workbook()
        sheet = workbook.create_sheet(title='Catalog',index=0)

        alignment = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
        alignment_right = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
        alignment_left = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        header_font = Font(name="Garamond",size="10",bold=False)
        table_header_font = Font(name="Garamond",size="10",bold=True)
        currency = 'USD' if self.currency == 'usd' else 'CAD'
        currency_symbol = '$'
        # ====================================================================
        # Company LOGO
        img = BytesIO()
        img.flush()
        img.write(base64.b64decode(self.env.companies[0].logo))
        image = openpyxl.drawing.image.Image(img)
        image.width = 550
        image.height = 55
        sheet.merge_cells('A1:F5')
        sheet.add_image(image, 'A2')
        # LOGO END
        sheet.merge_cells('G1:I5')
        address = []
        company = self.env.companies[0]
        if company.name:
            address.append(company.name)
            address.append('(A division of Tanzacan Tradelink Inc.)')
        if company.street:
            address.append(company.street)
        if company.street2:
            address.append(company.street2)
        if company.city:
            address.append(company.city)
        if company.state_id:
            address.append(company.state_id.name)
        if company.zip:
            address.append(company.zip)
        if company.country_id:
            address.append(company.country_id.name)
        address_line = '\n'.join(address)
        sheet.cell(row=1, column=7).value = address_line
        sheet.cell(row=1, column=7).font = header_font
        sheet.cell(row=1, column=7).alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
        # ====================================================================
        table_header_index = 7
        sheet.cell(row=table_header_index, column=1).value = 'Image'
        sheet.cell(row=table_header_index, column=2).value = 'Secondary Image'
        sheet.cell(row=table_header_index, column=3).value = 'Product'
        sheet.cell(row=table_header_index, column=4).value = 'CAT'
        sheet.cell(row=table_header_index, column=5).value = 'QTY'
        sheet.cell(row=table_header_index, column=6).value = 'WHOLESALE'
        sheet.cell(row=table_header_index, column=7).value = 'DISC %'
        sheet.cell(row=table_header_index, column=8).value = 'PRICE (%s)'%(currency or '')
        sheet.cell(row=table_header_index, column=9).value = 'SUBTOTAL'
        sheet.row_dimensions[table_header_index].height = 20

        sheet.cell(row=table_header_index, column=1).alignment = alignment
        sheet.cell(row=table_header_index, column=2).alignment = alignment
        sheet.cell(row=table_header_index, column=3).alignment = alignment
        sheet.cell(row=table_header_index, column=4).alignment = alignment
        sheet.cell(row=table_header_index, column=5).alignment = alignment
        sheet.cell(row=table_header_index, column=6).alignment = alignment
        sheet.cell(row=table_header_index, column=7).alignment = alignment
        sheet.cell(row=table_header_index, column=8).alignment = alignment
        sheet.cell(row=table_header_index, column=9).alignment = alignment

        sheet.cell(row=table_header_index, column=1).font = table_header_font
        sheet.cell(row=table_header_index, column=2).font = table_header_font
        sheet.cell(row=table_header_index, column=3).font = table_header_font
        sheet.cell(row=table_header_index, column=4).font = table_header_font
        sheet.cell(row=table_header_index, column=5).font = table_header_font
        sheet.cell(row=table_header_index, column=6).font = table_header_font
        sheet.cell(row=table_header_index, column=7).font = table_header_font
        sheet.cell(row=table_header_index, column=8).font = table_header_font
        sheet.cell(row=table_header_index, column=9).font = table_header_font

        total_products=0
        total_discount = 0.00
        total_subtotal = 0.00

        row_index = table_header_index+1
        restricted_products = self.env['product.product'].search([('geo_restriction','in',self.partner_id.country_id.ids)])
        img1,img2=False,False
        pricelist = self.env['product.pricelist'].search([('currency_id.name','=',currency),('name','ilike','Public Price')],limit=1)
        filter_by = 'available_qty_spt' if self.catalog_id.base_on_qty == 'available_qty' else 'qty_available'
        lines = self.catalog_id.line_ids
        if self.catalog_id.base_on_qty:
            lines = lines.filtered(lambda x: x.product_pro_id[filter_by] >= 1)
        for line in lines.sorted(lambda x: x.product_pro_id.variant_name):
            if line.product_pro_id not in restricted_products or line.product_id.available_qty_spt <= 0:
                product_name = line.product_pro_id.name_get()
                # for attribute in line.product_pro_id.product_template_attribute_value_ids:
                #     if attribute.attribute_id.name == 'Color':
                #         color_name = attribute.product_attribute_value_id.name.split('-')[0]
                #     if attribute.attribute_id.name == 'Eye Size':
                #         eye_size = attribute.product_attribute_value_id.name
                # product_name = str(str(line.product_pro_id.brand.name if line.product_pro_id.brand else '')+' '+str(line.product_pro_id.model.name if line.product_pro_id.model else '')+' '+color_name +' '+eye_size +str(line.product_pro_id.temple_size.name if line.product_pro_id.temple_size else '')) or ""
                if line.image_variant_1920:
                    img1 = BytesIO()
                    img1.flush()
                    img1.write(base64.b64decode(line.image_variant_1920))
                    image1 = openpyxl.drawing.image.Image(img1)
                    image1.width = 100
                    image1.height = 50
                    
                    sheet.add_image(image1,'A%s'%(row_index))
                
                if line.image_variant_1920:
                    img2 = BytesIO()
                    img2.flush()
                    img2.write(base64.b64decode(line.image_secondary_1920))
                    image2 = openpyxl.drawing.image.Image(img2)
                    image2.width = 100
                    image2.height = 50
                    
                    sheet.add_image(image2,'B%s'%(row_index))

            sheet.cell(row=row_index, column=3).value = product_name[0][1] or ''
            # sheet.cell(row=row_index, column=3).value = line.product_pro_id.variant_name or ''
            sheet.cell(row=row_index, column=4).value = line.product_pro_id.categ_id.name or ''
            sheet.cell(row=row_index, column=5).value = line.product_qty or 0
            total_products += line.product_qty
            wholesale = line.product_pro_id.price_wholesale_usd if currency == 'USD' else line.product_pro_id.price_wholesale
            sheet.cell(row=row_index, column=6).value = currency_symbol+str(round(wholesale,2) or 0.00)
            sheet.cell(row=row_index, column=7).value = round(line.discount,2) or 0.00
            
            price = pricelist.get_product_price(line.product_pro_id,line.product_qty,self.env['res.partner'].browse())
            if line.sale_type == 'on_sale':
                # if currency == 'USD':
                price = line.product_pro_id.on_sale_usd
                # else:
                #     price = line.product_pro_id.on_sale_cad
            if line.sale_type == 'clearance':
                # if currency  == 'USD':
                price = line.product_pro_id.clearance_usd
                # else:
                #     price = line.product_pro_id.clearance_cad
            if not price:
                # if currency == 'USD':
                #     price = line.product_pro_id.lst_price_usd
                # else:
                    price = line.product_pro_id.lst_price
            
            disc = (price * line.discount * 0.01) * line.product_qty
            sub_total = (price * line.product_qty) - disc
            
            total_discount += round(disc,2)
            total_subtotal+= round(sub_total,2)

            sheet.cell(row=row_index, column=8).value = currency_symbol+str(round(price - (price * line.discount / 100),2))
            sheet.cell(row=row_index, column=9).value = currency_symbol+str(round(sub_total,2))
        
            sheet.cell(row=row_index, column=3).alignment = alignment_left
            sheet.cell(row=row_index, column=4).alignment = alignment_left
            sheet.cell(row=row_index, column=5).alignment = alignment_right
            sheet.cell(row=row_index, column=6).alignment = alignment_right
            sheet.cell(row=row_index, column=7).alignment = alignment_right
            sheet.cell(row=row_index, column=8).alignment = alignment_right
            sheet.cell(row=row_index, column=9).alignment = alignment_right

            sheet.row_dimensions[row_index].height = 50
            row_index+=1

        footer_row = row_index+1
        sheet.cell(row=footer_row, column=8).value = 'Subtotal'
        sheet.cell(row=footer_row, column=9).value = currency_symbol+str(round(total_subtotal,2))
        sheet.cell(row=footer_row, column=9).alignment = alignment_right
        footer_row += 1
        sheet.cell(row=footer_row, column=8).value = 'Discount'
        sheet.cell(row=footer_row, column=9).value = currency_symbol+str(round(total_discount,2))
        sheet.cell(row=footer_row, column=9).alignment = alignment_right
        footer_row += 1
        sheet.cell(row=footer_row, column=8).value = 'Total (%s)'%(currency)
        sheet.cell(row=footer_row, column=9).value = currency_symbol+str(round(total_subtotal-total_discount,2))
        sheet.cell(row=footer_row, column=9).alignment = alignment_right
        footer_row += 1
        sheet.cell(row=footer_row, column=8).value = 'Total Quantity'
        sheet.cell(row=footer_row, column=9).value = total_products
        sheet.cell(row=footer_row, column=9).alignment = alignment_right
        footer_row += 1
        
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 5
        sheet.column_dimensions['E'].width = 5
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 8
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 12

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        img1.close() if img1 else None
        img2.close() if img2 else None
        self.file = base64.b64encode(data)
        return {
            "type": "ir.actions.act_url",
            "url": 'web/content/?model=kits.wizard.download.catalog.excel&download=True&field=file&id=%s&filename=%s.xlsx' % (self.id,'Catalog %s'%(self.catalog_id.name)),
            "target": "new",
            }

    def action_download_report_pdf(self):
        if self.partner_id:
            partner_catalog_id = self.catalog_id.filtered(lambda x:self.partner_id.id in x.partner_ids.ids)
            if partner_catalog_id:
                # return self.env.ref('tzc_sales_customization_spt.action_catalog_report_spt').report_action(partner_catalog_id)
                self.catalog_id.customer_id = self.partner_id.id
                return self.env.ref('tzc_sales_customization_spt.action_catalog_report_pdf').with_context(partner_id=self.partner_id).report_action(partner_catalog_id)
        else:
            raise UserError ('No Customer Selected.')

    def line_ordering_by_product(self):
        product_list = []
        user = self.env.user    
        for line in self.catalog_id.line_ids:
            if user.country_id.id not in line.product_pro_id.geo_restriction.ids:
                product_name = line.product_pro_id.name_get()[0][1]
                product_list.append(product_name.strip())
            else:
                product_list.append('')
        product_list = list(filter(None, product_list))
        product_list = list(set(product_list))
        product_list.sort()
        return product_list

    def line_product_dict(self,product_name):
        product_dict = {}
        user = self.env.user
        restricted_products = self.env['product.product'].search([('geo_restriction','in',user.country_id.ids)])
        catalog_id = self.catalog_id.line_ids.filtered(lambda line: line.product_pro_id.name_get()[0][1] == product_name)
        if catalog_id[0].product_pro_id.id not in restricted_products.ids:
            product_dict[product_name] = {'line_ids': [catalog_id[0]]}
            return product_dict
        else:
            product_dict[product_name] = {'line_ids': []}
            return product_dict
