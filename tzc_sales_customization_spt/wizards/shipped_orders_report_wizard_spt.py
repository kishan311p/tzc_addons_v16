from odoo import fields, models, api, _
from datetime import datetime,date
from dateutil import tz
from odoo.exceptions import UserError
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import base64
from io import BytesIO


class shipped_orders_report_wizard_spt(models.TransientModel):
    _name = 'shipped.orders.report.wizard.spt'
    _description = 'shipped orders Report'

    start_date = fields.Datetime("Start Date")
    end_date = fields.Datetime('End Date')
    # sale_order_ids = fields.Many2many('sale.order', 'shipped_orders_wizard_spt_sale_order_rel',
    #                                   'shipped_order_report_wizard_id', 'sale_order_id', 'Sale Orders')
    report_file = fields.Binary()

    @api.onchange("start_date",'end_date')
    def _check_date(self):
        if self.end_date:
            if self.start_date and (self.end_date <= self.start_date):
                raise UserError(
                    'You have entered Date and Time lesser or equal to Start date and time.')

    def shipping_orders_report_print(self):
        so_ids = self.get_sale_orders()
        return self.env.ref('tzc_sales_customization_spt.action_report_shipped_orders').report_action(self)

    def show_address(self,source):
        address=[]
        if source:
            if source.street:
                address.append(source.street)
            if source.street2:
                address.append(source.street2)
            if source.city:
                address.append(source.city)
            if source.state_id:
                address.append(source.state_id.name)
            if source.zip:
                address.append(source.zip)
        return ','.join(address)

    def shipping_orders_excel_report(self):
        active_id = self.id
        f_name = 'Shipped Orders'
        workbook = Workbook()
        sheet = workbook.create_sheet(title='Shipped Orders', index=0)
        bd = Side(style='thin', color="000000")
        bottom_border = Border(bottom=bd)
        top_bottom_border = Border(top=bd, bottom=bd)
        heading_font = Font(name="Garamond", size="10", bold=True)
        table_font = Font(name="Garamond", size="10", bold=False)
        align_left = Alignment(
            vertical="center", horizontal='left', text_rotation=0, wrap_text=True)
        align_right = Alignment(
            vertical="center", horizontal='right', text_rotation=0, wrap_text=True)

        table_header_row = 1
        sheet.row_dimensions[table_header_row].height = 30
        sheet.cell(row=table_header_row, column=1).value = 'Order Date'
        sheet.cell(row=table_header_row, column=2).value = "Order#"
        sheet.cell(row=table_header_row, column=3).value = "Shipping Date"
        sheet.cell(row=table_header_row, column=4).value = "Cutomer"
        sheet.cell(row=table_header_row, column=5).value = "Total Ordered Qty"
        sheet.cell(row=table_header_row, column=6).value = "Total Picked Qty"
        sheet.cell(row=table_header_row, column=7).value = 'Admin Fee'
        sheet.cell(row=table_header_row, column=8).value = "Shipping Cost"
        sheet.cell(row=table_header_row, column=9).value = "Shipping Provider"
        sheet.cell(row=table_header_row, column=10).value = "Tracking Number"
        sheet.cell(row=table_header_row, column=11).value = "Country"
        sheet.cell(row=table_header_row, column=12).value = "Shipping Address"
        sheet.cell(row=table_header_row, column=13).value = "Currency"
        sheet.cell(row=table_header_row, column=14).value = "Subtotal"

        sheet.cell(row=table_header_row, column=1).alignment = align_left
        sheet.cell(row=table_header_row, column=2 ).alignment = align_left
        sheet.cell(row=table_header_row, column=3).alignment = align_left
        sheet.cell(row=table_header_row, column=4 ).alignment = align_left
        sheet.cell(row=table_header_row, column=5 ).alignment = align_left
        sheet.cell(row=table_header_row, column=6 ).alignment = align_left
        sheet.cell(row=table_header_row, column=7 ).alignment = align_right
        sheet.cell(row=table_header_row, column=8 ).alignment = align_right
        sheet.cell(row=table_header_row, column=9 ).alignment = align_left
        sheet.cell(row=table_header_row, column=10 ).alignment = align_left
        sheet.cell(row=table_header_row, column=11 ).alignment = align_left
        sheet.cell(row=table_header_row, column=12 ).alignment = align_left
        sheet.cell(row=table_header_row, column=13 ).alignment = align_right
        sheet.cell(row=table_header_row, column=14 ).alignment = align_right

        for col in range(1, 15):
            sheet.cell(row=table_header_row,
                       column=col).border = top_bottom_border
            sheet.cell(row=table_header_row, column=col).font = heading_font

        report_data = self.get_report_data_by_query()
        row_index = table_header_row+1
        for data in report_data:
            if data[9] == '':
                sheet.row_dimensions[row_index].height = 35
            else:
                height = 3*len(data[9])/2
                sheet.row_dimensions[row_index].height = height if 3*len(data[9]) > 30 else 35
            sheet.cell(row=row_index, column=1).value = str(data[0])
            sheet.cell(row=row_index, column=2).value = data[1]
            sheet.cell(row=row_index, column=3).value = str(data[2])
            sheet.cell(row=row_index, column=4).value = data[3] if data[3] != None else ''
            sheet.cell(row=row_index, column=5).value = data[4]
            sheet.cell(row=row_index, column=6).value = data[5]
            sheet.cell(row=row_index, column=7).value = data[6]
            sheet.cell(row=row_index, column=8).value = data[7]
            sheet.cell(row=row_index, column=9).value = data[8]
            sheet.cell(row=row_index, column=10).value = data[9]
            sale_order = self.env['sale.order'].search([('name','=',data[1])])
            sheet.cell(row=row_index, column=11).value = sale_order.country_id.name if sale_order.country_id else ''
            address = self.env['res.partner'].search([('id','=',data[10])])
            sheet.cell(row=row_index, column=12).value = self.show_address(address)
            sheet.cell(row=row_index, column=13).value = data[11]
            sheet.cell(row=row_index, column=14).value = data[12]

            sheet.cell(row=row_index, column=1).font = table_font
            sheet.cell(row=row_index, column=2).font = table_font
            sheet.cell(row=row_index, column=3).font = table_font
            sheet.cell(row=row_index, column=4).font = table_font
            sheet.cell(row=row_index, column=5).font = table_font
            sheet.cell(row=row_index, column=6).font = table_font
            sheet.cell(row=row_index, column=7).font = table_font
            sheet.cell(row=row_index, column=8).font = table_font
            sheet.cell(row=row_index, column=9).font = table_font
            sheet.cell(row=row_index, column=10).font = table_font
            sheet.cell(row=row_index, column=11).font = table_font
            sheet.cell(row=row_index, column=12).font = table_font
            sheet.cell(row=row_index, column=13).font = table_font
            sheet.cell(row=row_index, column=14).font = table_font

            sheet.cell(row=row_index, column=1).border = bottom_border
            sheet.cell(row=row_index, column=2).border = bottom_border
            sheet.cell(row=row_index, column=3).border = bottom_border
            sheet.cell(row=row_index, column=4).border = bottom_border
            sheet.cell(row=row_index, column=5).border = bottom_border
            sheet.cell(row=row_index, column=6).border = bottom_border
            sheet.cell(row=row_index, column=7).border = bottom_border
            sheet.cell(row=row_index, column=8).border = bottom_border
            sheet.cell(row=row_index, column=9).border = bottom_border
            sheet.cell(row=row_index, column=10).border = bottom_border
            sheet.cell(row=row_index, column=11).border = bottom_border
            sheet.cell(row=row_index, column=12).border = bottom_border
            sheet.cell(row=row_index, column=13).border = bottom_border
            sheet.cell(row=row_index, column=14).border = bottom_border

            sheet.cell(row=row_index, column=1).alignment = align_left
            sheet.cell(row=row_index, column=2).alignment = align_left
            sheet.cell(row=row_index, column=3).alignment = align_left
            sheet.cell(row=row_index, column=4).alignment = align_left
            sheet.cell(row=row_index, column=5).alignment = align_left
            sheet.cell(row=row_index, column=6).alignment = align_left
            sheet.cell(row=row_index, column=7).alignment = align_right
            sheet.cell(row=row_index, column=8).alignment = align_right
            sheet.cell(row=row_index, column=9).alignment = align_left
            sheet.cell(row=row_index, column=10).alignment = align_left
            sheet.cell(row=row_index, column=11).alignment = align_left
            sheet.cell(row=row_index, column=12).alignment = align_left
            sheet.cell(row=row_index, column=13).alignment = align_right
            sheet.cell(row=row_index, column=14).alignment = align_right
            row_index += 1

        if not report_data:
            sheet.merge_cells("A"+str(row_index)+":M"+str(row_index))
            sheet.cell(row=row_index, column=1).value = 'There is no orderes between date ' + \
                str(self.start_date)+' and '+str(self.end_date)+' !'
            sheet.cell(row=row_index, column=1).font = table_font

        sheet.column_dimensions['A'].width = 17
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        sheet.column_dimensions['I'].width = 12
        sheet.column_dimensions['J'].width = 12
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 30
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.report_file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=shipped.orders.report.wizard.spt&download=true&field=report_file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }

    def get_report_data_by_query(self):
        query = '''SELECT SO.DATE_ORDER,
                    COALESCE(SO.NAME,'') AS ORDER,
                    SO.SHIPPED_DATE,
                    COALESCE(RP.NAME,'') AS NAME,
                    COALESCE(SO.ORDERED_QTY,0) AS ORDERED_QTY,
                    COALESCE(SO.PICKED_QTY,0) AS PICKED_QTY,
                    COALESCE(SO.AMOUNT_IS_ADMIN,0.00) AS AMOUNT_IS_ADMIN,
                    COALESCE(SO.AMOUNT_IS_SHIPPING_TOTAL,0.00) AS AMOUNT_IS_SHIPPING_COST,
                    COALESCE(SPS.NAME,'')AS SHIPPING_PROVIDER,
                    COALESCE(SP.TRACKING_NUMBER_SPT,'')AS TRACKING_NUMBER,
                    RP.ID,
                    COALESCE(RCU.NAME,'')AS CURRENCY,
                    COALESCE(SO.PICKED_QTY_ORDER_SUBTOTAL,'0.00')AS PICKED_QTY_ORDER_SUBTOTAL,
					SPT.CODE
                FROM SALE_ORDER AS SO
                INNER JOIN STOCK_PICKING AS SP ON SO.NAME = SP.ORIGIN
                INNER JOIN RES_PARTNER AS RP ON SO.PARTNER_ID = RP.ID
                INNER JOIN SHIPPING_PROVIDER_SPT AS SPS ON SP.SHIPPING_ID = SPS.ID
                INNER JOIN RES_CURRENCY AS RCU ON RCU.ID = SO.CURRENCY_ID
				INNER JOIN STOCK_PICKING_TYPE AS SPT ON SPT.ID = SP.PICKING_TYPE_ID
                WHERE SP.STATE = 'done' AND SPT.CODE = 'outgoing' '''
        
                    # RC.NAME AS COUNTY,
                # INNER JOIN RES_COUNTRY AS RC ON RC.ID = SO.COUNTRY_ID
        if self.start_date:
            query = query + " AND SO.DATE_ORDER >= '%s'" % (str(self.start_date))
        if self.end_date:
            query = query + " AND SO.DATE_ORDER <= '%s'" % (str(self.end_date))
        self.env.cr.execute(query)
        report_data=self.env.cr.fetchall()
        return report_data


    def get_sale_orders(self):
        tz_from, tz_to = tz.gettz(datetime.now().tzinfo), (tz.gettz(
            self.env.user.tz) or self.env.context.get('tz'))
        start_date = end_date = ''
        if self.start_date:
            start_date = self.start_date.replace(
                tzinfo=tz_from).astimezone(tz=tz_to)
        if self.end_date:
            end_date = self.end_date.replace(
                tzinfo=tz_from).astimezone(tz=tz_to)
        domain = [('picking_ids.state', '=', 'done')]
        if start_date:
            domain.append(('date_order', '>=', start_date))
        if end_date:
            domain.append(('date_order', '<=', end_date))
        if domain:
            so_ids = self.env['sale.order'].search(domain)
        return so_ids.sorted(lambda so: (so.date_order))
