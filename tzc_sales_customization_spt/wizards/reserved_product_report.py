from odoo import _, api, fields, models, tools
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import base64
from io import BytesIO
from datetime import datetime

class reserved_product_report(models.TransientModel):
    _name = "reserved.product.report"
    _description = "Reserved Product Report"

    product_sku = fields.Char('Products SKU\'s')
    report_file = fields.Binary()

    def action_export(self):
        active_id = self.id
        f_name = 'Reserved Product Report'
        workbook = Workbook()
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        heading_font = Font(name="Garamond", size="10", bold=True)
        table_font = Font(name="Garamond", size="10", bold=False)
        align_left = Alignment(
            vertical="center", horizontal='left', text_rotation=0, wrap_text=True)
        align_right = Alignment(
            vertical="center", horizontal='right', text_rotation=0, wrap_text=True)
        align_center = Alignment(
            vertical="center", horizontal='center', text_rotation=0, wrap_text=True)

        product_internal_ref = self.product_sku.split(',') if self.product_sku else ''
        for sku in list(dict.fromkeys(product_internal_ref)):
            sheet = workbook.create_sheet(title=sku, index=0)
            
            table_header_row = 1
            sheet.row_dimensions[table_header_row].height = 30
            sheet.cell(row=table_header_row, column=1).value = 'Order Reference'
            sheet.cell(row=table_header_row, column=2).value = "Customer"
            sheet.cell(row=table_header_row, column=3).value = "Creation Date"
            sheet.cell(row=table_header_row, column=4).value = "Updated On"
            sheet.cell(row=table_header_row, column=5).value = "Created by"
            sheet.cell(row=table_header_row, column=6).value = "Updated By"
            sheet.cell(row=table_header_row, column=7).value = 'Shipped Date'
            sheet.cell(row=table_header_row, column=7).value = "Salesperson"
            sheet.cell(row=table_header_row, column=8).value = "Order Source"
            sheet.cell(row=table_header_row, column=9).value = "Reserved Qty"
            sheet.cell(row=table_header_row, column=10).value = "Status"
            sheet.cell(row=table_header_row, column=11).value = "Payment Status"
            sheet.cell(row=table_header_row, column=12).value = "Has Payment Link"

            sheet.cell(row=table_header_row, column=1).alignment = align_center
            sheet.cell(row=table_header_row, column=2).alignment = align_center
            sheet.cell(row=table_header_row, column=3).alignment = align_center
            sheet.cell(row=table_header_row, column=4).alignment = align_center
            sheet.cell(row=table_header_row, column=5).alignment = align_center
            sheet.cell(row=table_header_row, column=6).alignment = align_center
            sheet.cell(row=table_header_row, column=7).alignment = align_center
            sheet.cell(row=table_header_row, column=8).alignment = align_center
            sheet.cell(row=table_header_row, column=9).alignment = align_center
            sheet.cell(row=table_header_row, column=10).alignment = align_center
            sheet.cell(row=table_header_row, column=11).alignment = align_center
            sheet.cell(row=table_header_row, column=12).alignment = align_center

            for col in range(1, 13):
                sheet.cell(row=table_header_row,
                        column=col).border = top_bottom_border
                sheet.cell(row=table_header_row, column=col).font = heading_font
            
            row_index = table_header_row+1
            sale_state = {'draft':'Quotation','sent':'Quotation Sent','received':'Quotation Received','sale':'Order Confirmed','in_scanning':'In Scanning','scanned':'Scanning Completed','scan':'Ready to Ship','shipped':'Shipped','draft_inv':'Draft Invoice','open_inv':'Invoiced','cancel':'Cancelled','merged':'Merged','done':'Locked','':''}
            payment_status = {'full':'Fully Paid','partial':'Partial Paid','over':'Over Paid','':''}
            query = F'''SELECT COALESCE(SO.NAME,'') AS ORDER,
                        COALESCE(RP.NAME,'')AS PARTNER,
                        SO.CREATE_DATE,
                        SO.UPDATED_ON,
                        COALESCE(R_P.NAME,'') AS CREATED_BY,
                        COALESCE(R_PA.NAME,'') AS UPDATED_BY,
                        COALESCE(R_PAR.NAME,'') AS SALESPERSON,
                        COALESCE(SO.SOURCE_SPT,'')AS ORDER_SOURCE,
                        COALESCE(SOL.PICKED_QTY,0)AS PICKED_QTY,
                        COALESCE(SO.STATE,'')AS SALE_STATE,
                        COALESCE(SO.PAYMENT_STATUS,'')AS PAYMENT_STATUS,
                        CASE
                            WHEN SO.IS_PAYMENT_LINK = FALSE THEN 0
                            WHEN SO.IS_PAYMENT_LINK = TRUE THEN 1
                        END AS IS_PAYMENT_LINK 
                    FROM SALE_ORDER AS SO
                    INNER JOIN RES_PARTNER AS RP ON SO.PARTNER_ID = RP.ID
                    INNER JOIN RES_USERS AS RU ON RU.ID = SO.CREATE_UID
                    INNER JOIN RES_PARTNER AS R_P ON RU.PARTNER_ID = R_P.ID
                    INNER JOIN RES_USERS AS R_U ON R_U.ID = SO.UPDATED_BY
                    INNER JOIN RES_PARTNER AS R_PA ON R_U.PARTNER_ID = R_PA.ID
                    INNER JOIN RES_USERS AS R_USER ON R_USER.ID = SO.USER_ID
                    INNER JOIN RES_PARTNER AS R_PAR ON R_USER.PARTNER_ID = R_PAR.ID
                    INNER JOIN SALE_ORDER_LINE AS SOL ON SOL.ORDER_ID = SO.ID
                    INNER JOIN PRODUCT_PRODUCT AS PP ON SOL.PRODUCT_ID = PP.ID
                    WHERE PP.DEFAULT_CODE = '{sku.strip()}' '''

            self.env.cr.execute(query)
            record_data = self.env.cr.fetchall()
            for order in record_data:
                sheet.row_dimensions[row_index].height = 22
                sheet.cell(row=row_index, column=1).value = order[0] 
                sheet.cell(row=row_index, column=2).value = order[1] 
                sheet.cell(row=row_index, column=3).value = datetime.strftime(order[2],'%m-%d-%Y %H:%M:%S')
                sheet.cell(row=row_index, column=4).value = datetime.strftime(order[3],'%m-%d-%Y %H:%M:%S')
                sheet.cell(row=row_index, column=5).value = order[4] 
                sheet.cell(row=row_index, column=6).value = order[5] 
                sheet.cell(row=row_index, column=7).value = order[6] 
                sheet.cell(row=row_index, column=8).value = order[7]
                sheet.cell(row=row_index, column=9).value = order[8]
                sheet.cell(row=row_index, column=10).value = sale_state[order[9]]
                sheet.cell(row=row_index, column=11).value = payment_status[order[10]]
                sheet.cell(row=row_index, column=12).value = order[11]

                sheet.cell(row=row_index, column=1).font = table_font
                sheet.cell(row=row_index, column=2).font = table_font
                sheet.cell(row=row_index, column=3).font = table_font
                sheet.cell(row=row_index, column=4).font = table_font
                sheet.cell(row=row_index, column=5).font = table_font
                sheet.cell(row=row_index, column=6).font = table_font
                sheet.cell(row=row_index, column=7).font = table_font
                sheet.cell(row=row_index, column=8).font = table_font
                sheet.cell(row=row_index, column=9).font = table_font
                sheet.cell(row=row_index, column=10).font = table_font
                sheet.cell(row=row_index, column=11).font = table_font
                sheet.cell(row=row_index, column=12).font = table_font

                sheet.cell(row=row_index, column=1).alignment = align_left
                sheet.cell(row=row_index, column=2).alignment = align_left
                sheet.cell(row=row_index, column=3).alignment = align_right
                sheet.cell(row=row_index, column=4).alignment = align_right
                sheet.cell(row=row_index, column=5).alignment = align_left
                sheet.cell(row=row_index, column=6).alignment = align_left
                sheet.cell(row=row_index, column=7).alignment = align_left
                sheet.cell(row=row_index, column=8).alignment = align_left
                sheet.cell(row=row_index, column=9).alignment = align_left
                sheet.cell(row=row_index, column=10).alignment = align_right
                sheet.cell(row=row_index, column=11).alignment = align_left
                sheet.cell(row=row_index, column=12).alignment = align_left

                row_index += 1
            
            sheet.column_dimensions['A'].width = 17
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 25
            sheet.column_dimensions['D'].width = 25
            sheet.column_dimensions['E'].width = 25
            sheet.column_dimensions['F'].width = 25
            sheet.column_dimensions['G'].width = 25
            sheet.column_dimensions['H'].width = 25
            sheet.column_dimensions['I'].width = 15
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 15
            sheet.column_dimensions['L'].width = 15

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.report_file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=reserved.product.report&download=true&field=report_file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }
