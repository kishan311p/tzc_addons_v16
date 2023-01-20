from odoo import models,fields,api,_
import base64
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from io import BytesIO
import os



class warning_spt_wizard(models.TransientModel):
    _name = "warning.spt.wizard"
    _description = 'Warning Wizard'

    message = fields.Text('Message')
    file = fields.Binary('File')
    file_name = fields.Char('File Name')

    success_partner_ids = fields.Many2many('res.partner','warning_spt_wizard_res_partner_rel_success','warning_spt_wizard_id','res_partner_id','Successfully Verified')
    failed_partner_ids = fields.Many2many('res.partner','warning_spt_wizard_res_partner_rel_failed','warning_spt_wizard_id','res_partner_id','Failed to Verify')
    verify_mail_success = fields.Text('Success Message')
    verify_mail_failed = fields.Text('Failed Message')

    def get_file(self):
        if self._context.get('image'):
            return self._image_product_report_xls(self._context.get('data'))
        else:
            if self._context.get('all_product'):
                active_id = self.id
                f = open('/tmp/All_Prodcts_Export.xlsx','rb')
                data = f.read()
                self.file = base64.b64encode(data)
                # wizard_id = wizard_obj.create({'file':base64.b64encode(data)})
                return {
                    'type': 'ir.actions.act_url',
                    'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, 'All Product Export'),
                    'target': 'self',
                }
            else:
                active_id = self.id
                f_name = 'Prodcts_Export'
                url='/tmp/Prodcts_Export.xlsx'
                f = open(url,'rb')
                data = f.read()
                self.file = base64.b64encode(data)
                return {
                    'type': 'ir.actions.act_url',
                    'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
                    'target': 'self',
                }

    def action_verify_email(self):
        f_name = 'Email Verification Failed'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Email Verification Failed", index=0)

        # sheet
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        table_header_font = Font(size=10, bold=True, name="Garamond")
        table_font = Font(size=10, bold=False, name="Garamond")
        alignment = Alignment(
            vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
        left_alignment = Alignment(
            vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        right_alignment = Alignment(
            vertical='center', horizontal='right', text_rotation=0, wrap_text=True)

        table_header_row = 1
        sheet.cell(row=table_header_row, column=1).value = 'Name'
        sheet.cell(row=table_header_row, column=2).value = 'Email'
        sheet.cell(row=table_header_row, column=3).value = 'Country'
        sheet.cell(row=table_header_row, column=4).value = 'Phone'
        sheet.cell(row=table_header_row, column=5).value = 'Territory'
        sheet.cell(row=table_header_row, column=6).value = 'B2B Flag'

        sheet.cell(row=table_header_row, column=1).font = table_header_font
        sheet.cell(row=table_header_row, column=2).font = table_header_font
        sheet.cell(row=table_header_row, column=3).font = table_header_font
        sheet.cell(row=table_header_row, column=4).font = table_header_font
        sheet.cell(row=table_header_row, column=5).font = table_header_font
        sheet.cell(row=table_header_row, column=6).font = table_header_font

        sheet.cell(row=table_header_row, column=1).alignment = left_alignment
        sheet.cell(row=table_header_row, column=2).alignment = left_alignment
        sheet.cell(row=table_header_row, column=3).alignment = left_alignment
        sheet.cell(row=table_header_row, column=4).alignment = left_alignment
        sheet.cell(row=table_header_row, column=5).alignment = left_alignment
        sheet.cell(row=table_header_row, column=6).alignment = left_alignment

        sheet.cell(row=table_header_row, column=1).border = top_bottom_border
        sheet.cell(row=table_header_row, column=2).border = top_bottom_border
        sheet.cell(row=table_header_row, column=3).border = top_bottom_border
        sheet.cell(row=table_header_row, column=4).border = top_bottom_border
        sheet.cell(row=table_header_row, column=5).border = top_bottom_border
        sheet.cell(row=table_header_row, column=6).border = top_bottom_border

        row_index = table_header_row+1
        for rec in self.failed_partner_ids:
            partner = self.env['res.partner'].browse(rec.id)
            col = 1 
            sheet.cell(row=row_index, column=1).value = partner.name or ""
            sheet.cell(row=row_index, column=1).font = table_font
            sheet.cell(row=row_index, column=1).alignment = left_alignment

            sheet.cell(row=row_index, column=2).value = partner.email or ""
            sheet.cell(row=row_index, column=2).font = table_font
            sheet.cell(row=row_index, column=2).alignment = left_alignment

            sheet.cell(row=row_index, column=3).value = partner.country_id.name or ""
            sheet.cell(row=row_index, column=3).font = table_font
            sheet.cell(row=row_index, column=3).alignment = left_alignment

            sheet.cell(row=row_index, column=4).value = partner.phone or ""
            sheet.cell(row=row_index, column=4).font = table_font
            sheet.cell(row=row_index, column=4).alignment = left_alignment
            
            sheet.cell(row=row_index, column=5).value = partner.territory.name or ""
            sheet.cell(row=row_index, column=5).font = table_font
            sheet.cell(row=row_index, column=5).alignment = left_alignment
            
            sheet.cell(row=row_index, column=6).value = dict(partner._fields['customer_type'].selection).get(partner.customer_type) or ""
            sheet.cell(row=row_index, column=6).font = table_font
            sheet.cell(row=row_index, column=6).alignment = left_alignment
            row_index += 1

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 10

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)

        active_id = self.id

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }

    def _image_product_report_xls(self,product_ids):

        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Products_Export_Temp.xlsm'
        wb = load_workbook(base_sample_file,read_only=False, keep_vba=True)
        wrksht = wb.active

        data_row = 2

        f_name = 'Products_Export'

        left_alignment = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        
        for product in product_ids:
            
            product = list(product)

            wrksht.row_dimensions[data_row].height = 220

            wrksht.cell(row=data_row, column=3).value = product[0] or ''
            wrksht.cell(row=data_row, column=3).alignment = left_alignment

            wrksht.cell(row=data_row, column=4).value = product[1] or ''
            wrksht.cell(row=data_row, column=4).alignment = left_alignment

            wrksht.cell(row=data_row, column=5).value = product[2]
            wrksht.cell(row=data_row, column=5).alignment = left_alignment

            wrksht.cell(row=data_row, column=6).value = product[3]
            wrksht.cell(row=data_row, column=6).alignment = left_alignment

            wrksht.cell(row=data_row, column=7).value = product[4]
            wrksht.cell(row=data_row, column=7).alignment = left_alignment

            wrksht.cell(row=data_row, column=8).value = product[5]
            wrksht.cell(row=data_row, column=8).alignment = left_alignment

            wrksht.cell(row=data_row, column=9).value = product[6]
            wrksht.cell(row=data_row, column=9).alignment = left_alignment

            wrksht.cell(row=data_row, column=10).value = product[7]
            wrksht.cell(row=data_row, column=10).alignment = left_alignment

            wrksht.cell(row=data_row, column=11).value = product[8]
            wrksht.cell(row=data_row, column=11).alignment = left_alignment

            wrksht.cell(row=data_row, column=12).value = product[9]
            wrksht.cell(row=data_row, column=12).alignment = left_alignment

            wrksht.cell(row=data_row, column=13).value = product[10]
            wrksht.cell(row=data_row, column=13).alignment = left_alignment

            wrksht.cell(row=data_row, column=14).value = product[11]
            wrksht.cell(row=data_row, column=14).alignment = left_alignment

            wrksht.cell(row=data_row, column=15).value = product[12]
            wrksht.cell(row=data_row, column=15).alignment = left_alignment
            
            wrksht.cell(row=data_row, column=16).value = product[13]
            wrksht.cell(row=data_row, column=16).alignment = left_alignment

            wrksht.cell(row=data_row, column=17).hyperlink = product[14]
            wrksht.cell(row=data_row, column=17).alignment = left_alignment

            wrksht.cell(row=data_row, column=18).hyperlink = product[15]
            wrksht.cell(row=data_row, column=18).alignment = left_alignment

            wrksht.cell(row=data_row, column=19).value = product[16]
            wrksht.cell(row=data_row, column=19).alignment = left_alignment

            wrksht.cell(row=data_row, column=20).value = product[17]
            wrksht.cell(row=data_row, column=20).alignment = left_alignment

            wrksht.cell(row=data_row, column=21).value = product[18]
            wrksht.cell(row=data_row, column=21).alignment = left_alignment

            wrksht.cell(row=data_row, column=22).value = product[19] or ''
            wrksht.cell(row=data_row, column=22).alignment = left_alignment

            wrksht.cell(row=data_row, column=23).value = product[20]
            wrksht.cell(row=data_row, column=23).alignment = left_alignment

            wrksht.cell(row=data_row, column=24).value = product[21]
            wrksht.cell(row=data_row, column=24).alignment = left_alignment

            wrksht.cell(row=data_row, column=25).value = product[22]
            wrksht.cell(row=data_row, column=25).alignment = left_alignment

            wrksht.cell(row=data_row, column=26).value = product[23]
            wrksht.cell(row=data_row, column=26).alignment = left_alignment

            wrksht.cell(row=data_row, column=27).value = product[24]
            wrksht.cell(row=data_row, column=27).alignment = left_alignment

            wrksht.cell(row=data_row, column=28).value = product[25]
            wrksht.cell(row=data_row, column=28).alignment = left_alignment

            wrksht.cell(row=data_row, column=29).value = product[26]
            wrksht.cell(row=data_row, column=29).alignment = left_alignment

            wrksht.cell(row=data_row, column=30).value = product[27]
            wrksht.cell(row=data_row, column=30).alignment = left_alignment

            wrksht.cell(row=data_row, column=31).value = product[28]
            wrksht.cell(row=data_row, column=31).alignment = left_alignment

            wrksht.cell(row=data_row, column=32).value = product[29]
            wrksht.cell(row=data_row, column=32).alignment = left_alignment

            wrksht.cell(row=data_row, column=33).value = product[30]
            wrksht.cell(row=data_row, column=33).alignment = left_alignment
            
            wrksht.cell(row=data_row, column=34).value = product[31]
            wrksht.cell(row=data_row, column=34).alignment = left_alignment

            wrksht.cell(row=data_row, column=35).value = product[32] or ''
            wrksht.cell(row=data_row, column=35).alignment = left_alignment

            wrksht.cell(row=data_row, column=36).value = product[33] or ''
            wrksht.cell(row=data_row, column=36).alignment = left_alignment

            wrksht.cell(row=data_row, column=37).value = product[34] or 0
            wrksht.cell(row=data_row, column=37).alignment = left_alignment

            wrksht.cell(row=data_row, column=38).value = product[35] or 0
            wrksht.cell(row=data_row, column=38).alignment = left_alignment

            wrksht.cell(row=data_row, column=39).value = product[36] or 0
            wrksht.cell(row=data_row, column=39).alignment = left_alignment

            wrksht.cell(row=data_row, column=40).value = product[37] or 0
            wrksht.cell(row=data_row, column=40).alignment = left_alignment

            data_row += 1

            wrksht.column_dimensions['C'].width = 20
        
        wrksht.column_dimensions['C'].width = 25
        wrksht.column_dimensions['D'].width = 30
        wrksht.column_dimensions['E'].width = 15
        wrksht.column_dimensions['F'].width = 10
        wrksht.column_dimensions['G'].width = 15
        wrksht.column_dimensions['H'].width = 10
        wrksht.column_dimensions['I'].width = 10
        wrksht.column_dimensions['J'].width = 10
        wrksht.column_dimensions['K'].width = 10
        wrksht.column_dimensions['L'].width = 10
        wrksht.column_dimensions['M'].width = 10
        wrksht.column_dimensions['N'].width = 20
        wrksht.column_dimensions['O'].width = 10
        wrksht.column_dimensions['P'].width = 10
        wrksht.column_dimensions['Q'].width = 25
        wrksht.column_dimensions['R'].width = 25
        wrksht.column_dimensions['S'].width = 10
        wrksht.column_dimensions['T'].width = 15
        wrksht.column_dimensions['U'].width = 15
        wrksht.column_dimensions['V'].width = 15
        wrksht.column_dimensions['W'].width = 15
        wrksht.column_dimensions['X'].width = 20
        wrksht.column_dimensions['Y'].width = 10
        wrksht.column_dimensions['Z'].width = 10
        wrksht.column_dimensions['AA'].width = 10
        wrksht.column_dimensions['AB'].width = 10
        wrksht.column_dimensions['AC'].width = 15
        wrksht.column_dimensions['AD'].width = 10
        wrksht.column_dimensions['AE'].width = 10
        wrksht.column_dimensions['AF'].width = 10
        wrksht.column_dimensions['AG'].width = 15
        wrksht.column_dimensions['AH'].width = 15
        wrksht.column_dimensions['AI'].width = 15
        wrksht.column_dimensions['AJ'].width = 15
        wrksht.column_dimensions['AK'].width = 15
        wrksht.column_dimensions['AL'].width = 15
        wrksht.column_dimensions['AM'].width = 15
        wrksht.column_dimensions['AN'].width = 15
        
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsm' % (self.id, f_name),
            'target': 'self',
        }
