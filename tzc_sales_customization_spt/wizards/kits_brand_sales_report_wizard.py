from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
from datetime import datetime
from dateutil import tz

class kits_brand_sales_report_wizard(models.TransientModel):
    _name = "kits.brand.sales.report.wizard"
    _description = "Kits Brand Sales Report Wizard"

    start_date = fields.Date("Start Date")
    end_date = fields.Date("End Date")
    brand = fields.Many2many('product.brand.spt','brand_product_brand_spt_rel','brand_id','product_brand_spt_id','Brand')
    country_ids = fields.Many2many('res.country','kits_brand_sales_country_rel','wizard_id','country_id','Country')
    file = fields.Binary('File')

    def validate_dates(self):
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise UserError(_("Start Date should be lesser than End Date."))

    def brand_sales_report(self):
        self.validate_dates()
        product_obj = self.env['product.product']
        cad_rate = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.on_sale_cad_spt','0'))

        f_name = 'Brand Sales Report'
        workbook = Workbook()
        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        right_alignment = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
        align_top = Alignment(vertical='top', horizontal='general', text_rotation=0, wrap_text=True, indent=0)
        query = """select 
                COALESCE(pb.name,'') as "Brand",
                COALESCE(pms.name,'') as "Model",
                COALESCE(pp.default_code,'') as "SKU",
                sum(COALESCE((select sum(sml.qty_done) from stock_move_line sml
                left join stock_move sm on sml.move_id = sm.id
                where sm.sale_line_id = sol.id
                ),0)) as "Quantity",
                COALESCE(avg(sol.unit_discount_price),0) as "Price",
                COALESCE(cur.name,'USD') as "Currency"
            from sale_order_line sol
                left join sale_order so on sol.order_id=so.id
                left join res_partner rp on so.partner_id=rp.id
                left join res_country rc on rp.country_id=rc.id
                left join product_product pp on sol.product_id=pp.id
                left join product_template pt on pp.product_tmpl_id=pt.id
                left join product_brand_spt pb on pp.brand=pb.id 
                left join res_currency cur on sol.currency_id=cur.id
                left join product_model_spt pms on pp.model = pms.id
                %s
            group by pb.name,pp.default_code,pms.name,"Currency"
            order by pb.name,pp.default_code
        """
        params = ['where so.state in %s'%(str(tuple(['shipped','draft_inv','open_inv','paid']))),' and sol.qty_delivered > 0']
        if self.start_date:
            params.append('and so.date_order::date >= \'%s\''%(str(self.start_date)))
        if self.end_date:
            params.append('and so.date_order::date <= \'%s\''%(str(self.end_date)))
        if self.brand:
            params.append('and %s'%('pp.brand in {}'.format(str(self.brand.ids).replace('[','(').replace(']',')')) if len(self.brand) else ''))
        if self.country_ids:
            params.append('and rp.country_id in {}'.format(str(self.country_ids.ids).replace('[','(').replace(']',')')))
        exec_query = query % (' '.join(params))
        self._cr.execute(exec_query)
        result = self._cr.fetchall()
        brands = []
        list(map(lambda x: brands.append(x[0]) if x[0] not in brands else None,result))

        if len(result):
            index = 0
            for data in brands:
                table_header = 1
                sheet = workbook.create_sheet(title=data, index=index)
                sheet.cell(row=table_header, column=1).value = 'Brand'
                sheet.cell(row=table_header, column=2).value = 'Model'
                sheet.cell(row=table_header, column=3).value = 'SKU'
                sheet.cell(row=table_header, column=4).value = 'Color Code'
                sheet.cell(row=table_header, column=5).value = 'Quantity'
                sheet.cell(row=table_header, column=6).value = 'Average Price'
                sheet.cell(row=table_header, column=7).value = 'Unit Price'
                sheet.cell(row=table_header, column=8).value = 'Subtotal'

                sheet.merge_cells('J1:M1')
                sheet.cell(row=1, column=10).value = 'Criteria'

                sheet.merge_cells('J2:M2')
                sheet.cell(row=2,column=10).value = 'Start Date : %s' %(str(self.start_date) if self.start_date else '') 
                
                sheet.merge_cells('J3:M3')
                sheet.cell(row=3,column=10).value = 'End Date : %s' %(str(self.end_date) if self.end_date else '')

                sheet.cell(row=4,column=10).value = 'Brands : '
                sheet.merge_cells('K4:M6')
                sheet.cell(row=4,column=11).value = ', '.join(self.brand.mapped('name')) if self.brand else ''
                sheet.cell(row=4, column=11).alignment = align_top

                sheet.cell(row=7,column=10).value = 'Countries : ' 
                sheet.merge_cells('K7:M9')
                sheet.cell(row=7,column=11).value = ', '.join(self.country_ids.mapped('name')) if self.country_ids else '' 
                sheet.cell(row=7, column=11).alignment = align_top

                for col in range(1, 11):
                    sheet.cell(row=table_header, column=col).border = top_bottom_border
                    sheet.cell(row=table_header, column=col).font = header_font

                sheet.column_dimensions['A'].width = 15
                sheet.column_dimensions['B'].width = 15
                sheet.column_dimensions['C'].width = 30
                sheet.column_dimensions['D'].width = 25
                sheet.column_dimensions['E'].width = 13
                sheet.column_dimensions['F'].width = 15
                sheet.column_dimensions['G'].width = 13
                sheet.column_dimensions['J'].width = 13
                sheet.column_dimensions['K'].width = 13

                row_index = table_header+1
                total_qty = 0.0
                total_price = 0.0
                print_dict = {}
                for result_line in [r for r in result if r[0] == data]:
                    price = result_line[4]
                    if result_line[5] == 'CAD':
                        try:
                            price = result_line[4]/cad_rate
                        except:
                            price = 0
                    subtotal = round(price * result_line[3],2)

                    total_qty += result_line[3]
                    
                    if not result_line[2] in print_dict:
                        print_dict[result_line[2]] = {'brand':result_line[0],'model':result_line[1],'sku':result_line[2],'qty':result_line[3],'price':price}
                    else:
                        print_dict[result_line[2]]['price'] = (print_dict[result_line[2]]['price']+price)/2
                        print_dict[result_line[2]]['qty'] = print_dict[result_line[2]]['qty']+result_line[3]

                for print_line in print_dict:
                    product_sku = self.env['product.product'].search([('active','in',(True,False)),('default_code','=',print_dict[print_line]['sku'])],limit=1)
                    sheet.cell(row=row_index, column=1).value = print_dict[print_line]['brand']
                    sheet.cell(row=row_index, column=2).value = print_dict[print_line]['model']
                    sheet.cell(row=row_index, column=3).value = print_dict[print_line]['sku']
                    sheet.cell(row=row_index, column=4).value = product_sku.manufacture_color_code
                    sheet.cell(row=row_index, column=5).value = print_dict[print_line]['qty']
                    sheet.cell(row=row_index, column=6).value = '{:,.2f}'.format(print_dict[print_line]['price'])
                    sheet.cell(row=row_index, column=6).alignment = right_alignment

                    sub_total = round(print_dict[print_line]['price']*print_dict[print_line]['qty'],2)
                    product_price = product_sku.lst_price_usd
                    if product_sku.sale_type == 'on_sale':
                        product_price = product_sku.on_sale_usd
                    if product_sku.sale_type == 'clearance':
                        product_price = product_sku.clearance_usd
                    sheet.cell(row=row_index, column=7).value = '{:,.2f}'.format(product_price)
                    sheet.cell(row=row_index, column=7).alignment = right_alignment
                    sheet.cell(row=row_index, column=8).value = '{:,.2f}'.format(sub_total)
                    sheet.cell(row=row_index, column=8).alignment = right_alignment
                    row_index += 1
                    total_price += sub_total
                index += 1

                sheet.cell(row=row_index, column=4).value = "Total"
                sheet.cell(row=row_index, column=4).font = header_font
                sheet.cell(row=row_index, column=4).alignment = right_alignment

                sheet.cell(row=row_index, column=5).value = total_qty
                sheet.cell(row=row_index, column=5).font = header_font

                sheet.cell(row=row_index, column=8).value = '{:,.2f}'.format(total_price)
                sheet.cell(row=row_index, column=8).font = header_font
                sheet.cell(row=row_index, column=8).alignment = right_alignment

                sheet.cell(row=row_index, column=9).value = "(USD)"
                sheet.cell(row=row_index, column=9).font = header_font
                sheet.cell(row=row_index, column=9).alignment = right_alignment

            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            self.file = base64.b64encode(data)
            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=kits.brand.sales.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,f_name),
                'target': 'self',
            }
        else:
            if self.start_date and not self.end_date:
                raise UserError(f'There are no order after {self.start_date.strftime("%m/%d/%Y")} date.')
            elif not self.start_date and self.end_date:
                raise UserError(f'There are no order before {self.end_date.strftime("%m/%d/%Y")} date.')
            elif self.start_date and self.end_date:
                raise UserError(f'There are no order between {self.start_date.strftime("%m/%d/%Y")} / {self.end_date.strftime("%m/%d/%Y")} date.')
            else:
                raise UserError(_('There are no orders.'))

    def specific_brand_excel_report(self):
        self.validate_dates()
        active_id = self.id
        f_name = 'Brand Country Report'
        workbook = Workbook()
        
        state_list = ['when so.state = \'{}\' then \'{}\''.format(status[0],status[1]) for status in self.env['sale.order']._fields['state'].selection]
        state = '(select case %s end) as "Order Status"'%('\n'.join(state_list))
        sale_type_list = ['when pp.sale_type = \'{}\' then \'{}\' '.format(sale_type[0],sale_type[1]) for sale_type in self.env['product.product']._fields['sale_type'].selection]
        sale_type_list.append('else \'Regular\'')
        sale_type = '(select case %s end) as "Sale Type"'%('\n'.join(sale_type_list))
        query = """select 
                    COALESCE(pp.default_code,'') as "SKU",
                    COALESCE(pb.name||' '||pms.name||' '||(SELECT SPLIT_PART(pp.manufacture_color_code,'-',1))||' '||COALESCE(pp.eye_size_compute::varchar,'00')|| ' ' ||COALESCE(pbss.name,'00'),'') || ' ' ||COALESCE(ptss.name,'00'),'' as "Product Name",
                    COALESCE(so.name,'') as "Order#",
                    so.date_order as "Order Date",
                    COALESCE(pms.name,'') as "Model",
                    COALESCE(pb.name,'') as "Brand",
                    %s,
                    %s,
                    COALESCE(sol.qty_delivered,0.0) as "Delivered Qty",
                    COALESCE(sol.discount::float,0.0) as "Discount",
                    COALESCE(sol.unit_discount_price::float,0.0) as "Unit Price",
                    COALESCE((sol.qty_delivered * sol.unit_discount_price)::float,0.0) as "Subtotal Amount",
                    COALESCE(cur.name,'USD') as "Currency",
                    COALESCE(rc.name->>'en_US','') as "country"

                from sale_order_line sol 
                    left join sale_order so on sol.order_id=so.id
                    left join res_partner rp on so.partner_id=rp.id
                    left join res_country rc on rp.country_id=rc.id
                    left join product_product pp on sol.product_id=pp.id
                    left join product_template pt on pp.product_tmpl_id=pt.id
                    left join product_brand_spt pb on pp.brand=pb.id 
                    left join res_currency cur on sol.currency_id=cur.id
                    left join product_model_spt pms on pp.model = pms.id
                    left join product_temple_size_spt ptss on pp.temple_size = ptss.id
                    left join product_bridge_size_spt pbss on pp.bridge_size = pbss.id
                     %s 
                    order by pb.name,pp.default_code,so.date_order"""
        params = ['where so.state in %s'%(str(tuple(['shipped','draft_inv','open_inv','paid']))),' and sol.qty_delivered > 0']
        start_date = ''
        end_date = ''
        from_tz, to_tz = (tz.gettz(datetime.now().tzinfo)), (tz.gettz(
            self.env.user.tz) or self.env.context.get('tz'))
        if self.start_date:
            start_date = self.start_date
        if self.end_date:
            end_date = self.end_date
        if start_date:
            params.append(" AND so.date_order >= '%s' " % str(start_date))
        if end_date:
            params.append("AND so.date_order <= '%s' " % str(end_date))
        if self.country_ids:
            params.append("AND rc.id in (%s) " % str(','.join(self.country_ids.mapped(lambda x : str(x.id)))))
        if self.brand:
            params.append("AND pb.id in (%s)" % str(','.join(self.brand.mapped(lambda x : str(x.id)))))
        q_vals = [state,sale_type,''.join(params)]
        print(query % tuple(q_vals))
        exec_q = query % tuple(q_vals)
        self.env.cr.execute(exec_q)
        report_data = self.env.cr.fetchall()
        brands = []
        list(map(lambda x: brands.append(x[5]) if x[5] not in brands else None,report_data))
        
        cad_rate = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.on_sale_cad_spt','0'))
        index = 0
        for brand in brands:
            sheet = workbook.create_sheet(title=brand.replace('/','-'), index=index)
            bd = Side(style='thin', color="000000")
            bottom_border = Border(bottom=bd)
            top_bottom_border = Border(top=bd, bottom=bd)
            heading_font = Font(name="Garamond", size="11", bold=True)
            table_font = Font(name="Garamond", size="10", bold=False)
            align_left = Alignment(vertical="center", horizontal='left', text_rotation=0, wrap_text=True)
            align_right = Alignment(vertical="center", horizontal='right', text_rotation=0, wrap_text=True)
            align_top = Alignment(vertical='top', horizontal='general', text_rotation=0, wrap_text=True, indent=0)

            # =============================== info table ===============================
            table_header_row = 1
            sheet.row_dimensions[table_header_row].height = 30
            sheet.cell(row=table_header_row, column=1).value = 'SKU'
            sheet.cell(row=table_header_row, column=2).value = "Product Name"
            sheet.cell(row=table_header_row, column=3).value = "Order#"
            sheet.cell(row=table_header_row, column=4).value = "Order Date"
            sheet.cell(row=table_header_row, column=5).value =  "Model"
            sheet.cell(row=table_header_row, column=6).value =  "Order Status"
            sheet.cell(row=table_header_row, column=7).value =  "Sale Type"
            sheet.cell(row=table_header_row, column=8).value =  "Country"
            sheet.cell(row=table_header_row, column=9).value =  "Delivered Qty"
            sheet.cell(row=table_header_row, column=10).value =  "Discount"
            sheet.cell(row=table_header_row, column=11).value =  "Our Price"
            sheet.cell(row=table_header_row, column=12).value =  "Subtotal Amount"

            sheet.merge_cells('O1:R1')
            sheet.cell(row=1, column=15).value = 'Criteria'
            sheet.cell(row=1, column=15).alignment = align_left

            sheet.merge_cells('O2:R2')
            sheet.cell(row=2,column=15).value = 'Start Date : %s' %(str(self.start_date) if self.start_date else '') 
            sheet.cell(row=2, column=15).alignment = align_left
            
            sheet.merge_cells('O3:R3')
            sheet.cell(row=3,column=15).value = 'End Date : %s' %(str(self.end_date) if self.end_date else '')
            sheet.cell(row=3, column=15).alignment = align_left

            sheet.cell(row=4,column=15).value = 'Brands : '
            sheet.cell(row=4, column=15).alignment = align_left
            sheet.merge_cells('P4:R6')
            sheet.cell(row=4,column=16).value = ', '.join(self.brand.mapped('name')) if self.brand else ''
            sheet.cell(row=4, column=16).alignment = align_top

            sheet.cell(row=7,column=15).value = 'Countries : ' 
            sheet.cell(row=7, column=15).alignment = align_left
            sheet.merge_cells('P7:R9')
            sheet.cell(row=7,column=16).value = ', '.join(self.country_ids.mapped('name')) if self.country_ids else '' 
            sheet.cell(row=7,column=16).alignment = align_top
            
            sheet.cell(row=table_header_row, column=1).alignment = align_left
            sheet.cell(row=table_header_row, column=2).alignment = align_left
            sheet.cell(row=table_header_row, column=3).alignment = align_left
            sheet.cell(row=table_header_row, column=4).alignment = align_left
            sheet.cell(row=table_header_row, column=5).alignment = align_left
            sheet.cell(row=table_header_row, column=6).alignment = align_left
            sheet.cell(row=table_header_row, column=7).alignment = align_left
            sheet.cell(row=table_header_row, column=8).alignment = align_left
            sheet.cell(row=table_header_row, column=9).alignment = align_left
            sheet.cell(row=table_header_row, column=10).alignment = align_left
            sheet.cell(row=table_header_row, column=11).alignment = align_right
            sheet.cell(row=table_header_row, column=12).alignment = align_right
            for col in range(1, 16):
                sheet.cell(row=table_header_row, column=col).border = top_bottom_border
                sheet.cell(row=table_header_row, column=col).font = heading_font
            
            row_index = table_header_row+1
            final_total = 0.00
            for data in [r for r in report_data if r[5] == brand]:
                # try:
                #     height = (3*len(data[1]))/2
                #     sheet.row_dimensions[row_index].height = height if len(data[1]) > 20 else 25
                # except:
                #     pass

                price = round(data[10],2)
                subtotal = data[11]
                if data[11] == 'CAD':
                    try:
                        price = round(data[10]/cad_rate,1)
                        subtotal = round(price*data[8],2)
                    except:
                        pass
                final_total += subtotal

                sheet.cell(row=row_index,column=1).value = data[0] or ''
                sheet.cell(row=row_index, column=2).value = data[1] or ''
                sheet.cell(row=row_index, column=3).value = data[2] or ''
                sheet.cell(row=row_index, column=4).value = data[3] or ''
                sheet.cell(row=row_index, column=5).value = data[4] or  ''
                sheet.cell(row=row_index, column=6).value = data[6] or ''
                sheet.cell(row=row_index, column=7).value = data[7] or ''
                sheet.cell(row=row_index, column=8).value = data[13] or ''
                sheet.cell(row=row_index, column=9).value = data[8] or 0.0
                sheet.cell(row=row_index, column=10).value = data[9] or 0.0
                sheet.cell(row=row_index, column=11).value = price
                sheet.cell(row=row_index, column=12).value = subtotal
                
                sheet.cell(row=row_index, column=1).border = bottom_border
                sheet.cell(row=row_index, column=1).font = table_font
                sheet.cell(row=row_index, column=2).border = bottom_border
                sheet.cell(row=row_index, column=2).font = table_font
                sheet.cell(row=row_index, column=3).border = bottom_border
                sheet.cell(row=row_index, column=3).font = table_font
                sheet.cell(row=row_index, column=4).border = bottom_border
                sheet.cell(row=row_index, column=4).font = table_font
                sheet.cell(row=row_index, column=5).border = bottom_border
                sheet.cell(row=row_index, column=5).font = table_font
                sheet.cell(row=row_index, column=6).border = bottom_border
                sheet.cell(row=row_index, column=6).font = table_font
                sheet.cell(row=row_index, column=7).border = bottom_border
                sheet.cell(row=row_index, column=7).font = table_font
                sheet.cell(row=row_index, column=8).border = bottom_border
                sheet.cell(row=row_index, column=8).font = table_font
                sheet.cell(row=row_index, column=9).border = bottom_border
                sheet.cell(row=row_index, column=9).font = table_font
                sheet.cell(row=row_index, column=10).border = bottom_border
                sheet.cell(row=row_index, column=10).font = table_font
                sheet.cell(row=row_index, column=11).border = bottom_border
                sheet.cell(row=row_index, column=11).font = table_font
                sheet.cell(row=row_index, column=12).border = bottom_border
                sheet.cell(row=row_index, column=12).font = table_font

                sheet.cell(row=row_index, column=1).alignment = align_left
                sheet.cell(row=row_index, column=2).alignment = align_left
                sheet.cell(row=row_index, column=3).alignment = align_left
                sheet.cell(row=row_index, column=4).alignment = align_left
                sheet.cell(row=row_index, column=5).alignment = align_left
                sheet.cell(row=row_index, column=6).alignment = align_left
                sheet.cell(row=row_index, column=7).alignment = align_left
                sheet.cell(row=row_index, column=8).alignment = align_left
                sheet.cell(row=row_index, column=9).alignment = align_left
                sheet.cell(row=row_index, column=10).alignment = align_left
                sheet.cell(row=row_index, column=11).alignment = align_right
                sheet.cell(row=row_index, column=12).alignment = align_right
                subtotal = round(float(data[10]),2) if data[10] else 0.00
                
                row_index += 1

                if not report_data:
                    sheet.merge_cells("A"+str(row_index)+":E"+str(row_index))
                    sheet.cell(row=row_index, column=1).value = 'There is no orderes between date ' + \
                        str(self.start_date)+' and '+str(self.end_date)+' !'
                    sheet.cell(row=row_index, column=1).font = table_font

            if report_data:
                row_final_total = row_index
                sheet.row_dimensions[row_final_total].height = 20
                sheet.cell(row=row_final_total, column=11).value = "Final Total :"
                sheet.cell(row=row_final_total, column=12).value= round(final_total,2)
                sheet.cell(row=row_final_total, column=13).value= "(USD)"
        
                sheet.cell(row=row_final_total, column=11).border = top_bottom_border
                sheet.cell(row=row_final_total, column=11).font = table_font
                sheet.cell(row=row_final_total, column=11).alignment = align_right
                sheet.cell(row=row_final_total, column=12).border = top_bottom_border
                sheet.cell(row=row_final_total, column=12).font = table_font
                sheet.cell(row=row_final_total, column=12).alignment = align_right
                sheet.cell(row=row_final_total, column=13).border = top_bottom_border
                sheet.cell(row=row_final_total, column=13).font = table_font
                sheet.cell(row=row_final_total, column=13).alignment = align_right

            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
            sheet.column_dimensions['H'].width = 15
            sheet.column_dimensions['I'].width = 15
            sheet.column_dimensions['J'].width = 15
            sheet.column_dimensions['K'].width = 15
            sheet.column_dimensions['L'].width = 15
            sheet.column_dimensions['O'].width = 13
            index +=1 

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=kits.brand.sales.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }
