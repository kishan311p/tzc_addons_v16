from odoo import _, api, fields, models, tools
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO,StringIO
import openpyxl
import os


class order_report_with_image_wizard(models.TransientModel):
    _name = 'order.report.with.image.wizard'
    _description = 'Order Report With Image'

    with_img = fields.Boolean('With Image ?')
    order_id = fields.Many2one('sale.order','Order')
    report_file = fields.Binary()

    def action_process_report(self):
        state = ""
        if self.order_id.state in ('draft','sent'):
            state = "Quotation"
        elif self.order_id.state not in ('draft','sent'):
            state = "Sales Order"

        for_order = ''
        if state == 'Sales Order':
            for_order = state.split(' ')[1].strip()

        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Order_templ.xlsm'
        wb = load_workbook(base_sample_file,read_only=False, keep_vba=True)
        wrksht = wb.active

        # workbook = Workbook()
        # sheet = workbook.create_sheet(title='%s - %s'%(for_order if for_order else state,self.order_id.name), index=0)
        base_sample_file_xlsx = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/order_sample_file.xlsx'
        wb_xlsx = load_workbook(base_sample_file_xlsx,read_only=False, keep_vba=False)
        sheet = wb_xlsx.active

        # f_name = '%s - %s'%(for_order if for_order else state,self.order_id.name)

        wrk_sheet = wrksht if self.with_img else sheet

        bd = Side(style='thin', color="000000")
        grey_bd = Side(style='thin', color="D3D3D3")
        all_border = Border(left=bd, top=bd, right=bd, bottom=bd)
        address_font = Font(name='Calibri', size=10, bold=False)
        table_header_font = Font(name='Calibri', size=11, bold=True)
        header_font = Font(name='Calibri', size=16, bold=True)
        name_header_font = Font(name="Calibri", size=10, bold=True)
        right_border = Border(right=bd)
        bottom_border = Border(bottom=bd)
        top_border = Border(top=bd)
        top_bottom_border = Border(top=bd, bottom=bd)
        alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0)
        alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0)
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
        alignment_wrapper = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)
        address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
        table_font = Font(size=11, bold=False, name="Calibri")

        # ------------------------------------------ Image ------------------------------------------
        # if not self.with_img:
        #     company_header_row = 1
        #     company_header_end = company_header_row+5
        #     width = 0
        #     height = 0
        #     if self.with_img:
        #         width = 500
        #         height = 60
        #         wrk_sheet.merge_cells('A'+str(company_header_row) + ':F'+str(company_header_end))
        #     else:
        #         width = 300
        #         height = 45
        #         wrk_sheet.merge_cells('A'+str(company_header_row) + ':E'+str(company_header_end))
        #     img = BytesIO()
        #     img.flush()
        #     img.write(base64.b64decode(self.env.companies[0].logo))
        #     try:
        #         image = openpyxl.drawing.image.Image(img)
        #         image.width = width
        #         image.height = height
        #         wrk_sheet.add_image(image, 'A'+str(company_header_row+1))
        #     except:
        #         pass
        # end_point = 11 if self.with_img else 10
            # for j in range(1, end_point):
            #     wrk_sheet.cell(row=company_header_end, column=j).border = Border(bottom=bd)
        # ------------------------------------ image end ------------------------------------
        add_col = 7 if self.with_img else 6
        # if self.with_img:
        #     wrk_sheet.merge_cells("G"+str(company_header_row) + ":J"+str(company_header_end))
        # else:
        # if not self.with_img:
        #     wrk_sheet.merge_cells("F"+str(company_header_row) + ":I"+str(company_header_end))
            
        # company_id = self.env.user.company_id
        # company_address = company_id.name+'\n'+'(A division of Tanzacan Tradelink Inc.)'+'\n'+self.order_id.create_address_line_for_sale(company_id)
        # wrk_sheet.cell(row=1, column=add_col).value = company_address
        # wrk_sheet.cell(row=1, column=add_col).alignment = address_alignment
        # wrk_sheet.cell(row=1, column=add_col).font = Font(name="Calibri", size=11, bold=False)
        # --------------------------------------------------------- Header Address end  ---------------------------------------------------------
        # ------------------------------------------------------------
        # Billing Address
        # ------------------------------------------------------------
        address_row = 9
        if not self.with_img:
            wrk_sheet.merge_cells("A"+str(address_row-1)+":D" + str(address_row-1))
            wrk_sheet.cell(row=address_row-1, column=1).value = "Billing Address:"
            wrk_sheet.cell(row=address_row-1, column=1).font = Font('Calibri',size=10, bold=True)

        wrk_sheet.merge_cells("A"+str(address_row)+":D" + str(address_row+6))  # added
        billing_address = self.env['sale.order'].create_address_line_for_sale(self.order_id.partner_id, take_name=True)
        wrk_sheet.cell(row=address_row, column=1).value = billing_address
        wrk_sheet.cell(row=address_row, column=1).alignment = address_alignment
        wrk_sheet.cell(row=address_row, column=1).font = address_font
        wrk_sheet.cell(row=address_row, column=1).font = address_font
        # -------------------------------------------------------------
        # Shipping Address
        # -------------------------------------------------------------
        if self.with_img:
            wrk_sheet.merge_cells("G"+str(address_row-1)+":J" + str(address_row-1))
        else:
            wrk_sheet.merge_cells("F"+str(address_row-1)+":I" + str(address_row-1))

        wrk_sheet.cell(row=address_row-1, column=add_col).value = "Shipping Address:"
        wrk_sheet.cell(row=address_row-1, column=add_col).font = Font('Calibri',size=10, bold=True)

        if self.with_img:
            wrk_sheet.merge_cells('G'+str(address_row)+':J'+str(address_row+6))
        else:
            wrk_sheet.merge_cells('F'+str(address_row)+':I'+str(address_row+6))

        shipping_address = self.env['sale.order'].create_address_line_for_sale(self.order_id.partner_shipping_id, take_name=True)
        wrk_sheet.cell(row=address_row, column=add_col).value = shipping_address
        wrk_sheet.cell(row=address_row, column=add_col).alignment = address_alignment
        wrk_sheet.cell(row=address_row, column=add_col).font = address_font
        # ---------------------------------------------- Name ----------------------------------------------
        name_row = address_row + 8
        wrk_sheet.merge_cells('A'+str(name_row)+':C'+str(name_row+1))
        wrk_sheet.cell(row=name_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        # wrk_sheet.cell(row=name_row, column=1).value = str("Quotation #"+self.order_id.name if self.order_id.name else 'Quotation ')
        wrk_sheet.cell(row=name_row, column=1).value = state + ' %s'%('# ' if self.order_id.name else '') + self.order_id.name
        wrk_sheet.cell(row=name_row, column=1).font = Font(name='Calibri', size=12, bold=True,color="666666")

        # ---------------------------------------------- date, salesperson, total qty ----------------------------------------------
        date_person_row = name_row + 3
        wrk_sheet.row_dimensions[date_person_row].height = 25
        wrk_sheet.row_dimensions[date_person_row+1].height = 25

        
        wrk_sheet.merge_cells("A"+str(date_person_row)+":A"+str(date_person_row+1))
        wrk_sheet.cell(row=date_person_row, column=1).value = str("%s Date:\n"%(for_order if for_order else "Quotation")+str(self.order_id.date_order.date() if self.order_id.date_order else ''))
        # sheet.cell(row=date_person_row, column=1).value = str("%s Date:\n"%(for_order if for_order else "Quotation")+str(self.order_id.date_order.strftime('%d/%m/%Y') if self.order_id.date_order else ''))
        # sheet.cell(row=date_person_row, column=1).value = str("Quotation Date:\n"+str(self.order_id.date_order.strftime('%d/%m/%Y') if self.order_id.date_order else ''))
        wrk_sheet.cell(row=date_person_row, column=1).font = name_header_font
        wrk_sheet.cell(row=date_person_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrk_sheet.merge_cells("B"+str(date_person_row)+":C"+str(date_person_row+1))
        wrk_sheet.cell(row=date_person_row, column=2).value = str("Salesperson:\n"+self.order_id.user_id.name or '')
        wrk_sheet.cell(row=date_person_row, column=2).font = name_header_font
        wrk_sheet.cell(row=date_person_row, column=2).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrk_sheet.merge_cells("D"+str(date_person_row)+":D"+str(date_person_row+1))
        wrk_sheet.cell(row=date_person_row, column=4).value = str("Total Qty:\n"+str(self.order_id.ordered_qty if self.order_id.ordered_qty else '') or '')
        wrk_sheet.cell(row=date_person_row, column=4).font = name_header_font
        wrk_sheet.cell(row=date_person_row, column=4).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        # ========================= Product Table ===========================
        table_header = date_person_row+3
        if self.with_img:
            wrk_sheet.merge_cells("A"+str(table_header)+":D"+str(table_header))
            wrk_sheet.cell(row=table_header, column=1).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=1).alignment = alignment
            wrk_sheet.cell(row=table_header, column=1).font = table_header_font
            wrk_sheet.cell(row=table_header, column=2).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=3).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=4).border = top_bottom_border

            wrk_sheet.cell(row=table_header, column=5).value = 'Product'
            wrk_sheet.cell(row=table_header, column=5).alignment = alignment
            wrk_sheet.cell(row=table_header, column=5).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=5).font = table_header_font

            wrk_sheet.cell(row=table_header, column=6).value = 'Cat'
            wrk_sheet.cell(row=table_header, column=6).alignment = alignment
            wrk_sheet.cell(row=table_header, column=6).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=6).font = table_header_font

            wrk_sheet.cell(row=table_header, column=7).value = 'Qty'
            wrk_sheet.cell(row=table_header, column=7).alignment = alignment
            wrk_sheet.cell(row=table_header, column=7).font = table_header_font
            wrk_sheet.cell(row=table_header, column=7).border = top_bottom_border

            wrk_sheet.cell(row=table_header, column=8).value = 'Disc.%'
            wrk_sheet.cell(row=table_header, column=8).alignment = alignment
            wrk_sheet.cell(row=table_header, column=8).font = table_header_font
            wrk_sheet.cell(row=table_header, column=8).border = top_bottom_border

            wrk_sheet.cell(row=table_header, column=9).value = 'Price'
            wrk_sheet.cell(row=table_header, column=9).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=9).alignment = alignment
            wrk_sheet.cell(row=table_header, column=9).font = table_header_font

            wrk_sheet.cell(row=table_header, column=10).value = 'Subtotal'
            wrk_sheet.cell(row=table_header, column=10).alignment = alignment
            wrk_sheet.cell(row=table_header, column=10).font = table_header_font
            wrk_sheet.cell(row=table_header, column=10).border = top_bottom_border
            
        else:
            wrk_sheet.cell(row=table_header, column=1).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=1).alignment = alignment
            wrk_sheet.cell(row=table_header, column=1).font = table_header_font
            wrk_sheet.cell(row=table_header, column=2).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=3).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=4).border = top_bottom_border

            wrk_sheet.merge_cells("A"+str(table_header)+":D"+str(table_header))
            wrk_sheet.cell(row=table_header, column=1).value = 'Product'
            wrk_sheet.cell(row=table_header, column=1).alignment = alignment
            wrk_sheet.cell(row=table_header, column=1).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=1).font = table_header_font

            wrk_sheet.cell(row=table_header, column=5).value = 'Cat'
            wrk_sheet.cell(row=table_header, column=5).alignment = alignment
            wrk_sheet.cell(row=table_header, column=5).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=5).font = table_header_font

            wrk_sheet.cell(row=table_header, column=6).value = 'Qty'
            wrk_sheet.cell(row=table_header, column=6).alignment = alignment
            wrk_sheet.cell(row=table_header, column=6).font = table_header_font
            wrk_sheet.cell(row=table_header, column=6).border = top_bottom_border

            wrk_sheet.cell(row=table_header, column=7).value = 'Disc.%'
            wrk_sheet.cell(row=table_header, column=7).alignment = alignment
            wrk_sheet.cell(row=table_header, column=7).font = table_header_font
            wrk_sheet.cell(row=table_header, column=7).border = top_bottom_border

            wrk_sheet.cell(row=table_header, column=8).value = 'Price'
            wrk_sheet.cell(row=table_header, column=8).border = top_bottom_border
            wrk_sheet.cell(row=table_header, column=8).alignment = alignment
            wrk_sheet.cell(row=table_header, column=8).font = table_header_font

            wrk_sheet.cell(row=table_header, column=9).value = 'Subtotal'
            wrk_sheet.cell(row=table_header, column=9).alignment = alignment
            wrk_sheet.cell(row=table_header, column=9).font = table_header_font
            wrk_sheet.cell(row=table_header, column=9).border = top_bottom_border

        row_index = table_header+1

        # ----------------------------- Table Data -----------------------------
        query = f'''SELECT 	COALESCE(PBS.NAME,' ') AS BRAND,
                            COALESCE(PMS.NAME,' ') AS MODEL,
                            COALESCE(KPCC.NAME,' ') AS COLOR,
                            COALESCE(PSS.NAME,' ') AS EYE,
                            COALESCE(PBSS.NAME,'00') AS BRIDGE,
                            COALESCE(PTSS.NAME,'00') AS TEMPLE,
                            COALESCE(PC.NAME,' ') AS CATEGORY,
                            COALESCE(SOL.PRODUCT_UOM_QTY,0) AS QTY,
                            COALESCE(SOL.DISCOUNT,0) AS DISCOUNT,
                            COALESCE(SOL.UNIT_DISCOUNT_PRICE,0) AS UNIT_DISCOUNT_PRICE,
                            COALESCE(SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY,00) AS SUBTOTAL_LINE,
                            COALESCE(PP.IMAGE_URL,'') AS IMAGE_URL,
                            COALESCE(PP.IMAGE_SECONDARY_URL,'') AS IMAGE_SECONDARY_URL,
                            COALESCE(SO.AMOUNT_WITHOUT_DISCOUNT,00) AS SUBTOTAL,
                            COALESCE(SO.AMOUNT_IS_SHIPPING_TOTAL,00) AS SHIPPING_COST,
                            COALESCE(SO.AMOUNT_IS_ADMIN,00) AS ADMIN_FEE,
                            COALESCE(SO.AMOUNT_DISCOUNT,00) AS DISCOUNT,
                            COALESCE(SO.AMOUNT_TAX,00) AS TAX,
                            COALESCE(RC.NAME,'') AS CURRENCY,
                            COALESCE(SO.AMOUNT_TOTAL,00) AS TOTAL,
                            COALESCE(SO.ORDERED_QTY,00) AS TOTAL_QTY,
                            COALESCE(SO.WEIGHT_TOTAL_KG,00) AS WEIGHT
                    FROM SALE_ORDER AS SO
                    LEFT JOIN SALE_ORDER_LINE AS SOL ON SOL.ORDER_ID = SO.ID
                    LEFT JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                    LEFT JOIN KITS_PRODUCT_COLOR_CODE AS KPCC ON PP.COLOR_CODE = KPCC.ID
                    LEFT JOIN PRODUCT_SIZE_SPT AS PSS ON PP.EYE_SIZE = PSS.ID
                    LEFT JOIN PRODUCT_BRAND_SPT AS PBS ON PP.BRAND = PBS.ID
                    LEFT JOIN PRODUCT_MODEL_SPT AS PMS ON PP.MODEL = PMS.ID
                    LEFT JOIN PRODUCT_BRIDGE_SIZE_SPT AS PBSS ON PBSS.ID = PP.BRIDGE_SIZE
                    LEFT JOIN PRODUCT_TEMPLE_SIZE_SPT AS PTSS ON PTSS.ID = PP.TEMPLE_SIZE
                    LEFT JOIN PRODUCT_CATEGORY AS PC ON PC.ID = PP.CATEG_ID
                    LEFT JOIN RES_CURRENCY AS RC ON RC.ID = SOL.CURRENCY_ID
                    WHERE ORDER_ID = {self.order_id.id} AND SOL.IS_ADMIN = false AND SOL.IS_SHIPPING_PRODUCT = false AND SOL.IS_GLOBAL_DISCOUNT = false '''
        self.env.cr.execute(query)
        record_data = self.env.cr.fetchall()
        if record_data:
            for data in record_data:
                product_name = ((data[0] if data[0] else '') + ' ' + (data[1] if data[1] else '') + ' ' + (data[2] if data[2] else '') + ' ' + (data[3] if data[3] else '') + ' ' + (data[4] if data[4] else '') + ' ' + (data[5] if data[5] else '')) or ""

            
                if self.with_img:
                    wrk_sheet.row_dimensions[row_index].height = 50

                    # --------------- First Image --------------- 
                    
                    # sheet.merge_cells("A"+str(row_index)+":B"+str(row_index))
                    # img = BytesIO()
                    # img.flush()
                    # try:
                    #     img.write(base64.b64decode(line.product_id.image_variant_1920))
                    #     image = openpyxl.drawing.image.Image(img)
                    #     image.width = 153
                    #     image.height = 61
                    #     sheet.add_image(image, 'A'+str(row_index))
                    # except:
                    #     pass
                    # --------------- Sec. Image --------------- 

                    # sheet.merge_cells("C"+str(row_index)+":D"+str(row_index))
                    # img = BytesIO()
                    # img.flush()
                    # try:
                    #     img.write(base64.b64decode(line.product_id.image_secondary))
                    #     image = openpyxl.drawing.image.Image(img)
                    #     image.width = 153
                    #     image.height = 61
                    #     sheet.add_image(image, 'C'+str(row_index))
                    # except:
                    #     pass

                    # ----------------------------- Other Data -----------------------------
                    wrk_sheet.cell(row=row_index, column=5).value = product_name
                    wrk_sheet.cell(row=row_index, column=5).alignment = alignment_wrapper
                    wrk_sheet.cell(row=row_index, column=6).value = data[6] or ''
                    wrk_sheet.cell(row=row_index, column=6).alignment = alignment
                    wrk_sheet.cell(row=row_index, column=7).value = data[7]
                    wrk_sheet.cell(row=row_index, column=7).alignment = alignment
                    wrk_sheet.cell(row=row_index, column=8).value = "{:,.2f}".format(data[8])
                    wrk_sheet.cell(row=row_index, column=8).alignment = alignment_right
                    wrk_sheet.cell(row=row_index, column=9).value = "$ {:,.2f}".format(data[9])
                    wrk_sheet.cell(row=row_index, column=9).alignment = alignment_right
                    wrk_sheet.cell(row=row_index, column=10).value = "$ {:,.2f}".format(round(data[9] * data[7],2))
                    wrk_sheet.cell(row=row_index, column=10).alignment = alignment_right
                    wrk_sheet.cell(row=row_index, column=11).hyperlink = data[11] or ''
                    wrk_sheet.cell(row=row_index, column=11).alignment = address_alignment
                    wrk_sheet.cell(row=row_index, column=12).hyperlink = data[12] or ''
                    wrk_sheet.cell(row=row_index, column=12).alignment = address_alignment
                    row_index += 1

                    wrk_sheet.column_dimensions['A'].width = 10
                    wrk_sheet.column_dimensions['D'].width = 9
                    wrk_sheet.column_dimensions['E'].width = 20
                else:
                    wrk_sheet.merge_cells("A"+str(row_index)+":D"+str(row_index))
                    wrk_sheet.cell(row=row_index, column=1).value = product_name
                    wrk_sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,wrap_text=True)
                    wrk_sheet.cell(row=row_index, column=5).value = data[6] or ''
                    wrk_sheet.cell(row=row_index, column=5).alignment = alignment
                    wrk_sheet.cell(row=row_index, column=6).value = data[7]
                    wrk_sheet.cell(row=row_index, column=6).alignment = alignment
                    wrk_sheet.cell(row=row_index, column=7).value = data[8]
                    wrk_sheet.cell(row=row_index, column=7).alignment = alignment_right
                    wrk_sheet.cell(row=row_index, column=8).value = "$ {:,.2f}".format(data[9])
                    wrk_sheet.cell(row=row_index, column=8).alignment = alignment_right
                    wrk_sheet.cell(row=row_index, column=9).value = "$ {:,.2f}".format(round(data[9] * data[7],2))
                    wrk_sheet.cell(row=row_index, column=9).alignment = alignment_right

                    wrk_sheet.row_dimensions[row_index].height = 27
                    row_index += 1

                    wrk_sheet.column_dimensions['A'].width = 10
            # ----------------------------- Total/Discount/shipping Cost/etc. ----------------------------- 
            footer_row = row_index+1
            col_num = 7 if self.with_img else 6
            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))


            wrk_sheet.cell(row=footer_row, column=col_num).value = "Subtotal"
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = "$ {:,.2f}".format(data[13]) or 0.0
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))

            wrk_sheet.cell(row=footer_row, column=col_num).value = "Shipping Cost"
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = "$ {:,.2f}".format(data[14]) or 0.0 # shipping
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))

            wrk_sheet.cell(row=footer_row, column=col_num).value = "Admin Fee"
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = "$ {:,.2f}".format(data[15]) or 0.0 # adminfee
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if abs(self.order_id.amount_discount):
                if self.with_img:
                    wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                    wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
                else:
                    wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                    wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))
                wrk_sheet.cell(row=footer_row, column=col_num).value = "Discount"
                wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
                wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
                wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
                wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
                wrk_sheet.cell(row=footer_row, column=col_num+2).value = "$ {:,.2f}".format(abs(data[16])) or 0.0 # discont
                wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
                wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
                wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
                wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
                footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))
                
            wrk_sheet.cell(row=footer_row, column=col_num).value = "Tax"
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = "$ {:,.2f}".format(data[17]) or 0.0 # taxes
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))

            wrk_sheet.cell(row=footer_row, column=col_num).value = "Total"
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = "({}) $ {:,.2f}".format(data[18],data[19]) # total
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))

            wrk_sheet.cell(row=footer_row, column=col_num).value = "Total Quantity"
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = int(data[20]) or 0.0 # TotalQuantity
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

            if self.with_img:
                wrk_sheet.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrk_sheet.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            else:
                wrk_sheet.merge_cells("F"+str(footer_row)+":G"+str(footer_row))
                wrk_sheet.merge_cells("H"+str(footer_row)+":I"+str(footer_row))
            
            wrk_sheet.cell(row=footer_row, column=col_num).value = "Total Weight (kg)"
            wrk_sheet.cell(row=footer_row, column=col_num).font = Font(size=11, bold=True, name="Calibri")
            wrk_sheet.cell(row=footer_row, column=col_num).alignment = alignment_left
            wrk_sheet.cell(row=footer_row, column=col_num).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+1).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+2).value = '{:,.2f}'.format(data[21]) or 0.0 # TotalWeight
            wrk_sheet.cell(row=footer_row, column=col_num+2).alignment = alignment_right
            wrk_sheet.cell(row=footer_row, column=col_num+2).font = table_font
            wrk_sheet.cell(row=footer_row, column=col_num+2).border = top_border
            wrk_sheet.cell(row=footer_row, column=col_num+3).border = top_border
            footer_row += 1

        save_wrksht = wb if self.with_img else wb_xlsx
        file_ext = '.xlsm' if self.with_img else '.xlsx'

        fp = BytesIO()
        save_wrksht.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.report_file = base64.b64encode(data)

        return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=order.report.with.image.wizard&download=true&field=report_file&id=%s&filename=%s' % (self.id, state +'-%s'%(self.order_id.name)+file_ext),
                'target': 'self',
            }

    def action_send_email(self):
        self.ensure_one()
        self.order_id.download_image_sent = False
        if self.with_img:
            self.order_id.download_image_sent = True
        verified = self.order_id.partner_verification()
        mail_template = self.env.ref('tzc_sales_customization_spt.mail_send_download_image_email')
        mail_template.sudo().with_context(active_id=self.order_id.id,active_model=self.order_id._name).send_mail(self.order_id.id,force_send=True,email_layout_xmlid='mail.mail_notification_light') if verified else None
