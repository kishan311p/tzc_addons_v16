# -*- coding: utf-8 -*-
# Part of SnepTech See LICENSE file for full copyright and licensing details.##
##############################################################################
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import openpyxl
from openpyxl.styles import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
import base64
import pandas
from io import BytesIO
from datetime import datetime
import re

class sales_report_for_sales_person_wizard_spt(models.TransientModel):
    _name = 'sales.report.for.sales.person.wizard.spt'
    _description = 'Set ETO Sale Method'

    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')    
    user_ids = fields.Many2many('res.users','sales_report_wizard_res_user_real','wizard_id','user_id',string='Sales Person')
    file = fields.Binary('File')

    
    def action_process(self):
        return self.env.ref('tzc_sales_customization_spt.action_report_sales_preson_spt').report_action(self)

    def get_report_data(self,user_id):
        sale_order_obj = self.env['sale.order']
        domain = []
        if not self.user_ids:
            raise UserError(_('Set Sales Person.'))
        else:
            domain.append(('user_id','in',user_id.ids))
        if self.start_date:
            domain.append(('date_order','>=',self.start_date))
        
        if self.end_date:
            domain.append(('date_order','<=',self.end_date))
        
        
        return sale_order_obj.search(domain).sorted(key=lambda x:x.partner_id.name)

    def get_excel_report_data(self):
        query = '''SELECT COALESCE(SO.NAME,'') AS NAME,
                    COALESCE(RP.NAME,'') AS SALES,
                    SO.DATE_ORDER,
                    COALESCE(SO.SOURCE_SPT,'') AS SOURCE,
                    COALESCE(SO.AMOUNT_TOTAL,0) AS TOTAL
                FROM SALE_ORDER AS SO
                INNER JOIN RES_PARTNER AS RP ON SO.PARTNER_ID = RP.ID
                WHERE'''
        for rec in self:
            if (self.user_ids):
                users = f'({self.user_ids.id})' if len(self.user_ids.ids)==1 else f'{tuple(self.user_ids.ids)}'
                query = query+' SO.USER_ID IN ' + users
                #  f'({self.user_ids.id})' if len(self.user_ids.ids)==1 else f'{tuple(self.user_ids.ids)}'
            if self.start_date:
                query = query+" AND SO.DATE_ORDER >= '%s'" % str(self.start_date)
            if self.end_date:
                query = query+" AND SO.DATE_ORDER <= '%s'" % str(self.end_date)
            self.env.cr.execute(query)
            report_data=self.env.cr.fetchall()
        return report_data

    def action_create_excel_report(self):
        for record in self:
            active_id= self.id
            f_name ='Sales Person Report'
            workbook = Workbook()
            for user in record.user_ids:
                user_name = user.name.replace('/','-')
                sheet = workbook.create_sheet(title=user_name,index=record.user_ids.ids.index(user.id))
                header_font = Font(name='Garamond', size=20, bold=True)
                header2_font = Font(name='Garamond', size=12, bold=True)
                header3_font = Font(name='Garamond', size=12)
                # bold_font = Font(bold=True)
                alignment = Alignment(horizontal='center',vertical='center',text_rotation=0)
                right_alignment = Alignment(horizontal='right',vertical='center',text_rotation=0)
                bd = Side(style='thin', color="000000")
                all_border = Border(left=bd, top=bd, right=bd, bottom=bd)
                # right_bottom_border = Border(right=bd, bottom=bd)
                right_border = Border(right=bd)
                bottom_border = Border(bottom=bd)
                top_border = Border(top=bd)
                left_border = Border(left=bd)
                sheet.cell(row=9, column=1).value = 'Sale Order'
                sheet.cell(row=9, column=2).value = 'Customers'
                sheet.cell(row=9, column=3).value = 'Date'
                sheet.cell(row=9, column=4).value = 'Medium'
                sheet.cell(row=9, column=5).value = 'Total'
                row_index = 11
                report_data = self.get_excel_report_data()
                for order in report_data:
                    sheet.cell(row=row_index, column=1).value = order[0]
                    sheet.cell(row=row_index, column=2).value = order[1]
                    sheet.cell(row=row_index, column=3).value = datetime.strftime(order[2],'%m-%d-%Y %H:%M:%S')
                    if order[3]:
                        medium = order[3]
                    # if order.catalog_id:
                    #     medium = 'Catalog'
                    # elif order.website_id:
                    #     medium = 'Website'
                        
                    # else:
                    #     pos_id = pos_obj.search([('sale_order_id','=',order.id)])
                    #     if pos_id:
                    #         medium = 'POS Order'
                    #     else:
                    #         medium = 'Normal'                    
                    sheet.cell(row=row_index, column=4).value = medium
                    sheet.cell(row=row_index, column=5).value = '{:,.2f}'.format(float(order[4]))
                    sheet.cell(row=row_index, column=5).alignment = right_alignment
                    sheet.cell(row=row_index, column=5).border = right_border                    
                    sheet.cell(row=row_index, column=1).border = left_border                    
                    row_index += 1
                row_index - 1
                for column_index in range(1,6):
                    sheet.cell(row=row_index, column=column_index).border = top_border   
                    sheet.cell(row=9, column=column_index).border = all_border 
                    sheet.cell(row=9, column=column_index).font =  header2_font
                    sheet.cell(row=9, column=column_index).alignment =  alignment
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 20
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 20
                sheet.column_dimensions['A'].alignment = alignment
                sheet.column_dimensions['B'].alignment = alignment
                sheet.column_dimensions['C'].alignment = alignment
                sheet.column_dimensions['D'].alignment = alignment
                sheet.column_dimensions['E'].alignment = alignment                
                # sheet.cell(row=1,column=1).value = user.name
                # sheet.merge_cells('A1:B4')
                # sheet.merge_cells('C1:C4')
                # sheet.cell(row=1,column=3).value = 'd'
                # sheet.merge_cells('D1:E4')
                # sheet.merge_cells('A7:E8')
                sheet.merge_cells('B9:B10')
                sheet.merge_cells('C9:C10')
                sheet.merge_cells('D9:D10')
                sheet.merge_cells('E9:E10')
                sheet.merge_cells('A9:A10')
                
                sheet.cell(row=1,column=1).value =  user.name
                # sheet.cell(row=1,column=1).border = all_border
                sheet.cell(row=1,column=1).alignment = alignment
                sheet.cell(row=1,column=1).font =  header2_font
                sheet.merge_cells('A1:B2')
                sheet.cell(row=3,column=1).value =  user.street+' '+user.street2 if user.street and user.street2 else user.street
                # sheet.cell(row=3,column=1).alignment = alignment
                sheet.cell(row=3,column=1).font =  header3_font
                sheet.merge_cells('A2:B2')
                sheet.cell(row=4,column=1).value =  user.city +' '+ user.state_id.name if user.city and user.state_id else user.state_id.name
                # sheet.cell(row=4,column=1).alignment = alignment
                sheet.cell(row=4,column=1).font =  header3_font
                sheet.merge_cells('A3:B3')
                sheet.cell(row=5,column=1).value =  user.country_id.name +' '+ user.zip if user.zip and user.country_id else user.country_id.name
                # sheet.cell(row=5,column=1).alignment = alignment
                sheet.cell(row=5,column=1).font =  header3_font
                sheet.merge_cells('A4:B4')
                sheet.cell(row=6,column=1).value = 'Tel:'+user.phone if user.phone else ' '
                # sheet.cell(row=6,column=1).alignment = alignment
                sheet.cell(row=6,column=1).font =  header3_font
                sheet.merge_cells('A5:B5')
                sheet.cell(row=7,column=1).value = 'Email:'+user.email if user.email else ' '
                # sheet.cell(row=7,column=1).alignment = alignment
                sheet.cell(row=7,column=1).font =  header3_font
                sheet.merge_cells('A6:B6')
                sheet.merge_cells('A7:B7')
                sheet.cell(row=1,column=1).border = top_border
                sheet.cell(row=1,column=3).border = left_border
                sheet.cell(row=2,column=3).border = left_border
                sheet.cell(row=3,column=3).border = left_border
                sheet.cell(row=4,column=3).border = left_border
                sheet.cell(row=5,column=3).border = left_border
                sheet.cell(row=6,column=3).border = left_border
                sheet.cell(row=7,column=3).border = left_border
                sheet.cell(row=7,column=1).border = bottom_border
                
            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            record.file = base64.b64encode(data) 
            
            return {
            'type' : 'ir.actions.act_url',
            'url':   'web/content/?model=sales.report.for.sales.person.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
             }
