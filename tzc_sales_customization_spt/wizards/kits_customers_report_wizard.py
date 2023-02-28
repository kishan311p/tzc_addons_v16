from re import S
from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO

class kits_customers_report_wizard(models.TransientModel):
    _name = "kits.customers.report.wizard"
    _description = "Kits Customers Report Wizard"

    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    file = fields.Binary()
    all_fields = fields.Boolean('All Field')

    def validate_dates(self):
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise UserError(_("Start Date should be lesser than End Date."))

    def action_customer_report(self):
        self.validate_dates()

        domain = [('is_customer','=',True)]
        if self.env.user.has_group('base.group_system') or self.env.user.has_group('marketing_automation.group_marketing_automation_user'):
            pass
        else:
            domain.append(('country_id','in',self.env.user.contact_allowed_countries.ids))
        
        if self.start_date:
            domain.append(('create_date','>=',self.start_date))
        if self.end_date:
            domain.append(('create_date','<=',self.end_date))

        contacts_ids = self.env['res.partner'].sudo().search(domain)

        f_name = 'Customers'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Customers", index=0)

        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        align_left = Alignment(vertical="center", horizontal='left', text_rotation=0, wrap_text=True)
        align_right = Alignment(vertical="center", horizontal='right', text_rotation=0, wrap_text=True)

        table_header = 1

        if self.all_fields:
            sheet.cell(row=table_header, column=1).value = 'Id'
            sheet.cell(row=table_header, column=2).value = 'Name'
            sheet.cell(row=table_header, column=3).value = 'Is Company'
            sheet.cell(row=table_header, column=4).value = 'Salesperson Internal Id'
            sheet.cell(row=table_header, column=5).value = 'Salesperson'
            sheet.cell(row=table_header, column=6).value = 'Active'
            sheet.cell(row=table_header, column=7).value = 'Phone'
            sheet.cell(row=table_header, column=8).value = 'Mobile'
            sheet.cell(row=table_header, column=9).value = 'Email'
            sheet.cell(row=table_header, column=10).value = 'Street'
            sheet.cell(row=table_header, column=11).value = 'Zip'
            sheet.cell(row=table_header, column=12).value = 'State'
            sheet.cell(row=table_header, column=13).value = 'Country'
            sheet.cell(row=table_header, column=14).value = 'Business Type'
            sheet.cell(row=table_header, column=15).value = 'Company Name'
            sheet.cell(row=table_header, column=16).value = 'Company Type'
            sheet.cell(row=table_header, column=17).value = 'Customer Type'
            sheet.cell(row=table_header, column=18).value = 'Activity State'
            sheet.cell(row=table_header, column=19).value = 'Journal Items'
            sheet.cell(row=table_header, column=20).value = 'Create Date'
            sheet.cell(row=table_header, column=21).value = 'Last Modified On'
            sheet.cell(row=table_header, column=22).value = 'Last Order Value'
        else:
            sheet.cell(row=table_header, column=1).value = 'Create Date'
            sheet.cell(row=table_header, column=2).value = 'Last Login'
            sheet.cell(row=table_header, column=3).value = 'Name'
            sheet.cell(row=table_header, column=4).value = 'Status'
            sheet.cell(row=table_header, column=5).value = 'Phone'
            sheet.cell(row=table_header, column=6).value = 'Street'
            sheet.cell(row=table_header, column=7).value = 'State'
            sheet.cell(row=table_header, column=8).value = 'Zip'
            sheet.cell(row=table_header, column=9).value = 'Country'
            sheet.cell(row=table_header, column=10).value = 'Territory'
            sheet.cell(row=table_header, column=11).value = 'Email'
            sheet.cell(row=table_header, column=12).value = 'Email Verified'
            sheet.cell(row=table_header, column=13).value = 'Is Granted Portal'
            sheet.cell(row=table_header, column=14).value = 'Salesperson'
            sheet.cell(row=table_header, column=15).value = '#Order'
            sheet.cell(row=table_header, column=16).value = '#Catalog'
            sheet.cell(row=table_header, column=17).value = 'Order Amount'

        loop_range = False
        if self.all_fields:
            loop_range = range(1,23)
        else:
            loop_range = range(1,18)

        for col in loop_range:
            sheet.cell(row=table_header, column=col).font = header_font
            sheet.cell(row=table_header, column=col).border = top_bottom_border

        row_index=table_header+1

        for partner in contacts_ids:


            query = f'''
                SELECT RP.ID,
                    COALESCE(RP.NAME,'') AS NAME,
                    RP.IS_COMPANY,
                    RP.ACTIVE,
                    RP.PHONE,
                    RP.MOBILE,
                    RP.EMAIL,
                    RP.STREET,
                    RP.STREET2,
                    RP.ZIP,
                    RCS.NAME AS STATE,
                    RC.NAME->>'en_US' AS COUNTRY,
                    RP.CUSTOMER_TYPE,
                    RP.CREATE_DATE,
                    RP.WRITE_DATE,
                    RP.CONTACT_NAME_SPT,
                    RP.LAST_ORDER_VALUE,
                    RCG.NAME AS TERRITORY,
                    RP.IS_EMAIL_VERIFIED,
	                RP.IS_GRANTED_PORTAL_ACCESS
                FROM RES_PARTNER AS RP
                LEFT JOIN RES_USERS AS RU ON RP.ID = RU.PARTNER_ID
                LEFT JOIN RES_COUNTRY_STATE AS RCS ON RP.STATE_ID = RCS.ID
                LEFT JOIN RES_COUNTRY AS RC ON RP.COUNTRY_ID = RC.ID
                LEFT JOIN RES_COUNTRY_GROUP AS RCG ON RC.TERRITORY_ID = RCG.ID
                WHERE RP.ID = {partner.id}
            '''
            self.env.cr.execute(query)
            records = self.env.cr.fetchall()

            for data in records:
                partner_adress = []
                if partner.street:
                    partner_adress.append(data[7])
                if partner.street2:
                    partner_adress.append(data[8])
                if self.all_fields:
                    sheet.cell(row=row_index, column=1).value = data[0] or ''
                    sheet.cell(row=row_index, column=1).alignment = align_left
                    sheet.cell(row=row_index, column=2).value = data[1] or '' 
                    sheet.cell(row=row_index, column=3).value = str(data[2]) 
                    sheet.cell(row=row_index, column=4).value = partner.user_id.internal_id if partner.user_id.internal_id else ''
                    sheet.cell(row=row_index, column=5).value = partner.user_id.name if partner.user_id else ''
                    sheet.cell(row=row_index, column=6).value = str(data[3])
                    sheet.cell(row=row_index, column=7).value = data[4] or ''
                    sheet.cell(row=row_index, column=8).value = data[5] or ''
                    sheet.cell(row=row_index, column=9).value = data[6] or ''
                    sheet.cell(row=row_index, column=10).value = ", ".join(partner_adress) if partner_adress else ''
                    sheet.cell(row=row_index, column=11).value = data[10] or ''
                    sheet.cell(row=row_index, column=12).value = data[11] or ''
                    sheet.cell(row=row_index, column=13).value = data[12] or ''
                    sheet.cell(row=row_index, column=14).value = ', '.join(partner.business_type_ids.mapped('name')) if partner.business_type_ids else ''
                    sheet.cell(row=row_index, column=15).value = partner.company_id.name if partner.company_id else ''
                    sheet.cell(row=row_index, column=16).value = dict(partner._fields['company_type'].selection).get(partner.company_type)
                    sheet.cell(row=row_index, column=17).value = dict(partner._fields['customer_type'].selection).get(data[13]) if data[13] else ''
                    sheet.cell(row=row_index, column=18).value = dict(partner._fields['activity_state'].selection).get(data[14]) if partner.activity_state else ''
                    sheet.cell(row=row_index, column=19).value = partner.journal_item_count if partner.journal_item_count else ''
                    sheet.cell(row=row_index, column=19).alignment = align_left
                    sheet.cell(row=row_index, column=20).value = data[15].strftime('%d-%m-%Y') or ''
                    sheet.cell(row=row_index, column=21).value = partner.search_read([('id','=',partner.id)],['__last_update'])[0].get('__last_update').strftime('%d-%m-%Y')
                    sheet.cell(row=row_index, column=22).value = data[16] or ''
                    sheet.cell(row=row_index, column=22).alignment = align_right
                    row_index += 1

                    sheet.column_dimensions['A'].width = 10
                    sheet.column_dimensions['B'].width = 25
                    sheet.column_dimensions['C'].width = 15
                    sheet.column_dimensions['D'].width = 25
                    sheet.column_dimensions['E'].width = 25
                    sheet.column_dimensions['F'].width = 10
                    sheet.column_dimensions['G'].width = 15
                    sheet.column_dimensions['H'].width = 15
                    sheet.column_dimensions['I'].width = 25
                    sheet.column_dimensions['J'].width = 30
                    sheet.column_dimensions['K'].width = 15
                    sheet.column_dimensions['L'].width = 15
                    sheet.column_dimensions['M'].width = 15
                    sheet.column_dimensions['N'].width = 15
                    sheet.column_dimensions['O'].width = 30
                    sheet.column_dimensions['P'].width = 15
                    sheet.column_dimensions['Q'].width = 15
                    sheet.column_dimensions['R'].width = 15
                    sheet.column_dimensions['S'].width = 15
                    sheet.column_dimensions['T'].width = 15
                    sheet.column_dimensions['U'].width = 20
                    sheet.column_dimensions['V'].width = 20
                else:
                    sheet.cell(row=row_index, column=1).value = partner.create_date.strftime("%d-%m-%Y") if partner.create_date else ''
                    sheet.cell(row=row_index, column=3).value = partner.name if partner.name else ''
                    sheet.cell(row=row_index, column=4).value = dict(self.env['res.partner']._fields['customer_type'].selection).get(partner.customer_type)
                    sheet.cell(row=row_index, column=5).value = partner.phone if partner.phone else ''
                    sheet.cell(row=row_index, column=6).value = ", ".join(partner_adress)
                    sheet.cell(row=row_index, column=7).value = partner.state_id.name if partner.state_id else ''
                    sheet.cell(row=row_index, column=8).value = partner.zip if partner.zip else ''
                    sheet.cell(row=row_index, column=9).value = partner.country_id.name if partner.country_id else ''
                    sheet.cell(row=row_index, column=10).value = partner.territory.name if partner.territory else ''
                    sheet.cell(row=row_index, column=11).value = partner.email if partner.email else ''
                    sheet.cell(row=row_index, column=12).value = "True" if partner.is_email_verified else "False"
                    sheet.cell(row=row_index, column=13).value = "True" if partner.is_granted_portal_access else "False"
                    sheet.cell(row=row_index, column=14).value = partner.user_id.name if partner.user_id else ''
                    sheet.cell(row=row_index, column=15).value = ", ".join(partner.sale_order_ids.mapped('name')) if partner.sale_order_ids else ''
                    sheet.cell(row=row_index, column=16).value = ", ".join(partner.catalog_ids.mapped('name')) if partner.catalog_ids else ''
                    sheet.cell(row=row_index, column=17).value = partner.total_invoiced or 0
                    row_index += 1

                    sheet.column_dimensions['A'].width = 15
                    sheet.column_dimensions['B'].width = 15
                    sheet.column_dimensions['C'].width = 30
                    sheet.column_dimensions['D'].width = 15
                    sheet.column_dimensions['E'].width = 25
                    sheet.column_dimensions['F'].width = 35
                    sheet.column_dimensions['G'].width = 15
                    sheet.column_dimensions['H'].width = 15
                    sheet.column_dimensions['I'].width = 18
                    sheet.column_dimensions['J'].width = 18
                    sheet.column_dimensions['K'].width = 25
                    sheet.column_dimensions['L'].width = 15
                    sheet.column_dimensions['M'].width = 18
                    sheet.column_dimensions['N'].width = 25
                    sheet.column_dimensions['O'].width = 30
                    sheet.column_dimensions['P'].width = 30
                    sheet.column_dimensions['Q'].width = 15

            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            self.file = base64.b64encode(data)

            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=kits.customers.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,f_name),
                'target': 'self',
            }
        
    def action_signup_user_xlsx_report(self):
        self.validate_dates()

        domain = []
        if not self.env.user.has_group('base.group_system') and not self.env.user.has_group('sales_team.group_sale_manager') and self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt'):
            domain.append(('country_id','in',self.env.user.contact_allowed_countries.ids))
            domain.append(('country_id','!=',False))
        if self.start_date:
            domain.append(('create_date', '>=', self.start_date))
        if self.end_date:
            domain.append(('create_date', '<=', self.end_date))

        users_ids = self.env['res.users'].search(domain)

        f_name = 'Signup Users'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Signup Users", index=0)

        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        align_left = Alignment(vertical="center", horizontal='left', text_rotation=0, wrap_text=True)

        table_header = 1
        if self.all_fields:
            sheet.cell(row=table_header, column=1).value = 'Id'
            sheet.cell(row=table_header, column=2).value = 'Name'
            sheet.cell(row=table_header, column=3).value = 'Is Company'
            sheet.cell(row=table_header, column=4).value = 'Salesperson Internal Id'
            sheet.cell(row=table_header, column=5).value = 'Salesperson'
            sheet.cell(row=table_header, column=6).value = 'Active'
            sheet.cell(row=table_header, column=7).value = 'Phone'
            sheet.cell(row=table_header, column=8).value = 'Mobile'
            sheet.cell(row=table_header, column=9).value = 'Email'
            sheet.cell(row=table_header, column=10).value = 'Street'
            sheet.cell(row=table_header, column=11).value = 'Zip'
            sheet.cell(row=table_header, column=12).value = 'State'
            sheet.cell(row=table_header, column=13).value = 'Country'
            sheet.cell(row=table_header, column=14).value = 'Business Type'
            sheet.cell(row=table_header, column=15).value = 'Company Name'
            sheet.cell(row=table_header, column=16).value = 'Company Type'
            sheet.cell(row=table_header, column=17).value = 'Customer Type'
            sheet.cell(row=table_header, column=18).value = 'Activity State'
            sheet.cell(row=table_header, column=19).value = 'Journal Items'
            sheet.cell(row=table_header, column=20).value = 'Create Date'
            sheet.cell(row=table_header, column=21).value = 'Last Modified On'
            sheet.cell(row=table_header, column=22).value = 'Last Order Value'
        else:
            sheet.cell(row=table_header, column=1).value = 'Signup Date'
            sheet.cell(row=table_header, column=2).value = 'Name'
            sheet.cell(row=table_header, column=3).value = 'Email'
            sheet.cell(row=table_header, column=4).value = 'Salesperson'
            sheet.cell(row=table_header, column=5).value = 'Country'
            sheet.cell(row=table_header, column=6).value = 'Territory'
            sheet.cell(row=table_header, column=7).value = 'City'
            sheet.cell(row=table_header, column=8).value = 'Phone'
            sheet.cell(row=table_header, column=9).value = 'Address'

        loop_range = False
        if self.all_fields:
            loop_range = range(1,23)
        else:
            loop_range = range(1,10)

        for col in loop_range:
            sheet.cell(row=table_header, column=col).font = header_font
            sheet.cell(row=table_header, column=col).border = top_bottom_border

        row_index=table_header+1

        for user in users_ids:

            user_adress = []
            if user.street:
                user_adress.append(user.street)
            if user.street2:
                user_adress.append(user.street2)

            if self.all_fields:
                sheet.cell(row=row_index, column=1).value = user.id if user.id else ''
                sheet.cell(row=row_index, column=1).alignment = align_left
                sheet.cell(row=row_index, column=2).value = user.name if user.name else '' 
                sheet.cell(row=row_index, column=3).value = str(user.is_company) 
                sheet.cell(row=row_index, column=4).value = user.user_id.internal_id if user.internal_id else ''
                sheet.cell(row=row_index, column=5).value = user.user_id.name if user.user_id else ''
                sheet.cell(row=row_index, column=6).value = str(user.active)
                sheet.cell(row=row_index, column=7).value = user.phone if user.phone else ''
                sheet.cell(row=row_index, column=8).value = user.mobile if user.mobile else ''
                sheet.cell(row=row_index, column=9).value = user.email if user.email else ''
                sheet.cell(row=row_index, column=10).value = ", ".join(user_adress) if user_adress else ''
                sheet.cell(row=row_index, column=11).value = user.zip if user.zip else ''
                sheet.cell(row=row_index, column=12).value = user.state_id.name if user.state_id else ''
                sheet.cell(row=row_index, column=13).value = user.country_id.name if user.country_id else ''
                sheet.cell(row=row_index, column=14).value = ', '.join(user.business_type_ids.mapped('name')) if user.business_type_ids else ''
                sheet.cell(row=row_index, column=15).value = user.company_id.name if user.company_id else ''
                sheet.cell(row=row_index, column=16).value = dict(user._fields['company_type'].selection(user)).get(user.company_type)
                sheet.cell(row=row_index, column=17).value = dict(user._fields['customer_type'].selection(user)).get(user.customer_type) if user.customer_type else ''
                sheet.cell(row=row_index, column=18).value = dict(user._fields['activity_state'].selection(user)).get(user.activity_state) if user.activity_state else ''
                sheet.cell(row=row_index, column=19).value = user.journal_item_count if user.journal_item_count else ''
                sheet.cell(row=row_index, column=19).alignment = align_left
                sheet.cell(row=row_index, column=20).value = user.create_date.strftime('%d-%m-%Y') if user.create_date else ''
                sheet.cell(row=row_index, column=21).value = user.search_read([('id','=',user.id)],['__last_update'])[0].get('__last_update').strftime('%d-%m-%Y')
                sheet.cell(row=row_index, column=22).value = user.last_order_value if user.last_order_value else ''
                sheet.cell(row=row_index, column=22).alignment = align_left
                row_index += 1
            
                sheet.column_dimensions['A'].width = 10
                sheet.column_dimensions['B'].width = 25
                sheet.column_dimensions['C'].width = 15
                sheet.column_dimensions['D'].width = 25
                sheet.column_dimensions['E'].width = 25
                sheet.column_dimensions['F'].width = 10
                sheet.column_dimensions['G'].width = 15
                sheet.column_dimensions['H'].width = 15
                sheet.column_dimensions['I'].width = 25
                sheet.column_dimensions['J'].width = 30
                sheet.column_dimensions['K'].width = 15
                sheet.column_dimensions['L'].width = 15
                sheet.column_dimensions['M'].width = 15
                sheet.column_dimensions['N'].width = 15
                sheet.column_dimensions['O'].width = 30
                sheet.column_dimensions['P'].width = 15
                sheet.column_dimensions['Q'].width = 15
                sheet.column_dimensions['R'].width = 15
                sheet.column_dimensions['S'].width = 15
                sheet.column_dimensions['T'].width = 15
                sheet.column_dimensions['U'].width = 20
                sheet.column_dimensions['V'].width = 20
            else:
                sheet.cell(row=row_index, column=1).value = user.create_date.strftime("%d-%m-%Y") if user.create_date else ''
                sheet.cell(row=row_index, column=2).value = user.name if user.name else ''
                sheet.cell(row=row_index, column=3).value = user.email if user.email else ''
                sheet.cell(row=row_index, column=4).value = user.user_id.name if user.user_id else ''
                sheet.cell(row=row_index, column=5).value = user.country_id.name if user.country_id.name else ''
                sheet.cell(row=row_index, column=6).value = user.territory.name if user.territory.name else ''
                sheet.cell(row=row_index, column=7).value = user.city if user.city else ''
                sheet.cell(row=row_index, column=8).value = user.phone if user.phone else ''
                sheet.cell(row=row_index, column=9).value = ", ".join(user_adress)
                row_index += 1

                sheet.column_dimensions['A'].width = 15
                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 30
                sheet.column_dimensions['D'].width = 25
                sheet.column_dimensions['E'].width = 15
                sheet.column_dimensions['F'].width = 15
                sheet.column_dimensions['G'].width = 15
                sheet.column_dimensions['H'].width = 18
                sheet.column_dimensions['I'].width = 30

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=kits.customers.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,f_name),
            'target': 'self',
        }
