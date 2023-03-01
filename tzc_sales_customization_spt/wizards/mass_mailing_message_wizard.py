from odoo import _, api, fields, models, tools
import base64
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from io import BytesIO
import os
from odoo.exceptions import UserError

class mass_mailing_message_wizard(models.TransientModel):
    _name = "mass.mailing.message.wizard"
    _description = "Mass Mailing Message Wiazrd"

    
    partner_ids = fields.Many2many('res.partner',string='Contact')
    email_partner_ids = fields.Many2many('res.partner','email_contact_res_partner_rel','wizard_id','partner_id',string=" Contact")
    message = fields.Char()
    none_mails_partner_ids = fields.Many2many('res.partner','none_mail_mailgun_varified_wizard_rel','wiz_id','prtnr_id',string='Partner')
    message = fields.Char()
    report_file = fields.Binary()

    def action_process(self):
        if self._context.get('raise_campaign'):
            form_view = self.env.ref('tzc_sales_customization_spt.marketing_campaing_form_view')
            campaign_id = self._context.get('campaign_id')
            if campaign_id:
                return{
                    'name': ('Campaign'),
                    'res_model': 'marketing.campaign',
                    'type': 'ir.actions.act_window',
                    'views': [(form_view.id, 'form')],
                    'res_id': campaign_id,
                    'target': 'current',
                }
            else:
                raise UserError ('Campaign not found.')
    
    def action_export(self):
        partner_ids = self.email_partner_ids.filtered(lambda x: x.mailgun_verification_status == 'rejected' and x.email)
        if not partner_ids:
            raise UserError('There is no unverified email in selected contact.')
        f_name = 'Unverified Email'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Unverified Email", index=0)
        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        left_alignment = Alignment(
            vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        right_alignment = Alignment(
            vertical='center', horizontal='right', text_rotation=0, wrap_text=True)

        header_row = 1
        # sheet.cell(row=header_row, column=1).value = 'ID'
        sheet.cell(row=header_row, column=1).value = 'Internal ID'
        sheet.cell(row=header_row, column=2).value = 'Name'
        sheet.cell(row=header_row, column=3).value = 'Is a Company'
        sheet.cell(row=header_row, column=4).value = 'Salesperson'
        sheet.cell(row=header_row, column=5).value = 'Sales Person ID'
        sheet.cell(row=header_row, column=6).value = 'Active'
        sheet.cell(row=header_row, column=7).value = 'Phone'
        sheet.cell(row=header_row, column=8).value = 'Mobile'
        sheet.cell(row=header_row, column=9).value = 'Email'
        sheet.cell(row=header_row, column=10).value = 'Street'
        sheet.cell(row=header_row, column=11).value = 'Zip'
        sheet.cell(row=header_row, column=12).value = 'State'
        sheet.cell(row=header_row, column=13).value = 'Country'
        sheet.cell(row=header_row, column=14).value = 'Business Type'
        sheet.cell(row=header_row, column=15).value = 'Customer Type'
        sheet.cell(row=header_row, column=16).value = 'ID'
        sheet.cell(row=header_row, column=17).value = 'Created on'
        sheet.cell(row=header_row, column=18).value = 'Last Modified on'

        sheet.cell(row=header_row, column=1).font = header_font
        sheet.cell(row=header_row, column=1).border = top_bottom_border
        sheet.cell(row=header_row, column=2).font = header_font
        sheet.cell(row=header_row, column=2).border = top_bottom_border
        sheet.cell(row=header_row, column=3).font = header_font
        sheet.cell(row=header_row, column=3).border = top_bottom_border
        sheet.cell(row=header_row, column=4).font = header_font
        sheet.cell(row=header_row, column=4).border = top_bottom_border
        sheet.cell(row=header_row, column=5).font = header_font
        sheet.cell(row=header_row, column=5).border = top_bottom_border
        sheet.cell(row=header_row, column=6).font = header_font
        sheet.cell(row=header_row, column=6).border = top_bottom_border
        sheet.cell(row=header_row, column=7).font = header_font
        sheet.cell(row=header_row, column=7).border = top_bottom_border
        sheet.cell(row=header_row, column=8).font = header_font
        sheet.cell(row=header_row, column=8).border = top_bottom_border
        sheet.cell(row=header_row, column=9).font = header_font
        sheet.cell(row=header_row, column=9).border = top_bottom_border
        sheet.cell(row=header_row, column=10).font = header_font
        sheet.cell(row=header_row, column=10).border = top_bottom_border
        sheet.cell(row=header_row, column=11).font = header_font
        sheet.cell(row=header_row, column=11).border = top_bottom_border
        sheet.cell(row=header_row, column=12).font = header_font
        sheet.cell(row=header_row, column=12).border = top_bottom_border
        sheet.cell(row=header_row, column=13).font = header_font
        sheet.cell(row=header_row, column=13).border = top_bottom_border
        sheet.cell(row=header_row, column=14).font = header_font
        sheet.cell(row=header_row, column=14).border = top_bottom_border
        sheet.cell(row=header_row, column=15).font = header_font
        sheet.cell(row=header_row, column=15).border = top_bottom_border
        sheet.cell(row=header_row, column=16).font = header_font
        sheet.cell(row=header_row, column=16).border = top_bottom_border
        sheet.cell(row=header_row, column=17).font = header_font
        sheet.cell(row=header_row, column=17).border = top_bottom_border
        sheet.cell(row=header_row, column=18).font = header_font
        sheet.cell(row=header_row, column=18).border = top_bottom_border

        row_index=header_row+1
        for partner in partner_ids:
            sheet.cell(row=row_index, column=1).value = partner.internal_id or ''
            sheet.cell(row=row_index, column=2).value = partner.name or ''
            sheet.cell(row=row_index, column=3).value = 'Yes' if partner.company_type == 'company' else 'No'
            sheet.cell(row=row_index, column=4).value = partner.user_id.name or ''
            sheet.cell(row=row_index, column=5).value = partner.user_id.partner_id.internal_id or ''
            sheet.cell(row=row_index, column=6).value = 'Yes' if partner.active else 'No'
            sheet.cell(row=row_index, column=7).value = partner.phone or ''
            sheet.cell(row=row_index, column=8).value = partner.mobile or ''
            sheet.cell(row=row_index, column=9).value = partner.email or ''
            sheet.cell(row=row_index, column=10).value = ','.join([l for l in [partner.street or '', partner.street2 or ''] if l])
            sheet.cell(row=row_index, column=11).value = partner.zip or ''
            sheet.cell(row=row_index, column=12).value = partner.state_id.name or ''
            sheet.cell(row=row_index, column=13).value = partner.country_id.name or ''
            sheet.cell(row=row_index, column=14).value = ','.join(partner.business_type_ids.mapped('name')) or ''
            sheet.cell(row=row_index, column=15).value = partner.customer_type or ''
            sheet.cell(row=row_index, column=16).value = partner._get_external_ids()[partner.id][0] if partner._get_external_ids()[partner.id] else ''
            sheet.cell(row=row_index, column=17).value = partner.create_date.strftime('%d-%m-%Y') or  ''
            sheet.cell(row=row_index, column=17).alignment = left_alignment
            sheet.cell(row=row_index, column=18).value = partner.write_date.strftime('%d-%m-%Y') or ''
            sheet.cell(row=row_index, column=18).alignment = left_alignment
            row_index += 1

        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 18
        sheet.column_dimensions['J'].width = 25
        sheet.column_dimensions['K'].width = 15
        sheet.column_dimensions['L'].width = 15
        sheet.column_dimensions['M'].width = 15
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['P'].width = 25
        sheet.column_dimensions['Q'].width = 20
        sheet.column_dimensions['R'].width = 20

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.report_file = base64.b64encode(data)

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=mass.mailing.message.wizard&download=true&field=report_file&id=%s&filename=%s.xlsx' % (self.id,f_name),
            'target': 'self',
        }
