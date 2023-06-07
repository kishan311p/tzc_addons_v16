from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
import io
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
import pandas as pd
from datetime import datetime, date, timedelta
import xlsxwriter
from lxml import etree
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import DEFAULT_FONT

class sales_commission_report_wizard(models.TransientModel):
    _name = 'sales.commission.report.wizard'
    _description = 'Sales Commission Report'

    def _get_salesmanagers(self):
        if self.env.user.has_group('base.group_system'):
            managers = self.env.ref('tzc_sales_customization_spt.group_sales_manager_spt').users
            return [('is_sales_manager','=',True),('id','in',managers.ids),'|',('active','=',True),('active','=',False)]
        else:
            return [('is_sales_manager','=',True)]
    
    # def _get_sales_person(self):
    #     return [('id','=',self.env.user.id),('is_salesperson','=',True),'|',('active','=',True),('active','=',False)]

    def _get_sales_person(self):
        user_id = self.env.user
        if not user_id.has_group('base.group_system'):
            return [('is_salesperson','=',True),('id','in',user_id.allow_user_ids.ids + user_id.ids),'|',('active','=',True),('active','=',False)]
        else:
            return [('is_salesperson','=',True)]

    start_date = fields.Date('Invoice Start Date')
    end_date = fields.Date('Invoice End Date')
    sales_person_ids = fields.Many2many('res.users', 'sales_commission_wizard_res_users_rel', 'wizard_id', 'user_id', string="Sales Persons",default=lambda self:self.env.user,domain=_get_sales_person)
    sales_manager_ids = fields.Many2many('res.users', 'sales_manager_commission_wizard_res_users_rel', 'wizard_id', 'user_id', string="Sales Manager",default=lambda self:self.env.user,domain=_get_salesmanagers)
    file = fields.Binary()
    commission_for = fields.Selection([('sales_person', 'Salesperson'),('sales_manager', 'Sales Manager')],default="sales_person")
    commission_is = fields.Selection([('draft','Not Paid'),('full','Fully Paid'),('partial','Partial Paid'),('over','Over Paid')],default='full',string="Commission Type")
    # commission_is = fields.Selection([('all','All'),('is_paid','Paid'),('is_unpaid','Unpaid')],default='all',string="Commission Type")
    apply_groups = fields.Selection([('admin','Administrator'),('sales_person','Salesperson'),('sales_manager','Sales Manager')],string="Apply Groups")


    @api.model
    def default_get(self, default_fields):
        res= super(sales_commission_report_wizard, self).default_get(default_fields)
        apply_groups =''
        if self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt'):
            apply_groups = 'sales_manager'
        if self.env.user.has_group('sales_team.group_sale_salesman'):
            apply_groups = 'sales_person'
        if self.env.user.has_group('base.group_system'):
            apply_groups = 'admin'
        res['apply_groups'] = apply_groups
        return res

    def get_users(self,return_users=False):
        users = False
        user_data = {}
        if self.commission_for == 'sales_person':
            users = self.with_context(active_test=False).sales_person_ids
        if self.commission_for == 'sales_manager':
            users = self.with_context(active_test=False).sales_manager_ids
        if return_users:
            return users.sorted(lambda x: x.name)
        for user in users.sorted(lambda x: x.name):
            if user.id in user_data.keys():
                user_data[user.id]['lines'].extend(self.get_commission(user))
            else:
                user_data[user.id] = {
                    'name':user.name,
                    'street':user.partner_id.street,
                    'street2':user.partner_id.street2,
                    'city':user.partner_id.city,
                    'state':user.partner_id.state_id.name,
                    'country':user.partner_id.country_id.name,
                    'phone':user.partner_id.phone,
                    'email':user.partner_id.email,
                    'start_date':self.start_date,
                    'end_date':self.end_date,
                    'commission_rule':user.commission_rule_id.name if self.commission_for == 'sales_person' else user.manager_commission_rule_id.name if self.commission_for == 'sales_manager' else '',
                    'lines':self.get_commission(user),
                    }
        
        return user_data

    def action_pdf_report(self):
        self.validate_dates()
        return self.env.ref('tzc_sales_customization_spt.action_sales_commission_pdf_report').report_action(self)

    def get_commission(self, sale_person):
        report_table_lines = []
        domain= [('user_id','=',sale_person.id)]
        if self.start_date:
            domain.append(('invoice_id.invoice_date', '>=', self.start_date))
        if self.end_date:
            domain.append(('invoice_id.invoice_date', '<=', self.end_date))
        if self.commission_is == 'draft':
            domain.append(('state','in',['draft']))
        elif self.commission_is == 'full':
            domain.append(('state','in',['full']))
        elif self.commission_is == 'partial':
            domain.append(('state','in',['partial']))
        elif self.commission_is == 'over':
            domain.append(('state','in',['over']))
        # if self.commission_is == 'all':
        #     domain.append(('state','in',['draft','paid']))
        
        invoice_ids = self.env['kits.commission.lines'].sudo().search(domain).mapped('invoice_id')

        # state = []
        # if self.commission_is == 'is_paid':
        #     state.append('full')
        #     state.append('partial')
        #     state.append('over')
        # elif self.commission_is == 'is_unpaid':
        #     state.append('draft')
        #     state.append('cancel')
        # elif self.commission_is == 'all':
        #     state.append('full')
        #     state.append('partial')
        #     state.append('over')
        #     state.append('draft')
        #     state.append('cancel')
        state = self.commission_is
        # if invoice_ids and ((sale_person.commission_rule_id and self.commission_for =='sales_person') or( sale_person.manager_commission_rule_id and self.commission_for =='sales_manager')):
        for invoice in invoice_ids:
            order = self.env['sale.order'].search([('invoice_ids','in',invoice.ids)],limit=1)
            commission = 0.0
            if self.commission_for == 'sales_person':
                commission = sum(invoice.commission_line_ids.filtered(lambda x: x.commission_for == 'saleperson' and x.user_id == sale_person and x.state in state).mapped('amount'))
            elif self.commission_for == 'sales_manager':
                commission = sum(invoice.commission_line_ids.filtered(lambda x: x.commission_for == 'manager' and x.user_id == sale_person and x.state in state).mapped('amount'))
            else:
                pass
            if commission:
                report_table_lines.append(self.data_dict_pdf(invoice,commission,dict(self._fields['commission_is'].selection).get(state)))
        return report_table_lines

    def create_address_line_for_commission(self, source_id, take_name=False):
        address = ''
        if take_name == True:
            if source_id.name:
                address += str(source_id.name)
            if source_id.street:
                if source_id.name:
                    address += '\n'+str(source_id.street)
                else:
                    address += source_id.street
        else:
            if source_id.street:
                address += str(source_id.street) 
        if source_id.street2:
            address += '\n'+str(source_id.street2)
        if source_id.city:
            address+= '\n'+str(source_id.city)
        if source_id.zip and take_name:
            address += ', '+source_id.zip
        if source_id.state_id:
            if take_name:
                address += '\n'+str(source_id.state_id.name)
            else:
                address += ' '+str(source_id.state_id.name)
        if source_id.country_id:
            if take_name:
                if source_id.state_id:
                    address += ', '+str(source_id.country_id.name)
                else:
                    address += '\n'+str(source_id.country_id.name)
            else:
                address += ' '+str(source_id.country_id.name)
                
        if source_id.zip and not take_name:
            address += ' '+source_id.zip
        address += '\nTel. '
        if source_id.phone:
            address += source_id.phone
        address += '\nEmail. '
        if source_id.email:
            address += source_id.email 
        return address

    def action_xls_report(self):
        self.validate_dates()
        workbook = Workbook()
        users = self.get_users(return_users=True)
        top_alignment = Alignment(vertical='top',horizontal='left')
        right_alignment = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
        center_alignment = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
        left_alignment = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        for sale_person in users:
            selected_state = {'draft':" in ('draft')",'full':" in ('full')",'partial': " in ('partial')",'over':" in ('over')"}
            commission_type = selected_state[self.commission_is]

            commi_for = {'sales_person':'saleperson','sales_manager':'manager'}
            commission_for = commi_for[self.commission_for]
            
            if self.start_date and self.end_date:
                set_date = "am.invoice_date >= '%s' and am.invoice_date <= '%s' and "%(self.start_date,self.end_date)
            elif self.end_date:
                set_date = "am.invoice_date <= '%s' and "%(self.end_date)
            elif self.start_date:
                set_date = "am.invoice_date >= '%s' and "%(self.start_date)
            else:
                set_date = ''
           
            query = f'''SELECT SO.DATE_ORDER,
                        SO.NAME,
                        COALESCE(AM.NAME,'') AS INVOICE_NAME,
                        COALESCE(RP.NAME,'') AS PARTNER_NAME,
                        COALESCE(RC.NAME->>'en_US','') AS COUNTRY,
                        COALESCE(RCG.NAME->>'en_US','') AS TERRITORY,
                        COALESCE(SO.ORDERED_QTY,0) AS ORDER_QTY,
                        COALESCE(R_CUR.NAME,'') AS CURRENCY,
                        COALESCE(ROUND(AM.AMOUNT_WITHOUT_DISCOUNT,2),0.0) AS GROSS_SALE,
                        COALESCE(ROUND(AM.AMOUNT_DISCOUNT,2),0.0) AS DISCOUNT,
                        COALESCE(ROUND(AM.AMOUNT_TAX,2),0.0) AS TAX,
                        COALESCE(ROUND(AM.AMOUNT_TOTAL,2),0.0) AS NET_SALE,
                        COALESCE(SUM(KCL.AMOUNT),00) AS COMMISION,
                        CASE 
                        WHEN KCL.state = 'partial' THEN 'Partial Paid'
                        WHEN KCL.state = 'full' THEN 'Full Paid'
                        WHEN KCL.state = 'over' THEN 'Over Paid'
                        WHEN KCL.state = 'draft' THEN 'Draft'
                        END AS STATUS
                        FROM KITS_COMMISSION_LINES AS KCL
                        left JOIN ACCOUNT_MOVE AS AM ON KCL.INVOICE_ID = AM.ID 
                        left JOIN SALE_ORDER AS SO ON SO.NAME = AM.INVOICE_ORIGIN
                        left JOIN RES_PARTNER AS RP ON SO.PARTNER_ID = RP.ID
                        left JOIN RES_COUNTRY AS RC ON RP.COUNTRY_ID = RC.ID
                        LEFT JOIN RES_COUNTRY_GROUP AS RCG ON RP.TERRITORY = RCG.ID
                        left JOIN RES_CURRENCY AS R_CUR ON AM.CURRENCY_ID = R_CUR.ID
                        WHERE {set_date}
                            KCL.STATE {commission_type}
                            AND KCL.USER_ID = {sale_person.id}
                            AND KCL.COMMISSION_FOR = '{commission_for}'
                        GROUP BY SO.DATE_ORDER,SO.NAME,RP.NAME,RC.NAME,RCG.NAME,SO.ORDERED_QTY,R_CUR.NAME,AM.AMOUNT_WITHOUT_DISCOUNT,AM.AMOUNT_DISCOUNT,AM.AMOUNT_TAX,AM.AMOUNT_TOTAL,AM.NAME,AM.PAYMENT_STATE,KCL.STATE,KCL.AMOUNT,SO.AMOUNT_PAID,SO.PICKED_QTY_ORDER_TOTAL '''

            self.env.cr.execute(query)
            record_data = self.env.cr.fetchall()
            total_lines = {}
            sheet_index = 0
            sheet = workbook.create_sheet(title=sale_person.name.replace('/','_'),index=sheet_index)
            _font = Font(name="Lato", sz=9)
            {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
            header_font = Font(name='Lato',size='9',bold=True)
            address_font = Font(name='Lato',size='9')
            bd = Side(style='thin', color="000000")
            top_border = Border(top=bd)


            address_alignment = Alignment(vertical='center', horizontal='left', text_rotation=0)
            sheet.merge_cells('A2:E6')
            img = BytesIO()
            img.flush()
            img.write(base64.b64decode(self.env.company.logo_web))
            image = openpyxl.drawing.image.Image(img)
            image.width = 400
            image.height = 60
            sheet.add_image(image, 'A3')
            sheet.merge_cells('G1:H1')
            sheet.cell(row=1,column=1).alignment=Alignment(vertical='center',horizontal='center')
            sheet.cell(row=1,column=7).font=Font(name='Lato',size='9',bold=True)
            sheet.cell(row=1,column=7).alignment = address_alignment
            sheet.cell(row=1,column=7).value=self.env.company.name
            sheet.merge_cells('G2:H2')
            sheet.cell(row=2,column=7).value='(A division of Tanzacan Tradelink Inc.)'
            sheet.cell(row=2,column=7).alignment = address_alignment
            sheet.cell(row=2,column=7).font=address_font
            sheet.merge_cells('G3:H3')
            sheet.cell(row=3,column=7).value=self.env.company.street
            sheet.cell(row=3,column=7).alignment = address_alignment
            sheet.cell(row=3,column=7).font=address_font
            city = self.env.company.city
            state = self.env.company.state_id.name
            country = self.env.company.country_id.name
            country_zip = self.env.company.zip
            sheet.merge_cells('G4:H4')
            sheet.cell(row=4,column=7).value='%s %s %s %s'%(city,state,country,country_zip)
            sheet.cell(row=4,column=7).font=address_font
            sheet.cell(row=4,column=7).alignment = address_alignment
            sheet.merge_cells('G5:H5')
            sheet.cell(row=5,column=7).value='Tel. %s'%(str(self.env.company.phone))
            sheet.cell(row=5,column=7).font=address_font
            sheet.cell(row=5,column=7).alignment = address_alignment
            sheet.merge_cells('G6:H6')
            sheet.cell(row=6,column=7).value='Email. %s'%(str(self.env.company.email))
            sheet.cell(row=6,column=7).alignment = address_alignment
            sheet.cell(row=6,column=7).font=address_font
            address = self.create_address_line_for_commission(sale_person, take_name=True)
            sheet.merge_cells('A8:C8')
            sheet.cell(row=8,column=1).value = 'Address:'
            sheet.cell(row=8, column=1).font = header_font
            sheet.merge_cells('A9:C15')
            sheet.cell(row=9, column=1).alignment = top_alignment
            sheet.cell(row=9,column=1).value = address

            sheet.cell(row=8,column=6).value = 'Start Date:'
            sheet.cell(row=8, column=6).font = header_font
            sheet.cell(row=8,column=7).value = '%s'%(str(self.start_date) if self.start_date else "")
            sheet.cell(row=9,column=6).value = 'End Date:'
            sheet.cell(row=9, column=6).font = header_font
            sheet.cell(row=9,column=7).value = '%s'%(str(self.end_date) if self.end_date else '')

            sheet.cell(row=10, column=6).value = 'Commission:'
            sheet.cell(row=10, column=6).font = header_font
            sheet.merge_cells('G10:H10')
            sheet.cell(row=10, column=7).value = '%s'%(sale_person.commission_rule_id.name if self.commission_for == 'sales_person' else sale_person.manager_commission_rule_id.name if self.commission_for == 'sales_manager' else '')
            sheet.cell(row=10, column=7).alignment = Alignment(wrap_text=True)
            sheet.cell(row=10, column=7).font = header_font
            sheet.merge_cells('D16:E16')
            table_header = 16
            sheet.cell(row=table_header, column=1).value = 'Date'
            sheet.cell(row=table_header, column=2).value = 'Order'
            sheet.cell(row=table_header, column=3).value = 'Invoice'
            sheet.cell(row=table_header, column=4).value = 'Customer'
            sheet.cell(row=table_header, column=6).value = 'Currency'
            sheet.cell(row=table_header, column=7).value = 'Status'
            sheet.cell(row=table_header, column=8).value = 'Commission'

            sheet.cell(row=table_header, column=1).font = header_font
            sheet.cell(row=table_header, column=1).alignment = center_alignment
            sheet.cell(row=table_header, column=2).font = header_font
            sheet.cell(row=table_header, column=2).alignment = center_alignment
            sheet.cell(row=table_header, column=3).font = header_font
            sheet.cell(row=table_header, column=3).alignment = center_alignment
            sheet.cell(row=table_header, column=4).font = header_font
            sheet.cell(row=table_header, column=4).alignment = left_alignment
            sheet.cell(row=table_header, column=6).font = header_font
            sheet.cell(row=table_header, column=6).alignment = center_alignment
            sheet.cell(row=table_header, column=7).font = header_font
            sheet.cell(row=table_header, column=7).alignment = center_alignment
            sheet.cell(row=table_header, column=8).font = header_font
            sheet.cell(row=table_header, column=8).alignment = right_alignment
           
            row_index=table_header+1

            paid_unpaid_total = {}
            for data in record_data:
                if data[12]:
                    sheet.cell(row=row_index, column=1).value = data[0].strftime("%d-%m-%Y")
                    sheet.cell(row=row_index, column=1).alignment = center_alignment
                    sheet.cell(row=row_index, column=2).value = data[1]
                    sheet.cell(row=row_index, column=2).alignment = center_alignment
                    sheet.cell(row=row_index, column=3).value = data[2]
                    sheet.cell(row=row_index, column=3).alignment = center_alignment
                    sheet.cell(row=row_index, column=4).value = data[3]
                    sheet.cell(row=row_index, column=4).alignment = left_alignment
                    merge_range='D%s:E%s'%(row_index,row_index)
                    sheet.merge_cells(merge_range)
                    sheet.cell(row=row_index, column=6).value = data[13]
                    sheet.cell(row=row_index, column=6).alignment = center_alignment
                    sheet.cell(row=row_index, column=7).value = data[7]
                    sheet.cell(row=row_index, column=7).alignment = center_alignment
                    sheet.cell(row=row_index, column=8).value = '$ {:,.2f}'.format(data[12])
                    sheet.cell(row=row_index, column=8).alignment = right_alignment
                    row_index+=1

                    if data[7] in total_lines:
                        total_lines[data[7]]['commission'] = round(total_lines[data[7]]['commission']+data[12],2)
                    else:
                        total_lines[data[7]] = {
                            'commission':data[12],
                        }

            sheet.merge_cells('A'+str(row_index)+':H'+str(row_index))
            total_row = row_index + 1
            for total_line in total_lines:
                # sheet.cell(row=total_row, column=7).value = total_lines[total_line]['qty'] or 0.00
                # sheet.cell(row=total_row, column=7).font = header_font
                sheet.merge_cells(merge_range)
                sheet.cell(row=total_row, column=7).value = "Total (%s)"%(total_line)
                sheet.cell(row=total_row, column=7).font = header_font
                sheet.cell(row=total_row, column=7).alignment = left_alignment
                sheet.cell(row=total_row, column=7).border = top_border
                # sheet.cell(row=total_row, column=9).value = '$ {:,.2f}'.format(total_lines[total_line]['gross_sale'] or 0.00)
                # sheet.cell(row=total_row, column=9).font = header_font
                # sheet.cell(row=total_row, column=9).alignment = right_alignment
                # sheet.cell(row=total_row, column=10).value = '$ {:,.2f}'.format(total_lines[total_line]['discount'] or 0.00)
                # sheet.cell(row=total_row, column=10).font = header_font
                # sheet.cell(row=total_row, column=10).alignment = right_alignment
                # sheet.cell(row=total_row, column=11).value = '$ {:,.2f}'.format(total_lines[total_line]['tax'] or 0.00)
                # sheet.cell(row=total_row, column=11).font = header_font
                # sheet.cell(row=total_row, column=11).alignment = right_alignment
                # sheet.cell(row=total_row, column=12).value = '$ {:,.2f}'.format(total_lines[total_line]['net_sale'] or 0.00)
                # sheet.cell(row=total_row, column=12).font = header_font
                # sheet.cell(row=total_row, column=12).alignment = right_alignment
                sheet.cell(row=total_row, column=8).value = '$ {:,.2f}'.format(total_lines[total_line]['commission'] or 0.00)
                sheet.cell(row=total_row, column=8).border = top_border
                # sheet.cell(row=total_row, column=7).font = header_font
                sheet.cell(row=total_row, column=8).alignment = right_alignment
                total_row += 1

            # total_row += 1
            # for paid_unpaid in paid_unpaid_total:
            #     if self.commission_is == 'is_paid':
            #         sheet.cell(row=total_row, column=13).value = "Paid Commission (%s)"%(paid_unpaid)
            #         sheet.cell(row=total_row, column=13).font = header_font
            #         sheet.cell(row=total_row, column=14).value = '$ {:,.2f}'.format(paid_unpaid_total[paid_unpaid]['paid'] or 0.00)
            #         sheet.cell(row=total_row, column=14).font = header_font
            #         sheet.cell(row=total_row, column=14).alignment = right_alignment
            #     if self.commission_is == 'is_unpaid':
            #         sheet.cell(row=total_row, column=13).value = "Unpaid Commission (%s)"%(paid_unpaid)
            #         sheet.cell(row=total_row, column=13).font = header_font
            #         sheet.cell(row=total_row, column=14).value = '$ {:,.2f}'.format(paid_unpaid_total[paid_unpaid]['unpaid'] or 0.00)
            #         sheet.cell(row=total_row, column=14).font = header_font
            #         sheet.cell(row=total_row, column=14).alignment = right_alignment
            #     if self.commission_is =='all':
            #         sheet.cell(row=total_row, column=13).value = "Paid Commission (%s)"%(paid_unpaid)
            #         sheet.cell(row=total_row, column=13).font = header_font
            #         sheet.cell(row=total_row, column=14).value = '$ {:,.2f}'.format(paid_unpaid_total[paid_unpaid]['paid'] or 0.00)
            #         sheet.cell(row=total_row, column=14).font = header_font
            #         sheet.cell(row=total_row, column=14).alignment = right_alignment
            #         total_row += 1
            #         sheet.cell(row=total_row, column=13).value = "Unpaid Commission (%s)"%(paid_unpaid)
            #         sheet.cell(row=total_row, column=13).font = header_font
            #         sheet.cell(row=total_row, column=14).value = '$ {:,.2f}'.format(paid_unpaid_total[paid_unpaid]['unpaid'] or 0.00)
            #         sheet.cell(row=total_row, column=14).font = header_font
            #         sheet.cell(row=total_row, column=14).alignment = right_alignment
            #     total_row += 1

            # Colomn size
            sheet.column_dimensions['A'].width = 11
            sheet.column_dimensions['B'].width = 11
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 10
            sheet.column_dimensions['F'].width = 13
            sheet.column_dimensions['G'].width = 11
            sheet.column_dimensions['H'].width = 21
            sheet.column_dimensions['I'].width = 17
            sheet.column_dimensions['J'].width = 15
            sheet.column_dimensions['K'].width = 15
            sheet.column_dimensions['L'].width = 15
            sheet.column_dimensions['M'].width = 30
            sheet.column_dimensions['N'].width = 25
            sheet_index+=1
        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=sales.commission.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,'sale_commission_report'),
            'target': 'self',
        }

    def validate_dates(self):
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise UserError(_("Start Date should be lesser than End Date."))

    def data_dict_pdf(self,order,commission,commission_is):
        data_dict = {}
        sale_order = self.env['sale.order'].search([('invoice_ids','in',order.ids)])
        invoice_num = ''
        # commission_is = ''
        # if order.inv_payment_status == 'full':
        #     commission_is = 'Fully Paid'
        # elif order.inv_payment_status == 'partial':
        #     commission_is = 'Partial Paid'
        # elif order.inv_payment_status == 'over':
        #     commission_is = 'Over Paid'
        # elif order.inv_payment_status == 'draft':
        #     commission_is = 'Draft'
        # else:
        #     commission_is = 'Unpaid'

        data_dict['gross_sale'] = order.amount_without_discount
        data_dict['date'] = sale_order.date_order.strftime("%d-%m-%Y")
        data_dict['discount'] = order.amount_discount
        data_dict['net_sale'] = order.amount_total
        data_dict['sale_order'] = sale_order.name
        data_dict['customer'] = order.partner_id.display_name
        data_dict['country'] = order.partner_id.country_id.name
        data_dict['territory'] = order.partner_id.territory.name
        data_dict['qty'] = sale_order.ordered_qty
        data_dict['tax'] = order.amount_tax
        data_dict['currency'] = order.currency_id.name
        data_dict['commission'] = round(commission, 2)
        data_dict['commission_is'] = commission_is

        if order.state == 'posted':
            invoice_num = invoice_num + order.name + ', '
        data_dict['invoice_order'] = invoice_num[:-2]

        return data_dict

    def get_currency_total(self,lines):
        currency_dict = {}
        for line in lines:
            currency = line['currency']
            if currency not in currency_dict.keys():
                currency_dict[currency] = {'gross_sale':line['gross_sale'],'discount':line['discount'],'tax':line['tax'],'net_sale':line['net_sale'],'commission':line['commission']}
            else:
                currency_dict[currency]['gross_sale'] = round(currency_dict[currency]['gross_sale']+line['gross_sale'],2)
                currency_dict[currency]['discount'] = round(currency_dict[currency]['discount']+line['discount'],2)
                currency_dict[currency]['tax'] = round(currency_dict[currency]['tax']+line['tax'],2)
                currency_dict[currency]['net_sale'] = round(currency_dict[currency]['net_sale']+line['net_sale'],2)
                currency_dict[currency]['commission'] = round(currency_dict[currency]['commission']+line['commission'],2)
        
        return currency_dict

    def get_paid_unpaid_commision_total(self,lines):
        paid_unpaid_total = {}
        for line in lines:
            if line['currency'] in paid_unpaid_total:
                paid_unpaid_total[line['currency']]['paid'] = paid_unpaid_total[line['currency']]['paid'] + line['commission'] if line['commission_is'] == 'Paid' else paid_unpaid_total[line['currency']]['paid']
                paid_unpaid_total[line['currency']]['unpaid'] = paid_unpaid_total[line['currency']]['unpaid'] + line['commission'] if line['commission_is'] == 'Unpaid' else paid_unpaid_total[line['currency']]['unpaid']
            else:
                paid_unpaid_total[line['currency']] = {'paid':line['commission'] if line['commission_is'] == 'Paid' else 0.0,'unpaid':line['commission'] if line['commission_is'] == 'Unpaid' else 0.0}

        return paid_unpaid_total
    
    @api.model
    def _fields_view_get(self, view_id=None, view_type=None, toolbar=False, submenu=False):
        res = super(sales_commission_report_wizard, self)._fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar, submenu=submenu)
        if self._context.get('kits_website_name'):
            doc = etree.XML(res['arch'])
            for node in doc.xpath("//field[@name='website_id']"):
                node.attrib['invisible'] = '1'
            res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    # def is_accessible_to(self,user):
    #     self = self.sudo()
    #     self.ensure_one()
    #     result = False
    #     if user:
    #         if user.has_group('base.group_system') or user.has_group('tzc_sales_customization_spt.group_sales_manager_spt') or user.has_group('sales_team.group_sale_salesman'):
    #             result = True
    #     return result
