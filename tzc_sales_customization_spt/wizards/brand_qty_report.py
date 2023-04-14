from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
from datetime import datetime, timedelta
from dateutil import tz
from io import BytesIO
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

import pandas as pd
import base64

class brand_qty_report_wizard(models.TransientModel):
    _name = 'brand.qty.report.wizard'
    _description = 'Monthly Brand Sales Report Wizard'

    years = {'2015':'2015'}

    def _get_years(self):
        last = max(self.years.keys())
        current = datetime.now().year
        if int(last) < current:
            self.years.update(dict([[str(i),str(i)] for i in range(int(last)+1,current+1)]))
        return [(year,year) for year in self.years]

    start_month = fields.Selection([('1', 'Jan'), ('2', 'Feb'), ('3', 'March'), ('4', 'April'), ('5', 'May'), ('6', 'Jun'), (
        '7', 'July'), ('8', 'Aug'), ('9', 'Sep'), ('10', 'Oct'), ('11', 'Nov'), ('12', 'Dec')], string="Month From ", required=True)
    end_month = fields.Selection([('1', 'Jan'), ('2', 'Feb'), ('3', 'March'), ('4', 'April'), ('5', 'May'), ('6', 'Jun'), (
        '7', 'July'), ('8', 'Aug'), ('9', 'Sep'), ('10', 'Oct'), ('11', 'Nov'), ('12', 'Dec')], string="Month To ", required=True)
    start_year = fields.Selection(_get_years,"Year From ", required=True)
    end_year = fields.Selection(_get_years,"Year To ", required=True)
    brand_ids = fields.Many2many('product.brand.spt',string='Brands')
    brand_id = fields.Many2one('product.brand.spt','Brand')
    file = fields.Binary()
    brand_selection = fields.Selection([('single','Single Brand'),('multiple','Multiple Brand')],'For ?',default='single')
    

    def action_print_report_file(self):
        if not (int(self.start_year) == datetime.now().year and int(self.start_month) == datetime.now().month ) and (not (self.start_month,self.start_year) > (self.end_month,self.end_year)):
            start = '%s-%s-01' % (self.start_year,self.start_month)
            end = '%s-%s-01' % (self.end_year, self.end_month)
            months = [i.strftime("%b-%Y") for i in pd.date_range(start=start, end=end, freq='MS')]
            tz_from, tz_to = tz.gettz(datetime.now().tzinfo), (tz.gettz(self.env.user.tz))
            alignment_right = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
            end_date = datetime.now().replace(tzinfo=tz_from).astimezone(tz=tz_to)
            workbook = Workbook()
            summary_data = []
            if self.brand_selection == 'multiple':
                if self.brand_ids:
                    for brand in self.brand_ids.sorted(lambda x:x.name,reverse=True):
                        header_font = Font(name='Calibri',size='11',bold=True)
                        bd = Side(style='thin', color="000000")
                        top_bottom_border = Border(top=bd,bottom=bd)
                        sheet_index = 0
                        table_header = 1
                        total_row = table_header + 1
                        sheet = workbook.create_sheet(title=brand.name,index=sheet_index)
                        sheet.cell(row=table_header, column=1).value = 'SKU'
                        sheet.cell(row=table_header, column=2).value = 'Quantity'
                        sheet.cell(row=table_header, column=3).value = 'Average Selling Unit Price'
                        sheet.cell(row=table_header, column=4).value = 'Sales'
                        sheet.cell(row=table_header, column=5).value = 'Month'
                        sheet.cell(row=table_header, column=6).value = "Total QTY = "
                        sheet.cell(row=table_header, column=8).value = brand.name
                        sheet.cell(row=table_header, column=9).value = months[0] + '  to  ' + months[-1]
                        
                        sheet.cell(row=table_header, column=1).font = header_font
                        sheet.cell(row=table_header, column=1).border = top_bottom_border
                        sheet.cell(row=table_header, column=1).alignment = alignment_right
                        sheet.cell(row=table_header, column=2).font = header_font
                        sheet.cell(row=table_header, column=2).border = top_bottom_border
                        sheet.cell(row=table_header, column=2).alignment = alignment_right
                        sheet.cell(row=table_header, column=3).font = header_font
                        sheet.cell(row=table_header, column=3).border = top_bottom_border
                        sheet.cell(row=table_header, column=3).alignment = alignment_right
                        sheet.cell(row=table_header, column=4).font = header_font
                        sheet.cell(row=table_header, column=4).border = top_bottom_border
                        sheet.cell(row=table_header, column=4).alignment = alignment_right
                        sheet.cell(row=table_header, column=5).font = header_font
                        sheet.cell(row=table_header, column=5).border = top_bottom_border
                        sheet.cell(row=table_header, column=5).alignment = alignment_right
                        sheet.cell(row=table_header, column=6).font = header_font
                        sheet.cell(row=table_header, column=6).border = top_bottom_border
                        sheet.cell(row=table_header, column=6).alignment = alignment_right
                        sheet.cell(row=table_header, column=7).font = header_font
                        sheet.cell(row=table_header, column=7).border = top_bottom_border
                        sheet.cell(row=table_header, column=7).alignment = alignment_right
                        sheet.cell(row=table_header, column=8).font = header_font
                        sheet.cell(row=table_header, column=8).border = top_bottom_border
                        sheet.cell(row=table_header, column=8).alignment = alignment_right
                        sheet.cell(row=table_header, column=9).font = header_font
                        sheet.cell(row=table_header, column=9).border = top_bottom_border
                        sheet.cell(row=table_header, column=9).alignment = alignment_right
                        total_qty = 0
                        for month in months:
                            year = month.split('-')[1]
                            long_month_name = month.split('-')[0]
                            m = datetime.strptime(long_month_name, "%b").month
                            start_date = datetime(int(year), m, 1)
                            start_date = start_date.replace(tzinfo=tz_from).astimezone(tz=tz_to)
                            if m + 1 > 12:
                                m = 0
                            end_date = datetime(int(year),m  + 1, 1) + timedelta(days=-1)
                            end_date = end_date.replace(tzinfo=tz_from).astimezone(tz=tz_to)
                            query =f'''
                            SELECT COALESCE(PP.DEFAULT_CODE,''),
                                    COALESCE(SUM(AML.QUANTITY),0) AS QUANTITY,
                                    COALESCE(ROUND((AML.PRICE_SUBTOTAL / AML.QUANTITY),2),00.00) AS AVERAGE_SALES,
                                    COALESCE(SUM(AML.PRICE_SUBTOTAL),00) AS SALES
                                FROM PRODUCT_PRODUCT AS PP
                                INNER JOIN ACCOUNT_MOVE_LINE AS AML ON AML.PRODUCT_ID = PP.ID
                                INNER JOIN PRODUCT_BRAND_SPT AS PBS ON PP.BRAND = PBS.ID
                                WHERE PBS.ID = {brand.id} AND AML.PARENT_STATE = 'posted' '''
                            if start_date:
                                query = query + " AND AML.CREATE_DATE >= '%s'" % (start_date)
                            if end_date:
                                query = query + " AND AML.CREATE_DATE <= '%s'" % (end_date)
                            query = query + " GROUP BY PP.DEFAULT_CODE,AML.PRICE_SUBTOTAL,AML.QUANTITY ORDER BY PP.DEFAULT_CODE"
                            self.env.cr.execute(query)
                            report_data = self.env.cr.fetchall()
                            for data in report_data:
                                sheet.cell(row=total_row, column=1).value = data[0]
                                sheet.cell(row=total_row, column=2).value = data[1]
                                sheet.cell(row=total_row, column=3).value = data[2]
                                sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
                                sheet.cell(row=total_row, column=4).value = data[3]
                                sheet.cell(row=total_row, column=4).alignment = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
                                sheet.cell(row=total_row, column=5).value = month or ''
                                sheet.cell(row=total_row, column=5).alignment = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
                                summary_data.append(data)                            
                                total_qty += data[1]
                                total_row += 1
                            sheet.cell(row=table_header, column=7).value = total_qty or 0
                        sheet.column_dimensions['A'].width = 30
                        sheet.column_dimensions['B'].width = 15
                        sheet.column_dimensions['C'].width = 25
                        sheet.column_dimensions['D'].width = 15
                        sheet.column_dimensions['E'].width = 20
                        sheet.column_dimensions['F'].width = 20
                        sheet.column_dimensions['G'].width = 20
                        sheet.column_dimensions['H'].width = 28
                        sheet.column_dimensions['I'].width = 28
                        sheet["F1"].fill = PatternFill("solid", start_color="65C18C")
                        sheet["G1"].fill = PatternFill("solid", start_color="65C18C")
                        sheet["H1"].fill = PatternFill("solid", start_color="00FFFF00")
                        sheet["I1"].fill = PatternFill("solid", start_color="00FFFF00")

                    header_font = Font(name='Calibri',size='11',bold=True)
                    bd = Side(style='thin', color="000000")
                    top_bottom_border = Border(top=bd,bottom=bd)
                    sheet_index = 0
                    table_header = 1
                    total_row = table_header + 1
                    sheet = workbook.create_sheet(title='Summary',index=-1)
                    sheet.cell(row=table_header, column=1).value = 'SKU'
                    sheet.cell(row=table_header, column=2).value = 'Quantity'
                    sheet.cell(row=table_header, column=3).value = 'Sales'

                    sheet.cell(row=table_header, column=1).font = header_font
                    sheet.cell(row=table_header, column=1).border = top_bottom_border
                    sheet.cell(row=table_header, column=1).alignment = alignment_right
                    sheet.cell(row=table_header, column=2).font = header_font
                    sheet.cell(row=table_header, column=2).border = top_bottom_border
                    sheet.cell(row=table_header, column=2).alignment = alignment_right
                    sheet.cell(row=table_header, column=3).font = header_font
                    sheet.cell(row=table_header, column=3).border = top_bottom_border
                    sheet.cell(row=table_header, column=3).alignment = alignment_right
                    for data in summary_data:
                        sheet.cell(row=total_row, column=1).value = data[0]
                        sheet.cell(row=total_row, column=2).value = data[1]
                        sheet.cell(row=total_row, column=3).value = data[3]
                        sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
                        total_row += 1
                    sheet.column_dimensions['A'].width = 30
                    sheet.column_dimensions['B'].width = 15
                    sheet.column_dimensions['C'].width = 15
                    sheet.sheet_properties.tabColor = 'ffb366'
                else:
                    raise UserError('Please select at least one brand.')
            else:
                if self.brand_id:
                    for brand in self.brand_id:
                        for month in months[::-1]:
                            year = month.split('-')[1]
                            long_month_name = month.split('-')[0]
                            m = datetime.strptime(long_month_name, "%b").month
                            start_date = datetime(int(year), m, 1)
                            start_date = start_date.replace(tzinfo=tz_from).astimezone(tz=tz_to)
                            if m + 1 > 12:
                                m = 0
                            end_date = datetime(int(year),m  + 1, 1) + timedelta(days=-1)
                            end_date = end_date.replace(tzinfo=tz_from).astimezone(tz=tz_to)
                            query = f'''SELECT COALESCE(PP.DEFAULT_CODE,''),
                                            COALESCE(SUM(AML.QUANTITY),0) AS QUANTITY,
                                            COALESCE(ROUND((AML.PRICE_SUBTOTAL / AML.QUANTITY),2)00.00) AS AVERAGE_SALES,
                                            COALESCE(SUM(AML.PRICE_SUBTOTAL),00) AS SALES
                                        FROM PRODUCT_PRODUCT AS PP
                                        INNER JOIN ACCOUNT_MOVE_LINE AS AML ON AML.PRODUCT_ID = PP.ID
                                        INNER JOIN PRODUCT_BRAND_SPT AS PBS ON PP.BRAND = PBS.ID
                                        WHERE PBS.ID = {brand.id}'''
                            if start_date:
                                query = query + " AND AML.CREATE_DATE >= '%s'" % (start_date)
                            if end_date:
                                query = query + " AND AML.CREATE_DATE <= '%s'" % (end_date)
                            query = query + " GROUP BY PP.DEFAULT_CODE,AML.PRICE_SUBTOTAL,AML.QUANTITY"
                            self.env.cr.execute(query)
                            report_data = self.env.cr.fetchall()

                            
                            header_font = Font(name='Calibri',size='11',bold=True)
                            bd = Side(style='thin', color="000000")
                            top_bottom_border = Border(top=bd,bottom=bd)
                            sheet_index = 0
                            table_header = 1
                            total_row = table_header + 1
                            sheet = workbook.create_sheet(title=month,index=sheet_index)
                            sheet.cell(row=table_header, column=1).value = 'SKU'
                            sheet.cell(row=table_header, column=2).value = 'Quantity'
                            sheet.cell(row=table_header, column=3).value = 'Average Selling Unit Price'
                            sheet.cell(row=table_header, column=4).value = 'Sales'
                            sheet.cell(row=table_header, column=5).value = 'Total QTY = '
                            sheet.cell(row=table_header, column=7).value = brand.name
                            sheet.cell(row=table_header, column=8).value = month
                            
                            sheet.cell(row=table_header, column=1).font = header_font
                            sheet.cell(row=table_header, column=1).border = top_bottom_border
                            sheet.cell(row=table_header, column=1).alignment = alignment_right
                            sheet.cell(row=table_header, column=2).font = header_font
                            sheet.cell(row=table_header, column=2).border = top_bottom_border
                            sheet.cell(row=table_header, column=2).alignment = alignment_right
                            sheet.cell(row=table_header, column=3).font = header_font
                            sheet.cell(row=table_header, column=3).border = top_bottom_border
                            sheet.cell(row=table_header, column=3).alignment = alignment_right
                            sheet.cell(row=table_header, column=4).font = header_font
                            sheet.cell(row=table_header, column=4).border = top_bottom_border
                            sheet.cell(row=table_header, column=4).alignment = alignment_right
                            sheet.cell(row=table_header, column=5).font = header_font
                            sheet.cell(row=table_header, column=5).border = top_bottom_border
                            sheet.cell(row=table_header, column=5).alignment = alignment_right
                            sheet.cell(row=table_header, column=6).font = header_font
                            sheet.cell(row=table_header, column=6).border = top_bottom_border
                            sheet.cell(row=table_header, column=6).alignment = alignment_right
                            sheet.cell(row=table_header, column=7).font = header_font
                            sheet.cell(row=table_header, column=7).border = top_bottom_border
                            sheet.cell(row=table_header, column=7).alignment = alignment_right
                            sheet.cell(row=table_header, column=8).font = header_font
                            sheet.cell(row=table_header, column=8).border = top_bottom_border
                            sheet.cell(row=table_header, column=8).alignment = alignment_right
                            avg_sale_price = 0.0
                            prod_qty = 0
                            total_quantity = 0
                            for data in report_data:
                                sheet.cell(row=total_row, column=1).value = data[0]
                                sheet.cell(row=total_row, column=2).value = data[1]
                                sheet.cell(row=total_row, column=3).value = data[2]
                                sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
                                sheet.cell(row=total_row, column=4).value = data[3]
                                sheet.cell(row=total_row, column=4).alignment = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
                                sheet.cell(row=total_row, column=5).value = month or ''
                                sheet.cell(row=total_row, column=5).alignment = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
                                summary_data.append(data)
                            
                                total_row += 1
                            sheet.cell(row=table_header, column=6).value = int(total_quantity) or 0.0
                            sheet.column_dimensions['A'].width = 30
                            sheet.column_dimensions['B'].width = 15
                            sheet.column_dimensions['C'].width = 25
                            sheet.column_dimensions['D'].width = 28
                            sheet.column_dimensions['E'].width = 20
                            sheet.column_dimensions['F'].width = 20
                            sheet.column_dimensions['G'].width = 28
                            sheet.column_dimensions['H'].width = 28
                            sheet["E1"].fill = PatternFill("solid", start_color="65C18C")
                            sheet["F1"].fill = PatternFill("solid", start_color="65C18C")
                            sheet["G1"].fill = PatternFill("solid", start_color="00FFFF00")
                            sheet["H1"].fill = PatternFill("solid", start_color="00FFFF00")
                        header_font = Font(name='Calibri',size='11',bold=True)
                        bd = Side(style='thin', color="000000")
                        top_bottom_border = Border(top=bd,bottom=bd)
                        sheet_index = 0
                        table_header = 1
                        total_row = table_header + 1
                        sheet = workbook.create_sheet(title='Summary',index=-1)
                        sheet.cell(row=table_header, column=1).value = 'SKU'
                        sheet.cell(row=table_header, column=2).value = 'Quantity'
                        sheet.cell(row=table_header, column=3).value = 'Sales'
                        sheet.cell(row=table_header, column=1).font = header_font
                        sheet.cell(row=table_header, column=1).border = top_bottom_border
                        sheet.cell(row=table_header, column=1).alignment = alignment_right
                        sheet.cell(row=table_header, column=2).font = header_font
                        sheet.cell(row=table_header, column=2).border = top_bottom_border
                        sheet.cell(row=table_header, column=2).alignment = alignment_right
                        sheet.cell(row=table_header, column=3).font = header_font
                        sheet.cell(row=table_header, column=3).border = top_bottom_border
                        sheet.cell(row=table_header, column=3).alignment = alignment_right

                        for data in summary_data:
                            sheet.cell(row=total_row, column=1).value = data[0]
                            sheet.cell(row=total_row, column=2).value = data[1]
                            sheet.cell(row=total_row, column=3).value = data[3]
                            sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='right', vertical='center', text_rotation=0)
                            total_row += 1
                        sheet.column_dimensions['A'].width = 30
                        sheet.column_dimensions['B'].width = 15
                        sheet.column_dimensions['C'].width = 15
                        sheet.sheet_properties.tabColor = 'ffb366'
                else:
                    raise UserError('Please select at least one brand.')

            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            self.file = base64.b64encode(data)

            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=brand.qty.report.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,'Monthly Brand Sales Report'),
                'target': 'self',
            }

        else:
            raise UserError(_('Please check values..'))
