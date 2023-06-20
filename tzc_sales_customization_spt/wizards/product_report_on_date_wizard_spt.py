# -*- coding: utf-8 -*-
# Part of SnepTech See LICENSE file for full copyright and licensing details.##
##############################################################################
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
from odoo.tools import float_compare
from datetime import timedelta,datetime
from dateutil import tz
import pandas as ps
class product_report_on_date_wizard_spt(models.TransientModel):
    _name = 'product.report.on.date.wizard.spt'
    _description = 'Set QTY Available'
    
    start_date = fields.Date('Date')
    file = fields.Binary('File')
    exclude_zero_qty = fields.Boolean('Exclude Zero Qty')


    def action_print_report_file(self):
        self.ensure_one()
        # tz_from = current system timezone,
        # tz_to = User Setting's timezone
        tz_from, tz_to = tz.gettz(datetime.now().tzinfo), (tz.gettz(self.env.user.tz) or self.env.context.get('tz'))
        if self.start_date:
            start_date = datetime.combine(self.start_date, datetime.min.time()) +timedelta(seconds=-1,days=+1)
            start_date = start_date.replace(tzinfo=tz_from).astimezone(tz=tz_to)

        product_ids = self.env['product.product'].search([])
        active_id = self.id
        f_name = 'Product_on_hand'  # FileName
        workbook = Workbook()
        sheet = workbook.create_sheet(
            title=str('Inventory At '+str(self.start_date)), index=0)  # sheet name

        # sheet
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        bottom_border = Border(bottom=bd)
        table_header_font = Font(size=10, bold=True, name="Garamond")
        table_font = Font(size=10, bold=False, name="Garamond")
        alignment = Alignment(
            vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
        left_alignment = Alignment(
            vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
     

        # -------------------------------- Table  --------------------------------
        table_header_row = 1
        sheet.cell(row=table_header_row, column=1).value = 'Brand'
        sheet.cell(row=table_header_row, column=2).value = 'Category'
        sheet.cell(row=table_header_row, column=3).value = 'Qty On Hand'

        sheet.cell(row=table_header_row, column=1).font = table_header_font
        sheet.cell(row=table_header_row, column=2).font = table_header_font
        sheet.cell(row=table_header_row, column=3).font = table_header_font

        sheet.cell(row=table_header_row, column=1).alignment = left_alignment
        sheet.cell(row=table_header_row, column=2).alignment = left_alignment
        sheet.cell(row=table_header_row, column=3).alignment = left_alignment

        sheet.cell(row=table_header_row, column=1).border = top_bottom_border
        sheet.cell(row=table_header_row, column=2).border = top_bottom_border
        sheet.cell(row=table_header_row, column=3).border = top_bottom_border

        row_index = table_header_row + 1
        # query = f'''SELECT COALESCE(PP.VARIANT_NAME,'') AS VARIANTNAME,
        #             COALESCE(PP.DEFAULT_CODE,'') AS DEFAULT_CODE,
        #             COALESCE(PBS.NAME,'') AS BRAND,
        #             COALESCE(PMS.NAME,'') AS MODEL,
        #             COALESCE(PC.NAME,'') AS CATEGORY,
        #             SUM(CASE
        #                 WHEN SL.USAGE = 'inventory'
        #                                     AND S.USAGE = 'internal' THEN SML.QTY_DONE
        #                 WHEN SL.USAGE = 'internal'
        #                                     AND S.USAGE = 'inventory' THEN - SML.QTY_DONE
        #                 WHEN SL.USAGE = 'internal'
        #                                     AND S.USAGE = 'customer' THEN - SML.QTY_DONE
        #                 WHEN SL.USAGE = 'internal'
        #                                     AND S.USAGE = 'supplier' THEN SML.QTY_DONE
        #                 WHEN SL.USAGE = 'customer'
        #                                     AND S.USAGE = 'internal' THEN SML.QTY_DONE
        #                 WHEN SL.USAGE = 'customer'
        #                                     AND S.USAGE = 'inventory' THEN - SML.QTY_DONE
        #             END) AS QTY,
        #             COALESCE(PP.LST_PRICE_USD,0.00) AS LST_PRICE_USD,
        #             COALESCE(SL.USAGE,'') AS INVE_IN,
        #             COALESCE(S.USAGE,'') AS INVE_OUT
        #         FROM STOCK_MOVE_LINE AS SML
        #         INNER JOIN STOCK_LOCATION AS SL ON SML.LOCATION_ID = SL.ID
        #         INNER JOIN STOCK_LOCATION AS S ON SML.LOCATION_DEST_ID = S.ID
        #         INNER JOIN PRODUCT_PRODUCT AS PP ON SML.PRODUCT_ID = PP.ID
        #         INNER JOIN PRODUCT_CATEGORY AS PC ON PP.CATEG_ID = PC.ID
        #         INNER JOIN PRODUCT_MODEL_SPT AS PMS ON PMS.ID = PP.MODEL
        #         INNER JOIN PRODUCT_BRAND_SPT AS PBS ON PBS.ID = PP.BRAND
        #         WHERE SML.PRODUCT_ID = {product_id.id}
        #             AND SML.STATE = 'done'
        #             AND SML.DATE <= '{str(self.start_date)}'
        #         GROUP BY PP.VARIANT_Ntable_header_rowAME,PC.NAME,PMS.NAME,SL.USAGE,S.USAGE,PP.DEFAULT_CODE,PBS.NAME,PP.LST_PRICE_USD ORDER BY PP.VARIANT_NAME'''
        query = f'''select 
                        pbs.name,
                        pc.name,
                        SUM(CASE
                            WHEN SL.USAGE = 'inventory'
                                                AND S.USAGE = 'internal' THEN SML.QTY_DONE
                            WHEN SL.USAGE = 'internal'
                                                AND S.USAGE = 'inventory' THEN - SML.QTY_DONE
                            WHEN SL.USAGE = 'internal'
                                                AND S.USAGE = 'customer' THEN - SML.QTY_DONE
                            WHEN SL.USAGE = 'internal'
                                                AND S.USAGE = 'supplier' THEN SML.QTY_DONE
                            WHEN SL.USAGE = 'customer'
                                                AND S.USAGE = 'internal' THEN SML.QTY_DONE
                            WHEN SL.USAGE = 'customer'
                                                AND S.USAGE = 'inventory' THEN - SML.QTY_DONE
                        END) AS QTY
                    from product_brand_spt pbs
                    left join product_product as pp on pbs.id = pp.brand
                    left join product_category as pc on pp.categ_id = pc.id
                    left join stock_move_line as sml on pp.id = sml.product_id
                    left join stock_location AS sl ON sml.location_id = sl.id
                    left join stock_location AS s ON sml.location_dest_id = s.id
                    where SML.STATE = 'done' AND SML.DATE <= '{str(self.start_date)}'
                    group by pbs.name,pc.name order by pbs.name'''

        self.env.cr.execute(query)
        record_data = self.env.cr.fetchall()
        on_date_products=[]
        # if self.exclude_zero_qty == True and sum(list(map(lambda rd: rd[5],record_data))) == 0.0:
        #     pass
        # else:
        for record in record_data:
            if self.exclude_zero_qty == True and record[2] == 0.0:
                pass
            else:
                sheet.cell(row=row_index,column=1).value = record[0]
                sheet.cell(row=row_index, column=1).font = table_font
                sheet.cell(row=row_index, column=1).alignment = left_alignment
                sheet.cell(row=row_index,column=2).value = record[1]
                sheet.cell(row=row_index, column=2).font = table_font
                sheet.cell(row=row_index, column=2).alignment = left_alignment
                sheet.cell(row=row_index,column=3).value = record[2]
                sheet.cell(row=row_index, column=3).font = table_font

                sheet.cell(row=row_index, column=1).border = bottom_border
                sheet.cell(row=row_index, column=2).border = bottom_border
                sheet.cell(row=row_index, column=3).border = bottom_border

                row_index += 1
                

        # -------------------------------- table end --------------------------------

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 18 
        sheet.column_dimensions['C'].width = 15 

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=product.report.on.date.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }
