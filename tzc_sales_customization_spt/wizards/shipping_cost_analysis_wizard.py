from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
import requests

class shipping_cost_analysis_wizard(models.TransientModel):
    _name = "shipping.cost.analysis.wizard"
    _description = "Shipping Cost Analysis Wizard"

    start_date = fields.Date()
    end_date = fields.Date()
    include_calculated_cost = fields.Boolean('Calculate shipping cost via API ?')
    file = fields.Binary()
    carrier_id = fields.Many2one('delivery.carrier','Shipping Method')

    def validate_dates(self):
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise UserError(_("Start Date should be lesser than End Date."))

    def action_xls_report(self):
        self.validate_dates()
        f_name = 'Shipping Cost Analysis'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Shipping Cost", index=0)
        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        left_alignment = Alignment(
            vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        right_alignment = Alignment(
            vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
        
        # Critria
        header_row = 0
        if self.carrier_id:
            header_row = 1
            # start date
            sheet.merge_cells('A'+str(header_row)+':B'+str(header_row))
            sheet.cell(row=header_row, column=1).value = 'Start Date : {}'.format(str(self.start_date) if self.start_date else '')
            sheet.cell(row=header_row, column=1).font = header_font

            # shipping provider name
            sheet.merge_cells('D'+str(header_row)+':G'+str(header_row))
            sheet.cell(row=header_row, column=4).value = 'Shipping Provider : %s'%(self.carrier_id.name or '')
            sheet.cell(row=header_row, column=4).font = header_font

            header_row += 1
            # end date
            sheet.merge_cells('A'+str(header_row)+':B'+str(header_row))
            sheet.cell(row=header_row, column=1).value = 'End Date : {}'.format(str(self.end_date) if self.end_date else '')
            sheet.cell(row=header_row, column=1).font = header_font

        table_header = header_row + 2
        sheet.cell(row=table_header, column=1).value = 'Shipping Date'
        sheet.cell(row=table_header, column=2).value = 'Order'
        sheet.cell(row=table_header, column=3).value = 'Order Status'
        sheet.cell(row=table_header, column=4).value = 'Salesperson'
        sheet.cell(row=table_header, column=5).value = 'Order Amount'
        sheet.cell(row=table_header, column=6).value = 'Currency'
        sheet.cell(row=table_header, column=7).value = 'Quantity'
        sheet.cell(row=table_header, column=8).value = 'Volume (m3)'
        sheet.cell(row=table_header, column=9).value = 'Actual Weight (kg)'
        sheet.cell(row=table_header, column=10).value = 'Calculated Weight (kg)'
        sheet.cell(row=table_header, column=11).value = 'Customer ID'
        sheet.cell(row=table_header, column=12).value = 'Customer'
        sheet.cell(row=table_header, column=13).value = 'Address'
        sheet.cell(row=table_header, column=14).value = 'City'
        sheet.cell(row=table_header, column=15).value = 'Zip'
        sheet.cell(row=table_header, column=16).value = 'Country'
        sheet.cell(row=table_header, column=17).value = 'Carrier'
        sheet.cell(row=table_header, column=18).value = 'Tracking Number'
        sheet.cell(row=table_header, column=19).value = 'Shipping Cost'
        sheet.cell(row=table_header, column=20).value = 'Carrier Method'
        sheet.cell(row=table_header, column=21).value = 'Calculated Shipping Cost'
        sheet.cell(row=table_header, column=22).value = 'Difference'
        sheet.cell(row=table_header, column=23).value = 'Error'

        sheet.cell(row=table_header, column=1).font = header_font
        sheet.cell(row=table_header, column=1).border = top_bottom_border
        sheet.cell(row=table_header, column=2).font = header_font
        sheet.cell(row=table_header, column=2).border = top_bottom_border
        sheet.cell(row=table_header, column=3).font = header_font
        sheet.cell(row=table_header, column=3).border = top_bottom_border
        sheet.cell(row=table_header, column=4).font = header_font
        sheet.cell(row=table_header, column=4).border = top_bottom_border
        sheet.cell(row=table_header, column=5).font = header_font
        sheet.cell(row=table_header, column=5).border = top_bottom_border
        sheet.cell(row=table_header, column=5).alignment = right_alignment
        sheet.cell(row=table_header, column=6).font = header_font
        sheet.cell(row=table_header, column=6).border = top_bottom_border
        sheet.cell(row=table_header, column=6).alignment = right_alignment
        sheet.cell(row=table_header, column=7).font = header_font
        sheet.cell(row=table_header, column=7).border = top_bottom_border
        sheet.cell(row=table_header, column=8).font = header_font
        sheet.cell(row=table_header, column=8).border = top_bottom_border
        sheet.cell(row=table_header, column=9).font = header_font
        sheet.cell(row=table_header, column=9).border = top_bottom_border
        sheet.cell(row=table_header, column=10).font = header_font
        sheet.cell(row=table_header, column=10).border = top_bottom_border
        sheet.cell(row=table_header, column=11).font = header_font
        sheet.cell(row=table_header, column=11).border = top_bottom_border
        sheet.cell(row=table_header, column=12).font = header_font
        sheet.cell(row=table_header, column=12).border = top_bottom_border
        sheet.cell(row=table_header, column=13).font = header_font
        sheet.cell(row=table_header, column=13).border = top_bottom_border
        sheet.cell(row=table_header, column=14).font = header_font
        sheet.cell(row=table_header, column=14).border = top_bottom_border
        sheet.cell(row=table_header, column=15).font = header_font
        sheet.cell(row=table_header, column=15).border = top_bottom_border
        sheet.cell(row=table_header, column=16).font = header_font
        sheet.cell(row=table_header, column=16).border = top_bottom_border
        sheet.cell(row=table_header, column=17).font = header_font
        sheet.cell(row=table_header, column=17).border = top_bottom_border
        sheet.cell(row=table_header, column=18).font = header_font
        sheet.cell(row=table_header, column=18).border = top_bottom_border
        sheet.cell(row=table_header, column=19).font = header_font
        sheet.cell(row=table_header, column=19).border = top_bottom_border
        sheet.cell(row=table_header, column=19).alignment = right_alignment
        sheet.cell(row=table_header, column=20).font = header_font
        sheet.cell(row=table_header, column=20).border = top_bottom_border
        sheet.cell(row=table_header, column=21).font = header_font
        sheet.cell(row=table_header, column=21).border = top_bottom_border
        sheet.cell(row=table_header, column=22).font = header_font
        sheet.cell(row=table_header, column=22).border = top_bottom_border
        sheet.cell(row=table_header, column=22).alignment = right_alignment
        sheet.cell(row=table_header, column=23).font = header_font
        sheet.cell(row=table_header, column=23).border = top_bottom_border

        row_index=table_header+1

        query = f'''SELECT 	COALESCE(SO.NAME,'') AS NAME,
                            COALESCE(SO.STATE,'')AS STATE,
                            SO.SHIPPED_DATE,
                            COALESCE(RP_USER.NAME,'') AS SALESPERSON,
                            CASE
                                WHEN SO.STATE = 'shipped' THEN SO.AMOUNT_TOTAL ELSE 00
                            END AS TOTAL_AMOUNT,
                            COALESCE(RC.NAME,'') AS CURRENCY,
                            COALESCE(SO.AMOUNT_IS_SHIPPING_TOTAL,00.00) AS AMOUNT_IS_SHIPPING_TOTAL,
                            COALESCE(SO.DELIVERED_QTY,0) AS DELIVERED_QTY,
                            COALESCE(RP.INTERNAL_ID,'') AS CUSTOMER_ID,
                            COALESCE(RP.NAME,'') AS CUSTOMER,
                            COALESCE(R_COU.NAME->>'en_US','') AS COUNTRY,
                            COALESCE(RP.CITY,'') AS CITY,
                            COALESCE(RP.ZIP,'') AS ZIP,
                            COALESCE(RP.STREET,'')AS STREET,
                            COALESCE(RP.STREET2,'') AS STREET2,
                            COALESCE(SUM(PP.LENGTH * PP.WIDTH * PP.HEIGHT/100),00) AS PRODUCT_VOLUME,
                            COALESCE(SO.WEIGHT_TOTAL_KG,00) AS CALCULATED_WEIGHT,
                            COALESCE(SO.ACTUAL_WEIGHT,00) AS ACTUAL_WEIGHT,
                            CASE
                                WHEN SP.STATE NOT IN ('cancel') AND SP.NAME LIKE 'WH/OUT%' THEN SPS.NAME ELSE ''
                            END AS CARRIER,
                            CASE
                                WHEN SP.STATE NOT IN ('cancel') AND SP.NAME LIKE 'WH/OUT%' THEN SP.TRACKING_NUMBER_SPT ELSE ''
                            END AS TRACKING_NO
                    FROM SALE_ORDER AS SO
                    INNER JOIN SHIPPING_PROVIDER_SPT AS SPS ON SPS.ID = SO.SHIPPING_ID
                    INNER JOIN RES_CURRENCY AS RC ON RC.ID = SO.CURRENCY_ID
                    INNER JOIN SALE_ORDER_LINE AS SOL ON SOL.ORDER_ID = SO.ID
                    INNER JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                    INNER JOIN RES_PARTNER AS RP ON RP.ID = SO.PARTNER_ID
                    INNER JOIN RES_COUNTRY AS R_COU ON R_COU.ID = RP.COUNTRY_ID
                    INNER JOIN STOCK_PICKING AS SP ON SP.ORIGIN = SO.NAME
                    INNER JOIN RES_USERS AS RU ON RU.ID = SO.USER_ID
                    INNER JOIN RES_PARTNER AS RP_USER ON RP_USER.ID = RU.PARTNER_ID
                    WHERE SO.STATE IN ('shipped','draft_inv','open_inv','paid') AND SPS.NAME = 'Pick Up' '''
        if self.start_date:
            query = query + " AND SO.DATE_ORDER >= '%s'" % (str(self.start_date))
        if self.end_date:
            query = query + " AND SO.DATE_ORDER <= '%s'" % (str(self.end_date))
        query = query + " GROUP BY SO.NAME,SO.STATE,SO.SHIPPED_DATE,SO.AMOUNT_TOTAL,RC.NAME,SO.AMOUNT_IS_SHIPPING_TOTAL,SO.DELIVERED_QTY,SO.ID,RP.INTERNAL_ID,RP.NAME,RP.CITY,RP.ZIP,R_COU.NAME,RP.STREET,RP.STREET2,SP.STATE,SP.NAME,SPS.NAME,SP.TRACKING_NUMBER_SPT,RP_USER.NAME"
        self.env.cr.execute(query)
        record_data = self.env.cr.fetchall()
        for data in record_data:
            order = self.env['sale.order'].search([('name','=',data[0])])
            sheet.cell(row=row_index, column=1).value = data[2].strftime("%d-%m-%Y") if data[2] else ''
            sheet.cell(row=row_index, column=2).value = data[0]
            sheet.cell(row=row_index, column=3).value = dict((order._fields['state'].selection)).get(data[1])
            sheet.cell(row=row_index, column=4).value = data[3]
            sheet.cell(row=row_index, column=5).value = "{:,.2f}".format(data[4])
            sheet.cell(row=row_index, column=5).alignment = right_alignment
            sheet.cell(row=row_index, column=6).value = data[5]
            sheet.cell(row=row_index, column=6).alignment = right_alignment
            sheet.cell(row=row_index, column=7).value = data[7]
            sheet.cell(row=row_index, column=7).alignment = left_alignment
            sheet.cell(row=row_index, column=8).value = round(data[15],2)
            sheet.cell(row=row_index, column=8).alignment = left_alignment
            sheet.cell(row=row_index, column=9).value = round(data[16],2)
            sheet.cell(row=row_index, column=9).alignment = left_alignment
            sheet.cell(row=row_index, column=10).value = round(data[17],2)
            sheet.cell(row=row_index, column=10).alignment = left_alignment
            sheet.cell(row=row_index, column=11).value = data[8]
            sheet.cell(row=row_index, column=12).value = data[9]
            sheet.cell(row=row_index, column=13).value = '{},'.format(data[13])+'{}'.format(data[14])
            sheet.cell(row=row_index, column=13).alignment = left_alignment
            sheet.cell(row=row_index, column=14).value = data[11]
            sheet.cell(row=row_index, column=14).alignment = left_alignment
            sheet.cell(row=row_index, column=15).value = data[12]
            sheet.cell(row=row_index, column=15).alignment = left_alignment
            sheet.cell(row=row_index, column=16).value = data[10]
            sheet.cell(row=row_index, column=16).alignment = left_alignment
            sheet.cell(row=row_index, column=17).value = data[18] if data[18] else "None"
            sheet.cell(row=row_index, column=18).value = data[19]
            order_data = {}
            if self.carrier_id:
                response = self.get_shipping_method(order)
                order_data['carrier_method'] = self.carrier_id.name or ''
                if response:
                    if not response['error_message'] and response['price'] != False and response['success'] == True:
                        order_data['calcu_shipping_cost'] = round(response['price'],2)
                        order_data['differ'] = abs(data[6] - order_data['calcu_shipping_cost']) if order_data['calcu_shipping_cost'] else 0.0
                    else:
                        order_data['calcu_shipping_cost'] = 'N\A'
                        order_data['differ'] = 0.0
                    if response.get('error_message'):
                        product_name = ',\n'.join(order.order_line.filtered(lambda x: not x.product_id.weight and not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount).mapped('product_id.display_name'))
                        # FEDEX Api error.
                        if len(response['error_message'].split(':')) >= 3 and int(response['error_message'].split(':')[1][1:]) == 809:
                            if product_name:
                                error_msg += "These products have no Weight : \n" + product_name
                            undelivered_products = ',\n'.join(order.order_line.filtered(lambda x: not x.picked_qty and not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount).mapped('product_id.display_name'))
                            if undelivered_products:
                                error_msg += "These products are not picked yet : \n" + undelivered_products
                            self.create_error_log(order,error_msg)
                            order_data['error'] = error_msg
                        elif len(response['error_message'].split(':')) >= 3 and (int(response['error_message'].split(':')[1][1:]) == 521 or int(response['error_message'].split(':')[1][1:]) == 522 or int(response['error_message'].split(':')[1][1:]) == 711 or int(response['error_message'].split(':')[1][1:]) == 868):
                            self.create_error_log(order,response.get('error_message'))
                            order_data['error'] = response['error_message'].split(':')[-1].strip() 
                        # UPS Api error.
                        elif 'missing field' in (response['error_message']).lower():
                            self.create_error_log(order,response.get('error_message').split('\n')[1][1:-1])
                            order_data['error'] = response.get('error_message').split('\n')[1][1:-1]
                        elif 'the estimated price cannot be computed because the weight of your product' in (response['error_message']).lower():
                            if order.order_line.filtered(lambda x: not x.product_id.weight and not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount):
                                error_msg += "These products have no Weight : \n" + product_name
                            self.create_error_log(order,error_msg)
                            order_data['error'] = error_msg
                        elif 'the requested service is unavailable between the selected locations' in (response['error_message']).lower():
                            self.create_error_log(order,response.get('error_message').split(':')[1][1:])
                            order_data['error'] = response['error_message'].split(':')[1][1:]
                        else:
                            order_data['error'] = response.get('error_message')
                else:
                    order_data['calcu_shipping_cost'] = 'N\A'
                    order_data['differ'] = 0.0
            sheet.cell(row=row_index, column=19).value = "{:,.2f}".format(data[6]) or 0.0
            sheet.cell(row=row_index, column=19).alignment = right_alignment
            sheet.cell(row=row_index, column=20).value = order_data['carrier_method']
            sheet.cell(row=row_index, column=20).alignment = left_alignment
            sheet.cell(row=row_index, column=21).value = "{:,.2f}".format(order_data['calcu_shipping_cost']) if type(order_data['calcu_shipping_cost']) == float else order_data['calcu_shipping_cost']
            sheet.cell(row=row_index, column=21).alignment = right_alignment
            sheet.cell(row=row_index, column=22).value = "{:,.2f}".format(order_data['differ']) or 0.0
            sheet.cell(row=row_index, column=22).alignment = right_alignment
            sheet.cell(row=row_index, column=23).value = order_data.get('error')
            sheet.cell(row=row_index, column=23).alignment = left_alignment
            row_index += 1
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 24
        sheet.column_dimensions['K'].width = 25
        sheet.column_dimensions['L'].width = 25
        sheet.column_dimensions['M'].width = 30
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['P'].width = 15
        sheet.column_dimensions['Q'].width = 15
        sheet.column_dimensions['R'].width = 25
        sheet.column_dimensions['S'].width = 15
        sheet.column_dimensions['T'].width = 38
        sheet.column_dimensions['U'].width = 17
        sheet.column_dimensions['V'].width = 15
        sheet.column_dimensions['W'].width = 40
        
        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=shipping.cost.analysis.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (self.id,f_name),
            'target': 'self',
        }

    def get_shipping_method(self,order):
        shipping_method = self.carrier_id
        res = False
        if shipping_method and hasattr(shipping_method, '%s_rate_shipment' % shipping_method.delivery_type.lower()):
            res = getattr(shipping_method, '%s_rate_shipment' % shipping_method.delivery_type.lower())(order)

        return res

    def create_error_log(self,order,error_msg):
        street = [order.partner_id.street if order.partner_id.street else '']
        address = street.extend([order.partner_id.street2 if order.partner_id.street2 else ''])
        vals = {'so_date':order.date_order,
                'sale_order':order.name,
                'customer_id':order.partner_id.internal_id,
                'customer_name':order.partner_id.name,
                'email':order.partner_id.email,
                'city':order.partner_id.city,
                'state_id':order.partner_id.state_id.id if order.partner_id.state_id else False,
                'postal_code':order.partner_id.zip,
                'country_id':order.partner_id.country_id.id if order.partner_id.country_id else False,
                'street':''.join(street),
                'shipping_method_id':self.carrier_id.id or '',
                'error':error_msg}
        self.env['kits.shipping.error.log'].create(vals)

    # def _get_postal_code(self,country,state,zip,city):
    #     address_obj = self.env['kits.country.data.import']
    #     if not zip:
    #         if city:
    #             zip = address_obj.search([('city','ilike',city)]).filtered(lambda x:x.city.lower() == city.lower()).zip
    #         elif state:
    #             state_id = address_obj.search([('state_id.name','ilike',state.name)])
    #             zip = state_id[0].zip
    #         elif country:
    #             url = "https://countriesnow.space/api/v0.1/countries/capital"
    #             response = requests.post(url,{"country": country.name})
    #             if response.json()['error'] is not True:
    #                 capital_state = address_obj.search([('state_id.name','ilike',response.json()['data']['capital'])])
    #                 if capital_state:
    #                     zip = capital_state[0].zip
    #     return zip
