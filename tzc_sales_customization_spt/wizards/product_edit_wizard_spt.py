# -*- coding: utf-8 -*-
# Part of SnepTech See LICENSE file for full copyright and licensing details.##
##############################################################################
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
from datetime import datetime

class product_edit_wizard_spt(models.TransientModel):
    _name = 'product.edit.wizard.spt'
    _description = 'Edit Products Wizard'
    _rec_name = 'file_name'


    file = fields.Binary('File')
    file_name = fields.Char('File Name')
    line_ids = fields.One2many('product.edit.line.wizard.spt','product_edit_wizard_id','Lines')

    def action_generate_excel_report(self):
        active_id = self.id
        f_name = 'Product_eidt_%s'% datetime.strftime(datetime.now(),'%d-%m-%Y HH:MM:SS')  # FileName
        workbook = Workbook()
        sheet = workbook.create_sheet(
            title="Import Product", index=0) 

        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        heading_font = Font(name="Garamond", size="10", bold=True)

        table_header_row = 1
        for col in range(1,11):
            sheet.cell(row=table_header_row,
                       column=col).border = top_bottom_border
            sheet.cell(row=table_header_row, column=col).font = heading_font


        sheet.cell(row=table_header_row, column=1).value = 'Id'
        sheet.cell(row=table_header_row, column=2).value = 'Color Name'
        sheet.cell(row=table_header_row, column=3).value = 'Secondary Color Name'
        sheet.cell(row=table_header_row, column=4).value = 'Lense Color Name'
        sheet.cell(row=table_header_row, column=5).value = 'Rim Type'
        sheet.cell(row=table_header_row, column=6).value = 'Shape'
        sheet.cell(row=table_header_row, column=7).value = 'Material'
        sheet.cell(row=table_header_row, column=8).value = 'Flex Hinges'
        sheet.cell(row=table_header_row, column=9).value = 'Qty'
        sheet.cell(row=table_header_row, column=10).value = 'Gender'
  
        for product in self.line_ids:
            if product.is_edit == 'yes':
                table_header_row +=1
                sheet.cell(row=table_header_row, column=1).value = product.product_id.default_code or "N/A"
                sheet.cell(row=table_header_row, column=2).value = product.product_color_name.name or "N/A"
                sheet.cell(row=table_header_row, column=3).value = product.secondary_color_name.name or "N/A"
                sheet.cell(row=table_header_row, column=4).value = product.lense_color_name.name or "N/A"
                sheet.cell(row=table_header_row, column=5).value = product.rim_type.name or "N/A"
                sheet.cell(row=table_header_row, column=6).value = ', '.join(product.shape.mapped('name')) or "N/A"
                sheet.cell(row=table_header_row, column=7).value = ', '.join(product.material.mapped('name')) or "N/A"
                sheet.cell(row=table_header_row, column=8).value = dict(product.product_id._fields['flex_hinges'].selection).get(product.flex_hinges)
                sheet.cell(row=table_header_row, column=9).value = product.onhand_qty or "N/A"
                sheet.cell(row=table_header_row, column=10).value = 'F' if product.gender=='female' else 'M' if product.gender=='male' else  'M/F' if product.gender == 'm/f' else "N/A" 
            
        sheet.column_dimensions['A'].width = 17
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 12
        sheet.column_dimensions['I'].width = 12
        sheet.column_dimensions['J'].width = 12
        
    
        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        self.file = base64.b64encode(data)
        self.file_name = f_name
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=product.edit.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id, f_name),
            'target': 'self',
        }


class product_edit_line_wizard_spt(models.TransientModel):
    _name = 'product.edit.line.wizard.spt'
    _description = 'Edit Products Line Wizard'
    _rec_name = 'product_id'

    product_id = fields.Many2one('product.product','Product')
    product_color_name =  fields.Many2one('product.color.spt','Color Name',required=True)
    secondary_color_name =  fields.Many2one('product.color.spt','Secondary Color Name')
    # shape = fields.Many2one('product.shape.spt','Shape')
    shape = fields.Many2many('product.shape.spt','product_edit_wizard_product_shape_spt_rel','wizard_id','product_id','Shapes')
    lense_color_name =  fields.Many2one('product.color.spt','Lense Color Name')
    rim_type = fields.Many2one('product.rim.type.spt','Rim Type')
    # material = fields.Many2one('product.material.spt','Material')
    material = fields.Many2many('product.material.spt','product_edit_wizard_product_material_spt_rel','wizard_id','product_id','Materials')
    gender = fields.Selection([('male','M'),('female','F'),('m/f','M/F')], string='Gender')
    is_edit = fields.Selection([('yes','Yes'),('no','No')],'Is It Edited?')
    # image_1 = fields.Image('Primary Image', max_width=512, max_height=512)
    # image_1_zoom = fields.Image('Primary Image Zoom',related="product_id.image_variant_1920")
    # image_2 = fields.Image('Secondary Image', max_width=512, max_height=512)
    # image_2_zoom = fields.Image('Secondary Image Zoom',related="product_id.image_secondary")
    image_1 = fields.Char('  Primary Image')
    image_2 = fields.Char('Secondary Image')
    product_edit_wizard_id = fields.Many2one('product.edit.wizard.spt','Edit')
    onhand_qty = fields.Float('On Hand Qty')
    flex_hinges = fields.Selection([('yes','Yes'),('no','No')],'Flex Hinges',related="product_id.flex_hinges")
    


    @api.onchange('product_id','product_color_name','lense_color_name','rim_type','material','gender','shape','secondary_color_name','flex_hinges')
    def onchange_date(self):
        for record in self:
            record.is_edit = 'yes'
