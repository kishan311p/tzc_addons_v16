from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from io import BytesIO

import datetime
import os
import base64

class marketing_campaign(models.Model):
    _inherit = "marketing.campaign"

    mailing_list_ids = fields.Many2many('mailing.list','marketing_campaign_mailing_list','campaign_id','list_id','Mailing List')
    execution_datetime = fields.Datetime('Execution Date & Time')
    sale_order_count = fields.Integer('#Sale Order',compute='_compute_button_count')
    monetization_count = fields.Integer('Sale Order Count',compute='_compute_button_count')

    file_name = fields.Char('File ')

    count_sent = fields.Integer()
    count_open = fields.Integer(compute="_compute_count_sent")
    count_reply = fields.Integer(compute="_compute_count_sent")
    count_click = fields.Integer(compute="_compute_count_sent")
    count_balcklisted = fields.Integer(compute="_compute_count_sent")
    count_failed = fields.Integer(compute="_compute_count_sent")
    count_bounce = fields.Integer(compute="_compute_count_sent")
    count_unsubscribe = fields.Integer(compute="_compute_count_sent")
    count_received = fields.Integer(compute="_compute_count_sent")
    mailgun_failed_count = fields.Integer(compute="_compute_count_sent")

    @api.depends('marketing_activity_ids')
    def _compute_count_sent(self):
        for rec in self:
            count_open = 0
            count_clicked = 0
            count_reply = 0
            count_balcklisted = 0
            count_bounce = 0
            count_received = 0
            # mailgun_failed_count = 0
            contacts = self.env['campaign.report.contacts'].search([('campaign_id','=',rec.id)])
            for con in contacts:
                if con.opened:
                    count_open += 1
                if con.replied:
                    count_reply += 1
                # if con.ignored:
                #     count_balcklisted += 1
                # if con.bounced:
                #     count_bounce += 1
                if con.received:
                    count_received += 1
                # if con.failed_by_mailgun:
                #     mailgun_failed_count += 1
                if con.clicked:
                    count_clicked += 1
               
            rec.count_open = count_open
            rec.count_reply = count_reply
            rec.count_balcklisted = count_balcklisted
            rec.count_bounce = count_bounce
            rec.count_received = count_received
            rec.count_sent = len(self.env['marketing.participant'].search([('campaign_id','=',rec.id)]))
            rec.count_click = count_clicked
            # rec.count_click = self.env['campaign.report.contacts'].search([('campaign_id','=',rec.id)]).filtered(lambda x:x.clicked)
            # rec.count_unsubscribe = self.env['mailing.unsubscribe.list'].search_count([('marketing_campaign_id','=',rec.id)])
            rec.count_failed = len(contacts.filtered(lambda x: x['state'] == 'failed'))
            rec.mailgun_failed_count = len(contacts.filtered(lambda x:x.state in ('ignored','failed_mailgun')))

    def action_start_campaign(self):
        monetization_obj = self.env['campaign.monetization.spt']
        trace_obj = self.env['marketing.trace']
        participant_obj = self.env['marketing.participant']
        participants = participant_obj.search([('is_test','=',True),('campaign_id','=',self.id)])
        if participants:
            participants.mapped('trace_ids').mapped('mailing_trace_ids').mapped('links_click_ids').unlink()
            participants.mapped('trace_ids').mapped('mailing_trace_ids').unlink()
            participants.mapped('trace_ids').unlink()
            participants.unlink()
        res = super(marketing_campaign,self).action_start_campaign()
        for record in self:
            if record.execution_datetime:
                if record.execution_datetime >= datetime.now():
                    participants_ids = record.sync_participants() or record.participant_id
                    trace_ids = trace_obj.search([('participant_id','in',participants_ids.ids)])
                    for trace_id in trace_ids.filtered(lambda x: x.state == 'scheduled'):
                        trace_id.schedule_date = record.execution_datetime
                else:
                    raise UserError(_('Execution datetime is before the current datetime'))
        return res

    @api.onchange('mailing_list_ids')
    def _onchange_mailing_list_ids_spt(self):
        for record in self:
            record.model_id = False
            record.unique_field_id = False
            record.domain = '[]'
            if record.mailing_list_ids:
                record.model_id = self.env.ref('mass_mailing.model_mailing_contact').id
                record.unique_field_id = self.env.ref('mass_mailing.field_mailing_contact__email').id
                record.domain = str([('list_ids','in',record.mailing_list_ids.ids)])
                print(record.domain)


    def sync_participants(self):
        res = super(marketing_campaign, self).sync_participants()
        blacklist_participants = res.filtered(lambda participant: participant.model_name == 'mailing.contact' and participant.resource_ref.email in self.env['mail.blacklist'].search([]).mapped('email'))
        blacklist_participants.unlink()
        monetization_obj = self.env['campaign.monetization.spt']
        for participant in res - blacklist_participants:
            if participant.resource_ref and participant.resource_ref._name == 'res.partner':
                if not monetization_obj.search([('participant_id','=',participant.id)],limit=1):
                    monetization_obj.create({
                        'name' : participant.resource_ref.name,
                        'email' : participant.resource_ref.email,
                        'internal_id' : participant.resource_ref.internal_id,
                        'salesperson_id' : participant.resource_ref.user_id.id,
                        'odoo_contact_id' : participant.resource_ref.id,
                        'before_source' : 'odoo_contact',
                        'before_status_type' : participant.resource_ref.customer_type,
                        'campaign_id' : self.id,
                        'participant_id' :participant.id,
                    })
            else:
                if not monetization_obj.search([('participant_id','=',participant.id)],limit=1):
                    monetization_obj.create({
                        'name' : participant.resource_ref.name,
                        'email' : participant.resource_ref.email,
                        'internal_id' : participant.resource_ref.internal_id,
                        'odoo_contact_id' : participant.resource_ref.odoo_contact_id.id,
                        'before_source' : participant.resource_ref.source,
                        'before_prospect_level' : participant.resource_ref.prospect_level,
                        'before_status_type' : participant.resource_ref.status_type,
                        'before_action_type' : participant.resource_ref.action_type,
                        'before_orders' : participant.resource_ref.orders,
                        # 'before_promo_code_ids' : [(6,0,participant.resource_ref.promo_code_ids.ids)],
                        'campaign_id' : participant.campaign_id.id,
                        'participant_id' :participant.id,
                    })
        return res
    
    def _compute_button_count(self):
        sale_order_obj = self.env['sale.order']
        # monetization_obj = self.env['campaign.monetization.spt']
        for record in self:
            if record.model_name in ['mailing.contact','res.partner']:
                record.sale_order_count = len(sale_order_obj.search([('partner_id','in',record.mailing_list_ids.mapped('contact_ids.odoo_contact_id').ids)]))
                # record.monetization_count = len(monetization_obj.search([('campaign_id','in',record.ids)]))
            else:
                record.sale_order_count = 0 
                record.monetization_count = 0 

    def action_view_sale_report_spt(self):
        for record in self:
            if record.model_name == 'mailing.contact':
                order_ids = self.env['sale.report'].search([('partner_id','in',record.mailing_list_ids.mapped('contact_ids.odoo_contact_id').ids)])
                if order_ids:
                    return{
                        'type': 'ir.actions.act_window',
                        'name': _('Orders'),
                        'res_model': 'sale.report',
                        'view_mode': 'dashboard,pivot,graph',
                        'domain': [('id', 'in', order_ids.ids)],
                        'context': {'search_default_group_by_partner_id': 1},
                    }
                else:
                    raise UserError(_('Sale orders not found.'))
            else:
                raise UserError(_('Sale orders not found.'))

    def action_generate_excel_report(self):

        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/export_contact_base_template.xlsx'
        wb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
        wrksht = wb.active

        # active_id = self.id
        # workbook = Workbook()

        # sheet = workbook.create_sheet(title="Making Contact", index=0)

        f_name = 'MA_%s'% self.name  # FileName

        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        heading_font = Font(name="Garamond", size="10", bold=True)

        # table_header_row = 1
        # for col in range(1,15):
        #     wrksht.cell(row=table_header_row,
        #                column=col).border = top_bottom_border
        #     wrksht.cell(row=table_header_row, column=col).font = heading_font


        # wrksht.cell(row=table_header_row, column=1).value = 'Name'
        # wrksht.cell(row=table_header_row, column=2).value = 'Email'
        # wrksht.cell(row=table_header_row, column=3).value = 'Country'
        # wrksht.cell(row=table_header_row, column=4).value = 'Territory'
        # wrksht.cell(row=table_header_row, column=5).value = 'Tags'
        # wrksht.cell(row=table_header_row, column=6).value = 'Last Updated On'
        # wrksht.cell(row=table_header_row, column=7).value = 'Source'
        # wrksht.cell(row=table_header_row, column=8).value = 'Prospect Level'
        # wrksht.cell(row=table_header_row, column=9).value = 'Status Type'
        # wrksht.cell(row=table_header_row, column=10).value = 'Action Type'
        # wrksht.cell(row=table_header_row, column=11).value = '#Order'
        # wrksht.cell(row=table_header_row, column=12).value = 'Internal Id'
        # wrksht.cell(row=table_header_row, column=13).value = 'Odoo Contact Id'
        # wrksht.cell(row=table_header_row, column=14).value = 'Promotion Coupons'

        table_header_row = 2
        campaign_report_contacts_ids = self.env['campaign.report.contacts'].search([('campaign_id','=',self.id)])
        for contact in self.participant_ids:
            campaign_contact = campaign_report_contacts_ids.filtered(lambda x:x.email == contact.resource_ref.email)
            wrksht.cell(row=table_header_row, column=1).value = contact.resource_ref.name or ''
            wrksht.cell(row=table_header_row, column=2).value = contact.resource_ref.email or ''
            wrksht.cell(row=table_header_row, column=3).value = contact.resource_ref.phone or ''
            wrksht.cell(row=table_header_row, column=4).value = ', '.join([contact.resource_ref.street,contact.resource_ref.street2]) if contact.resource_ref.street2 else contact.resource_ref.street
            wrksht.cell(row=table_header_row, column=5).value = contact.resource_ref.city
            wrksht.cell(row=table_header_row, column=6).value = contact.resource_ref.state_id.name if contact.resource_ref.state_id else ''
            wrksht.cell(row=table_header_row, column=7).value = contact.resource_ref.country_id.name if contact.resource_ref.country_id else ''
            wrksht.cell(row=table_header_row, column=8).value = contact.resource_ref.territory.name if contact.resource_ref.territory else ''
            wrksht.cell(row=table_header_row, column=9).value = 'B2C' if contact.resource_ref.customer_type == 'b2c' else 'B2B-Regular' if contact.resource_ref.customer_type == 'b2b_regular' else 'B2B-Fs' if contact.resource_ref.customer_type == 'b2b_fs' else  ''
            wrksht.cell(row=table_header_row, column=10).value = contact.resource_ref.user_id.name
            wrksht.cell(row=table_header_row, column=11).value = contact.resource_ref.sale_order_count
            wrksht.cell(row=table_header_row, column=12).value = 1 if campaign_contact.sent else 0
            wrksht.cell(row=table_header_row, column=13).value = campaign_contact.sent or ''
            wrksht.cell(row=table_header_row, column=14).value = 1 if campaign_contact.received else 0
            wrksht.cell(row=table_header_row, column=15).value = campaign_contact.received or ''
            wrksht.cell(row=table_header_row, column=16).value = 1 if campaign_contact.clicked else 0
            wrksht.cell(row=table_header_row, column=17).value = campaign_contact.clicked or ''
            wrksht.cell(row=table_header_row, column=18).value = 1 if campaign_contact.exception else 0
            wrksht.cell(row=table_header_row, column=19).value = 1 if campaign_contact.failed_by_mailgun else 0
            table_header_row += 1
        
        # wrksht.column_dimensions['A'].width = 17
        # wrksht.column_dimensions['B'].width = 20
        # wrksht.column_dimensions['C'].width = 12
        # wrksht.column_dimensions['D'].width = 10
        # wrksht.column_dimensions['E'].width = 15
        # wrksht.column_dimensions['F'].width = 15
        # wrksht.column_dimensions['G'].width = 12
        # wrksht.column_dimensions['H'].width = 12
        # wrksht.column_dimensions['I'].width = 12
        # wrksht.column_dimensions['J'].width = 12
        # wrksht.column_dimensions['K'].width = 10
        # wrksht.column_dimensions['L'].width = 10
        # wrksht.column_dimensions['M'].width = 30
        # wrksht.column_dimensions['N'].width = 10

        
        
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        # self.file = base64.b64encode(data)
        self.file_name = f_name

        wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, f_name),
            'target': 'self',
        }

    # def action_campaign_monetization(self):
    #     contact_obj = self.env['mailing.contact']
    #     monetization_ids = self.env['campaign.monetization.spt'].search([('campaign_id', 'in', self.ids)])
    #     for monetization in monetization_ids:
    #         contact = contact_obj.search([('email','=',monetization.email)],limit=1)
    #         monetization.write({
    #             'after_source' : contact.source,
    #             'after_prospect_level' : contact.prospect_level,
    #             'after_status_type' : contact.status_type,
    #             'after_action_type' : contact.action_type,
    #             'after_orders' : contact.orders,
    #             'after_promo_code_ids' : [(6,0,contact.promo_code_ids.ids)],
    #         })
    #     return{
    #             'type': 'ir.actions.act_window',
    #             'name': _('Campaign Monetization'),
    #             'res_model': 'campaign.monetization.spt',
    #             'view_mode': 'tree,form',
    #             'domain': [('campaign_id', 'in', self.ids)],
    #         }
    
    def action_show_clicked_contact(self):
        self.ensure_one()
        clicks = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x.clicked)
        # clicks = self.env['link.tracker.click'].search([('campaign_id', 'in', self.utm_campaign_id.ids)])
        return {
            'name':_("Mailing Contact"),
            'type':'ir.actions.act_window',
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            'views':[(self.env.ref('tzc_sales_customization_spt.campaign_report_contacts_tree_view').id,'tree'),
                    (self.env.ref('tzc_sales_customization_spt.campaign_report_contacts_form_view').id,'form'),
                    ],
            'domain':[('id','=',clicks.ids)],
            'target':'self',
        }

    def action_show_sent_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)])
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    
    def action_show_opened_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x.opened)
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    
    def action_show_reply_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x['state'] == 'replied')
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    
    def action_show_bounced_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x['state'] == 'bounced')
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    def action_show_blacklisted_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x['state'] == 'ignored')
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }

    def action_show_failed_contacts(self):
        self.ensure_one()
        trace_ids = self.env['mailing.trace'].search([('campaign_id', '=', self.id),('trace_status','=','error')])
        contact_ids = self.env['campaign.report.contacts'].search([('trace_id', 'in', trace_ids.ids)])
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    
    def action_show_received_contacts(self):
        self.ensure_one()
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x['state'] == 'received')
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }

    def action_show_unsubscribe_contacts(self):
        self.ensure_one()
        return {
            "name":_("Mailing Unsubscribe List"),
            "type":"ir.actions.act_window",
            'res_model':"mailing.unsubscribe.list",
            'view_mode':'tree',
            "domain":[('marketing_campaign_id','=',self.id)],
            'target':'self',
        }

    def action_show_mailgun_failed_contacts(self):
        self.ensure_one()
        # contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id)]).filtered(lambda x: x.failed_by_mailgun or x.ignored)
        contact_ids = self.env['campaign.report.contacts'].search([('campaign_id', '=', self.id),('state','in',['ignored','failed_mailgun'])])
        return {
            "name":_("Mailing Contact"),
            "type":"ir.actions.act_window",
            'res_model':"campaign.report.contacts",
            'view_mode':'tree,form',
            "domain":[('id','in',contact_ids.ids)],
            'target':'self',
        }
    
    def action_resend_campaign(self):
        ctx = {
            'default_model': 'res.partner',
            'default_res_id': self.env['res.partner'].search([],limit=1).id,
            'default_composition_mode': 'comment',
            'mark_so_as_sent': True,
            'custom_layout': "mail.mail_notification_light",
            'force_email': True,
            'default_body':self.marketing_activity_ids[0].mass_mailing_id.body_html,
            'campaign':True,
            'resend':True,
            'raise_campaign':True,
            'subject':self.name,
            'default_attachment_ids':[(6,0,self.marketing_activity_ids.mass_mailing_id.attachment_ids.ids)] if self.marketing_activity_ids else False
        }
        return {
                'name': _('Send Mail'),
                'type': 'ir.actions.act_window',
                'res_model': 'mail.compose.message',
                'binding_model':"marketing.campaign",
                'view_mode' : 'form',
                'target': 'new',
                'context':ctx
            }
    
    def action_show_mail_previwe(self):
        self.ensure_one()
        form_view = self.env.ref('tzc_sales_customization_spt.tzc_mailing_mailing_view_form_marketing_activity')
        return{
            'name': ('Preview'),
            'res_model': 'mailing.mailing',
            'type': 'ir.actions.act_window',
            'views': [(form_view.id, 'form')],
            'res_id': self.marketing_activity_ids[0].mass_mailing_id.id,
            'target': 'current',
        }

    def unlink(self):
        for rec in self:
            same_campaing = self.env['mail.compose.message'].search([('campaign_name','=',rec.name)])
            if same_campaing:
                for campaign in same_campaing:
                    campaign.unlink()
        return super(marketing_campaign,self).unlink()

    def execute_activities(self):
        for campaign in self:
            if campaign._context.get('from_bulk_mail'):
                campaign.marketing_activity_ids.execute()