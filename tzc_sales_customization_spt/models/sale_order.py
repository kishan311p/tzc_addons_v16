from odoo import api, Command, fields, models, SUPERUSER_ID, _
import hashlib
import re
from odoo.exceptions import UserError
from dateutil.relativedelta import relativedelta
from datetime import datetime
from werkzeug.urls import url_encode
from datetime import date,timedelta,datetime
from dateutil import tz
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
import os
from lxml import etree
import logging
from odoo import api, fields, models, SUPERUSER_ID, _
from odoo.http import request
from odoo.addons.website.models import ir_http
from odoo.exceptions import UserError
from odoo.modules import get_module_resource
import os
from os.path import isfile, join
import pandas as pd
from odoo.osv import expression

_logger = logging.getLogger(__name__)

field_list = ['partner_id','date_order','payment_term_id','include_cases','no_of_cases','case_weight_gm','order_line',
              'note','sale_order_option_ids','user_id','sale_manager_id','team_id','client_order_ref','picking_policy','commitment_date',
              'fiscal_position_id','origin','opportunity_id','campaign_id','medium_id','source_id','next_execution_date','street','street2',
              'country_id','city','postal_code','state_id','mark_as_paid_by_user','payment_link_approved_by','order_approved_by','require_signature',
              'require_payment','signed_by','signed_on','signature','approve_by_salesperson','approve_by_salesmanager','approve_by_admin','payment_link',
              'amount_paid','due_amount','payment_status','state']

NO_TRACKING_FIELDS = ['amount_total','amount_untaxed','amount_without_discount','amount_discount']

class sale_order(models.Model):
    _inherit = 'sale.order'

    def _get_salesmangers(self):
        manager_users = self.env.ref('tzc_sales_customization_spt.group_sales_manager_spt').sudo().users
        return [('id','in',manager_users.ids)]

    def get_sale_manager(self):
        user_obj = self.env['res.users'].sudo()
        manager_users = self.env.ref('tzc_sales_customization_spt.group_sales_manager_spt').sudo().users
        for record in self:
            manager = manager_users.filtered(lambda user: record.partner_id.country_id in user.contact_allowed_countries)
            if not manager:
                manager = user_obj.search([('allow_user_ids','in',record.user_id.ids)],limit=1)
            if not manager:
                default_salesperson = self.env['ir.config_parameter'].sudo().get_param('default_sales_person_id')
                try:
                    default_salesperson = int(default_salesperson)
                    manager = user_obj.browse(default_salesperson)
                except:
                    pass
            return manager.id if len(manager) <= 1 else manager[0].id

    sale_manager_id = fields.Many2one('res.users','Sales Manager',default=get_sale_manager,domain=_get_salesmangers)

    # kits_abadon_card_order
    next_execution_date = fields.Datetime('Next Execution Date',copy=False)
    abandoned_reason = fields.Char('Reason',compute="_get_reson_message")

    company_id = fields.Many2one('res.company')

    #kits_shipping_cost
    shipping_id = fields.Many2one('shipping.provider.spt',default=False)
    estimate_shipping_cost = fields.Float('Shipping Cost ')
    actual_weight = fields.Float('Actual Weight (kg)')
    kits_carrier_tracking_ref = fields.Char('Tracking Reference',compute="_compute_carrier_tracking_ref",compute_sudo=True,store=True)
    glass_weight_kg = fields.Float('Weight for items (kg)' ,compute='_compute_weight_kits',compute_sudo=True,store=True,help="The total weight of picked products.")
    weight_total_kg = fields.Float('Total Weight (kg)',compute='_compute_weight_kits',compute_sudo=True,store=True)
    case_weight_kg = fields.Float('Weight for cases (kg)',compute="_calculate_case_weight_kg",compute_sudo=True,store=True)
    order_approved_by = fields.Many2one('res.partner','Order Approved By')
    is_abandoned_cart = fields.Boolean('Abandoned Cart', compute='_compute_abandoned_cart', search='_search_abandoned_cart')
    cart_recovery_email_sent = fields.Boolean('Cart recovery email already sent')
    report_token = fields.Char('Report Access Token')

    @api.depends('case_weight_gm','no_of_cases','include_cases')
    def _calculate_case_weight_kg(self):
        context = self._context.copy()
        context.update({'custom':True})
        self.env.context = context
        for record in self:
            picking_id = self.env['stock.picking'].search([('id','in',record.picking_ids.ids)])
            no_of_cases = sum(picking_id.move_ids_without_package.filtered(lambda x:x.product_id.sale_type != 'clearance').mapped('quantity_done'))
            if record.include_cases:
                record.case_weight_kg = 0.001 * record.case_weight_gm * no_of_cases
            else:
                record.case_weight_kg = 0.0
            picking_id.no_of_cases = no_of_cases if picking_id else None

    @api.depends('case_weight_gm','state','case_weight_kg','no_of_cases','glass_weight_kg','weight_total_kg','order_line','order_line.picked_qty','order_line.product_id.weight')
    def _compute_weight_kits(self):
        for record in self:
            total_weight_kg = 0.0
            if record.include_cases:
                total_weight_kg += record.case_weight_kg
            lines = record.order_line.filtered(lambda x: x.picked_qty)
            weight = round(sum(list(map(lambda x: x.product_id.weight * x.picked_qty,lines))),2)
            if record.state in ('draft','sent','salesperson_confirmation','received','sale','in_scanning'):
                weight = round(sum(list(map(lambda x: x.product_id.weight * x.product_uom_qty,record.order_line))),2)
            record.glass_weight_kg = weight
            record.weight_total_kg  = round(total_weight_kg + weight,2)

    @api.depends('picking_ids','picking_ids.carrier_tracking_ref','picking_ids.picking_type_id','write_date')
    def _compute_carrier_tracking_ref(self):
        for record in self:
            picking = record.picking_ids.filtered(lambda p: p.state != 'cancel' and p.picking_type_id.code == 'outgoing')[:1]
            record.kits_carrier_tracking_ref = picking.carrier_tracking_ref

    # @api.depends('picking_ids','picking_ids.shipping_id')
    # def compute_shipping_id(self):
    #     for record in self:
    #         picking = record.picking_ids.filtered(lambda x: x.state != 'cancel' and 'WH/OUT' in x.name)
    #         record.shipping_id = picking.shipping_id.id

    def action_check_shipping_cost(self):
        if self.state not in ['scan','shipped','draft_inv','open_inv','cancel','merged','done']:
            view_id = self.env.ref('delivery.choose_delivery_carrier_view_form').id
            return {
                'name': _('Check Shipping Cost'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'choose.delivery.carrier',
                'view_id': view_id,
                'views': [(view_id, 'form')],
                'target': 'new',
                'context': {
                    'default_order_id': self.id,
                    'default_weight_of_case': self.case_weight_kg,
                    'default_weight_of_glasses': self.glass_weight_kg,
                    'estimated_shipping_cost':True,
                    'get_product_uom':True,
                }
            }
        else:
            raise UserError('You can not add shipping cost after order ready to ship.')
        # pass
        

    def send_shipment_ready_email_to_salesperson_spt(self):
        #create attachment
        self.ensure_one()
        if self.state == 'scanned':
            picking_id = self.picking_ids.filtered(lambda x:x.state != 'cancel')
            template_id = self.env.ref('tzc_sales_customization_spt.tzc_picking_ready_notification_to_salesperson_spt')
            template_id.send_mail(picking_id.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")
            ready_to_ship_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_order_ready_to_ship')
            ready_to_ship_template_id.send_mail(self.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")
            picking_id.write({'state':'assigned'})
            self.write({'state': 'scan'})
            self.order_approved_by = self.env.user.partner_id.id
            picking_id.get_picking_order_values()
            return True
        else:
            raise UserError ('This order is in "In Scanning" state, You can\'t approve in scanning order.')

    @api.model
    def _fields_view_get(self,view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(sale_order,self)._fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar, submenu=submenu)
        if view_type == 'form':
            doc = etree.fromstring(res['arch'])
            is_admin = self.env.user.has_group('base.group_system')
            is_manager = self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt')
            is_sales_person = self.env.user.is_salesperson
            if is_admin or is_manager or is_sales_person:
                for scanned_msg in doc.xpath('//div[@id="scanned_msg"]'):
                    scanned_msg.attrib['invisible'] = '0'
                for revert_msg in doc.xpath('//div[@id="revert_msg"]'):
                    revert_msg.attrib['invisible'] = '0'
            else:
                for scanned_msg in doc.xpath('//div[@id="scanned_msg"]'):
                    scanned_msg.attrib['invisible'] = '1'
                for revert_msg in doc.xpath('//div[@id="revert_msg"]'):
                    revert_msg.attrib['invisible'] = '1'
            invoice_action = self.env.ref('sale.action_view_sale_advance_payment_inv').id
            if invoice_action:
                for  invoice_node in doc.xpath('//button[@name='+str(invoice_action)+']'):
                    if invoice_node.attrib and 'attrs' in invoice_node.attrib.keys():
                        if 'groups' in invoice_node.attrib.keys():
                            invoice_node.attrib['groups'] = invoice_node.attrib['groups']+',account.group_account_invoice'
                        else:
                            invoice_node.attrib['groups'] = 'account.group_account_invoice'
                for tax_id in doc.xpath("//field[@name='order_line']/tree/field[@name='tax_id']"):
                    tax_id.attrib['readonly'] = '0' if self.env.user.has_group('base.group_system') else "1"
                for manager_id in doc.xpath("//field[@name='sale_manager_id']"):
                    manager_id.attrib['readonly'] = '0' if self.env.user.has_group('base.group_system') else '1'
                res['arch'] = etree.tostring(doc, encoding='unicode')
        return res


    # kits_bambora_payment
    approve_by_salesperson = fields.Boolean('Approve By Salesperson',copy=False)
    approve_by_salesmanager = fields.Boolean('Approve By Sales Manager',copy=False)
    approve_by_admin = fields.Boolean('Approve By Administrator',copy=False)
    payment_link = fields.Char('Payment Link',copy=False)
    is_payment_link = fields.Boolean("Has Payment Link",compute="_compute_is_payment_link",store=True)
    is_paid = fields.Boolean('Is Paid ?',track_visibility='onchange',copy=False)
    mark_as_paid_by_user = fields.Many2one('res.users','Order Mark As Paid By',copy=False)
    payment_link_approved_by = fields.Many2one('res.users','Payment Approved By',copy=False)
    paid_amount = fields.Float('Paid Amount',copy=False)
    amount_paid = fields.Float('Paid Amount',copy=False,compute="_get_amount_paid",store=True)
    payment_ids = fields.One2many('order.payment','order_id','Payment',copy=False)
    payment_status = fields.Selection([('full','Fully Paid'),('partial','Partial Paid'),('over','Over Paid')],'Payment Status',compute="_compute_payment_status",copy=False,store=True)
    due_amount = fields.Float('Amount Due',compute="_compute_amount_due",store=True,copy=False)

    #kits_package_product

    #tzc_sales_customize_apt
    def action_quotation_send(self):
        ''' Opens a wizard to compose an email, with relevant mail template loaded by default '''
        self.ensure_one()
        template_id =self._find_mail_template()
        # template_id = self.env.ref('tzc_sales_customization_spt.mail_template_notify_customer_order_completion') if self.state in ['draft_inv','open_inv','scan','shipped'] else self._find_mail_template()
        lang = self.env.context.get('lang')
        template_id = self.env['mail.template'].browse(template_id.id)
        if template_id.lang:
            lang = template_id._render_lang(self.ids)[self.id]
        ctx = {
            'default_model': 'sale.order',
            'default_res_id': self.id,
            'default_use_template': bool(template_id),
            'default_template_id': template_id.id if template_id else None,
            'default_composition_mode': 'comment',
            'mark_so_as_sent': True,
            'default_email_layout_xmlid': 'mail.mail_notification_light',
            'proforma': self.env.context.get('proforma', False),
            'force_email': True,
            'model_description': self.with_context(lang=lang).type_name,
        }
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'views': [(False, 'form')],
            'view_id': False,
            'target': 'new',
            'context': ctx,
        }

    def _get_default_case_weight_gm(self):
        weight = self.env['ir.config_parameter'].sudo().get_param('tzc_case_weight_gm')
        if isinstance(weight, str):
            try:
                weight = eval(weight)
            except:
                weight = 0.0
        return weight

    state = fields.Selection(selection_add=[
               ('draft', 'Quotation'),('sent', 'Quotation Sent'),('received', 'Quotation Received'),('salesperson_confirmation', 'Salesperson Confirmation'),('sale', 'Order Confirmed'),('in_scanning','In Scanning'),('scanned','Scanning Completed'),('scan', 'Ready to Ship'),('shipped', 'Shipped'),('draft_inv', 'Draft Invoice'),('open_inv', 'Invoiced'),('cancel', 'Cancelled'),('merged', 'Merged'),('done', 'Locked')], string='Status', readonly=True, copy=False, index=True, tracking=3, default='draft')

    catalog_id = fields.Many2one('sale.catalog', ondelete='set null', string='Corresponding Catalog')
    amount_is_shipping_total = fields.Monetary(string='Shipping Cost', store=True, readonly=True,compute_sudo=True, compute='_amount_all', tracking=4)
    shipping_cost = fields.Float(string="Original Shipping Cost",store=True, readonly=True,compute_sudo=True, compute='_amount_all')
    amount_is_admin = fields.Monetary(string='Admin Fee', store=True, readonly=True,compute_sudo=True, compute='_amount_all', tracking=4)
    amount_discount = fields.Monetary(string='Discount', store=True, readonly=True,compute_sudo=True, compute='_amount_all', tracking=4)
    delivered_qty = fields.Integer('Delivered Quantity ',compute_sudo=True,compute="_amount_all",store=True)
    ordered_qty = fields.Integer('Ordered Quantity ',compute_sudo=True,compute="_amount_all",store=True)
    picked_qty = fields.Integer('Quantity Shipped ',compute_sudo=True,compute="_amount_all",store=True)
    amount_without_discount = fields.Monetary(string='Subtotal',compute_sudo=True,compute='_amount_all',  store=True,tracking=4)
    global_discount = fields.Monetary('Additional Discount',compute_sudo=True, store=True, readonly=True, compute='_amount_all', tracking=4)
    # check_line_qty = fields.Integer('Line Qty',compute="_compute_check_line_qty")
    # report_file = fields.Binary()
    applied_promo_code = fields.Char("Applied Promo Code")
    shipped_date = fields.Datetime(string="Shipped Date",compute="_compute_shipped_date",store=True,compute_sudo=True)
    source_spt = fields.Char("Order Source ",compute="_compute_order_source",store=True,compute_sudo=True)
    invoice_name = fields.Char("Invoice",compute="_compute_invoice_name",store=True,compute_sudo=True)
    partner_shipping_id = fields.Many2one('res.partner','Delivery Address')

    picked_qty_order_subtotal = fields.Monetary('picked Subtotal', tracking='32',compute="_amount_all",store=True,compute_sudo=True)
    picked_qty_order_total = fields.Monetary('picked Total', tracking='33',compute="_amount_all",store=True,compute_sudo=True)
    picked_qty_order_tax = fields.Monetary('Tax', tracking='34',compute="_amount_all",store=True,compute_sudo=True)
    picked_qty_order_discount = fields.Monetary('picked Discount', tracking='35',scompute="_amount_all",store=True,compute_sudo=True)
    create_uid_spt = fields.Many2one('res.users', 'Created by User',compute="_compute_create_uid_spt", index=True, readonly=True)
    count_backup_order = fields.Integer('Original Order',compute="_get_compute_message",store=True)
    free_shipping = fields.Boolean('Free Shippping')
    country_id = fields.Many2one('res.country','Country',compute="_compute_country_id",store=True,compute_sudo=True)
    street = fields.Char('Street',related='partner_id.street')
    street2 = fields.Char('Street2',related='partner_id.street2')
    city = fields.Char(related="partner_id.city")
    postal_code = fields.Char(related="partner_id.zip")
    state_id = fields.Many2one('res.country.state','State',related="partner_id.state_id")
    
    delivery_street = fields.Char('Delivery Address ',related='partner_shipping_id.street')
    delivery_street2 = fields.Char('Street2 ',related='partner_shipping_id.street2')
    delivery_city = fields.Char('City ',related="partner_shipping_id.city")
    delivery_postal_code = fields.Char('Zip Code ',related="partner_shipping_id.zip")
    delivery_state_id = fields.Many2one('res.country.state','State ',related="partner_shipping_id.state_id")
    delivery_country_id = fields.Many2one('res.country','Country ',related="partner_shipping_id.country_id")

    include_cases = fields.Boolean('Include Cases ?',default=True)
    no_of_cases = fields.Integer('#Cases',compute="get_no_of_cases",store=True)
    case_weight_gm = fields.Float('Weight for cases (gm)',default=_get_default_case_weight_gm)
    is_confirm_by_saleperson = fields.Boolean()
    download_image_sent = fields.Boolean()
    message = fields.Char(compute="_get_compute_message",store=True)
    is_same_delivery_address = fields.Boolean(compute='_get_delivery_address',default=False,store=True)
    reward_amount = fields.Float(compute='_compute_base_boolean_fields',store=True)
    delivery_set = fields.Boolean(compute='_compute_base_boolean_fields',store=True)
    is_all_service = fields.Boolean("Service Product", compute="_compute_base_boolean_fields",store=True)
    updated_on = fields.Datetime("Updated On")
    updated_by = fields.Many2one('res.users',"Updated By")
    is_picking_set = fields.Boolean('Is Picking Set (Flag)',help="Flag for set order id in delivery order.",compute='_set_order_id')
    is_not_product_aval_qty_flag = fields.Boolean(compute="_get_product_avail_qty")
    cart_discount_availability = fields.Boolean()
    cart_to_price = fields.Char()
    cart_from_price = fields.Char()
    cart_discount = fields.Char()
    currency_name = fields.Char('Currency Name',related="currency_id.name")
    website_id = fields.Many2one('kits.b2b.website')
    merged_order= fields.Boolean()
    merge_reference = fields.Many2many("sale.order","merged_order_sale_order_rel","merge_order_id","order_id","Merge Order of")
    customer_credit = fields.Char('Customer Credit')
    def _get_currency_id(self):
        return self.partner_id.preferred_currency.id
    
    b2b_currency_id = fields.Many2one('res.currency',default=_get_currency_id ,string=' Currency')
    currency_id = fields.Many2one(related='b2b_currency_id',depends=["b2b_currency_id"],store=True, precompute=True, ondelete="restrict")


    def compute_all(self):
        for record in self:
            record._amount_all()
        return True
        # pass
    
    # def action_confirm(self):
    #     res = super(sale_order, self).action_confirm()
    #     for so in self:
    #         currency_rate = self.env['kits.b2b.multi.currency.mapping'].search([('currency_id','=',so.b2b_currency_id.id)],limit =1).currency_rate
    #         if currency_rate:
    #             for sol in so.line_ids:
    #                 sol.b2b_currency_rate = currency_rate
    #     return res

    def action_view_sale_advance_payment_inv(self):
        if self.state != 'shipped':
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else:
            return self.env['sale.advance.payment.inv'].with_context(active_ids =self.ids).create({
                'advance_payment_method': 'delivered',
            }).create_invoices()

    # def action_revert_order_to_quotation(self):
        # Method to reverse order to quotation and copy order lines of original orde.
        # if self.state not in ['sent','received','sale','in_scanning','scanned','scan']:
        #     return {
        #         'type': 'ir.actions.client',
        #         'tag': 'display_notification',
        #         'params': {
        #                 'title': 'Something is wrong.',
        #                 'message': 'Please reload your screen.',
        #                 'sticky': True,
        #             }
        #         }
        # else:
        #     self.sale_order_cancel_spt()
        #     self.action_draft()
        #     order_id = self.env['sale.order.backup.spt'].search([('order_id','=',self.id)],limit=1,order="id desc")
        #     for order_line in self.order_line:
        #         if order_line.product_id and order_line.product_id.id in order_id.line_ids.product_id.ids:
        #             line=order_id.line_ids.search([('product_id','=',order_line.product_id.id),('order_backup_id','=',order_id.id)])
        #             order_line.write({
        #                 'product_uom_qty':line.product_uom_qty,
        #                 'price_unit':line.price_unit,
        #                 'unit_discount_price':line.unit_discount_price,
        #                 'fix_discount_price':line.fix_discount_price,
        #                 'discount':line.discount,
        #                 'tax_id':line.tax_id,
        #                 'is_global_discount':line.is_global_discount,
        #                 'is_fs':line.is_fs,
        #                 'is_admin':line.is_admin,
        #                 'is_shipping_product':line.is_shipping_product,
        #                 'is_promotion_applied':line.is_promotion_applied,
        #                 'sale_type':line.sale_type,
        #             })
        #         else:
        #             order_line.unlink()
        # pass

    def sale_order_cancel_spt(self):
        if len(self) <= 1 and self.state in ['cancel']:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else: 
            cancel_order_ids = self.filtered(lambda x:x.state == 'cancel')
            if not cancel_order_ids:
                for record in self:
                    if record.state in ['draft','sent','received','sale'] or (record.state not in ['draft','sent','received','sale'] and self.env.user.has_group('tzc_sales_customization_spt.group_cancel_sale_order_rule_spt')):
                        picking_ids = self.env['stock.picking'].search([('sale_id','=',record.id)])
                        if picking_ids:
                            return_ids = self.env['stock.return.picking'].search([('picking_id','in',picking_ids.ids)])
                            if return_ids:
                                for stock_return in return_ids:
                                    stock_return.packing_id.with_context(cancel_delivery=True).picking_cancel_spt()
                            else:
                                for picking in picking_ids:
                                    picking.with_context(cancel_delivery=True).picking_cancel_spt()
                        if record.invoice_ids:
                            for invoice_id in record.invoice_ids:
                                if invoice_id.state == 'paid':
                                    invoice_id.move_id.line_ids.remove_move_reconcile()
                                    invoice_id.action_cancel()
                                else:
                                    invoice_id.action_cancel()
                                # invoice_id.is_commission_paid = False
                        self.env['order.payment'].sudo().search([('order_id','=',record.id)]).unlink()
                        record.write({
                            'state' : 'cancel',
                            'is_paid' : False,
                            'mark_as_paid_by_user' : False,
                            'payment_link_approved_by' : False,
                            'payment_link' : False,
                            'approve_by_salesperson' : False,
                            'approve_by_salesmanager' : False,
                            'approve_by_admin' : False,
                            'amount_paid' : 0.0,
                            'due_amount' : 0.0,
                            'payment_status' : False,
                        })
                    else:
                        message = 'You don\'t have access to cancel order'
                        raise UserError(_(message))

            else:
                return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }


    def action_update_order_qty(self):
        for line in self.order_line.filtered(lambda x: not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount):
            if line.product_uom_qty > line.product_id.available_qty_spt:
                if line.product_id.available_qty_spt > 0.0:
                    line.product_uom_qty = line.product_id.available_qty_spt
                else:    
                    line.product_uom_qty = 0.0
                    line.price_unit = 0.0
                self.is_not_product_aval_qty_flag = False

                line.product_uom_change()
                line.product_uom_change_spt()

    def order_revert(self):
        for line in self.order_line.filtered(lambda x:not x.product_id.is_shipping_product and not x.product_id.is_admin):
            if not line.product_id.is_global_discount:
                line.discount = 0.0
                line.product_id_change()
                line._onchange_discount_spt()
                line._onchange_fix_discount_price_spt()
                line._onchange_unit_discounted_price_spt()
            else:
                line.unlink()

    def action_open_remove_product_wizard(self):
        self.ensure_one()
        if self.state in ['draft','sent','salesperson_confirmation']:
            wizard_id = self.env['remove.product.spt'].create({'partner_id':self.partner_id.id,'sale_id' : self.id,'product_ids':[(6,0,self.order_line.mapped('product_id').ids)]})
            return {
                'name': 'Remove Items',
                'view_mode': 'form',
                'target': 'new',
                'res_id':wizard_id.id,
                'res_model': 'remove.product.spt',
                'type': 'ir.actions.act_window',
                'context':{
                    "default_sale_id":self.id,
                    "default_partner_id":self.partner_id.id,
                    }
                }  
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        # return {
        #     "name":_("Remove Items"),
        #     "type":"ir.actions.act_window",
        #     "res_model":"remove.product.spt",
        #     "view_mode":"form",
        #     "context":{
        #         "default_sale_id":self.id,
        #         "default_partner_id":self.partner_id.id,
        #     },
        #     "target":"new",
        # }

    def action_discount_wizard(self):
        self.ensure_one()
        if self.state in ['draft','sent','received','salesperson_confirmation','in_scanning','sale','scan','scanned','shipped']:
            if 'draft' not in self.mapped('invoice_ids.state') and  'posted' not in self.mapped('invoice_ids.state'):
                return {
                    'name': 'Bulk Discount',
                    'view_mode': 'form',
                    'target': 'new',
                    'context':{'default_sale_id':self.id,'default_base_on':'brand','default_apply_on':'fix'},
                    'res_model': 'discount.on.sale.order.line.wizard.spt',
                    'type': 'ir.actions.act_window',
                }
            else:
                raise UserError(_("You can not give discount after create invoice."))
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }

    @api.model
    def default_get(self, default_fields):
        partner_obj = self.env['res.partner']
        rec = super(sale_order, self).default_get(default_fields)
        if 'partner_id' in rec.keys():
            pass
        if 'partner_id' in rec.keys():
            partner_id = False
            partner_id = partner_obj.browse(rec['partner_id'])
            if partner_id and partner_id.user_id:
                rec['user_id'] = partner_id.user_id.id
        return rec

    def _get_unavailable_package_ids(self):
        self.ensure_one()
        restricted_packages = self.package_order_lines.filtered(lambda x: x.availability == 'out_of_stock')
        if restricted_packages:
            return {
                'name':"Restricted Packages",
                'type':'ir.actions.act_window',
                'res_model':'kits.package.restriction',
                'view_mode':'form',
                'context':{'default_order_id':self.id,'default_restricted_package_ids':[(6,0,restricted_packages.ids)]},
                'target':'new',
            }
        return True

    def merge_order_lines(self):
        product_id_dict = {}
        for record in self:
            #Merge same product lines
            for line in record.order_line.filtered(lambda x: not x.is_pack_order_line):
                is_line = True
                if line.product_id in product_id_dict.keys():
                    order_line = product_id_dict[line.product_id]
                    order_line.update({'product_uom_qty':order_line.product_uom_qty + line.product_uom_qty,
                                       'price_unit':round(line.price_unit,2),
                                       'unit_discount_price':line.unit_discount_price,
                                       'discount':line.discount,
                                       'fix_discount_price':line.fix_discount_price,
                                       'picked_qty_subtotal':line.picked_qty_subtotal})
                    line.unlink()
                    is_line = False
                else:
                    product_id_dict[line.product_id] = line
                if is_line and line.product_uom_qty == 0.0:
                    line.unlink()

    def unlink(self):
        for rec in self:
            if self.env.user.has_group('base.group_system'):
                if rec.state == 'cancel' or rec.state == 'draft':
                    rec.picking_ids.unlink()
                    rec.invoice_ids.unlink()
                else:
                    raise UserError('Order must be in Cancel or Quotation.')
            else:
                raise UserError('Only administrator can delete order.')

        return super(sale_order,self).unlink()

    def write(self,vals):
        update = self.env['ir.model']._updated_data_validation(field_list,vals,self._name)
        if update:
            vals.update({'updated_by':self.env.user.id,'updated_on':datetime.today()})
        res = super(sale_order,self).write(vals)
        if vals.get('sale_manager_id'):
            self.invoice_ids.filtered(lambda inv: inv.state != 'cancel')._onchange_user_id()
        return res
    # def action_confirm(self):
        # if self.state in ['draft','sent','received']:
        #     self = self.sudo()
        #     if self.website_id and self.state == 'draft' and self.env.user.has_group('base.group_user'):
        #         self.write({'is_confirm_by_saleperson':True})
        #     # catalog_obj = self.env['sale.catalog']
        #     # sol_obj = self.env['sale.order.line']
        #     # catalog_obj.connect_server()
        #     # method = catalog_obj.get_method('action_confirm')
        #     # if method['method']:
        #         # localdict = {'self': self,'_':_,}
        #         # exec(method['method'], localdict)
        #         # record = localdict['record']
        #     order_backup_obj = self.env['sale.order.backup.spt']
        #     for record in self:
        #         geo_restriction_list = []
        #         backup_order_line_list =[]
        #         #Merge same product lines
        #         record.merge_order_lines()
        #         #checked stock
        #         # record.check_stock_spt()
        #         for line in record.order_line:
        #             backup_order_line_list.append((0,0,{
        #                 'product_id':line.product_id.id,
        #                 'product_uom_qty':line.product_uom_qty,
        #                 'price_unit':line.price_unit,
        #                 'unit_discount_price':line.unit_discount_price,
        #                 'fix_discount_price':line.fix_discount_price,
        #                 'discount':line.discount,
        #                 'name':line.name,
        #                 'sale_type':line.sale_type,
        #                 'tax_id':[(6,0,line.tax_id.ids)],
        #             })) 
        #             if line.product_id.type == 'product' and record.partner_id.country_id.id in line.product_id.geo_restriction.ids:
        #                 geo_restriction_list.append(line.id)
        #         order_backup_obj.create({
        #             'name' : record.name,
        #             'partner_id' : record.partner_id.id,
        #             'partner_invoice_id' : record.partner_invoice_id.id,
        #             'partner_shipping_id' : record.partner_shipping_id.id,
        #             'currency_id' : record.pricelist_id.currency_id.id,
        #             'payment_term_id' : record.payment_term_id.id,
        #             'date_order' : record.date_order,
        #             'applied_promo_code' : record.applied_promo_code,
        #             'line_ids' : backup_order_line_list,
        #             'order_id': record.id,
        #             'user_id': self.env.user.id,
        #             'pricelist_id' : record.pricelist_id.id
        #         })

        #         if geo_restriction_list and not record._context.get('allow_restricted'):
        #             order_lines = self.env['sale.order.line'].browse(geo_restriction_list)
        #             products=[]
        #             if len(order_lines) > 0 and not self._context.get('on_consign_wizard'):
        #                 products = order_lines.sorted(lambda x: x.product_id.variant_name).ids
        #                 return {
        #                 'name': record.name,
        #                 'type': 'ir.actions.act_window',
        #                 'view_type': 'form',
        #                 'view_mode': 'form',
        #                 'res_model': 'geo.restriction.message.wizard.spt',
        #                 'target': 'new',
        #                 'context': {'default_order_line_ids': [(6,0,products)] }
        #             }
        #     on_consign_product_ids = self.order_line.filtered(lambda x:x.product_id.on_consignment and x.product_uom_qty > x.product_id.actual_stock)
        #     if on_consign_product_ids and not self._context.get('on_consign_wizard'):
        #         for line in on_consign_product_ids:
        #             line.product_id.assign_qty = line.product_uom_qty or 0.0 
        #         return {
        #             'name': _('Product Minimum Stock Alert.'),
        #             'type': 'ir.actions.act_window',
        #             'view_type': 'form',
        #             'view_mode': 'form',
        #             'res_model': 'on.consignment.product.message.wizard',
        #             'target': 'new',
        #             'context': {'default_product_ids': [(6,0,on_consign_product_ids.mapped('product_id').ids)],'default_order_id':self.id }
        #         }
        #     if record.state in ['draft'] and not record.website_id and not record.catalog_id:
        #         template_id = self.env.ref('tzc_sales_customization_spt.mail_template_notify_customer_quotation_create').sudo()
        #         # template_id.send_mail(res_id=record.id,force_send=True,notif_layout="mail.mail_notification_light")
        #     return super(sale_order, self).action_confirm()
        # else:
        #     return {
        #         'type': 'ir.actions.client',
        #         'tag': 'display_notification',
        #         'params': {
        #                 'title': 'Something is wrong.',
        #                 'message': 'Please reload your screen.',
        #                 'sticky': True,
        #             }
        #         }
    #     # pass

    def button_open_quick_scan(self):
        self.ensure_one()
        if self.state in ['draft','sent','received','salesperson_confirmation']:
            wizard_id = self.env['sale.barcode.order.spt'].create({'partner_id':self.partner_id.id,'sale_id' : self.id})
            return {
                'name': 'Scan Order',
                'view_mode': 'form',
                'target': 'new',
                'res_id':wizard_id.id,
                'res_model': 'sale.barcode.order.spt',
                'type': 'ir.actions.act_window',
                'context':{'default_partner_id':self.partner_id.id,'default_sale_id' : self.id },
                }
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        # pass

    def action_add_shipping_cost(self):
        return {
            "name":_("Add Shipping Cost"),
            "type":"ir.actions.act_window",
            "res_model":"add.shipping.cost.wizard",
            "view_mode":"form",
            "context":{"default_sale_order_id_kits":self.id,},
            "target":"new",
        }

    def action_add_admin_fee(self):
        return {
            "name":_("Add Admin Fee"),
            "type":"ir.actions.act_window",
            "res_model":"add.admin.fee.wizard",
            "view_mode":"form",
            "context":{"default_kits_so_id":self.id,},
            "target":"new",
        }

    def action_add_aditional_discount(self):
        return {
            "name":_("Add Aditional Discount"),
            "type":"ir.actions.act_window",
            "res_model":"add.aditional.discount.wizard",
            "view_mode":"form",
            "context":{"default_sale_order_id":self.id,},
            "target":"new",
        }

    def _set_order_id(self):
        picking_obj = self.env['stock.picking']
        for rec in self:
            rec.is_picking_set = False
            if rec.state not in ['draft','sent','received']:
                if not rec.picking_ids.filtered(lambda x:x.state != 'cancel'):
                    picking_id = picking_obj.search([('origin','=',rec.name)])
                    if picking_id:
                        picking_id.sale_id = rec.id
                        rec.is_picking_set = True

    @api.depends('picking_ids','picking_ids.state','picking_ids.date_done')
    def _compute_shipped_date(self):
        for record in self:
            record.shipped_date = None
            picking_id = record.picking_ids.filtered(lambda x:x.state == 'done')
            if picking_id  and picking_id.date_done :
                record.shipped_date = picking_id.date_done

    def _get_reward_lines(self):
        self.ensure_one()
        return self.order_line.filtered(lambda line: line.is_reward_line)
        # pass

    @api.depends('order_line')
    def _compute_base_boolean_fields(self):
        for order in self:
            order.reward_amount = sum([line.price_subtotal for line in order._get_reward_lines()])
            order.is_all_service =  all(order.order_line.mapped(lambda x:x.product_id.type == 'service'))
            order.delivery_set = any(order.order_line.mapped('is_delivery'))
        # pass

    @api.depends('partner_id','partner_shipping_id')
    def _get_delivery_address(self):
        for rec in self:
            if rec.partner_id == rec.partner_shipping_id:
                rec.is_same_delivery_address = True
            else:
                rec.is_same_delivery_address = False

    @api.depends('order_line')
    def _amount_all(self):
        """
        Compute the total amounts of the SO.
        """
        picking_line_obj = self.env['stock.move']
        product_list = []
        for order in self:
            global_discount = 0.0 
            amount_is_admin = 0.0 
            amount_without_discount = 0.0
            amount_discount = 0.0
            amount_is_shipping_total = 0.0
            amount_untaxed = 0.0
            amount_tax = 0.0
            picked_qty_order_subtotal = 0.00
            unit_per_tax = 0.00
            total_discount_per_unit = 0.00
            picked_qty_order_tax = 0.00
            picked_qty_order_discount = 0.00
            picked_qty = 0
            delivered_qty = 0
            ordered_qty = 0
            shipping_cost = 0.00
            for line in  range(len(order.order_line)):
                line =  order.order_line[line]
                unit_per_tax = 0.0
                if line.is_shipping_product:
                    amount_is_shipping_total += line.unit_discount_price * line.product_uom_qty or 0.00
                    shipping_cost += line.price_unit * line.product_uom_qty or 0.00
                elif line.is_global_discount:
                    global_discount += line.price_subtotal
                elif line.is_admin:
                    amount_is_admin += line.price_subtotal
                
                else:
                    if line.product_id.type != 'service':
                        delivered_qty += line.qty_delivered
                        ordered_qty += line.product_uom_qty
                    if line.price_tax and line.product_uom_qty:
                        unit_per_tax = round(line.price_tax/line.product_uom_qty,2)
                        amount_tax = amount_tax + line.picked_qty_subtotal*line.tax_id.amount/100
                        # amount_tax = amount_tax + round(unit_per_tax * line.product_uom_qty,2)
                    amount_untaxed += line.price_subtotal
                    amount_discount +=  ((line.product_uom_qty * line.price_unit) - line.price_subtotal)
                    amount_without_discount += line.product_uom_qty * line.price_unit
    
                if line.product_id.id not in product_list:
                    picking_ids =picking_line_obj.search([('picking_id.state','!=','cancel'),('picking_id','in',order.picking_ids.ids),('product_id','=',line.product_id.id)])
                    for picking in picking_ids:
                        if line.product_id.type != 'service':
                            picked_qty += picking.quantity_done
                    product_list.append(line.product_id.id)


                if line.product_id.type != 'service' and line.picked_qty:
                        picked_qty_order_subtotal = picked_qty_order_subtotal + (line.price_unit * line.picked_qty)

                        unit_per_tax = round(line.price_tax/line.product_uom_qty,2)
                        # picked_qty_order_tax = picked_qty_order_tax + round(unit_per_tax * line.picked_qty,2)
                        picked_qty_order_tax = picked_qty_order_tax + line.picked_qty_subtotal*line.tax_id.amount/100

                        total_discount_per_unit = line.price_unit - line.unit_discount_price
                        picked_qty_order_discount = round(picked_qty_order_discount + (total_discount_per_unit * line.picked_qty),2)
            amount_discount += abs(global_discount)
            picked_qty_order_discount += round(abs(global_discount),2)
        
            order.update({
                'delivered_qty': delivered_qty,
                'picked_qty': picked_qty,
                'ordered_qty': ordered_qty,
                'amount_untaxed': amount_untaxed,
                'amount_without_discount': amount_without_discount,
                'amount_discount': amount_discount,
                'amount_tax': round(amount_tax,2),
                'amount_total': amount_without_discount + amount_tax + amount_is_shipping_total+amount_is_admin-amount_discount,
                'amount_is_shipping_total' : amount_is_shipping_total,
                'shipping_cost' : shipping_cost,
                'amount_is_admin' : amount_is_admin,
                'global_discount' : - global_discount,
                'picked_qty_order_subtotal' : round(picked_qty_order_subtotal,2),
                'picked_qty_order_tax' : round(picked_qty_order_tax,2),
                'picked_qty_order_discount' : round(picked_qty_order_discount,2),
                'picked_qty_order_total' : round(amount_is_admin + amount_is_shipping_total + picked_qty_order_subtotal + picked_qty_order_tax - picked_qty_order_discount ,2)
            })


    @api.depends('website_id', 'date_order', 'order_line', 'state', 'partner_id')
    def _compute_abandoned_cart(self):
        for order in self.sudo():
            # a quotation can be considered as an abandonned cart if it is linked to a website,
            # is in the 'draft' state and has an expiration date
            if order.state == 'draft' and order.date_order:
                public_partner_id = order.partner_id
                # by default the expiration date is 1 hour if not specified on the website configuration
                abandoned_delay = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0')) or 1.0
                abandoned_datetime = datetime.utcnow() - relativedelta(hours=abandoned_delay)
                order.is_abandoned_cart = bool(order.date_order <= abandoned_datetime and order.partner_id != public_partner_id and order.order_line)
            else:
                order.is_abandoned_cart = False

    def _compute_create_uid_spt(self):
        for record in self:
            if record.catalog_id or  record.website_id:
                record.create_uid_spt = record.partner_id.user_id.id
            else:
                record.create_uid_spt = record.create_uid.id 

    @api.depends('invoice_ids.name')
    def _compute_invoice_name(self):
        invoice_obj = self.env['account.move']
        for record in self:
            invoice_name = ''
            invoice_ids = invoice_obj.search([('state','not in',['cancel']),('id','in',record.invoice_ids.ids)])
            if record.state in ['open_inv','paid'] and invoice_ids:
                invoice_name = invoice_ids[0].name
            record.invoice_name = invoice_name

    @api.constrains('partner_id')
    def _check_partner_id(self):
        partner_list = [user.partner_id.id for user in self.env['res.users'].search([('active','=',False)]) if user.has_group('base.group_public')] + self.env.ref('base.group_public').users.mapped('partner_id.id') 
        for record in self:
            if record.partner_id.id in partner_list:
                raise UserError(_("Public type users can not be a customer."))

    @api.depends('state','catalog_id','website_id')
    def _compute_order_source(self):
        for record in self:
            source_spt = 'Manually'
            if record.catalog_id:
                source_spt = 'Catalog'
            if record.website_id:
                source_spt = 'Website'
        record.write({'source_spt':source_spt})

    @api.depends('picked_qty','ordered_qty')
    def get_no_of_cases(self):
        for rec in self:
            rec.no_of_cases = 0
            if rec.picked_qty:
                no_of_cases = sum(self.env['stock.picking'].browse(rec.picking_ids.ids).filtered(lambda x:x.state != 'cancel').move_ids_without_package.filtered(lambda x:x.product_id.sale_type != 'clearance').mapped('quantity_done'))
                rec.no_of_cases = no_of_cases
            elif rec.ordered_qty:
                no_of_cases = sum(rec.order_line.filtered(lambda x:x.product_id.sale_type != 'clearance').mapped('product_uom_qty'))
                rec.no_of_cases = no_of_cases

    @api.depends('picking_ids')
    def _compute_picking_ids(self):
        stock_move_obj = self.env['stock.picking']
        for order in self:
            if order.picking_ids:
                order.delivery_count = len(order.picking_ids.filtered(lambda x: x.state != 'cancel'))
            else:
                picking_id = stock_move_obj.search([('origin','=',order.name),('state','!=','cancel')],limit=1)
                order.picking_ids = [(4,(picking_id.id))] if picking_id else None
                order.delivery_count = len(picking_id) if picking_id else 0

    @api.onchange('include_cases')
    def _onchange_include_cases(self):
        for record in self:
            weight = self.env['ir.config_parameter'].sudo().get_param('tzc_case_weight_gm')
            picking_id =self.env['stock.picking'].search([('id','in',record.picking_ids.ids)])
            if not record.case_weight_gm:
                record.case_weight_gm = weight
            if picking_id:
                picking_id.include_cases = record.include_cases

    @api.onchange('partner_shipping_id', 'partner_id', 'company_id','partner_id.country_id','pricelist_id')
    def onchange_partner_shipping_id_kits(self):
        fiscal_position_obj = self.env['account.fiscal.position']
        for record in self:
            record.pricelist_id = record.partner_id.b2b_pricelist_id if record.partner_id else False
            record.b2b_currency_id = record.partner_id.preferred_currency.id
            if record.partner_id and record.partner_id.country_id:
                 record.fiscal_position_id = self.env['account.fiscal.position'].with_context(force_company=record.company_id.id or self.env.user.company_id.id)._get_fiscal_position(record.partner_id)
            else:
                fiscal_position_id = fiscal_position_obj.sudo().search([('country_id','=',False)]).id
                record.fiscal_position_id = fiscal_position_id
            record.order_line._compute_tax_id()

    def check_stock_for_quotaion_spt(self,sale_id,order_lines):
        warning = False
        product_obj = self.env['product.product'].sudo()
        if sale_id:
            # sale = self.env['sale.order'].sudo().browse(sale_id)
            # sale_order_line_obj = self.env['sale.order.line'].sudo()
            # for product in sale.order_line.mapped('product_id'):
            #     total_order_qty = 0
            #     for line in sale_order_line_obj.search([('order_id','=',sale.id),('product_id.is_shipping_product','=',False),('product_id.is_admin','=',False),('product_id','=',product.id)]):
            #         total_order_qty = total_order_qty + line.product_uom_qty
            #     if total_order_qty > 0.0 and (total_order_qty > product.available_qty_spt):
            #         warning = True
            #         break
            for line_vals in order_lines:
                if 'product_id' in line_vals:
                    product = product_obj.browse(int(line_vals['product_id']))
                    if 'quantity' in  line_vals:
                        order_qty = line_vals['quantity'] or 0
                    else:
                        order_qty = 0
                    available_qty = product.available_qty_spt - product.minimum_qty if product.on_consignment else product.available_qty_spt
                    if order_qty > 0 and order_qty > available_qty:
                        warning = True
                        break
        return warning

    def check_stock_of_orders(self, order_lines):
        order_data = []
        for line in order_lines:
            product_id = self.env['product.product'].search([('id','=',line.get('product_id'))])
            qty = product_id.available_qty_spt - product_id.minimum_qty if product_id.on_consignment else product_id.available_qty_spt
            data = {
                'product_id':line.get('product_id'),
                'quantity': 0 if qty<=0 else qty,
                'unit':product_id.uom_name,
                'warning': True if line.get('quantity') > qty else False,
            }
            order_data.append(data)

        return order_data

    def accept_quotaion_spt(self, order_id, order_lines):
        product_obj = self.env['product.product']
        if order_id:
            sale_order = self.env['sale.order'].search([('id','=',order_id)])
            for line in order_lines:
                product_id  = product_obj.search([('id', '=', line['product_id'])])
                if product_id in sale_order.order_line.mapped('product_id'):
                    line_id = sale_order.order_line.filtered(lambda x:x.product_id == product_id)
                    qty = line['quantity']
                    line_id.write({'product_uom_qty':qty,
                                   'discount':float(line.get('discount'))})
                    line_id._onchange_discount_spt()
                    line_id._onchange_fix_discount_price_spt()
                    line_id._onchange_unit_discounted_price_spt()
                else:
                    fix_discount_price =  round((float(line.get('price')) * float(line.get('discount')))/100,2)
                    order_lines_vals = [(0,0,{
                                'product_id' : product_id.id,
                                'product_uom_qty' : line['quantity'],
                                'discount': float(line.get('discount')),
                                'fix_discount_price': fix_discount_price,
                                'unit_discount_price': round(float(line.get('price')) - fix_discount_price,2),
                                'sale_type': product_id.sale_type,
                            })]

                    sale_order.write({
                        'order_line':order_lines_vals
                    })

            for line in sale_order.order_line.filtered(lambda x:not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount):
                if line.product_id.id not in [int(i.get('product_id')) for i in order_lines]:
                    line.unlink()

            sale_order.write({'state':'received'})
            template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_quotation_confirmation_spt')
            template_id.send_mail(self.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")

            return True

    @api.onchange('partner_id')
    def onchange_partner_id(self):
        """
        Update the following fields when the partner is changed:
        - Pricelist
        - Payment terms
        - Invoice address
        - Delivery address
        - Sales Team
        """
        if not self.partner_id:
            self.update({
                'partner_invoice_id': False,
                'partner_shipping_id': False,
                'fiscal_position_id': False,
            })
            return

        addr = self.partner_id.address_get(['delivery', 'invoice'])
        partner_user = self.partner_id.user_id or self.partner_id.commercial_partner_id.user_id
        values = {
            'pricelist_id': self.partner_id.b2b_pricelist_id and self.partner_id.b2b_pricelist_id.id or False,
            'payment_term_id': self.partner_id.property_payment_term_id and self.partner_id.property_payment_term_id.id or False,
            'partner_invoice_id': addr['invoice'],
            'partner_shipping_id': addr['delivery'],
        }
        user_id = partner_user.id
        if not self.env.context.get('not_self_saleperson'):
            user_id = user_id or self.env.context.get('default_user_id', self.env.uid)
        if user_id and self.user_id.id != user_id:
            values['user_id'] = user_id

        if self.env['ir.config_parameter'].sudo().get_param('account.use_invoice_terms') and self.env.company.invoice_terms:
            values['note'] = self.with_context(lang=self.partner_id.lang).env.company.invoice_terms
        if not self.env.context.get('not_self_saleperson') or not self.team_id:
            values['team_id'] = self.env['crm.team'].with_context(
                default_team_id=self.partner_id.team_id.id
            )._get_default_team_id(domain=['|', ('company_id', '=', self.company_id.id), ('company_id', '=', False)], user_id=user_id)
        self.update(values)

    @api.onchange('partner_id')
    def onchange_partner_id_user_id_spt(self):
        for record in self:
            for line in record.order_line:
                line.with_context(partner_change=True).product_uom_change()
            record.onchange_partner_id()
            if record.partner_id and record.partner_id.user_id:
                record.user_id = record.partner_id.user_id.id
                if record._origin.id:
                    return {'warning':{'message':"Customer is changed, Please Verify 'Salesperson' and 'Salesmanager'.",'title':'Warning'}}

    def excel_order_report(self):
        return {
                'name': 'Freight Invoice',
                'view_mode': 'form',
                'target': 'new',
                'context':{'default_sale_id':self.id,},
                'res_model': 'sale.order.report.wizard.spt',
                'type': 'ir.actions.act_window',
            }

    def excel_report(self):
        # catalog_obj = self.env['sale.catalog']
        # catalog_obj.connect_server()
        # method = catalog_obj.get_method('excel_report')
        # if method['method']:
        #     localdict = {'self': self,'Workbook':Workbook,'Font':Font,'Alignment':Alignment,'Side':Side,'Border':Border,'BytesIO':BytesIO,'base64':base64,'openpyxl':openpyxl}
        #     exec(method['method'], localdict)
        #     active_id = localdict['active_id']
        #     f_name = localdict['f_name']
        
        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Abbreviate Report Sample.xlsx'
        wb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
        wrksht = wb.active

        # active_id = self.id
        f_name = 'Freight-Inv-'+str(self.name or '')  # 
        address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0)
        alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0)

        address_font = Font(name='Lato', size=9, bold=False)
        table_font = Font(name='Lato', size=9, bold=False)

        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
        alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0)
        # alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0)

        bd = Side(style='thin', color="d2d4d4")
        tp_bd = Side(style='thin', color="000000")
        all_border = Border(left=Side(style='thin', color="d2d4d4"), 
                            right=Side(style='thin', color="d2d4d4"), 
                            top=Side(style='thin', color="d2d4d4"), 
                            bottom=Side(style='thin', color="d2d4d4"))
        bottom_border = Border(bottom=bd)
        top_border = Border(top=tp_bd)
        name_header_font = Font(name="Lato", size=9, bold=True)
        bank_detail_font = Font(name="Lato", size=9, bold=False)
        


        # ------------------------------------------------------------
        # Billing Address
        # ------------------------------------------------------------
        address_row = 9
        billing_address = self.create_address_line_for_sale(self.partner_invoice_id, take_name=True)
        wrksht.merge_cells("A"+str(address_row)+":E" + str(address_row+6))  # added
        wrksht.cell(row=address_row, column=1).value = billing_address
        wrksht.cell(row=address_row, column=1).alignment = address_alignment
        wrksht.cell(row=address_row, column=1).font = address_font

        # -------------------------------------------------------------
        # Shipping Address
        # -------------------------------------------------------------
        wrksht.merge_cells('G'+str(address_row)+':J'+str(address_row+6))
        shipping_address = self.create_address_line_for_sale(self.partner_shipping_id, take_name=True)
        wrksht.cell(row=address_row, column=7).value = shipping_address
        wrksht.cell(row=address_row, column=7).alignment = address_alignment
        wrksht.cell(row=address_row, column=7).font = address_font

        # ---------------------------------------------------------
        # name
        # ---------------------------------------------------------
        name_row = address_row + 7
        wrksht.cell(row=name_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        invoice = self.invoice_ids.filtered(lambda x: x.state != 'cancel')
        if len(invoice):
            invoice = invoice if len(invoice) else invoice[0]
        wrksht.cell(row=name_row, column=1).value = str("Invoice " + invoice.name if invoice else 'Invoice %s'%(self.name))
        wrksht.cell(row=name_row, column=1).font = Font(name='Lato', size=14, bold=False,color="666666")

        # ----------------------------------------------------------------
        # Date, Salesperson
        # ----------------------------------------------------------------
        date_person_row = name_row + 3
        # wrksht.row_dimensions[date_person_row].height = 25
        # wrksht.row_dimensions[date_person_row+1].height = 25

        wrksht.cell(row=date_person_row-1, column=1).value = str("Invoice Date:" if invoice.invoice_date else '') if invoice else "Order Date:"

        wrksht.cell(row=date_person_row, column=1).value = str(str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y'))

        # wrksht.cell(row=date_person_row, column=1).font = name_header_font
        # wrksht.cell(row=date_person_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=2).value = self.name or ''
        # wrksht.cell(row=date_person_row, column=2).font = name_header_font
        # wrksht.cell(row=date_person_row, column=2).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''
        # wrksht.cell(row=date_person_row, column=9).font = name_header_font
        # wrksht.cell(row=date_person_row, column=9).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        
        picking = self.picking_ids.filtered(lambda x: 'WH/OUT' in x.name and x.state == 'done')
        picking = picking if len(picking) == 1 else picking[0] if len(picking) else False
        wrksht.cell(row=date_person_row, column=5).value = str(picking.shipping_id.name if picking and picking.shipping_id else '') or ''
        # wrksht.cell(row=date_person_row, column=7).font = name_header_font
        # wrksht.cell(row=date_person_row, column=7).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=7).value = str(picking.tracking_number_spt if picking and picking.tracking_number_spt else '') or ''
        # wrksht.cell(row=date_person_row, column=6).font = name_header_font
        # wrksht.cell(row=date_person_row, column=6).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
       
        wrksht.cell(row=date_person_row, column=9).value = self.no_of_cases or ''
        # wrksht.cell(row=date_person_row, column=5).font = name_header_font
        # wrksht.cell(row=date_person_row, column=5).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=10).value = self.b2b_currency_id.name or ''
        # wrksht.cell(row=date_person_row, column=8).font = name_header_font
        # wrksht.cell(row=date_person_row, column=8).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        # ========================= Product Table ===========================
        table_header = date_person_row+2
        
        query = f'''SELECT	PP.VARIANT_NAME,
                            PC.NAME,
                            PP.MATERIAL,
                            PP.HS_CODE,
                            SOL.PICKED_QTY,
                            SOL.UNIT_DISCOUNT_PRICE,
                            (SOL.UNIT_DISCOUNT_PRICE * SOL.PICKED_QTY) AS SUBTOTAL,
                            CASE
                                            WHEN SOL.IS_SHIPPING_PRODUCT = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS SHIPPING_CHARGE,
                            CASE
                                            WHEN SOL.IS_ADMIN = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS ADMIN_FEE,
                            CASE
                                            WHEN SOL.IS_GLOBAL_DISCOUNT = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS ADDITIONAL_DISCOUNT,
                            CASE
                                            WHEN SOL.PRODUCT_UOM_QTY > 0
                                                                AND SOL.PRICE_TAX > 0 THEN SOL.PRICE_TAX / SOL.PRODUCT_UOM_QTY * SOL.PICKED_QTY
                                            ELSE 0
                            END AS PRICE_TAX
                            FROM SALE_ORDER AS SO
                            INNER JOIN SALE_ORDER_LINE AS SOL ON SOL.ORDER_ID = SO.ID
                            INNER JOIN STOCK_PICKING AS SP ON SP.ORIGIN = SO.NAME
                            INNER JOIN SHIPPING_PROVIDER_SPT AS SPS ON SP.SHIPPING_ID = SPS.ID
                            INNER JOIN RES_CURRENCY AS RC ON RC.ID = SOL.CURRENCY_ID
                            INNER JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                            INNER JOIN PRODUCT_CATEGORY AS PC ON PC.ID = PP.CATEG_ID
                            WHERE SO.ID ={self.id}'''
        self.env.cr.execute(query)
        record_data = self.env.cr.fetchall()
        row_index = table_header+1
        total_quantity = 0
        sub_total = 0
        shipping_cost = 0
        admin_fee = 0
        additional_discount = 0
        taxes = 0

        for data in record_data:
            total_quantity += data[4]
            sub_total += data[5] * data[4]
            shipping_cost += data[7]
            admin_fee += data[8]
            additional_discount += data[9]
            taxes += data[10] 
            # product name
            if data[4] <= 0:
                pass
            else:
                wrksht.merge_cells("A"+str(row_index)+":D"+str(row_index))
                wrksht.cell(row=row_index, column=1).value = data[0]
                # category name
                wrksht.cell(row=row_index, column=5).value = data[1]
                # material
                wrksht.cell(row=row_index, column=6).value = data[2]
                # HS Code
                wrksht.cell(row=row_index, column=7).value = data[3]
                # product qty
                wrksht.cell(row=row_index, column=8).value = data[4]
                # discount
                # wrksht.cell(row=row_index, column=6).value = dict_data.discount
                # price
                wrksht.cell(row=row_index, column=9).value = "$ {:,.2f}".format(data[5])
                # subtotal
                wrksht.cell(row=row_index, column=10).value = "$ {:,.2f}".format(data[6])
                
                # font of lines
                wrksht.cell(row=row_index, column=1).font = table_font
                wrksht.cell(row=row_index, column=5).font = table_font
                wrksht.cell(row=row_index, column=6).font = table_font
                wrksht.cell(row=row_index, column=7).font = table_font
                wrksht.cell(row=row_index, column=8).font = table_font
                wrksht.cell(row=row_index, column=9).font = table_font
                wrksht.cell(row=row_index, column=10).font = table_font
                
                # border of lines
                wrksht.cell(row=row_index, column=1).border = bottom_border
                wrksht.cell(row=row_index, column=2).border = bottom_border
                wrksht.cell(row=row_index, column=3).border = bottom_border
                wrksht.cell(row=row_index, column=4).border = bottom_border
                wrksht.cell(row=row_index, column=5).border = bottom_border
                wrksht.cell(row=row_index, column=6).border = bottom_border
                wrksht.cell(row=row_index, column=7).border = bottom_border
                wrksht.cell(row=row_index, column=8).border = bottom_border
                wrksht.cell(row=row_index, column=9).border = bottom_border
                wrksht.cell(row=row_index, column=10).border = bottom_border
                
                # alignment of lines
                wrksht.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
                wrksht.cell(row=row_index, column=5).alignment = alignment
                wrksht.cell(row=row_index, column=6).alignment = alignment
                wrksht.cell(row=row_index, column=7).alignment = alignment
                wrksht.cell(row=row_index, column=8).alignment = alignment
                wrksht.cell(row=row_index, column=9).alignment = alignment_right
                wrksht.cell(row=row_index, column=10).alignment = alignment_right

                wrksht.row_dimensions[row_index].height = 20
                row_index += 1
        # ======================== Product table end ========================

        # ----------------------------------------------------
        # Above table total Quantity
        # ----------------------------------------------------
        wrksht.cell(row=date_person_row, column=8).value = total_quantity or 0
        # wrksht.cell(row=date_person_row, column=4).font = name_header_font
        # wrksht.cell(row=date_person_row, column=4).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        footer_row = row_index+1

        # ===================== Bank Details =========================
        # if invoice and invoice.get_html_field_val(invoice.company_id.bank_details):
        #     wrksht.merge_cells("A"+str(footer_row)+":E"+str(footer_row))
        #     wrksht.cell(row=footer_row, column=1).value = "Bank Transfer Details"
        #     wrksht.cell(row=footer_row, column=1).font = Font(name='Lato', size=10, bold=True)
        #     wrksht.cell(row=footer_row, column=1).alignment = alignment_left
        #     for i in range(1,6):
        #         wrksht.cell(row=footer_row, column=i).border = Border(left=Side(style='thin', color="d2d4d4"),
        #                                                               right=Side(style='thin', color="d2d4d4"), 
        #                                                               top=Side(style='thin', color="d2d4d4"))

        #     wrksht.merge_cells("A"+str(footer_row+1)+":E"+str(footer_row+6))
        #     wrksht.row_dimensions[footer_row+6].height = 67
        #     bank_details = BeautifulSoup(invoice.company_id.bank_details,"html.parser")
        #     # wrksht.cell(row=footer_row+1, column=1).value = '\n'.join([i.strip() for i in bank_details.get_text().split('\n') if len(i.strip()) > 0])
        #     wrksht.cell(row=footer_row+1, column=1).value = self.bank_details()
        #     wrksht.cell(row=footer_row+1, column=1).font = bank_detail_font
        #     wrksht.cell(row=footer_row+1, column=1).alignment = address_alignment
        #     bank_details_row = footer_row+1
        #     for row in range(footer_row+1,footer_row+7):
        #         for col in range(1,6):
        #             wrksht.cell(row=row, column=col).border = Border(left=Side(style='thin', color="d2d4d4"),
        #                                                               right=Side(style='thin', color="d2d4d4"), 
        #                                                               bottom=Side(style='thin', color="d2d4d4"))

        #         bank_details_row += 1

        # # ===================== Footer right =========================
        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Subtotal"
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(sub_total)
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Shipping"
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(shipping_cost)  # shipping
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Admin Fee"
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(admin_fee)  # adminfee
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1


        if abs(additional_discount):
            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
            wrksht.cell(row=footer_row, column=7).value = "Discount"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(abs(additional_discount))  # discont
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = table_font
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Tax"
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(taxes)  # taxes
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Total"
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "({}) $ {:,.2f}".format(self.b2b_currency_id.name,taxes + sub_total+shipping_cost+admin_fee-abs(additional_discount)) # total
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Total Quantity"
        wrksht.cell(row=footer_row, column=7).font = Font(name='Lato', size=9, bold=True)
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = int(total_quantity)  # TotalQuantity
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1
        
        # ========================== footer close ========================
        bank_details_wb = wb.get_sheet_by_name('Bank Details')
        bank_sheet_row = 11

        bank_details_wb.cell(row=bank_sheet_row,column=3).value = invoice.name if invoice  else self.name
        bank_details_wb.cell(row=bank_sheet_row+1,column=1).value = str("Invoice Date:" if invoice.invoice_date else '') if invoice else "Order Date:"
        bank_details_wb.cell(row=bank_sheet_row+1,column=3).value = str(str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y'))
        bank_details_wb.cell(row=bank_sheet_row+2,column=3).value = self.b2b_currency_id.name or ''
        bank_details_wb.cell(row=bank_sheet_row+3,column=3).value = "({}) $ {:,.2f}".format(self.b2b_currency_id.name,taxes + sub_total+shipping_cost+admin_fee-abs(additional_discount))
        bank_details_wb.cell(row=bank_sheet_row+4,column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''
        # total width 70 perfect for excel->pdf

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, f_name),
            'target': 'self',
        }

    def excel_report_line(self):
        
        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Assorted Report Sample.xlsx'
        workb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
        wrksht = workb.active
        wrksht = workb.get_sheet_by_name(workb.sheetnames[0])

        bd = Side(style='thin', color="d2d4d4")
        tp_bd = Side(style='thin', color="000000")
        # all_border = Border(left=Side(style='thin', color="d2d4d4"), 
        #                     right=Side(style='thin', color="d2d4d4"), 
        #                     top=Side(style='thin', color="d2d4d4"), 
        #                     bottom=Side(style='thin', color="d2d4d4"))
        address_font = Font(name='Lato', size=9, bold=False)
        # name_header_font = Font(name="Lato", size=9, bold=True)
        bottom_border = Border(bottom=bd)
        top_border = Border(top=tp_bd)
        alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0)
        alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0)
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
        address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
        table_font = Font(size=9, bold=False, name="Lato")
        bank_detail_font = Font(name="Lato", size=7, bold=False)
        
        # ------------------------------------------------------------
        # Billing Address
        # ------------------------------------------------------------
        address_row = 9

        wrksht.merge_cells("A"+str(address_row)+":E" + str(address_row+6))  # added
        billing_address = self.create_address_line_for_sale(self.partner_invoice_id, take_name=True)
        wrksht.cell(row=address_row, column=1).value = billing_address
        wrksht.cell(row=address_row, column=1).alignment = address_alignment
        wrksht.cell(row=address_row, column=1).font = address_font

        # -------------------------------------------------------------
        # Shipping Address
        # -------------------------------------------------------------
        wrksht.merge_cells('G'+str(address_row)+':J'+str(address_row+6))
        shipping_address = self.create_address_line_for_sale(self.partner_shipping_id, take_name=True)
        wrksht.cell(row=address_row, column=7).value = shipping_address
        wrksht.cell(row=address_row, column=7).alignment = address_alignment
        wrksht.cell(row=address_row, column=7).font = address_font

        # ---------------------------------------------- Name ----------------------------------------------
        name_row = address_row + 7
        wrksht.merge_cells('A'+str(name_row)+':D'+str(name_row))
        wrksht.cell(row=name_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        invoice = self.invoice_ids.filtered(lambda x: x.state != 'cancel')
        if len(invoice):
            invoice = invoice if len(invoice) else invoice[0]
        wrksht.cell(row=name_row, column=1).value = str("Invoice "+invoice.name if invoice else 'Invoice %s'%(self.name))
        # if invoice:
        wrksht.cell(row=name_row, column=1).font = Font(name='Lato', size=14, bold=False,color='666666')
        # else:
        #     wrksht.cell(row=name_row, column=1).font = Font(name='Lato', size=12, bold=False,color='666666')


        # ---------------------------------------------- date, salesperson, total qty ----------------------------------------------
        date_person_row = name_row + 3
        
        # wrksht.row_dimensions[date_person_row].height = 25
        # wrksht.row_dimensions[date_person_row+1].height = 25

        wrksht.cell(row=date_person_row-1, column=1).value = str("Invoice Date:" if invoice.invoice_date else 'Order Date')

        wrksht.cell(row=date_person_row, column=1).value = str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y') or ''
        # wrksht.cell(row=date_person_row, column=1).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=2).value = self.name or ''
        # wrksht.cell(row=date_person_row, column=2).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=2).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        # wrksht.merge_cells("I"+str(date_person_row)+":J"+str(date_person_row+1))
        wrksht.cell(row=date_person_row, column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''
        # wrksht.cell(row=date_person_row, column=9).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=9).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        picking = self.picking_ids.filtered(lambda x: 'WH/OUT' in x.name and x.state == 'done')
        picking = picking if len(picking) == 1 else picking[0] if len(picking) else False
        # wrksht.merge_cells("F"+str(date_person_row)+":F"+str(date_person_row+1))
        wrksht.cell(row=date_person_row, column=5).value = str(picking.shipping_id.name if picking and picking.shipping_id else '') or ''
        # wrksht.cell(row=date_person_row, column=7).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=7).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=7).value = str(picking.tracking_number_spt if picking and picking.tracking_number_spt else '') or ''
        # wrksht.cell(row=date_person_row, column=6).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=6).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        wrksht.cell(row=date_person_row, column=9).value = self.no_of_cases or ''
        # wrksht.cell(row=date_person_row, column=5).font = name_header_font
        # wrksht.cell(row=date_person_row, column=5).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)


        # wrksht.merge_cells("G"+str(date_person_row)+":G"+str(date_person_row+1))
        wrksht.cell(row=date_person_row, column=10).value = self.b2b_currency_id.name or ''
        # wrksht.cell(row=date_person_row, column=8).font = Font(name="Lato", size=10, bold=True)
        # wrksht.cell(row=date_person_row, column=8).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

        table_header = date_person_row+2

        query = f'''SELECT SO.NAME,
                        ROUND(SUM(SOL.PRICE_UNIT) / SUM(SOL.PRODUCT_UOM_QTY),2) AS TOTAL_AMOUNT,
                        ROUND(SUM(SOL.FIX_DISCOUNT_PRICE) / SUM(SOL.PRODUCT_UOM_QTY)) AS TOTAL_DISCOUNT,
                        SUM(SOL.UNIT_DISCOUNT_PRICE * SOL.PICKED_QTY) AS TOTAL_PRICE,
                        SUM((SOL.PRICE_TAX / SOL.PRODUCT_UOM_QTY) * SOL.PICKED_QTY) AS TAX,
                        SUM(SOL.PICKED_QTY) AS QTY,
                        CASE
                                        WHEN PC.NAME = 'E'
                                                            AND PT.TYPE != 'service'
                                                            AND SOL.PICKED_QTY > 0 THEN 'Assorted Eyeglasses'
                                        WHEN PC.NAME = 'S'
                                                            AND PT.TYPE != 'service'
                                                            AND SOL.PICKED_QTY > 0 THEN 'Assorted Sunglasses'
                                        WHEN PC.NAME not in ('E','S')
                                                            AND PT.TYPE != 'service'
                                                            AND SOL.PICKED_QTY > 0 THEN 'Assorted Other'
                        END AS TYPE
                    FROM SALE_ORDER AS SO
                    INNER JOIN SALE_ORDER_LINE AS SOL ON SO.ID = SOL.ORDER_ID
                    INNER JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                    INNER JOIN PRODUCT_TEMPLATE AS PT ON PT.ID = PP.PRODUCT_TMPL_ID
                    INNER JOIN PRODUCT_CATEGORY AS PC ON PC.ID = PP.CATEG_ID
                    WHERE SO.ID = {self.id}
                    GROUP BY TYPE,
                        SO.NAME,
                        PC.NAME,
                        SOL.PICKED_QTY,
                        SOL.PRICE_UNIT'''
        self.env.cr.execute(query)
        record_data = self.env.cr.fetchall()
        row_index = table_header+1
        sub_total = 0.00
        tax = 0.00
        total_quantity = 0
        for data in record_data:
            if (data[6] == None):
                pass
            else:
                total_quantity += data[5]
                height = int((3*len(data[6]))/2) if len(data) > 20 else 20
                wrksht.row_dimensions[row_index].height = height

                wrksht.merge_cells("A"+str(row_index)+":E"+str(row_index))
                wrksht.cell(row=row_index, column=1).value = data[6]
                wrksht.cell(row=row_index, column=1).font = table_font
                wrksht.cell(row=row_index, column=1).border = bottom_border
                wrksht.cell(row=row_index, column=2).border = bottom_border
                wrksht.cell(row=row_index, column=3).border = bottom_border
                wrksht.cell(row=row_index, column=4).border = bottom_border
                wrksht.cell(row=row_index, column=5).border = bottom_border
                
                wrksht.cell(row=row_index, column=6).value = data[5]
                wrksht.cell(row=row_index, column=6).font = table_font
                wrksht.cell(row=row_index, column=6).border = bottom_border

                
                wrksht.merge_cells("G"+str(row_index)+":H"+str(row_index))
                wrksht.cell(row=row_index, column=7).value = "$ {:,.2f}".format(round(data[1]/data[5],2))
                wrksht.cell(row=row_index, column=7).font = table_font
                wrksht.cell(row=row_index, column=7).border = bottom_border
                wrksht.cell(row=row_index, column=8).border = bottom_border

                wrksht.merge_cells("I"+str(row_index)+":J"+str(row_index))
                wrksht.cell(row=row_index, column=9).value = "$ {:,.2f}".format(round(data[3] , 2))
                wrksht.cell(row=row_index, column=9).font = table_font
                wrksht.cell(row=row_index, column=9).border = bottom_border
                wrksht.cell(row=row_index, column=10).border = bottom_border
                
                sub_total += round(data[3], 2)
                tax = round(data[4], 2) + tax

                wrksht.cell(row=row_index, column=1).alignment = alignment_left
                wrksht.cell(row=row_index, column=6).alignment = alignment
                wrksht.cell(row=row_index, column=7).alignment = alignment_right
                wrksht.cell(row=row_index, column=8).alignment = alignment_right
                wrksht.cell(row=row_index, column=9).alignment = alignment_right

                row_index += 1

        # ========================= Table end =========================
        # wrksht.merge_cells("D"+str(date_person_row)+":D"+str(date_person_row+1))
        wrksht.cell(row=date_person_row, column=8).value = str(int(total_quantity))
        # wrksht.cell(row=date_person_row, column=4).font = name_header_font
        # wrksht.cell(row=date_person_row, column=4).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        footer_row = row_index+1
        
        # ===================== Bank Details =========================
        # if invoice and invoice.get_html_field_val(invoice.company_id.bank_details):
        #     wrksht.merge_cells("A"+str(footer_row)+":E"+str(footer_row))
        #     wrksht.cell(row=footer_row, column=1).value = "Bank Transfer Details"
        #     wrksht.cell(row=footer_row, column=1).font = Font(name='Lato', size=10, bold=True)
        #     wrksht.cell(row=footer_row, column=1).alignment = alignment_left
        #     for i in range(1,6):
        #         wrksht.cell(row=footer_row, column=i).border = Border(left=Side(style='thin', color="d2d4d4"),
        #                                                               right=Side(style='thin', color="d2d4d4"), 
        #                                                               top=Side(style='thin', color="d2d4d4"))

        #     wrksht.merge_cells("A"+str(footer_row+1)+":E"+str(footer_row+6))
        #     wrksht.row_dimensions[footer_row+6].height = 67
        #     bank_details = BeautifulSoup(invoice.company_id.bank_details,"html.parser")
        #     wrksht.cell(row=footer_row+1, column=1).value = self.bank_details()
        #     # wrksht.cell(row=footer_row+1, column=1).value = '\n'.join([i.strip() for i in bank_details.get_text().split('\n') if len(i.strip()) > 0])
        #     wrksht.cell(row=footer_row+1, column=1).font = bank_detail_font
        #     wrksht.cell(row=footer_row+1, column=1).alignment = address_alignment
        #     bank_details_row = footer_row+1
        #     for row in range(footer_row+1,footer_row+7):
        #         for col in range(1,6):
        #             wrksht.cell(row=row, column=col).border = Border(left=Side(style='thin', color="d2d4d4"),
        #                                                               right=Side(style='thin', color="d2d4d4"), 
        #                                                               bottom=Side(style='thin', color="d2d4d4"))

        #         bank_details_row += 1
                
        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = 'Subtotal'
        wrksht.cell(row=footer_row, column=7).font =  Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(round(self.picked_qty_order_subtotal - self.picked_qty_order_discount,2))
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20


        footer_row += 1
        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = 'Shipping Cost'
        wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(self.amount_is_shipping_total)
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = 'Admin Fee'
        wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(self.amount_is_admin)
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1


        if abs(self.global_discount):
            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))
        
            wrksht.cell(row=footer_row, column=7).value = "Discount"
            wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(self.picked_qty_order_discount)
            wrksht.cell(row=footer_row, column=9).font = table_font
            wrksht.cell(row=footer_row, column=9).alignment = alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = 'Tax'
        wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border

        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(self.picked_qty_order_tax)
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = 'Total'
        wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(self.picked_qty_order_total)
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
        wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

        wrksht.cell(row=footer_row, column=7).value = "Total Qty"
        wrksht.cell(row=footer_row, column=7).font = Font(size=9, bold=True, name="Lato")
        wrksht.cell(row=footer_row, column=7).alignment = alignment_left
        wrksht.cell(row=footer_row, column=7).border = top_border
        wrksht.cell(row=footer_row, column=8).border = top_border
        wrksht.cell(row=footer_row, column=9).alignment = alignment_right
        wrksht.cell(row=footer_row, column=9).font = table_font
        wrksht.cell(row=footer_row, column=9).value = int(total_quantity)
        wrksht.cell(row=footer_row, column=9).border = top_border
        wrksht.cell(row=footer_row, column=10).border = top_border
        wrksht.row_dimensions[footer_row].height = 20
        footer_row += 1

        bank_details_wb = workb.get_sheet_by_name('Bank Details')
        bank_sheet_row = 11

        bank_details_wb.cell(row=bank_sheet_row,column=3).value = invoice.name if invoice else self.name
        bank_details_wb.cell(row=bank_sheet_row+1,column=1).value = str("Invoice Date:" if invoice.invoice_date else '') if invoice else "Order Date:"
        bank_details_wb.cell(row=bank_sheet_row+1,column=3).value = str(str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y'))
        bank_details_wb.cell(row=bank_sheet_row+2,column=3).value = self.b2b_currency_id.name or ''
        bank_details_wb.cell(row=bank_sheet_row+3,column=3).value = '(' + self.b2b_currency_id.name + ') ' +"$ {:,.2f}".format(self.picked_qty_order_total)
        bank_details_wb.cell(row=bank_sheet_row+4,column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''
        
        fp = BytesIO()
        workb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        return base64.b64encode(data)

    def excel_abbreviate_report(self):

        for rec in self:
            # active_id = rec.id

            base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Abbreviate Report Sample.xlsx'
            wb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
            wrksht = wb.active

            f_name = 'Abbreviate-Inv-'+str(rec.name or '')  # FileName
            address_font = Font(name='Lato', size=9, bold=False)
            table_font = Font(name='Lato', size=9, bold=False)
            name_header_font = Font(name="Lato", size=9, bold=True)
            bank_detail_font = Font(name="Lato", size=7, bold=False)

            alignment = Alignment(horizontal='center', vertical='center', text_rotation=0)
            address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0)
            alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0)

            # sheet = workbook.create_sheet(title="excel "+str(rec.name).replace('/','-'), index=0)  # sheet name

            bd = Side(style='thin', color="d2d4d4")
            tp_bd = Side(style='thin', color="000000")
            all_border = Border(left=Side(style='thin', color="d2d4d4"), 
                            right=Side(style='thin', color="d2d4d4"), 
                            top=Side(style='thin', color="d2d4d4"), 
                            bottom=Side(style='thin', color="d2d4d4"))
            bottom_border = Border(bottom=bd)
            top_border = Border(top=tp_bd)

            # ------------------------------------------------------------
            # Billing Address abbreviate
            # ------------------------------------------------------------
            address_row = 9
            billing_address = self.create_address_line_for_sale(self.partner_invoice_id, take_name=True)
            wrksht.cell(row=address_row, column=1).value = billing_address
            wrksht.cell(row=address_row, column=1).alignment = address_alignment
            wrksht.cell(row=address_row, column=1).font = address_font
            # ------------------------------------------------------------
            # Shipping Address  abbreviate
            # ------------------------------------------------------------
            shipping_address = self.create_address_line_for_sale(self.partner_shipping_id, take_name=True)
            wrksht.cell(row=address_row, column=7).value = shipping_address
            wrksht.cell(row=address_row, column=7).alignment = address_alignment
            wrksht.cell(row=address_row, column=7).font = address_font
            # -------------------------- set border to addresss ----------------------------
            # ---------------------------------------------------------
            # name abbreviate
            # ---------------------------------------------------------
            name_row = address_row + 7
            wrksht.cell(row=name_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
            invoice = self.invoice_ids.filtered(lambda x: x.state != 'cancel')
            if len(invoice):
                invoice = invoice if len(invoice) else invoice[0]
            wrksht.cell(row=name_row, column=1).value = str("Invoice "+invoice.name if invoice else 'Invoice %s'%(self.name))
            # if invoice:
            wrksht.cell(row=name_row, column=1).font = Font(name='Lato', size=14, bold=False,color='666666')
            # else:
            #     wrksht.cell(row=name_row, column=1).font = Font(name='Lato', size=14, bold=False,color='666666')


            # ----------------------------------------------------------------
            # Date, Salesperson abbreviate
            # ----------------------------------------------------------------
            date_person_row = name_row + 3
            wrksht.cell(row=date_person_row-1, column=1).value = str("Invoice Date:" if invoice.invoice_date else 'Order Date')
            wrksht.cell(row=date_person_row, column=1).value = str(str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y'))
            # wrksht.cell(row=date_person_row, column=1).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

            wrksht.cell(row=date_person_row, column=2).value = self.name or ''
            # wrksht.cell(row=date_person_row, column=2).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=2).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

            wrksht.cell(row=date_person_row, column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''
            # wrksht.cell(row=date_person_row, column=9).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=9).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
            
            picking = self.picking_ids.filtered(lambda x: 'WH/OUT' in x.name and x.state == 'done')
            picking = picking if len(picking) == 1 else picking[0] if len(picking) else False
            wrksht.cell(row=date_person_row, column=5).value = str(picking.shipping_id.name if picking and picking.shipping_id else '') or ''
            # wrksht.cell(row=date_person_row, column=7).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=7).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

            wrksht.cell(row=date_person_row, column=7).value = str(picking.tracking_number_spt if picking and picking.tracking_number_spt else '') or ''
            # wrksht.cell(row=date_person_row, column=6).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=6).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
            
            wrksht.cell(row=date_person_row, column=9).value = str(self.no_of_cases) or ''
            # wrksht.cell(row=date_person_row, column=5).font = name_header_font
            # wrksht.cell(row=date_person_row, column=5).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)


            wrksht.cell(row=date_person_row, column=10).value = str(self.b2b_currency_id.name or '')
            # wrksht.cell(row=date_person_row, column=8).font = Font(name="Lato", size=10, bold=True)
            # wrksht.cell(row=date_person_row, column=8).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)

            # ========================= Product Table ===========================
            table_header = date_person_row+2

            query = f'''SELECT CONCAT(PP.DEFAULT_CODE,' ',PMS.NAME,' ',KPCC.NAME,' ',PSS.NAME),
                            PC.NAME,
                            PP.MATERIAL,
                            PP.HS_CODE,
                            SOL.PICKED_QTY,
                            SOL.UNIT_DISCOUNT_PRICE,
                            (SOL.UNIT_DISCOUNT_PRICE * SOL.PICKED_QTY) AS SUBTOTAL,
                            CASE
                                            WHEN SOL.IS_SHIPPING_PRODUCT = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS SHIPPING_CHARGE,
                            CASE
                                            WHEN SOL.IS_ADMIN = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS ADMIN_FEE,
                            CASE
                                            WHEN SOL.IS_GLOBAL_DISCOUNT = TRUE THEN SOL.UNIT_DISCOUNT_PRICE * SOL.PRODUCT_UOM_QTY
                                            ELSE 0
                            END AS ADDITIONAL_DISCOUNT,
                            CASE
                                            WHEN SOL.PRODUCT_UOM_QTY > 0
                                                                AND SOL.PRICE_TAX > 0 THEN SOL.PRICE_TAX / SOL.PRODUCT_UOM_QTY * SOL.PICKED_QTY
                                            ELSE 0
                            END AS PRICE_TAX,
                            SO.PICKED_QTY
                        FROM SALE_ORDER AS SO
                        INNER JOIN SALE_ORDER_LINE AS SOL ON SOL.ORDER_ID = SO.ID
                        INNER JOIN STOCK_PICKING AS SP ON SP.ORIGIN = SO.NAME
                        INNER JOIN SHIPPING_PROVIDER_SPT AS SPS ON SP.SHIPPING_ID = SPS.ID
                        INNER JOIN RES_CURRENCY AS RC ON RC.ID = SOL.CURRENCY_ID
                        INNER JOIN PRODUCT_PRODUCT AS PP ON PP.ID = SOL.PRODUCT_ID
                        INNER JOIN PRODUCT_CATEGORY AS PC ON PC.ID = PP.CATEG_ID
                        INNER JOIN PRODUCT_MODEL_SPT AS PMS ON PMS.ID = PP.MODEL
                        INNER JOIN KITS_PRODUCT_COLOR_CODE AS KPCC ON KPCC.ID = PP.COLOR_CODE
                        INNER JOIN PRODUCT_SIZE_SPT AS PSS ON PSS.ID = PP.EYE_SIZE
                        WHERE SO.ID ={self.id}'''
            self.env.cr.execute(query)
            record_data = self.env.cr.fetchall()
            row_index = table_header+1
            sub_total = 0
            total_quantity = 0
            shipping_cost = 0
            admin_fee = 0
            additional_discount = 0
            # discount_total = 0
            taxes = 0

            # for line in range(len(self.order_line)):
            #     line = self.order_line[line]
            #     taxes_total = 0
            #     taxes_total += round(line.price_tax/line.product_uom_qty,2) if line.product_uom_qty and line.price_tax else 0.0
            #     taxes += round(taxes_total * line.picked_qty, 2)
            #     if line.product_id.is_shipping_product:
            #         shipping_cost += line.unit_discount_price or 0
            #     elif line.product_id.is_admin:
            #         admin_fee += line.unit_discount_price or 0
            #     elif line.product_id.is_global_discount:
            #         additional_discount += line.unit_discount_price or 0
            #     else:
            #         if line.picked_qty > 0:
            #             # discount_total += round((line.price_unit - line.unit_discount_price) * line.picked_qty, 2)
            #             sub_total += (line.unit_discount_price*line.picked_qty)
            #             taxes_total = 0
            #             # for attribute in line.product_id.product_template_attribute_value_ids:
            #             #     if attribute.attribute_id.name == 'Color':
            #             #         color_name = attribute.product_attribute_value_id.name.split('-')[0]
            #             #     if attribute.attribute_id.name == 'Eye Size':
            #             #         eye_size = attribute.product_attribute_value_id.name
            #             color_name = line.product_id.color_code.name
            #             eye_size = line.product_id.eye_size.name

            #             product_name = line.product_id.default_code.split('-')[1]+" "+line.product_id.model.name+" "+color_name+" "+eye_size or ""
            #             if product_name in orders_list.keys():
            #                 orders_list[product_name][0].append(line)
            #             else:
            #                 orders_list[product_name] = [[line], product_name]

            for data in record_data:
                height = (2*len(data[0])) if len(data[0]) > 30 else 30
                wrksht.row_dimensions[row_index].height = height
                total_quantity += data[4]
                sub_total += data[5] * data[4]
                shipping_cost += data[7]
                admin_fee += data[8]
                additional_discount += data[9]
                taxes += data[10]
                if data[4] <= 0:
                    pass
                else: 
                    # product name
                    wrksht.merge_cells("A"+str(row_index)+':'+"D"+str(row_index))
                    wrksht.cell(row=row_index, column=1).value = data[0]
                    # HS Code
                    wrksht.cell(row=row_index, column=7).value = data[3]
                    # Material
                    wrksht.cell(row=row_index, column=6).value = data[2]
                    # category name
                    wrksht.cell(row=row_index, column=5).value = data[1]
                    # product qty
                    wrksht.cell(row=row_index, column=8).value = data[4]
                    # discount
                    # wrksht.cell(row=row_index, column=6).value = dict_data.discount
                    # price
                    wrksht.cell(row=row_index, column=9).value = "$ {:,.2f}".format(data[5])
                    # subtotal
                    wrksht.cell(row=row_index, column=10).value = "$ {:,.2f}".format(data[6])

                    wrksht.cell(row=row_index, column=1).font = table_font
                    wrksht.cell(row=row_index, column=5).font = table_font
                    wrksht.cell(row=row_index, column=6).font = table_font
                    wrksht.cell(row=row_index, column=7).font = table_font
                    wrksht.cell(row=row_index, column=8).font = table_font
                    wrksht.cell(row=row_index, column=9).font = table_font
                    wrksht.cell(row=row_index, column=10).font = table_font

                    wrksht.cell(row=row_index, column=1).border = bottom_border
                    wrksht.cell(row=row_index, column=2).border = bottom_border
                    wrksht.cell(row=row_index, column=3).border = bottom_border
                    wrksht.cell(row=row_index, column=4).border = bottom_border
                    wrksht.cell(row=row_index, column=5).border = bottom_border
                    wrksht.cell(row=row_index, column=6).border = bottom_border
                    wrksht.cell(row=row_index, column=7).border = bottom_border
                    wrksht.cell(row=row_index, column=8).border = bottom_border
                    wrksht.cell(row=row_index, column=9).border = bottom_border
                    wrksht.cell(row=row_index, column=10).border = bottom_border

                    wrksht.cell(row=row_index, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
                    wrksht.cell(row=row_index, column=5).alignment = alignment
                    wrksht.cell(row=row_index, column=6).alignment = alignment
                    wrksht.cell(row=row_index, column=7).alignment = alignment
                    wrksht.cell(row=row_index, column=8).alignment = alignment
                    wrksht.cell(row=row_index, column=9).alignment = alignment_right
                    wrksht.cell(row=row_index, column=10).alignment = alignment_right
                    
                    wrksht.row_dimensions[row_index].height = 20

                    row_index += 1
            # ======================== Product table end ========================

            # ----------------------------------------------------
            # Above table total Quantity Abbreviate
            # ----------------------------------------------------
            wrksht.cell(row=date_person_row, column=8).value = str(int(total_quantity))
            # wrksht.cell(row=date_person_row, column=4).font = name_header_font
            # wrksht.cell(row=date_person_row, column=4).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
            footer_row = row_index+1

            # ===================== Bank Details =========================
            # if invoice and invoice.get_html_field_val(invoice.company_id.bank_details):
            #     wrksht.merge_cells("A"+str(footer_row)+":E"+str(footer_row))
            #     wrksht.cell(row=footer_row, column=1).value = "Bank Transfer Details"
            #     wrksht.cell(row=footer_row, column=1).font = Font(name='Lato', size=10, bold=True)
            #     wrksht.cell(row=footer_row, column=1).alignment = alignment_left
            #     for i in range(1,6):
            #         wrksht.cell(row=footer_row, column=i).border = Border(left=Side(style='thin', color="d2d4d4"),
            #                                                           right=Side(style='thin', color="d2d4d4"), 
            #                                                           top=Side(style='thin', color="d2d4d4"))

            #     wrksht.merge_cells("A"+str(footer_row+1)+":E"+str(footer_row+6))
            #     wrksht.row_dimensions[footer_row+6].height = 67
            #     bank_details = BeautifulSoup(invoice.company_id.bank_details,"html.parser")
            #     wrksht.cell(row=footer_row+1, column=1).value = self.bank_details()
            #     # wrksht.cell(row=footer_row+1, column=1).value = '\n'.join([i.strip() for i in bank_details.get_text().split('\n') if len(i.strip()) > 0])
            #     wrksht.cell(row=footer_row+1, column=1).font = bank_detail_font
            #     wrksht.cell(row=footer_row+1, column=1).alignment = address_alignment
            #     bank_details_row = footer_row+1
            #     for row in range(footer_row+1,footer_row+7):
            #         for col in range(1,6):
            #             wrksht.cell(row=row, column=col).border = Border(left=Side(style='thin', color="d2d4d4"),
            #                                                           right=Side(style='thin', color="d2d4d4"), 
            #                                                           bottom=Side(style='thin', color="d2d4d4"))
            #         bank_details_row += 1

            # ===================== Footer right =========================
            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Subtotal"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(sub_total)
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Shipping"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(shipping_cost)  # shipping
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Admin Fee"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(admin_fee)  # adminfee
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1


            if abs(additional_discount):
                wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
                wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

                wrksht.cell(row=footer_row, column=7).value = "Discount"
                wrksht.cell(row=footer_row, column=7).alignment = alignment_left
                wrksht.cell(row=footer_row, column=7).border = top_border
                wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
                wrksht.cell(row=footer_row, column=8).border = top_border
                wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(abs(additional_discount))  # discont
                wrksht.cell(row=footer_row, column=9).alignment = alignment_right
                wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
                wrksht.cell(row=footer_row, column=9).border = top_border
                wrksht.cell(row=footer_row, column=10).border = top_border
                wrksht.row_dimensions[footer_row].height = 20
                footer_row += 1

            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Tax"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "$ {:,.2f}".format(taxes)  # taxes
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Total"
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = "({}) $ {:,.2f}".format(self.b2b_currency_id.name,sub_total+shipping_cost+admin_fee - abs(additional_discount)+taxes)  # total
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

            wrksht.merge_cells("G"+str(footer_row)+":H"+str(footer_row))
            wrksht.merge_cells("I"+str(footer_row)+":J"+str(footer_row))

            wrksht.cell(row=footer_row, column=7).value = "Total Quantity"
            wrksht.cell(row=footer_row, column=7).font = Font(name="Lato", size=9, bold=True)
            wrksht.cell(row=footer_row, column=7).alignment = alignment_left
            wrksht.cell(row=footer_row, column=7).border = top_border
            wrksht.cell(row=footer_row, column=8).border = top_border
            wrksht.cell(row=footer_row, column=9).value = int(total_quantity)  # TotalQuantity
            wrksht.cell(row=footer_row, column=9).alignment = alignment_right
            wrksht.cell(row=footer_row, column=9).font = Font(name="Lato", size=9, bold=False)
            wrksht.cell(row=footer_row, column=9).border = top_border
            wrksht.cell(row=footer_row, column=10).border = top_border
            wrksht.row_dimensions[footer_row].height = 20
            footer_row += 1

            bank_details_wb = wb.get_sheet_by_name('Bank Details')
            bank_sheet_row = 11

            bank_details_wb.cell(row=bank_sheet_row,column=3).value = invoice.name if invoice  else self.name
            bank_details_wb.cell(row=bank_sheet_row+1,column=1).value = str("Invoice Date:" if invoice.invoice_date else '') if invoice else "Order Date:"
            bank_details_wb.cell(row=bank_sheet_row+1,column=3).value = str(str(invoice.invoice_date.strftime('%d/%m/%Y') if invoice.invoice_date else '') if invoice else self.date_order.strftime('%d/%m/%Y'))
            bank_details_wb.cell(row=bank_sheet_row+2,column=3).value = self.b2b_currency_id.name or ''
            bank_details_wb.cell(row=bank_sheet_row+3,column=3).value = "({}) $ {:,.2f}".format(self.b2b_currency_id.name,taxes + sub_total+shipping_cost+admin_fee-abs(additional_discount))
            bank_details_wb.cell(row=bank_sheet_row+4,column=3).value = self.payment_term_id.name or 'Immediate Payment' or ''

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})

            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, f_name),
                'target': 'self',
            }


    def open_import_order_line_wizard(self):
        self.ensure_one()
        return {
            "name":_("Import Items"),
            "type":"ir.actions.act_window",
            "res_model":"import.order.lines.wizard.spt",
            "view_mode":"form",
            "context":{"default_so_id":self.id,},
            "target":"new",
        }

    # @api.onchange('user_id')
    # def _onchange_user_id(self):
    #     for record in self:
    #         manager = self.env['res.users'].search([('allow_user_ids','in',record.user_id.ids)],limit=1)
    #         if manager and record.user_id:
    #             record.sale_manager_id = manager.id

    @api.depends('order_line.invoice_lines')
    def _get_invoiced(self):
        # The invoice_ids are obtained thanks to the invoice lines of the SO
        # lines, and we also search for possible refunds created directly from
        # existing invoices. This is necessary since such a refund is not
        # directly linked to the SO.
        for order in self:
            invoices = order.sudo().order_line.invoice_lines.move_id.filtered(lambda r: r.move_type in ('out_invoice', 'out_refund'))
            order.invoice_ids = invoices
            order.invoice_count = len(invoices.filtered(lambda x: x.state != 'cancel'))

    def create_address_line_for_sale(self, source_id, take_name=False):
        # catalog_obj = self.env['sale.catalog']
        # catalog_obj.connect_server()
        # method = catalog_obj.get_method('create_address_line_for_sale')
        # if method['method']:
        #     localdict = {'self': self,'source_id':source_id,'take_name':take_name}
        #     exec(method['method'], localdict)
        # Param source_id : Partner record.
        address = ''
        if take_name == True:
            if source_id.name:
                address += str(source_id.name)
            if source_id.street:
                if source_id.name:
                    address += '\n'+str(source_id.street)
                else:
                    address += source_id.street
        else:
            if source_id.street:
                address += str(source_id.street) 
        if source_id.street2:
            address += '\n'+str(source_id.street2)
        if source_id.city:
            address+= '\n'+str(source_id.city)
        if source_id.zip and take_name:
            address += ', '+source_id.zip
        if source_id.state_id:
            if take_name:
                address += '\n'+str(source_id.state_id.name)
            else:
                address += ' '+str(source_id.state_id.name)
        if source_id.country_id:
            if take_name:
                if source_id.state_id:
                    address += ', '+str(source_id.country_id.name)
                else:
                    address += '\n'+str(source_id.country_id.name)
            else:
                address += ' '+str(source_id.country_id.name)
                
        if source_id.zip and not take_name:
            address += ' '+source_id.zip
        address += '\nTel. '
        if source_id.phone:
            address += source_id.phone
        address += '\nEmail. '
        if source_id.email:
            address += source_id.email 
        # return localdict['address']
        return address

    def partner_verification(self):
        verified = False
        if self.partner_id and self.partner_id.user_ids:
            if self.partner_id.user_ids[0].has_group('base.group_user') or self.partner_id.customer_type == 'b2b_regular':
                if self.partner_id.mail_notification:
                    verified = True
        return verified

    def action_order_xls_report(self):
        return {
            "name":_("Quotation Excel Report"),
            "type":"ir.actions.act_window",
            "res_model":"order.report.with.image.wizard",
            "view_mode":"form",
            "context":{"default_order_id":self.id,},
            "target":"new",
        }

    @api.depends('partner_id.country_id','partner_id')
    def _compute_country_id(self):
        for record in self:
            record.country_id = record.partner_id.country_id.id

    def _get_product_avail_qty(self):
        for rec in self:
            rec.is_not_product_aval_qty_flag = False
            if rec.state in ('draft','sent'):
                for line in self.order_line.filtered(lambda x: not x.product_id.is_admin and not x.product_id.is_shipping_product and not x.product_id.is_global_discount):
                    if line.product_uom_qty != 0.0 and line.product_uom_qty > line.product_id.available_qty_spt:
                        rec.is_not_product_aval_qty_flag = True
                        break

    @api.depends('order_line','order_line.product_id.virtual_available','order_line.product_uom_qty','website_id','state')
    def _get_compute_message(self):
        backup_order_obj = self.env['sale.order.backup.spt']
        self._get_product_avail_qty()
        for rec in self:
            rec.count_backup_order = int(len(backup_order_obj.search([('order_id','=',rec.id)],limit=1)))
            rec.message = False
            if rec.website_id and rec.state == 'draft':
                for line in range(len(rec.order_line)):
                    line = rec.order_line[line]
                    if line.product_uom_qty > line.product_id.available_qty_spt:
                        rec.message = "This order comes from an abandoned cart. The quantities may not be correct. Please check the red icons in the order lines and verify the quantities with the warehouse."
                        break

    @api.model
    def create(self, vals):
        res = super(sale_order, self).create(vals)
        for record in range(len(res)):
            record = res[record]
            record.no_of_cases = record.ordered_qty
            if record.partner_id and record.partner_id.signup_from_website and record.partner_id.customer_type == 'b2c':
                raise UserError(_('You cannot create order without Customer verification.'))
            # PDF Links
            pdf_links = self.env['ir.model'].sudo().generate_report_access_link(
                'sale.order',
                record.id,
                'sale.action_report_saleorder',
                record.partner_id.id,
                'pdf'
            )
            url = ''
            if pdf_links.get('success') and pdf_links.get('url'):
                url = pdf_links.get('url')

            if record.state in ['draft','sent'] and record.catalog_id and not record.website_id:
                mail_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_sales_person_sale_order_confirm_manully_spt').sudo()
                mail_template_id.with_context(pdf_url=url).send_mail(res_id=record.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")
                if record.partner_id:
                    verified = record.partner_verification()
                    quotation_template_id = self.env.ref('sale.email_template_edi_sale')
                    quotation_template_id.with_context(proforma=False,pdf_url=url).send_mail(record.id,force_send=True,email_layout_xmlid="mail.mail_notification_light") if verified else None
            if record.state == 'draft' and not record.catalog_id and record.website_id:
                mail_template_id = self.env.ref('tzc_sales_customization_spt.tzc_start_adding_into_cart_notification_to_salesperson_spt').sudo()
                recipients = record.user_id.partner_id.ids if record.user_id and record.user_id.partner_id else []
                mail_template_id.with_context.get(pdf_url=url).send_mail(res_id=record.id,force_send=True,email_values={'recipient_ids':[(6,0,recipients)]},email_layout_xmlid="mail.mail_notification_light")
        return res

    #kits_abadon_card_order
    def _get_reson_message(self):
        abandon_time = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0'))
        for rec in self:
            rec.abandoned_reason = False
            if rec.state != 'draft':
                rec.abandoned_reason = 'Quotation is confirmed.'
            elif rec.create_date and rec.create_date >= fields.Datetime.now() - timedelta(hours=abandon_time):
                rec.abandoned_reason = 'The quotation is less than %sh old.'%(int(abandon_time))
            elif rec.next_execution_date and rec.next_execution_date >= fields.Datetime.now():
                rec.abandoned_reason = 'Email is sent within %s past days.'%(int(self.env['ir.config_parameter'].sudo().get_param('kits_abandon_cart_order.kits_abandone_mail_delay','0')))

    def count_days(self):
        delta = date.today() - self.create_date.date()
        return delta.days

    def cron_send_abondand_order(self):
        
        now = fields.Datetime.now()
        abandon_time = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0'))
        updated_time = now - timedelta(hours=abandon_time)

        order_ids = self.env['sale.order'].search([('state','in',['draft']),('create_date','<=',updated_time)])
        order_ids.with_context(cron=True).recovery_mail_salesperson()
        order_ids.with_context(cron=True).recovery_mail_customer()

    def recovery_mail_salesperson(self):
        salesperson_tmp_id = self.env.ref('tzc_sales_customization_spt.mail_template_salesperson_abandone_cart')
        mail_delay_date = fields.Datetime.now() + timedelta(days=int(self.env['ir.config_parameter'].sudo().get_param('kits_abandon_cart_order.kits_abandone_mail_delay','0')))
        order_ids = self

        if self.env.context.get('cron') and self.env.context.get('cron') == True:

            order_ids = self.order_filter().filtered(lambda x: not x.next_execution_date or (x.next_execution_date and x.next_execution_date <= fields.Datetime.now()))

        for rec in order_ids:
            if self._context.get('active_id'):
                self._context['active_id'] = rec.id
            salesperson_tmp_id.send_mail(rec.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")
            rec.write({"next_execution_date":mail_delay_date})

    def recovery_mail_customer(self):
        customer_tmp_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_cart_recovery_spt')
        mail_delay_date = fields.Datetime.now() + timedelta(days=int(self.env['ir.config_parameter'].sudo().get_param('kits_abandon_cart_order.kits_abandone_mail_delay','0')))
        abandon_time = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0'))

        order_ids = self.filtered(lambda x: not x.next_execution_date or (x.next_execution_date and x.next_execution_date <= fields.Datetime.now()))

        if self._context.get('active_ids') and len(self._context.get('active_ids')) <= 1 and not self._context.get('active_domain'):
            view_type = self._fields_view_get()
            if view_type.get('type') and view_type.get('type') == 'form' and not self.env.context.get('cron'):
                if self.state != 'draft':
                    raise UserError('Quotation is confirmed.')
                elif self.create_date and self.create_date >= fields.Datetime.now() - timedelta(hours=abandon_time):
                    raise UserError('The quotation is less than %sh old.'%(int(abandon_time)))
                elif self.next_execution_date and self.next_execution_date >= fields.Datetime.now():
                    raise UserError('Email is sent within %s past days.'%(int(self.env['ir.config_parameter'].sudo().get_param('kits_abandon_cart_order.kits_abandone_mail_delay','0'))))
                else:
                    lang = self.env.context.get('lang')
                    ctx = {
                        'default_model': 'sale.order',
                        'default_res_id': self.ids[0],
                        'default_use_template': bool(customer_tmp_id),
                        'default_template_id': customer_tmp_id.id,
                        'default_composition_mode': 'comment',
                        'mark_so_as_sent': False,
                        'custom_layout': "mail.mail_notification_light",
                        'force_email': True,
                        'default_no_auto_thread':False,
                        'model_description': self.with_context(lang=lang).type_name,
                        'cart_recovery':True,
                        'next_execution_date':mail_delay_date,
                    }
                    return {
                        'type': 'ir.actions.act_window',
                        'view_mode': 'form',
                        'res_model': 'mail.compose.message',
                        'views': [(False, 'form')],
                        'view_id': False,
                        'target': 'new',
                        'context': ctx,
                    }
        else:
            if self.env.context.get('cron') and self.env.context.get('cron') == True:
                order_ids = self.order_filter().filtered(lambda x: not x.next_execution_date or (x.next_execution_date and x.next_execution_date <= fields.Datetime.now()))
            
            order_ids = order_ids.filtered(lambda x:x.state == 'draft')
            order_ids = order_ids.filtered(lambda x:x.create_date and x.create_date <= datetime.now() - timedelta(hours=abandon_time))
           
            invalid_order_id = self - order_ids

            return {
                'name': 'Warning Message',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'warning.message.wizard',
                'target': 'new',
                'context': {
                            'default_invalid_orders': [(6,0,invalid_order_id.ids)],
                            'default_valid_orders': [(6,0,order_ids.ids)],
                            'mail_template_id':customer_tmp_id.id,
                            'next_date':mail_delay_date}
            }
        
    def recovery_mail_warehouse(self):
        tmp_id = self.env.ref('tzc_sales_customization_spt.mail_template_for_warehouse_abandone_cart')

        user_id = self.env['res.users'].search([('is_warehouse','=',True)],limit=1)
        for rec in self:
            tmp_id = tmp_id.with_context(name=user_id.name,active_id=rec.id)
            tmp_id.email_to = user_id.email
            tmp_id.send_mail(rec.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")

    def order_filter(self):
        now = fields.Datetime.now()
        abandon_time = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0'))
        updated_time = now - timedelta(hours=abandon_time)

        order_ids = self.filtered(lambda x: x.state in ('draft') and x.create_date <= updated_time)

        return order_ids



     # kits_bambora_payment
    @api.depends('payment_ids','due_amount','amount_paid','payment_status')
    def _get_amount_paid(self):
        for rec in self:
            rec.amount_paid = sum(self.env['order.payment'].sudo().search([('order_id','=',rec.id),('state','=','approve')]).mapped('amount')) or 0.0
            # rec._compute_payment_status()

    @api.depends('picked_qty_order_total','amount_paid','is_paid')
    def _compute_amount_due(self):
        for rec in self:
            # rec.due_amount = 0.0
            # if rec.amount_paid:
            rec.due_amount = rec.picked_qty_order_total - rec.amount_paid
            # rec._compute_payment_status()
    
    @api.depends('due_amount','amount_paid','picked_qty_order_total','payment_ids')
    def _compute_payment_status(self):
        for rec in self:
            rec.payment_status = False
            if rec.amount_paid:
                if round(rec.amount_paid,2) < round(rec.picked_qty_order_total,2):
                    rec.payment_status = 'partial'
                elif round(rec.amount_paid,2) == round(rec.picked_qty_order_total,2):
                    rec.payment_status = 'full'
                elif round(rec.amount_paid,2) > round(rec.picked_qty_order_total,2):
                    rec.payment_status = 'over'

    @api.depends('payment_link')
    def _compute_is_payment_link(self):
        for rec in self:
            if rec.payment_link:
                rec.is_payment_link = True
            else:
                rec.is_payment_link = False

    def action_order_payment_approval(self):
        user_id = False
        for rec in self:
            if self.env.user.has_group('base.group_system'):
                rec.approve_by_admin = True
                user_id = self.env.user.id
            elif self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt'):
                rec.approve_by_salesmanager = True
                user_id = self.env.user.id
            elif self.env.user.has_group('sales_team.group_sale_salesman'):
                rec.approve_by_salesperson = True
                user_id = self.env.user.id
            rec.payment_link_approved_by = user_id

    def action_generate_payment_link(self):
        for rec in self:
            if rec.payment_link:
                regenerate_message = "Payment link is already generated.\nAre you sure you want to regenerate new payment link?"
                return {
                    'name':_('Confirmation Wizard'),
                    'type':'ir.actions.act_window',
                    'res_model':'kits.message.wizard',
                    'view_mode':'form',
                    'context':{'default_kits_message':regenerate_message,'default_order_id':rec.id},
                    'target':'new',
                }
            else:
                if rec.approve_by_admin or (rec.approve_by_admin and (rec.approve_by_salesperson or rec.approve_by_salesmanager)):
                    link = rec.generate_link(rec)
                    if link:
                        rec.payment_link = link
                        return{
                            'name':_('Generate a Payment Link'),
                            'type':'ir.actions.act_window',
                            'res_model':'kits.generate.payment.link.wizard',
                            'view_mode':'form',
                            'context':{'default_kits_payment_link':link,'default_sale_order_id':rec.id},
                            'target':'new',
                        }
                    else:
                        raise UserError('Something went wrong.. Payment link can\'t generate.\n\n- Please check bambora account details.')
                else:
                    raise UserError(f'Order {rec.name} is not approved for payment, Please contact your administrator.')

    def generate_link(self,order):
        company_id = order.company_id if order.company_id else False
        url = ''
        merchand_id,hash_value = '',''

        if company_id and company_id.account_type == 'sand_box':
            merchand_id = company_id.sand_box_merchant_id_usd if order.b2b_currency_id.name == 'USD' else company_id.sand_box_merchant_id_cad
            hash_value = company_id.sand_box_hash_value_usd if order.b2b_currency_id.name == 'USD' else company_id.sand_box_hash_value_cad
        elif company_id and company_id.account_type == 'production':
            merchand_id = company_id.production_merchant_id_usd if order.b2b_currency_id.name == 'USD' else company_id.production_merchant_id_cad
            hash_value = company_id.production_hash_value_usd if order.b2b_currency_id.name == 'USD' else company_id.production_hash_value_cad

        if not merchand_id or not hash_value:
            raise UserError('Please add missing payment provider detail in company.\n\n- Merchant ID & Hash-Value.')
        else:
            hash = hashlib.sha1(str('merchant_id=').encode('utf-8') + str(merchand_id).encode('utf-8') + str(hash_value).encode('utf-8')).hexdigest()
            url += "https://web.na.bambora.com/scripts/payment/payment.asp?merchant_id=%s&hashValue=%s&trnType=%s"%(merchand_id,hash,'P')

        if url:
            if order:
                url += "&trnOrderNumber=%s"%(order.name)
            if order.picked_qty_order_total:
                if not order.amount_paid:
                    url += "&trnAmount=%s"%(round(order.picked_qty_order_total,2))
                else:
                    if order.due_amount and order.due_amount > 0.0:
                        url += "&trnAmount=%s"%(round(order.due_amount,2))
                    else:
                        raise UserError('You can not generate a payment link because you have no remaining amount.')    
            if order.partner_id:
                name = re.sub('[@_!#$%^&*()<>?/\|}{~:;]',' ',order.partner_id.name)
                url += "&ordName=%s"%(name)
            if order.partner_id.email:
                url += "&ordEmailAddress=%s"%(order.partner_id.email)
            if order.partner_id.phone:
                url += "&ordPhoneNumber=%s"%(order.partner_id.phone)
            if order.partner_id.street:
                street_1 = order.partner_id.street
                if '%' in order.partner_id.street or '?' in order.partner_id.street or '+' in order.partner_id.street:
                    street_1 = re.sub('[%?+]','', order.partner_id.street)
                if '#' in street_1:
                    street_1 = street_1.replace('#','No. ')
                if '&' in street_1:
                    street_1 = street_1.replace('&','And ')
                url += "&ordAddress1=%s"%(street_1) if street_1 else None
            if order.partner_id.street2:
                street_2 = order.partner_id.street2
                if '%' in order.partner_id.street2 or '?' in order.partner_id.street2 or '+' in order.partner_id.street2:
                    street_2 = re.sub('[%?+]','', order.partner_id.street2)
                if '#' in street_2:
                    street_2 = street_2.replace('#','No. ')
                if '&' in street_2:
                    street_2 = street_2.replace('&','And ')
                url += "&ordAddress2=%s"%(street_2) if street_2 else None
            if order.partner_id.city:
                city = order.partner_id.city
                if '%' in order.partner_id.city or '?' in order.partner_id.city or '+' in order.partner_id.city:
                    city = re.sub('[%?+]','', order.partner_id.city)
                if '#' in city:
                    city = city.replace('#','No. ')
                if '&' in city:
                    city = city.replace('&','And ')
                url += "&ordCity=%s"%(city) if city else None
            if order.partner_id.zip:
                zip = order.partner_id.zip
                if '%' in order.partner_id.zip or '?' in order.partner_id.zip or '+' in order.partner_id.zip:
                    zip = re.sub('[%?+]','', order.partner_id.zip)
                if '#' in zip:
                    zip = zip.replace('#','No. ')
                if '&' in zip:
                    zip = zip.replace('&','And ')
                url += "&ordPostalCode=%s"%(zip) if zip else None
            if order.partner_id.state_id:
                url += "&ordProvince=%s"%(order.partner_id.state_id.code)
            if order.partner_id.country_id:
                url += "&ordCountry=%s"%(order.partner_id.country_id.code)

        return url

    def send_payment_link_mail(self):
        if self.payment_link:
            mail_context = {
                'default_model': 'sale.order',
                'default_res_id': self.ids[0],
                'default_template_id': self.env.ref('tzc_sales_customization_spt.mail_template_customer_send_payment_link').id,
                'default_partner_ids': self.partner_id.ids,
                'default_use_template':True,
                'default_composition_mode': 'comment',
                'mark_so_as_sent': True,
                'force_email': True,
                'default_email_layout_xmlid':'mail.mail_notification_light'
            }
            return {
                'name': _('Send Mail'),
                'type': 'ir.actions.act_window',
                'res_model': 'mail.compose.message',
                'view_mode' : 'form',
                'target': 'new',
                'context':mail_context
            }
        else:
            raise UserError('There is no payment link generated for this order.')

    def mail_subject(self,order):
        subject = 'Payment request for your order (%s%s)'%(order.name,'/'+order.invoice_ids.name if order.invoice_ids else '')
        return  subject

    def action_is_paid(self):
        if self.state in ('scanned','scan','shipped','draft_inv','open_inv'):
            return {
                    'name': _('Paid Amount'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'paid.amount.wizard',
                    'view_mode' : 'form',
                    'target': 'new',
                    'context':{
                        'default_order_id':self.id,
                        'default_date':datetime.now()
                    }
                }
        else:
            raise UserError('You can not mark this order as paid before scanning completed.')

    def action_is_unpaid(self):
        for rec in self:
            if rec.state in ('scanned','scan','shipped','draft_inv','open_inv'):
                cancel_order_ids = rec.filtered(lambda x:x.state == 'cancel')
                invoice_id = rec.invoice_ids.filtered(lambda invoice:invoice.state not in ('cancel','draft'))
                if cancel_order_ids:
                    raise UserError('You can\'t unpay canceled orders.')
                else:
                    rec.is_paid = False
                    # invoice_id.write({'is_commission_paid':False})
                rec.mark_as_paid_by_user = False
                rec.paid_amount = 0.0
            else:
                raise UserError('You can not mark this order as unpaid before scanning completed.')

    def action_revert_to_scanned(self):
        for rec in self:
            rec.picking_ids.filtered(lambda x:x.state != 'cancel').write({'state':'scanned'})
            rec.order_approved_by = False
            rec.state = 'scanned'


    def _force_lines_to_invoice_policy_order(self):
        for line in self.order_line:
            if self.state in  ['sale', 'done','in_scanning','scanned','scan','shipped','draft_inv','open_inv']:
                line.qty_to_invoice = line.product_uom_qty - line.qty_invoiced
            else:
                line.qty_to_invoice = 0

    @api.model
    def get_payment_data(self,payment_id):
        payment_ids = self.env['sale.order'].browse(payment_id).payment_ids
        payment_list=[]
        if payment_ids:
            for payment in payment_ids.sorted(lambda x:x.create_date if x.create_date else datetime.now() - relativedelta(years=1000)):
                payment_dict={}
                payment_dict.update({'payments':{'create_date':payment.create_date if payment.create_date else False,'amount':payment.amount,'state':payment.state,'currency_icon':payment.order_id.b2b_currency_id.symbol,'is_manual_paid':payment.is_manual_paid}})
                payment_list.append(payment_dict)
        return payment_list


    def action_update_customer(self):
        for record in self:
            if record.partner_id:
                invoice_ids = record.invoice_ids.filtered(lambda invoice: invoice.state != 'cancel')
                picking_ids = record.picking_ids.filtered(lambda picking: picking.state in ['assigned','done'])
                if picking_ids:
                    raise UserError(_("For changing customer you have to cancel delivery first."))
                elif invoice_ids:
                    raise UserError(_("For changing customer you have to cancel invoice first."))
                else:

                    return {
                        'name': record.name,
                        'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'update.partner.in.order.wizard.spt',
                        'target': 'new',
                        'context': {'default_sale_id': record.id,'default_partner_id':record.partner_id.id }
                    }

    #kits_package_product
    package_order = fields.Boolean('Pack Order',compute="_compute_package_order",store=True,compute_sudo=True)
    # package_order = fields.Boolean('Package Order')
    package_order_status = fields.Selection([('available','Available'),('out_of_stock','Out of stock')],compute="_compute_package_order_status")
    package_order_lines =fields.One2many('kits.package.order.line','order_id','Package Order Line')

    @api.depends('package_order_lines','count_backup_order','package_order_lines.order_id','package_order_lines.backup_order')
    def _compute_package_order(self):
        backup_Obj = self.env['sale.order.backup.spt']
        for record in self:
            package_order = True if len(record.package_order_lines) > 0 else False
            if record.count_backup_order:
                backup_orders = backup_Obj.search([('order_id','=',record.id)],order="id desc",limit=1)
                package_order = backup_orders.pack_order_backup
            record.package_order = package_order
    
    def action_draft(self):
        if self.state != 'cancel':
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else:
            return super(sale_order,self).action_draft()

    def action_revert_order_to_quotation(self):
        if self.state not in ['sent','received','sale','in_scanning','scanned','scan']:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else:
            order_id = self.env['sale.order.backup.spt'].search([('order_id','=',self.id)],limit=1,order="id desc")
            if self.state not in ['sent','received','sale','in_scanning','scanned','scan']:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                            'title': 'Something is wrong.',
                            'message': 'Please reload your screen.',
                            'sticky': True,
                        }
                    }
            else:
                self.sale_order_cancel_spt()
                self.action_draft()
                order_id = self.env['sale.order.backup.spt'].search([('order_id','=',self.id)],limit=1,order="id desc")
                for order_line in self.order_line:
                    if order_line.product_id and order_line.product_id.id in order_id.line_ids.product_id.ids:
                        line=order_id.line_ids.search([('product_id','=',order_line.product_id.id),('order_backup_id','=',order_id.id)])
                        order_line.write({
                            'product_uom_qty':line.product_uom_qty,
                            'price_unit':line.price_unit,
                            'unit_discount_price':line.unit_discount_price,
                            'fix_discount_price':line.fix_discount_price,
                            'discount':line.discount,
                            'tax_id':line.tax_id,
                            'is_global_discount':line.is_global_discount,
                            'is_fs':line.is_fs,
                            'is_admin':line.is_admin,
                            'is_shipping_product':line.is_shipping_product,
                            'is_promotion_applied':line.is_promotion_applied,
                            'sale_type':line.sale_type,
                        })
                    else:
                        order_line.unlink()
            # to revert pack lines
            for pack_line in order_id.backup_package_lines:
                line = self.package_order_lines.filtered(lambda pol: pol.product_id == pack_line.product_id)
                vals={
                    'order_id':self.id,
                    'product_id':pack_line.product_id.id,
                    'qty':pack_line.qty,
                    'sale_price':pack_line.sale_price,
                    'discount_amount':pack_line.discount_amount,
                    'pack_price':pack_line.pack_price,
                }
                if line:
                    line.write(vals)
                else:
                    self.env['kits.package.order.line'].create(vals)
        
    @api.depends('package_order_lines')
    def _compute_package_order_status(self):
        for record in self:
            if all([each.availability == 'available' for each in record.package_order_lines]):
                record.package_order_status = 'available'
            else:
                record.package_order_status = 'out_of_stock'

    def check_stock_spt(self):
        sale_order_line_obj = self.env['sale.order.line']
        for record in self:
            warning_message = ""
            for product in record.order_line.mapped('product_id'):
                total_order_qty = 0
                for line in sale_order_line_obj.search([('order_id','=',record.id),('product_id.type','!=','service'),('product_id.is_global_discount','=',False),('product_id.is_shipping_product','=',False),('product_id.is_admin','=',False),('product_id','=',product.id)]):
                    total_order_qty = total_order_qty + line.product_uom_qty
                if total_order_qty > 0.0 and (total_order_qty > product.available_qty_spt):
                    warning_message += "Product %s having only %s quantity in stock,You can not add %s quantity.\n" % (product.display_name,int(product.available_qty_spt),int(total_order_qty))
                
            if warning_message:
                raise UserError(_(warning_message))

    def action_confirm(self):
        sol_obj = self.env['sale.order.line']
        order_backup_obj = self.env['sale.order.backup.spt']
        state_list = self.mapped(lambda so : so.state in ['draft','sent','received'])
        if any(state_list):
            for record in self:
                record._get_unavailable_package_ids()
                restricted_package_lines = record.package_order_lines.filtered(lambda x: x.availability == 'out_of_stock')
                pack_sale_lines = []
                if restricted_package_lines and not self.env.context.get('package_allow'):
                    return {
                        'name':_("Unavailable Package"),
                        'type':'ir.actions.act_window',
                        'res_model':'kits.package.restriction',
                        'view_mode':"form",
                        'context':{'default_restricted_package_ids':[(6,0,restricted_package_lines.mapped('product_id').ids)],'default_order_id':record.id},
                        'target':'new',
                    }
                geo_restriction_list = []
                backup_order_line_list =[]
                # Unpack Packed Product Lines
                for pack_line in (record.package_order_lines-restricted_package_lines) if not self._context.get('package_allow') else record.package_order_lines:
                    for product_line in pack_line.product_id.product_line_ids:
                        pack_sale_line = sol_obj.search([('order_id','=',record.id),('product_id','=',product_line.product_id.id),('package_id','=',pack_line.product_id.id)])
                        if not pack_sale_line:
                            pack_sale_line = sol_obj.create({
                                'order_id':record.id or False,
                                'product_id':product_line.product_id.id or False,
                                'product_uom_qty': product_line.qty * pack_line.qty or 0.0,
                                'package_id': pack_line.product_id.id or False,
                                'is_pack_order_line': True,
                                'package_line_id': pack_line.id or False,
                                'price_unit': record.partner_id.b2b_pricelist_id.get_product_price(product_line.product_id,product_line.qty * pack_line.qty,record.partner_id),
                                'sale_type':product_line.product_id.sale_type,
                            })
                            if record.partner_id.b2b_pricelist_id.currency_id.name == 'CAD':
                                pack_sale_line.write({'unit_discount_price':product_line.cad_price or 0.00})
                            else:
                                pack_sale_line.write({'unit_discount_price':product_line.usd_price or 0.00})
                        pack_sale_line._onchange_unit_discounted_price_spt()
                        pack_sale_lines.append(pack_sale_line.id) if pack_sale_line.id and pack_sale_line.id not in pack_sale_lines else None
                # backup order line for Packed Product Lines
                for line in record.order_line.filtered(lambda x: not x.is_pack_order_line or not x.package_id):
                    backup_order_line_list.append((0,0,{
                        'product_id':line.product_id.id,
                        'product_uom_qty':line.product_uom_qty,
                        'price_unit':line.price_unit,
                        'unit_discount_price':line.unit_discount_price,
                        'fix_discount_price':line.fix_discount_price,
                        'discount':line.discount,
                        'name':line.name,
                        'sale_type':line.sale_type,
                        'tax_id':[(6,0,line.tax_id.ids)],
                        'is_pack_order_line':bool(len(line.package_id)),
                        'package_id':line.package_id.id or False,
                    }))
                    if line.product_id.type == 'product' and record.partner_id.country_id.id in line.product_id.geo_restriction.ids:
                        geo_restriction_list.append(line.id)
                order_lines = self.env['sale.order.line'].browse(geo_restriction_list)
                products=[]
                if not record._context.get('on_consign_wizard') and not record._context.get('allow_restricted'):
                    if len(order_lines) > 0:
                        products = order_lines.sorted(lambda x: x.product_id.variant_name).ids
                        return {
                        'name': record.name,
                        'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'geo.restriction.message.wizard.spt',
                        'target': 'new',
                        'context': {'default_order_line_ids': [(6,0,products)] }
                    }
                if len(record.package_order_lines):
                    backup_order_id = order_backup_obj.search([('order_id','=',record.id)],order="id desc",limit=1)
                    record.action_sync_backup_order(backup_order_id)
            if record.state in ['draft','sent','received']:
                record = self.sudo()
                order_backup_obj.create({
                    'name' : record.name,
                    'partner_id' : record.partner_id.id,
                    'partner_invoice_id' : record.partner_invoice_id.id,
                    'partner_shipping_id' : record.partner_shipping_id.id,
                    'currency_id' : record.pricelist_id.currency_id.id,
                    'payment_term_id' : record.payment_term_id.id,
                    'date_order' : record.date_order,
                    'applied_promo_code' : record.applied_promo_code,
                    'line_ids' : backup_order_line_list,
                    'order_id': record.id,
                    'user_id': self.env.user.id,
                    'pricelist_id' : record.pricelist_id.id
                })
                if record.website_id and record.state == 'draft' and record.env.user.has_group('base.group_user'):
                    record.write({'is_confirm_by_saleperson':True})
                    geo_restriction_list = []
                    backup_order_line_list =[]
                    #Merge same product lines
                    record.merge_order_lines()
                    #checked stock
                    for line in record.order_line:
                        backup_order_line_list.append((0,0,{
                            'product_id':line.product_id.id,
                            'product_uom_qty':line.product_uom_qty,
                            'price_unit':line.price_unit,
                            'unit_discount_price':line.unit_discount_price,
                            'fix_discount_price':line.fix_discount_price,
                            'discount':line.discount,
                            'name':line.name,
                            'sale_type':line.sale_type,
                            'tax_id':[(6,0,line.tax_id.ids)],
                        }))
                        if line.product_id.type == 'product' and record.partner_id.country_id.id in line.product_id.geo_restriction.ids:
                            geo_restriction_list.append(line.id)

                if geo_restriction_list and not record._context.get('allow_restricted'):
                    order_lines = self.env['sale.order.line'].browse(geo_restriction_list)
                    products=[]
                    if len(order_lines) > 0 and not self._context.get('on_consign_wizard'):
                        products = order_lines.sorted(lambda x: x.product_id.variant_name).ids
                        return {
                        'name': record.name,
                        'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'geo.restriction.message.wizard.spt',
                        'target': 'new',
                        'context': {'default_order_line_ids': [(6,0,products)] }
                    }
                on_consign_product_ids = self.order_line.filtered(lambda x:x.product_id.on_consignment and x.product_uom_qty > x.product_id.actual_stock)
                if on_consign_product_ids and not self._context.get('on_consign_wizard'):
                    for line in on_consign_product_ids:
                        line.product_id.assign_qty = line.product_uom_qty or 0.0
                    return {
                        'name': _('Product Minimum Stock Alert.'),
                        'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'on.consignment.product.message.wizard',
                        'target': 'new',
                        'context': {'default_product_ids': [(6,0,on_consign_product_ids.mapped('product_id').ids)],'default_order_id':self.id }
                    }
                if record.state in ['draft'] and not record.website_id and not record.catalog_id:
                    template_id = self.env.ref('tzc_sales_customization_spt.mail_template_notify_customer_quotation_create').sudo()
                    template_id.send_mail(res_id=record.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")
                res = super(sale_order, self).action_confirm()
                picking_ids = self.mapped('picking_ids')
                if picking_ids:
                    PickingObj = self.env['stock.picking']
                    for picking in  picking_ids:
                        if picking not in ['cancel', 'done'] and picking.state == 'assigned':
                            PickingObj |= picking
                    if PickingObj.ids:
                        PickingObj.do_unreserve()
                        
                return res
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                            'title': 'Something is wrong.',
                            'message': 'Please reload your screen.',
                            'sticky': True,
                        }
                    }

    def line_ordering_by_product(self):
        product_list = []
        product_list = self.order_line.mapped(lambda x:x and x.product_id.name_get()[0][1].strip()) if  self.order_line else []
        # for line in range(len(self.order_line)):
        #     line = self.order_line[line]
        #     product_name = line.product_id.name_get()[0][1].split('(')
        #     product_list.append(product_name[0])
        # product_list = list(set(product_list))
        product_list.sort()
        return product_list

    def line_product_dict(self,product_name):
        product_dict = {}
        # for line in range(len(self.order_line)):
        #     line = self.order_line[line]
        #     line_dict = {} 
        #     product_name = line.product_id.name_get()[0][1].split('(')[0].strip()
        #     if line.product_id.name in product_dict.keys():
        #         product_dict[product_name]['line_ids'].append(line)  
        #     else:
        #         line_dict['line_ids'] = [line]
        #         product_dict[product_name] = line_dict
        product_dict[product_name] = {'line_ids': self.order_line.filtered(lambda x:x.product_id.name_get()[0][1].strip() == product_name)}
        return product_dict

    def get_access_token_spt(self):
        self.ensure_one()
        auth_param = url_encode(self.partner_id.signup_get_auth_param()[self.partner_id.id])
        return auth_param

    @api.depends('order_line.product_uom_qty', 'order_line.product_id','package_order_lines.qty')
    def _compute_cart_info(self):
        vals = super(sale_order,self)._compute_cart_info()
        for order in self:
            order.cart_quantity += int(sum(order.mapped('package_order_lines.qty')))            
        return vals

    def action_merge_quotation_spt(self):
        so_obj=self.env['sale.order']
        sol_obj = self.env['sale.order.line']
        package_line_obj = self.env['kits.package.order.line']
        if not len(self) <= 1:
            partner_id = self.mapped('partner_id')
            if len(partner_id) == 1:
                if not len(self.mapped('b2b_currency_id')) > 1:
                    if all(i in ('draft','sent','received') for i in self.mapped('state')):
                        so_id = self[0] if len(self) > 1 else self if self else so_obj
                        self = self-so_id
                        diff_price = []
                        for line in self.mapped('order_line'):
                            sale_line_id = so_id.order_line.filtered(lambda x: x.product_id.id == line.product_id.id and x.is_fs == line.is_fs and x.is_promotion_applied == line.is_promotion_applied and x.sale_type == line.sale_type)
                            if sale_line_id and round(sale_line_id.unit_discount_price,2) != round(line.unit_discount_price,2):
                                if sale_line_id.product_id.name not in diff_price:
                                    diff_price.append(sale_line_id.product_id.name)
                        if not diff_price:
                            for sol in self.mapped('order_line'):
                                if not sol.is_pack_order_line or not sol.package_id:
                                    sale_line_id = so_id.order_line.filtered(lambda x: x.product_id.id == sol.product_id.id and x.is_fs == sol.is_fs and x.is_promotion_applied == sol.is_promotion_applied and x.sale_type == sol.sale_type)
                                    if sale_line_id and sale_line_id.id:
                                        sale_line_id.write({
                                            'product_uom_qty':(sale_line_id.product_uom_qty+sol.product_uom_qty),
                                            'price_unit':round((sale_line_id.price_unit+sol.price_unit)/2,2),
                                            'unit_discount_price':round((sale_line_id.unit_discount_price+sol.unit_discount_price)/2,2),
                                            'discount':round((sale_line_id.discount+sol.discount)/2,2),
                                            })
                                    else:
                                        sale_line_id = sol_obj.create({
                                            'order_id':so_id.id,
                                            'name':sol.name,
                                            'product_id':sol.product_id.id,
                                            'product_uom_qty':sol.product_uom_qty,
                                            'price_unit':sol.price_unit,
                                            'unit_discount_price':sol.unit_discount_price,
                                            'discount':sol.discount,
                                            'product_uom':sol.product_uom.id,
                                            'is_fs':sol.is_fs,
                                            'sale_type':sol.sale_type,
                                            'is_promotion_applied':sol.is_promotion_applied
                                        })
                                    sale_line_id._onchange_unit_discounted_price_spt()
                        else:
                            raise UserError('Below products having different price.\n\n%s'%('\n'.join(i for i in diff_price)))
                        for pack_line in self.mapped('package_order_lines'):
                            if pack_line.product_id in so_id.package_order_lines.mapped('product_id'):
                                package_line = so_id.package_order_lines.filtered(lambda x: x.product_id == pack_line.product_id)
                                package_line.write({
                                    'qty':package_line.qty+pack_line.qty,
                                    'sale_price':round((package_line.sale_price+pack_line.sale_price)/2,2),
                                    'pack_price':round((package_line.pack_price+pack_line.pack_price)/2,2),
                                    })
                                package_line._onchange_pack_price()
                            else:
                                package_line = package_line_obj.create({'product_id':pack_line.product_id.id,'qty':pack_line.qty,'sale_price':pack_line.sale_price,'pack_price':pack_line.pack_price,'order_id':so_id.id,})
                                package_line._onchange_pack_price()
                        so_id._compute_package_order()
                        so_id.write({'merged_order':True,'merge_reference':[(6,0,so_id.merge_reference.ids+self.ids+self.mapped('merge_reference').ids)]})
                        self.sale_order_cancel_spt()
                        self.sudo().write({'state':'merged'})
                        return {
                            'name': 'Original Order',
                            'view_type': 'form',
                            'view_mode': 'form',
                            'target': 'current',
                            'res_model': 'sale.order',
                            'type': 'ir.actions.act_window',
                            'res_id': so_id.id,
                        }
                    else:
                        raise UserError(_("You can only merge quotations which are in 'Quotation, Quotation Sent, Quotation Received' state."))
                else:
                    raise UserError(_("Currency must be same for all selected quotations."))
            else:
                raise UserError(_("Please select quotation of same customer."))
        else:
            raise UserError("You have to select more than one quotation.")
    
    # def merge_order_lines(self):
    #     product_id_dict = {}
    #     for record in self:
    #         #Merge same product lines
    #         for line in record.order_line.filtered(lambda x: not x.is_pack_order_line):
    #             is_line = True
    #             if line.product_id in product_id_dict.keys():
    #                 order_line = product_id_dict[line.product_id]
    #                 order_line.update({'product_uom_qty':order_line.product_uom_qty + line.product_uom_qty})
    #                 line.unlink()
    #                 is_line = False
    #             else:
    #                 product_id_dict[line.product_id] = line
    #             if is_line and line.product_uom_qty == 0.0:
    #                 line.unlink()
    
    # def _get_unavailable_package_ids(self):
    #     self.ensure_one()
    #     restricted_packages = self.package_order_lines.filtered(lambda x: x.availability == 'out_of_stock')
    #     if restricted_packages:
    #         return {
    #             'name':"Restricted Packages",
    #             'type':'ir.actions.act_window',
    #             'res_model':'kits.package.restriction',
    #             'view_mode':'form',
    #             'context':{'default_order_id':self.id,'default_restricted_package_ids':[(6,0,restricted_packages.ids)]},
    #             'target':'new',
    #         }
    #     return True

    def action_get_backup_order(self):
        self.ensure_one()
        order_id = self.env['sale.order.backup.spt'].search([('order_id','=',self.id)],order="id desc",limit=1)
        return {
            'name': 'Original Order',
            'view_type': 'form',
            'view_mode': 'form',
            'target': 'current',
            'res_model': 'sale.order.backup.spt',
            'type': 'ir.actions.act_window',
            'res_id': order_id.id,
        }
    
    def action_sync_backup_order(self,backup_order):
        for record in self:
            backup_order.pack_order_backup = True
            backup_pack_line_list =self.env['kits.package.order.line']
            backup_package_line_obj = self.env['kits.package.order.line']
            for so_pack_line in record.package_order_lines:
                bpl = backup_package_line_obj.create({
                    'product_id':so_pack_line.product_id.id,
                    'pack_price':so_pack_line.pack_price,
                    'discount_amount':so_pack_line.discount_amount,
                    'sale_price':so_pack_line.sale_price,
                    'qty':so_pack_line.qty,
                })
                backup_pack_line_list |= bpl
            backup_pack_line_list.write({'backup_order':backup_order.id})

    # kits_picking_return
    kits_credit_payment_ids = fields.Many2many('account.payment','sale_order_credit_payments_rel','sale_order_id','credit_payment_id','Credit Notes')
    # count_kits_credit_notes = fields.Integer(compute="_count_kits_credit_notes")
    # count_kits_return_order = fields.Integer(compute="_count_kits_return_picking")

    # @api.depends('kits_credit_payment_ids','count_kits_return_order','picking_ids','picking_ids.is_return_picking')
    # def _count_kits_credit_notes(self):
    #     for record in self:
    #         record.count_kits_credit_notes = len(record.kits_credit_payment_ids)
    #     # pass

    # @api.depends('picking_ids','picking_ids.is_return_picking')
    # def _count_kits_return_picking(self):
    #     for record in self:
    #         record.count_kits_return_order = len(record.picking_ids.filtered(lambda x: not x.product_returned and x.kits_return_picking))
    #     # pass

    # def action_get_kits_credit_notes(self):
    #     notes = self.kits_credit_payment_ids
    #     action =  {
    #         'name':_('Credit Notes'),
    #         'type':'ir.actions.act_window',
    #         'res_model':'account.payment',
    #         'view_mode':'form',
    #         'context':{'default_is_return_credit':True},
    #         'target':'self',
    #     }
    #     if len(notes) == 1:
    #         action['res_id'] = notes.id
    #     else:
    #         action['view_mode'] = 'tree,form'
    #         action['domain'] = [('id','in',notes.ids)]
    #     return action

    # def action_get_kits_return_pickings(self):
    #     pickings = self.picking_ids.filtered(lambda x: not x.product_returned and x.kits_return_picking)
    #     action = {
    #         'name':_("Return Orders"),
    #         'type':'ir.actions.act_window',
    #         'res_model':"stock.picking",
    #         'view_mode':'form',
    #         'target':'self',
    #     }
    #     if len(pickings) == 1:
    #         action['res_id']=pickings.id
    #     else:
    #         action['view_mode'] = 'tree,form'
    #         action['domain'] = [('id','in',pickings.ids)]
    #     return  action

    def action_create_credit_note_kits(self):
        if self.env.user.has_group('account.group_account_invoice'):
            return {
                'name':_("Credit Note"),
                'type':'ir.actions.act_window',
                'res_model':"kits.create.credit.note.wizard",
                'view_mode':'form',
                'context':{'default_sale_id':self.id},
                'target':'new',
            }
        else:
            raise UserError('Only billing user can create credit note.')


    # theme_tzc_enterprice
    def recompute_coupon_lines(self):
        for order in self:
            if not order.applied_coupon_ids.on_order_line and order.applied_coupon_ids:
            # if not order.code_promo_program_id.on_order_line and order.code_promo_program_id:
                order._remove_invalid_reward_lines()
                order._create_new_no_code_promo_reward_lines()
                order._update_existing_reward_lines()

    def _cart_update(self, product_id=None, line_id=None, add_qty=0, set_qty=0, **kwargs):
        """ Add or set product quantity, add_qty can be negative """
        self.ensure_one()
        product_context = dict(self.env.context)
        product_context.setdefault('lang', self.sudo().partner_id.lang)
        SaleOrderLineSudo = self.env['sale.order.line'].sudo().with_context(product_context)
        website = self.env['website'].get_current_website()
        # change lang to get correct name of attributes/values
        product_with_context = self.env['product.product'].with_context(product_context)
        if not product_id:
            raise UserError(_("The given product does not exist therefore it cannot be added to cart."))
        product = product_with_context.browse(int(product_id))
        values = {}
        try:
            if add_qty:
                add_qty = int(add_qty)
        except ValueError:
            add_qty = 1
        try:
            if set_qty:
                set_qty = int(set_qty)
        except ValueError:
            set_qty = 0
        quantity = 0
        order_line = False
        if self.state != 'draft':
            request.session['sale_order_id'] = None
            # raise UserError(_('It is forbidden to modify a sales order which is not in draft status.'))
            return True
        
        # Create line if no line with product_id can be located
        if not line_id:
            order_line = self._cart_find_product_line(product_id, line_id, **kwargs)[:1]
        
        if not order_line:
            if not product:
                raise UserError(_("The given combination does not exist therefore it cannot be added to cart."))
            product_id = product.id
            # Get order line in case line id exist  
            if line_id:
                order_line = self._cart_find_product_line(product_id, line_id, **kwargs)[:1]
            if not order_line:
                values = self._website_product_id_change(self.id, product_id, qty=1)
                # if product.eto_sale_method in ['fs'] and self.partner_id.customer_type in ['b2b_fs']:
                #     values['is_fs'] = True
                
                # create the line
                if product.on_consignment and product.actual_stock:
                    if  int(product.actual_stock)  >= values['product_uom_qty'] :
                        order_line = SaleOrderLineSudo.create(values)
                    else:
                        # raise UserError(_("Sorry, Temporarily out of stock."))
                        self.warning_stock = _("Some products have sold out and your cart has been updated. Please don't delay processing your order.")
                        values['warning'] = self.warning_stock
                else:
                    order_line = SaleOrderLineSudo.create(values)
                try:
                    order_line._compute_tax_id()
                except ValidationError as e:
                    # The validation may occur in backend (eg: taxcloud) but should fail silently in frontend
                    _logger.debug("ValidationError occurs during tax compute. %s" % (e))
                if add_qty:
                    add_qty -= 1

        # compute new quantity
        if set_qty:
            quantity = set_qty
        elif add_qty is not None:
            quantity = order_line.product_uom_qty + (add_qty or 0)

        # Remove zero of negative lines and Remove unpublshed lines and remove out of stock lines 
        available_qty_spt = order_line.product_id.with_context(warehouse=website.warehouse_id.id).available_qty_spt
        minimum_qty_spt = order_line.product_id.with_context(warehouse=website.warehouse_id.id).minimum_qty
        if order_line.product_id.with_context(warehouse=website.warehouse_id.id).on_consignment:
            available_qty_spt = available_qty_spt - minimum_qty_spt
        if quantity <= 0 or available_qty_spt <= 0 or not order_line.product_id.is_published_spt:
            linked_line = order_line.linked_line_id
            if available_qty_spt <= 0 or not order_line.product_id.is_published_spt:
                self.warning_stock = _("Some products have sold out and your cart has been updated. Please don't delay processing your order.")
                values['warning'] = self.warning_stock
            order_line.unlink()
            # values['warning'] = self.warning_stock
            if linked_line:
                # update description of the parent
                linked_product = product_with_context.browse(linked_line.product_id.id)
                linked_line.name = linked_line.get_sale_order_line_multiline_description_sale(linked_product)
        else:
            # update line
            no_variant_attributes_price_extra = [
                ptav.price_extra for ptav in order_line.product_no_variant_attribute_value_ids]
            values = self.with_context(no_variant_attributes_price_extra=tuple(
                no_variant_attributes_price_extra))._website_product_id_change(self.id, product_id, qty=quantity)
            if self.pricelist_id.discount_policy == 'with_discount' and not self.env.context.get('fixed_price'):
                order = self.sudo().browse(self.id)
                product_context.update({
                    'partner': order.partner_id,
                    'quantity': quantity,
                    'date': order.date_order,
                    'pricelist': order.pricelist_id.id,
                    'force_company': order.company_id.id,
                })
            product_with_context = self.env['product.product'].with_context(
                product_context)
            product = product_with_context.browse(product_id)
            partner_price_list = self.env.user.partner_id.b2b_pricelist_id
            values['price_unit'] = partner_price_list.get_product_price(
                product, quantity, self.env.user.partner_id)
            pricelist_price = partner_price_list.get_product_price(product,values.get('product_uom_qty'),self.env.user.partner_id)
            order_currency = order.b2b_currency_id if order else self.env.user.currency_id
            sale_price = 0.00
            if product.sale_type:
                if product.sale_type == 'on_sale':
                    if partner_price_list.currency_id.name == 'CAD':
                        pricelist_price = product.on_sale_cad
                    else:
                        pricelist_price = product.on_sale_usd

                if product.sale_type == 'clearance':
                    if partner_price_list.currency_id.name == 'CAD':
                        pricelist_price = product.clearance_cad
                    else:
                        pricelist_price = product.clearance_usd
                
                if order_currency:
                    if order_currency.name.lower() == 'usd':
                        sale_price = order_line.product_id.lst_price
                    if order_currency.name.lower() == 'cad':
                        sale_price = order_line.product_id.lst_price

                values['price_unit'] = round(sale_price, 2)
                values['unit_discount_price'] = round(pricelist_price, 2)
                values['discount'] = round(100 - (pricelist_price / sale_price) * 100, 2)
                values['fix_discount_price'] = round((sale_price * values['discount'])/100, 2)
                values['sale_type'] = product.sale_type
            else:
                if order_currency:
                    if order_currency.name.lower() == 'usd':
                        sale_price = order_line.product_id.lst_price
                    if order_currency.name.lower() == 'cad':
                        sale_price = order_line.product_id.lst_price

                values['price_unit'] = round(sale_price, 2)
                values['unit_discount_price'] = round(pricelist_price, 2)
                values['discount'] = round(100 - (pricelist_price / sale_price) * 100, 2)
                values['fix_discount_price'] = round((sale_price * values['discount'])/100, 2)

            active_inflation = self.env['kits.inflation'].search([('is_active','=',True)])
            inflation_id = self.env['kits.inflation.rule'].search([('country_id','in',self.env.user.partner_id.country_id.ids),('brand_ids','in',product.brand.ids),('inflation_id','=',active_inflation.id)])
            inflation_rule_id = inflation_id[-1] if inflation_id else False
            is_inflation = False

            if inflation_rule_id:
                if active_inflation.from_date and active_inflation.to_date:
                    if active_inflation.from_date <= datetime.now().date() and active_inflation.to_date >= datetime.now().date():
                        is_inflation = True
                elif active_inflation.from_date:
                    if active_inflation.from_date <= datetime.now().date():
                        is_inflation = True
                elif active_inflation.to_date:
                    if active_inflation.to_date >= datetime.now().date():
                        is_inflation = True
                else:
                    if not active_inflation.from_date:
                        is_inflation = True
                    if not active_inflation.to_date:
                        is_inflation = True
                
                if is_inflation:
                    inflation_discount_price = round(values['unit_discount_price'] + (values['unit_discount_price'] * inflation_rule_id.inflation_rate / 100),2)
                    price_unit = round(values['price_unit'] + (values['price_unit'] * inflation_rule_id.inflation_rate / 100),2)
                    fix_discount_price = round((100 * (values['price_unit'] - inflation_discount_price) / values['price_unit']),2)
                    values['unit_discount_price'] = inflation_discount_price
                    values['price_unit'] = price_unit
                    values['fix_discount_price'] = round(values['price_unit'] - inflation_discount_price,2)
                    if product.sale_type:
                        values['discount'] = fix_discount_price

            active_fest_id = self.env['tzc.fest.discount'].search([('is_active','=',True)])
            special_disocunt_id = self.env['kits.special.discount'].search([('country_id','in',self.env.user.partner_id.country_id.ids),('brand_ids','in',product.brand.ids),('tzc_fest_id','=',active_fest_id.id)])
            price_rule_id = special_disocunt_id[-1] if special_disocunt_id else False
            applicable = False

            if price_rule_id:
                if active_fest_id.from_date and active_fest_id.to_date:
                    if active_fest_id.from_date <= datetime.now().date() and active_fest_id.to_date >= datetime.now().date():
                        applicable = True
                elif active_fest_id.from_date:
                    if active_fest_id.from_date <= datetime.now().date():
                        applicable = True
                elif active_fest_id.to_date:
                    if active_fest_id.to_date >= datetime.now().date():
                        applicable = True
                else:
                    if not active_fest_id.from_date:
                        applicable = True
                    if not active_fest_id.to_date:
                        applicable = True
                
                if applicable:
                    special_discount_price = round(values['unit_discount_price'] - (values['unit_discount_price'] * price_rule_id.discount / 100),2)
                    fix_discount_price = round((100 * (values['price_unit'] - special_discount_price) / values['price_unit']),2)
                    values['unit_discount_price'] = special_discount_price
                    values['discount'] = fix_discount_price
                    values['fix_discount_price'] = round(values['price_unit'] - special_discount_price,2)
                    values['is_special_discount'] = True

            order_line.write(values)
            # self.recompute_coupon_lines()
            line_id = order_line.id
            for order_line in self.order_line:
                if order_line.product_id.type == 'product' and (order_line.id == line_id) and order_line.product_id.inventory_availability in ['always', 'threshold']:
                    cart_qty = sum(self.order_line.filtered(
                        lambda p: p.product_id.id == order_line.product_id.id).mapped('product_uom_qty'))
                    # The quantity should be computed based on the warehouse of the website, not the
                    # warehouse of the SO.
                    available_qty_spt = order_line.product_id.with_context(
                        warehouse=website.warehouse_id.id).available_qty_spt
                    if product.on_consignment:
                        available_qty_spt = available_qty_spt - product.minimum_qty
                    if available_qty_spt < 0:
                        available_qty_spt = 0
                        
                    if cart_qty > available_qty_spt:
                        qty = available_qty_spt
                        order_line.product_uom_qty = qty
                        values.update({
                            'line_id': order_line.id, 'quantity': order_line.product_uom_qty,
                        })
                        # Make sure line still exists, it may have been deleted in super()_cartupdate because qty can be <= 0
                        if order_line.exists() and order_line.product_uom_qty >= 1:
                            order_line.warning_stock = _('Only %s pcs are available now.') % (
                                int(order_line.product_uom_qty))
                            values['warning'] = order_line.warning_stock
                        else:
                            self.warning_stock = _(
                                "Some products have sold out and your cart has been updated. Please don't delay processing your order.")
                            values['warning'] = self.warning_stock
                    else:
                        values['quantity'] = quantity
            order_line.name = order_line.get_sale_order_line_multiline_description_sale(
                product)
            values['order_line'] = order_line.id

        return values

    @api.depends('order_line')
    def _compute_website_order_line(self):
        for order in self:
            sale_order_line = self.env["sale.order.line"]
            sorted_order_line_list = sorted(order.order_line, key=lambda l: l.product_id.variant_name)
            for line in sorted_order_line_list:
                sale_order_line |= line
            order.website_order_line = sale_order_line


    # download images
    def download_images_excel(self):
        path = get_module_resource('catalog','static')
        list_of_files = [f for f in os.listdir(path) if isfile(join(path, f))]

        df = pd.DataFrame(list_of_files)
        df.to_excel(f'{path[:-7]}/Product Images.xlsx', header=False ,index=False)

        byte_data = 0
        with open(f'{path[:-7]}/Product Images.xlsx', "rb") as xlfile:
            byte_data = xlfile.read()
        
        attachment = self.env['ir.attachment'].create({
                        'name': 'Product Images.xlsx',
                        'datas': base64.b64encode(byte_data),
                        'type': 'binary',
                        'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    })

        return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=ir.attachment&download=true&field=datas&id=%s&filename=%s' % (attachment.id,'Product Images.xlsx'),
                'target': 'self',
            }

    def action_view_invoice(self):
        invoices = self.invoice_ids.filtered(lambda x:x.state != 'cancel') if self._context.get('show_invoice') else self.mapped('invoice_ids')
        action = self.env.ref('account.action_move_out_invoice_type').read()[0]
        if len(invoices) > 1:
            action['domain'] = [('id', 'in', invoices.ids)]
        elif len(invoices) == 1:
            form_view = [(self.env.ref('account.view_move_form').id, 'form')]
            if 'views' in action:
                action['views'] = form_view + [(state,view) for state,view in action['views'] if view != 'form']
            else:
                action['views'] = form_view
            action['res_id'] = invoices.id
        else:
            action = {'type': 'ir.actions.act_window_close'}

        context = {
            'default_type': 'out_invoice',
        }
        if len(self) == 1:
            context.update({
                'default_partner_id': self.partner_id.id,
                'default_partner_shipping_id': self.partner_shipping_id.id,
                'default_invoice_payment_term_id': self.payment_term_id.id or self.partner_id.property_payment_term_id.id or self.env['account.move'].default_get(['invoice_payment_term_id']).get('invoice_payment_term_id'),
                'default_invoice_origin': self.name,
                'default_user_id': self.user_id.id,
            })
        action['context'] = context
        return action
    def _search_abandoned_cart(self, operator, value):
        abandoned_delay = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_abandoned_delay', default='1.0'))
        abandoned_datetime = fields.Datetime.to_string(datetime.utcnow() - relativedelta(hours=abandoned_delay))
        abandoned_domain = expression.normalize_domain([
            ('date_order', '<=', abandoned_datetime),
            ('state', '=', 'draft'),
            ('partner_id', '!=', self.env.ref('base.public_partner').id),
            ('order_line', '!=', False)
        ])
        abandoned_domain = expression.normalize_domain(abandoned_domain)
        # is_abandoned domain possibilities
        if (operator not in expression.NEGATIVE_TERM_OPERATORS and value) or (operator in expression.NEGATIVE_TERM_OPERATORS and not value):
            return abandoned_domain
        return expression.distribute_not(['!'] + abandoned_domain)  # negative domain

    def action_recovery_email_send(self):
        for order in self:
            order._portal_ensure_token()
        composer_form_view_id = self.env.ref('mail.email_compose_message_wizard_form').id

        template_id = self._get_cart_recovery_template()

        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'view_id': composer_form_view_id,
            'target': 'new',
            'context': {
                'default_composition_mode': 'mass_mail' if len(self.ids) > 1 else 'comment',
                'default_email_layout_xmlid': 'mail.mail_notification_layout_with_responsible_signature',
                'default_res_id': self.ids[0],
                'default_model': 'sale.order',
                'default_use_template': bool(template_id),
                'default_template_id': template_id,
                'website_sale_send_recovery_email': True,
                'active_ids': self.ids,
            },
        }


    def get_sorted_lines(self,lines):
        return lines.sorted(lambda x: x.product_id.variant_name)
        
    def _get_cart_recovery_template(self):
        """
        Return the cart recovery template record for a set of orders.
        If they all belong to the same website, we return the website-specific template;
        otherwise we return the default template.
        If the default is not found, the empty ['mail.template'] is returned.
        """
        # to changes
        template = eval(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.cart_recovery_mail_template', default='None'))
        if not template:
            template = self.env.ref('tzc_sales_customization_spt.mail_template_sale_cart_recovery', raise_if_not_found=False)
            if template:
                return template.id
            else:
                return self.env['mail.template']
        else:
            return template 

    def get_order_portal_url(self):
        website_id = self.env['kits.b2b.website'].search([],limit=1)
        if website_id and website_id.url:
            return website_id.url

    def get_email_subject(self):
        subject = 'Your pro-forma for order %s is ready!'%self.name
        if not self._context.get('proforma'):
            subject =  '%s has sent you a quotation!: %s'%(self.user_id.name,self.name)
        
        return subject
