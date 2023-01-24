from odoo import fields, models, api,_
from odoo.exceptions import ValidationError,UserError
from odoo.tools.float_utils import float_compare, float_is_zero, float_round
from base64 import b64decode
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from io import BytesIO
from datetime import datetime
import math
import base64
import ast
import os
import ast
from lxml import etree

field_list = ['move_ids_without_package','partner_id','user_id','scheduled_date','origin','shipping_id','include_cases','no_of_cases','carrier_id','calulate_shipping_cost',
              'tracking_number_spt','package_contain','total_box','weight','weight_unit','package_type_id','height','width','kits_length','ship_date','carriage_value','currency_id',
              'shipment','transportation_to','duties_taxes','notes','customer_ref','shipment_purpose','commercial_invoice','export_export','b13a','exemption','note','ups_no_of_package',
              'general_desc','street','street_2','city','state_id','country_id','zip_code','company_name','phone','phone_ext','fright','state']
class stock_picking(models.Model):
    _inherit = 'stock.picking'

    def _get_default_include_cases(self):
        order = self.sale_id
        if not order:
            order = self.env['sale.order'].search([('picking_ids','in',self.ids)])
        return order.include_cases
    
    def action_open_return_picking_wizard_kits(self):
        return {
            'name':_('Return Product'),
            'type':'ir.actions.act_window',
            'res_model':'kits.wizard.return.picking',
            'view_mode':'form',
            'view_id':self.env.ref('tzc_sales_customization_spt.kits_wizard_return_picking_form_view').id,
            'context':{'default_picking_id':self.id,'default_total_qty':self.delivered_qty},
            'target':'new',
            }

    def action_create_credit_note(self):
        if self.env.user.has_group('account.group_account_invoice'):
            return {
                'name':_("Credit Note"),
                'type':'ir.actions.act_window',
                'res_model':"kits.create.credit.note.wizard",
                'view_mode':'form',
                'context':{'default_sale_id':self.sale_id.id,'default_picking_id':self.id},
                'target':'new',
            }
        else:
            raise UserError('Only billing user can create credit note.')

    def action_show_return_pickings_kits(self):
        pickings = self.sale_id.picking_ids.filtered(lambda x: x.kits_return_picking)
        action = {
            'name':_("Return Ordes"),
            'type':'ir.actions.act_window',
            'res_model':"stock.picking",
            'view_mode':'form',
            'target':'self',
        }
        if len(pickings) == 1:
            action['res_id']=pickings.id
        else:
            action['view_mode']='tree,form'
            action['domain'] = [('id','in',pickings.ids)]
        return action

    def _get_rec_name(self):
        for rec in self:
            rec_name = '%s (%s)'%(rec.origin,rec.name)
            rec.delivery_name = rec_name

    def _get_default_no_of_cases(self):
        order = self.sale_id
        if not order:
            order = self.env['sale.order'].search([('picking_ids','in',self.ids)])
        cases = order.no_of_cases
        return cases
    
    # sale_website_id = fields.Many2one('website', string='Sale Order Website',compute="_compute_check_in_sale_order")
   

    # @api.depends('sale_id')  
    # def _compute_check_in_sale_order(self):
    #     website = False
    #     for record in self:
    #         record.sale_website_id = record.sale_id.website_id.id if record.sale_id.website_id else website 
    
    state = fields.Selection(selection_add=[
        ('draft', 'Draft'),
        ('waiting', 'Waiting Another Operation'),
        ('confirmed', 'Waiting'),
        ('in_scanning', 'In Scanning'),
        ('scanned','Scanning Completed'),
        ('assigned', 'Ready To Ship'),
        ('done', 'Shipped'),
        ('cancel', 'Cancelled'),
    ], string='Status',
        copy=False, index=True, tracking=True,default='confirmed' ,
        help="Delivery in Which state",readonly=False)
    ordered_qty = fields.Integer('Order Quantity',compute="_compute_qty",store=True)
    delivered_qty = fields.Integer('Picked Quantity',compute="_compute_qty",store=True)
    source_spt =  fields.Char('source',related="sale_id.source_spt")
    order_qty_count_spt = fields.Integer('Order Picked Quantity',compute="_compute_qty",store=True)
    include_cases = fields.Boolean('Include Cases ?',default=_get_default_include_cases)
    no_of_cases = fields.Integer('#Cases',default=_get_default_no_of_cases,compute='_get_no_of_cases',store=True)
    partner_street = fields.Char('Street',related='partner_id.street')
    partner_street2 = fields.Char('Street2',related='partner_id.street2')
    partner_city = fields.Char(" City",related="partner_id.city")
    partner_postal_code = fields.Char(related="partner_id.zip")
    partner_state_id = fields.Many2one('res.country.state',' State',related="partner_id.state_id")
    partner_country_id = fields.Many2one('res.country',' Country',related="partner_id.country_id")
    show_update_button = fields.Boolean()
    updated_on = fields.Datetime("Updated On ")
    updated_by = fields.Many2one('res.users',"Updated By ")
    preiviews_scanning_products_data = fields.Char()
    delivery_data = fields.Char()
    delivery_name = fields.Char(compute="_get_rec_name",string="Name")
    is_multiple_delivery = fields.Boolean("Is Multiple Deliver",help="This is a flag for multiple delivery of same order.",compute="_check_order_delivery",store=True,compute_sudo=True)
    show_package_scan = fields.Boolean(compute="_compute_show_package_scan")
    package_order = fields.Boolean('Package Order',compute="_compute_show_package_scan")
    package_order_status = fields.Selection([('available','Available'),('out_of_stock','Out of stock')],string="Package Status",compute="_compute_show_package_scan")

    product_returned = fields.Boolean('Product Returned ?')
    product_scraped = fields.Boolean('Product Scrapped ?')
    credit_note_created  = fields.Boolean('Credit note ?')
    count_return_order = fields.Integer('Return Orders',compute="_count_return_pickings")
    kits_return_picking = fields.Boolean(compute="_compute_kits_return_picking",store=True,compute_sudo=True)
    move_lines = fields.One2many('stock.move', 'picking_id', string="Stock Moves", copy=True)

    actual_weight = fields.Float('Actual Weight (kg)',related="weight")
    weight_of_cases = fields.Float('Calculated Weight for cases (kg)',compute="_compute_weight_of_cases",store=True,compute_sudo=True)
    weight_total_kg = fields.Float('Total Weight (kg)',compute="_compute_weight_of_cases",store=True,compute_sudo=True)
    calulate_shipping_cost = fields.Float('Shipping Cost ')
    shipping_id = fields.Many2one('shipping.provider.spt', ondelete='set null', string='Shipping Provider')
    tracking_number_spt = fields.Char('Tracking Number')
    shipping_label = fields.Binary('Shipping Label')
    file_name = fields.Char()
    is_fedex = fields.Boolean()
    is_ups = fields.Boolean()
    provider = fields.Char("Provider ",compute="_get_provider",store=True)
    is_provider_pick_up = fields.Boolean(default=False)

    # Shipping Details 
    
    # Recipient Info Fields.
    recipient_id = fields.Many2one('res.partner',"Recipient ID")
    country_id = fields.Many2one('res.country',"Country")
    company_name = fields.Char('Company Name')
    street = fields.Char('Address 1')
    street_2 = fields.Char('Address 2')
    zip_code = fields.Char('Postal Code')
    state_id = fields.Many2one('res.country.state','State')
    city = fields.Char('City')
    phone = fields.Char('Telephone')
    phone_ext = fields.Char('Ext.')
    tax = fields.Char('Tax ID/EIN')
    location = fields.Char('Location #')

    # Packge & Shipment Detail Fields.
    package_contain = fields.Selection([('document','Document'),('commodity','Commodity/Merchandise')],'Package Contains',default="commodity")
    num_of_package = fields.Integer('Number of packages',default=1,compute="_depends_default_package")
    total_box = fields.Integer('Total Boxes',default=1)
    identical_package = fields.Boolean('Identical packages')
    weight = fields.Float('Total Weight',related="weight_total_kg")
    weight_unit = fields.Selection([('kg','KG'),('lb','LB')],default='kg')
    package_type_id = fields.Many2one('product.packaging','Package Type')
    height = fields.Integer('Height (cm)')
    width = fields.Integer('Width (cm)')
    kits_length = fields.Integer("Length (cm)")
    ship_date = fields.Datetime('Ship Date',default=fields.Date.today())
    carriage_value = fields.Monetary('Total Carriage value',related="sale_id.picked_qty_order_subtotal",currency_field='currency_id')
    currency_id = fields.Many2one('res.currency',related="sale_id.currency_id")

    # Sender Info Fields
    currecnt_sender_name = fields.Many2one('res.company',compute="_depends_default_package")
    kits_street = fields.Char()
    kits_street_1 = fields.Char()
    kits_country_id = fields.Many2one('res.country')
    kits_state_id = fields.Many2one('res.country.state')
    kits_zip_code = fields.Char()
    kits_city = fields.Char()

    # Billing Details
    transportation_to =fields.Selection([('sender','Sender'),('recipient','Recipient'),('third_party','Third Party'),('collect','Collect')],"Bill transportation to",default="sender")
    duties_taxes = fields.Float('Bill Duties/taxes/fees')
    notes = fields.Char('Department Notes')
    customer_ref = fields.Char('Customer Reference')

    # Additional Ref.
    order_id = fields.Many2one('sale.order',"P.O. number",related="sale_id")
    shipment = fields.Char('Shipment ID')

    # Customs Info.
    shipment_purpose = fields.Selection([('gift','Gift'),('sample','Sample'),('repair_return','Repair And Return'),('personal','Personal Effects'),('sold','Sold'),('not_sold','Not Sold')],"Shipment Purpose",default="sold")

    # Customs Document
    commercial_invoice = fields.Selection([('COMMERCIAL_INVOICE','My Own Commercial Invoice'),('PRO_FORMA_INVOICE','My Own Profoma'),('CERTIFICATE_OF_ORIGIN','FedEx-Generated Commercial Invoice'),('NAFTA_CERTIFICATE_OF_ORIGIN','FedEx-Generated Proforma')],"Commercial Invoice/Proforma",default="COMMERCIAL_INVOICE")
    export_export = fields.Char('Export Permit #')
    b13a = fields.Selection([('NOT_REQUIRED','No B13A Required'),('MANUALLY_ATTACHED','Manual B13A Attached'),('FILED_ELECTRONICALLY','CAED Electronic B13A'),('SUMMARY_REPORTING','B13A Summary Reporting')],"B13A",default="NOT_REQUIRED")
    exemption = fields.Char('Exemption #')

    # UPS Shipment
    fright = fields.Boolean('Freight')
    general_desc = fields.Char('General Desc. of Goods')
    num_of_pieces = fields.Float('No of Pices')
    shipped_pack_as = fields.Selection([('carboy','Carboy(s)'),('pallet','Pallet(s)'),('skid','Skid(s)'),('tote','Tote(s)')],"Shipped pack as",default='pallet')
    num_of_pack = fields.Float('No of packs')
    pack_type = fields.Selection([('loose','Loose'),('other','Other')],"Pack type",default='loose')
    ups_no_of_package = fields.Integer(' No of Package',default=1)
    warning = fields.Char()
    
    def set_order_status(self):
        sale_ids = self.env['sale.order'].search([('state','in',['sale','in_scanning','scanned','scan','shipped'])])
        for sale in sale_ids:
            picking = sale.picking_ids.filtered(lambda x:x.state !='cancel')
            if picking and sale.state != picking.state :
                if picking.state == 'done' and sale.state not in ['draft_inv','open_inv','cancel','shipped']:
                    invoice_id = sale.invoice_ids.filtered(lambda x:x.state != 'cancel')
                    if invoice_id:
                        if invoice_id.state == 'draft' and sale.state != 'draft_inv':
                            sale.state = 'draft_inv'
                        if  invoice_id.state == 'posted' and sale.state != 'open_inv':
                            sale.state = 'open_inv'
                    else:
                        sale.state = 'shipped'
                if picking.state == 'assigned' and sale.state != 'scan':
                    sale.state = 'scan'
                if picking.state == 'scanned' and sale.state != 'scanned':
                    sale.state = 'scanned'
                if picking.state == 'in_scanning' and sale.state != 'in_scanning':
                    sale.state = 'in_scanning'
                if picking.state == 'cancel' and sale.state != 'cancel':
                    sale.state = 'cancel'
                if picking.state == 'confirmed' and sale.state != 'sale':
                    sale.state = 'sale'

    def _depends_default_package(self):
        for rec in self:
            if not rec.num_of_package:
                rec.num_of_package = 1
            if not rec.total_box:
                rec.total_box = 1
            if not rec.currecnt_sender_name:
                company_id = self.env.company
                rec.currecnt_sender_name = company_id.id
                rec.kits_street = company_id.street
                rec.kits_street_1 = company_id.street2
                rec.kits_country_id = company_id.country_id.id
                rec.kits_state_id = company_id.state_id.id
                rec.kits_zip_code = company_id.zip
                rec.kits_city = company_id.city

    @api.depends('shipping_id')
    def _get_provider(self):
        for rec in self:
            rec.provider = False
            if rec.shipping_id and rec.shipping_id.provider:
                rec.provider = rec.shipping_id.provider
                
    @api.depends('sale_id','sale_id.include_cases','sale_id.no_of_cases','sale_id.case_weight_kg','shipping_weight')
    def _compute_weight_of_cases(self):
        for record in self:
            order = record.sale_id or self.env['sale.order'].search([('picking_ids','in',record.ids)],limit=1)
            weight_of_cases = 0.0
            if order:
                weight_of_cases = order.case_weight_kg
            record.weight_of_cases = weight_of_cases
            record.weight_total_kg = round(round(record.shipping_weight,2)+weight_of_cases,2)

    @api.depends('name','is_return_picking','sale_id','sale_id.picking_ids')
    def _compute_kits_return_picking(self):
        for record in self:
            record.kits_return_picking = record.is_return_picking or 'IN' in record.name

    @api.depends('state','sale_id','sale_id.count_kits_return_order')
    def _count_return_pickings(self):
        for record in self:
            record.count_return_order = record.sale_id.count_kits_return_order

            
    @api.depends('move_ids_without_package','move_ids_without_package.package_id')
    def _compute_show_package_scan(self):
        for record in self:
            record.show_package_scan = True if len(record.move_ids_without_package.filtered(lambda x: x.package_id)) else False
            record.package_order = record.sale_id.package_order or record.show_package_scan
            record.package_order_status = record.sale_id.package_order_status
            
    @api.depends('ordered_qty','delivered_qty')
    def _get_no_of_cases(self):
        for rec in self:
            rec.no_of_cases = 0
            if rec.delivered_qty:
                rec.no_of_cases = rec.delivered_qty
            elif rec.ordered_qty:
                rec.no_of_cases = rec.ordered_qty

    @api.depends('sale_id.picking_ids')
    def _check_order_delivery(self):
        for rec in self:
            rec.is_multiple_delivery = False
            if rec.sale_id:
                cancel_picking_ids = rec.sale_id.picking_ids.filtered(lambda x:x.state == 'cancel')
                if cancel_picking_ids:
                    rec.is_multiple_delivery = True
    
    @api.depends('move_ids_without_package','sale_id','move_ids_without_package.sale_line_id.product_uom_qty','move_ids_without_package.product_uom_qty','move_ids_without_package.quantity_done')
    def _compute_qty(self):
        for record in self:
            delivered_qty = 0
            ordered_qty = 0
            qty_total = 0
            for line in range(len(record.move_ids_without_package)):
                line = record.move_ids_without_package[line]
                try:
                    delivered_qty += line.quantity_done
                    ordered_qty += line.product_uom_qty
                    if line.sale_line_id:
                        if line.quantity_done > line.sale_line_id.product_uom_qty:
                            qty_total = 1
                    else:
                            qty_total = 1

                except:
                    pass
            record.order_qty_count_spt = qty_total
            record.ordered_qty = ordered_qty
            record.delivered_qty = delivered_qty
            record.sale_id._amount_all()

    def all_picking_cancle_spt(self):
        for rec in self:
            rec.with_context(cancel_delivery=True).picking_cancel_spt()

    def picking_cancel_spt(self):
        if self.state in ['scanned', 'in_scanning', 'done', 'confirmed', 'assigned']:
            self = self.with_context(force_delete=True)
            picking_data = ast.literal_eval(
                self.delivery_data) if self.delivery_data else {}
            for stock_picking in self:
                if self._context.get('cancel_delivery'):
                    picking_data.update({stock_picking.id: {}})
                    for line in stock_picking.move_ids_without_package:
                        if picking_data[stock_picking.id].get(line.product_id.id):
                            picking_data[stock_picking.id].get(line.product_id.id).update(
                                {
                                    'demand': picking_data[stock_picking.id].get(line.product_id.id).get('demand') + line.product_uom_qty,
                                    'done': picking_data[stock_picking.id].get(line.product_id.id).get('done') + line.quantity_done
                                }
                            )
                        else:
                            picking_data[stock_picking.id].update(
                                {
                                    line.product_id.id: {
                                        'demand': line.product_uom_qty, 'done': line.quantity_done}
                                }
                            )
                    stock_picking.delivery_data = picking_data
                stock_picking.state = 'cancel'
                stock_picking.sale_id.write(
                    {'state': 'received' if stock_picking.sale_id.source_spt != 'Manually' else 'draft'})
                stock_picking.move_lines.stock_quant_update_spt()
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                }
            }
    def action_product_qty_delivery_order_cancel(self):
        try:
            product_wizard_obj = self.env['stock.change.product.qty']
            for record in self:
                product_dict = {}
                # product record with qty
                for line in range(len( record.move_ids_without_package)):
                    line =  record.move_ids_without_package[line]
                    if line.quantity_done:
                        product_dict[line.product_id ] = line.qty_available

                record.all_picking_cancle_spt()            
                if product_dict:
                    for line in product_dict.keys():
                        product_wizard_id = product_wizard_obj.with_context(create=True,default_origin=record.origin+'-'+record.name if record.origin else record.name).create({
                        'product_id' : line.id,
                        'product_tmpl_id' : line.product_tmpl_id.id,
                        'new_quantity' : product_dict[line] if product_dict[line] > 0 else 0

                    })
                    product_wizard_id.change_product_qty()
        except Exception as e:
            raise UserError(_(e.args))

    def excel_report_line_stock(self):
        line_obj = self.env['sale.order.line']
        stock_dict = {}
        active_id= self.id
        f_name ='Case-report-%s'%(self.sale_id.name)
        base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Case Report.xlsx'
        wb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
        wrksht = wb.active
        bd = Side(style='thin', color="000000")
        all_border = Border(left=bd, top=bd, right=bd, bottom=bd)
        right_border = Border(right=bd)
        bottom_border = Border(top=bd)
        address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
        address_font = Font(name='Calibri', size=10, bold=False)
        all_font = Font(size=12, bold=True)
        for record in self.sale_id:
            brand_list=[]
            brand_ids = record.order_line.mapped('product_id.brand')
            for brand in brand_ids:
                line_ids = line_obj.search([('product_id.type','!=','service'),('product_id.brand','=',brand.id),('order_id','=',record.id)])
                for line in line_ids:
                    categ_dict = {}
                    if line.picked_qty:
                        if line.product_id.brand.name in stock_dict.keys():
                            if line.product_id.categ_id.name  in stock_dict[line.product_id.brand.name].keys():
                                if line.product_id.sale_type in stock_dict[line.product_id.brand.name][line.product_id.categ_id.name]['sale_type'].keys():
                                    stock_dict[line.product_id.brand.name][line.product_id.categ_id.name]['sale_type'][line.product_id.sale_type] = stock_dict[line.product_id.brand.name][line.product_id.categ_id.name]['sale_type'][line.product_id.sale_type] +line.picked_qty
                                else:
                                    stock_dict[line.product_id.brand.name][line.product_id.categ_id.name]['sale_type'].update({line.product_id.sale_type:line.picked_qty})
                            else:
                                stock_dict[line.product_id.brand.name][line.product_id.categ_id.name] = {'categ_id': line.product_id.categ_id.name, 'sale_type':{line.product_id.sale_type: line.picked_qty}}
                        else:
                            categ_dict[line.product_id.categ_id.name] = {'categ_id': line.product_id.categ_id.name,'sale_type':{line.product_id.sale_type: line.picked_qty}}
                            stock_dict[line.product_id.brand.name] = categ_dict
        
        # --------------------------------------------------------- Header Address end  ---------------------------------------------------------
        # ------------------------------------------------------------
        # Billing Address
        # ------------------------------------------------------------
        address_row = 1
        billing_address = self.env['sale.order'].create_address_line_for_sale(self.partner_id, take_name=True)
        wrksht.cell(row=address_row+1, column=1).value = billing_address
        wrksht.cell(row=address_row+1, column=1).alignment = address_alignment
        wrksht.cell(row=address_row+1, column=1).font = address_font
        wrksht.cell(row=address_row+1, column=1).font = address_font
        # -------------------------------------------------------------
        # Shipping Address
        # -------------------------------------------------------------
        shipping_address = self.env['sale.order'].create_address_line_for_sale(self.sale_id.partner_shipping_id, take_name=True)
        wrksht.cell(row=address_row+1, column=5).value = shipping_address
        wrksht.cell(row=address_row+1, column=5).alignment = address_alignment
        wrksht.cell(row=address_row+1, column=5).font = address_font
        # ---------------------------------------------- Name ----------------------------------------------
        name_row = address_row + 10
        wrksht.cell(row=name_row, column=1).alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True)
        wrksht.cell(row=name_row, column=1).value = self.delivery_name
        wrksht.cell(row=name_row, column=1).font = Font(name='Calibri', size=12, bold=True)

        wrksht.cell(row=14, column=1).border = all_border
        wrksht.cell(row=14, column=2).border = all_border
        wrksht.cell(row=14, column=3).border = all_border
        wrksht.cell(row=14, column=4).border = all_border
        wrksht.cell(row=14, column=5).border = all_border
        wrksht.cell(row=14, column=6).border = all_border
        wrksht.cell(row=14, column=7).border = all_border
        wrksht.cell(row=14, column=1).font = all_font
        wrksht.cell(row=14, column=5).font = all_font
        wrksht.cell(row=14, column=6).font = all_font
        wrksht.cell(row=14, column=7).font = all_font

        row_index = 15
        brand_list = list(stock_dict.keys())
        brand_list.sort()
        total_regular = 0
        total_on_sale = 0
        total_clearance = 0
        for data in brand_list:
            for brand_data in stock_dict[data]:
                for sale_type in stock_dict[data][brand_data]['sale_type']:
                    product_sale_type = ''
                    if sale_type == 'on_sale':
                        product_sale_type = 'On Sale'
                        total_on_sale += stock_dict[data][brand_data]['sale_type'][sale_type]
                    elif sale_type == 'clearance':
                        total_clearance += stock_dict[data][brand_data]['sale_type'][sale_type]
                        product_sale_type = 'Clearance'
                    else:
                        product_sale_type = 'Regular'
                        total_regular += stock_dict[data][brand_data]['sale_type'][sale_type]
                    wrksht.merge_cells('A'+str(row_index)+':D'+str(row_index))
                    wrksht.cell(row=row_index, column=1).value = data
                    wrksht.cell(row=row_index, column=5).value = stock_dict[data][brand_data]['categ_id']
                    wrksht.cell(row=row_index, column=6).value = product_sale_type  
                    wrksht.cell(row=row_index, column=7).value = stock_dict[data][brand_data]['sale_type'][sale_type]
                    wrksht.cell(row=row_index, column=7).border = right_border
                    row_index += 1
        
        wrksht.cell(row=row_index+1, column=6).value = 'Total Clearance'
        wrksht.cell(row=row_index+1, column=7).value = total_clearance
        wrksht.cell(row=row_index+2, column=6).value = 'Total On Sale'
        wrksht.cell(row=row_index+2, column=7).value = total_on_sale
        wrksht.cell(row=row_index+3, column=6).value = 'Total Regular'
        wrksht.cell(row=row_index+3, column=7).value = total_regular
        wrksht.cell(row=row_index+4, column=6).value = 'Total'
        wrksht.cell(row=row_index+4, column=7).value = total_regular+total_clearance+total_on_sale

        for col in range(1,8):
            wrksht.cell(row=row_index, column=col).border = bottom_border
        wrksht.cell(row=row_index, column=6).border = bottom_border
        wrksht.cell(row=row_index, column=7).border = bottom_border

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})

        f_name = f_name
        return {
            'type' : 'ir.actions.act_url',
            'url':   'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, f_name),
            'target': 'self',
            }

    def update_sale_order_spt(self):
        context_spt = dict(self.env.context)
        context_spt.update({'no_create_spt':True})
        order_line_obj = self.env['sale.order.line'].with_context(context_spt)
        for record in self:
            if record.state in ['confirmed','in_scanning','scanned','assigned']:
                for line in range(len(record.move_ids_without_package)):
                    line = record.move_ids_without_package[line]
                    if line.sale_line_id:
                        #update the quontity in related sale order line
                        if line.quantity_done > line.product_uom_qty:
                            line.sale_line_id.product_uom_qty = line.quantity_done
                            line.product_uom_qty = line.quantity_done
                    else:
                        if not record.sale_id:
                            record.sale_id = record.sale_id.search([('name','=',self.origin)]).id
                        price_unit = record.sale_id.pricelist_id.get_product_price(line.product_id, line.quantity_done, record.sale_id.partner_id)
                        if line.product_id.sale_type == 'on_sale' and record.sale_id and record.sale_id.pricelist_id and record.sale_id.pricelist_id.currency_id:
                            if record.sale_id.pricelist_id.currency_id.name == 'CAD':
                                price_unit = line.product_id.on_sale_cad
                            if record.sale_id.pricelist_id.currency_id.name == 'USD':
                                price_unit = line.product_id.on_sale_usd
                        
                        if line.product_id.sale_type == 'clearance' and record.sale_id and record.sale_id.pricelist_id and record.sale_id.pricelist_id.currency_id:
                            if record.sale_id.pricelist_id.currency_id.name == 'CAD':
                                price_unit = line.product_id.clearance_cad
                            if record.sale_id.pricelist_id.currency_id.name == 'USD':
                                price_unit = line.product_id.clearance_usd

                        if line.quantity_done:
                            # create new sale order line
                            order_line_id = order_line_obj.sudo().create({
                                'order_id': record.sale_id.id,
                                'product_id': line.product_id.id,
                                'product_uom_qty': line.quantity_done,
                                'product_uom': line.product_id.uom_id.id,
                                'name': line.product_id.display_name,
                                'price_unit': price_unit,
                                'sale_type':  line.product_id.sale_type,
                            })
                            if order_line_id:
                                order_line_id.product_id_change()
                                order_line_id._onchange_discount_spt()
                                order_line_id._onchange_unit_discounted_price_spt()
                                line.write({'sale_line_id':order_line_id.id,'product_uom_qty':line.quantity_done})
                record.sale_id._amount_all()
        return {
        'type': 'ir.actions.client',
        'tag': 'display_notification',
        'params': {
                'title': 'Something is wrong.',
                'message': 'Please reload your screen.',
                'sticky': True,
            }
        }


    def picked_order_qty_spt(self):
        for record in self:
            if record.state in ['in_scanning','confirmed']:
                if not record.sale_id:
                    record.sale_id = record.get_order_id(self)
                record.preiviews_scanning_products_data = record.get_scanned_product()
                if record.ordered_qty >= record.delivered_qty:
                    for line in range(len(record.move_ids_without_package)):
                        line = record.move_ids_without_package[line]
                        line.quantity_done = line.product_uom_qty
                    record.sale_id.write({'state':'in_scanning'})
                    record.write({'state': 'in_scanning'})
                    # record.sale_id.with_context(fulfilled=True).send_shipment_ready_email_to_salesperson_spt()
                    record.sale_id._amount_all()
                else:
                    raise UserError(_('If user clicks when fulfilled is higher than ordered qty.'))
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                    'title': 'Something is wrong.',
                    'message': 'Please reload your screen.',
                    'sticky': True,
                }
            }

    def action_restore_preiviews_scanning(self):
        for rec in self:
            if rec.state in ['in_scanning']:
                if not rec.sale_id:
                    rec.sale_id = rec.get_order_id(self)
                if rec.id and rec.preiviews_scanning_products_data:
                    data = ast.literal_eval(rec.preiviews_scanning_products_data) or ''
                    if data and data.get(rec.id):
                        for product_data in data.get(rec.id):
                            product_id = self.env['product.product'].browse([product_data])
                            line_id = rec.move_ids_without_package.filtered(lambda x:x.product_id.id == product_id.id)
                            if line_id:
                                line_id.quantity_done = data.get(rec.id).get(product_id.id)
                        rec.sale_id.write({'updated_by':self.env.user.id,'updated_on':datetime.now()}) if rec.sale_id else None
                                
                self.message_post(body='Revert Fullfil.')
                self.state = 'confirmed' if not any(self.move_ids_without_package.mapped('quantity_done')) else 'in_scanning'
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                    'title': 'Something is wrong.',
                    'message': 'Please reload your screen.',
                    'sticky': True,
                }
            }


    def action_update_order(self):
        self.ensure_one()
        if self.state in ['in_scanning','confirmed']:
            if not self.sale_id:
                self.sale_id = self.get_order_id(self)
            return{
                'name': _('Add Order'),
                'type': 'ir.actions.act_window',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'kits.update.picking.wizard',
                'target': 'new',
                'context' : {
                    'default_partner_id' : self.sale_id.partner_id.id,
                    'default_picking_id' : self.id,
                }
            }
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }

# ============================== pick list==============================
    def excel_picking_order_report(self):
        for rec in self:
            active_id = self.id
            f_name = 'Order Report Stock Picking'  # FileName
            base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/Pick List Sample.xlsx'
            wb = load_workbook(base_sample_file,read_only=False, keep_vba=False)
            wrksht = wb.active

            #wrksht name
            header_font = Font(name='Calibri', size=16, bold=True)
            table_font = Font(name='Calibri', size=10, bold=False)
            alignment = Alignment(horizontal='center',vertical='center', text_rotation=0,wrap_text=True)
            address_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', text_rotation=0,wrap_text=True)
            alignment_right = Alignment(horizontal='right', vertical='center', text_rotation=0,wrap_text=True)
            bd = Side(style='thin', color="D3D3D3")
            bottom_border = Border(bottom=bd)
            
            # ------------------------------------------------------------
            # Delivery Address
            # ------------------------------------------------------------
            delivery_address = self.create_address_line_for_sale(self.partner_id,take_name=True)
            wrksht.cell(row=8, column=1).value = delivery_address
            wrksht.cell(row=8, column=1).alignment = address_alignment
            wrksht.cell(row=8, column=1).font = Font(name="Calibri",size=10,bold=False)
            # ==================================== 
            # Name 
            # ====================================
            wrksht.cell(row=16, column=1).value = str(self.delivery_name or '')
            wrksht.cell(row=16, column=1).font = Font(name='Calibri', size=12, bold=True)
            wrksht.cell(row=16, column=1).alignment = alignment_left
            
            # ==================================== 
            # Date, source
            # ====================================
            wrksht.cell(row=19, column=1).value = str("Scheduled Date:\r\n"+str(self.scheduled_date.date()) or '')
            wrksht.cell(row=19, column=1).font = Font(name='Calibri', size=10, bold=True)
            wrksht.cell(row=19, column=1).alignment = alignment_left
            wrksht.cell(row=19, column=2).value = str("Source Document:\r\n"+str(self.origin )if self.origin else '')
            wrksht.cell(row=19, column=2).font = Font(name='Calibri', size=10, bold=True)
            wrksht.cell(row=19, column=2).alignment = alignment_left

            header_row = 23
            row_index = header_row+1
            total_quantity = 0
            orders_dict = {}
            orders = []
            for line in range(len(self.move_ids_without_package)):
                line = self.move_ids_without_package[line]
                if line.product_uom_qty > 0:
                    total_quantity += line.product_uom_qty
                    orders.append([line.product_id.default_code,line.product_id.name_get()[0][1],line.product_id.categ_id.name, line.product_uom_qty])
                    orders_dict.update({line.product_id.default_code:[line.product_id.default_code,line.product_id.name_get()[0][1],line.product_id.categ_id.name, line.product_uom_qty]})

            for order in sorted(orders):
                wrksht.cell(row=row_index, column=1).value = order[0]
                wrksht.cell(row=row_index, column=2).value = order[1]
                wrksht.cell(row=row_index,column=3).value = order[2]
                wrksht.cell(row=row_index,column=4).value = order[3]
                wrksht.row_dimensions[row_index].height = 27
                wrksht.cell(row=row_index, column=1).border = bottom_border
                wrksht.cell(row=row_index, column=2).border = bottom_border
                wrksht.cell(row=row_index, column=3).border = bottom_border
                wrksht.cell(row=row_index, column=4).border = bottom_border
                
                wrksht.cell(row=row_index, column=1).font = table_font
                wrksht.cell(row=row_index, column=2).font = table_font
                wrksht.cell(row=row_index, column=3).font = table_font
                wrksht.cell(row=row_index, column=4).font = table_font

                wrksht.cell(row=row_index, column=1).alignment = alignment_left
                wrksht.cell(row=row_index, column=2).alignment = alignment_left
                wrksht.cell(row=row_index, column=3).alignment = alignment
                wrksht.cell(row=row_index, column=4).alignment = alignment_right
                row_index += 1
            
            # -------------- 
            # total QTY 
            # -------------- 
            wrksht.cell(row=19, column=3).value = str("Total QTY:\n"+str(int(total_quantity)))
            wrksht.cell(row=19, column=3).font = Font(name='Calibri', size=10, bold=True)
            wrksht.cell(row=19, column=3).alignment = alignment_left

            footer_row = row_index+1
            wrksht.cell(row=footer_row, column=3).value = "Total QTY :"
            wrksht.cell(row=footer_row, column=3).font = Font(name='Calibri', size=11, bold=False)
            wrksht.cell(row=footer_row, column=3).alignment = alignment_right
            wrksht.cell(row=footer_row, column=4).value = int(total_quantity)
            wrksht.cell(row=footer_row, column=4).font = table_font
    
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})
        # rec.file = base64.b64encode(data) 
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, f_name),
            'target': 'self',
        }

    def action_scanned(self):
        error_msg = ''
        for rec in self:
            if rec.state in ['in_scanning']:
                for line in range(len(rec.move_ids_without_package)):
                    line = rec.move_ids_without_package[line]
                    rec.check_duplicate_move(line)
                if not rec.shipping_id:
                    error_msg = 'Please select shipping provider.'
                elif not rec.sale_id.order_line.filtered(lambda x:x.product_id.is_shipping_product) and rec.shipping_id and rec.shipping_id.name.lower() != 'pick up':
                    error_msg = 'Please calculate and add the shipping cost to the order.'

                if error_msg:
                    raise UserError(error_msg)

                template_id = self.env.ref('tzc_sales_customization_spt.mail_template_notify_salesperson_order_scanned')
                user_id = self.env['res.users'].search([('is_warehouse','=',True)],limit=1)
                rec.write({'state':'scanned'})
                rec.sale_id.write({'state':'scanned'})
                rec.get_picking_order_values()
                if user_id and user_id.email:
                    template_id.send_mail(rec.id,force_send=True,notif_layout="mail.mail_notification_light")
                rec.sale_id._amount_all()
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                            'title': 'Something is wrong.',
                            'message': 'Please reload your screen.',
                            'sticky': True,
                        }
                    }

    def set_qty_by_script(self):
        inventory_dict = {}
        product_dict = {}
        product_wizard_obj = self.env['stock.change.product.qty']
        for record in self:
            if record.state in ['assigned']:
                if not record.sale_id:
                    record.sale_id = record.get_order_id(self)

                # product record with qty
                for line in range(len(record.move_ids_without_package)):
                    line = record.move_ids_without_package[line]
                    if line.product_id in product_dict.keys():
                        product_dict[line.product_id ] = line.quantity_done + product_dict[line.product_id ]
                    else:
                        product_dict[line.product_id ] = line.quantity_done
                # product dict loop
                for product_id in product_dict.keys():
                    if product_id.type != 'service':
                        # qty = product_dict[product_id]
                        qty = product_id.minimum_qty + product_dict[product_id] if product_id.on_consignment else product_dict[product_id]
                        if qty > product_id.qty_available:
                            inventory_dict[product_id] =  {'qty' : qty}
                            inventory_dict[product_id]['location_id'] = product_id.property_stock_inventory.id
                            inventory_dict[product_id]['product_uom_id'] = product_id.uom_id.id
                            inventory_dict[product_id]['before_qty_on_hand'] = product_id.qty_available
                            inventory_dict[product_id]['before_available_qty'] = product_id.available_qty_spt
                            inventory_dict[product_id]['before_reserved_qty'] = product_id.reversed_qty_spt
                            record.sale_id.write({'updated_by':self.env.user.id,'updated_on':datetime.now()})
                if inventory_dict:
                    for line in inventory_dict.keys():
                        product_wizard_id = product_wizard_obj.with_context(create=True).create({
                            'product_id' : line.id,
                            'product_tmpl_id' : line.product_tmpl_id.id,
                            'new_quantity' : inventory_dict[line]['qty']

                        })
                        product_wizard_id.change_product_qty()
                        self.create_update_qty_log(line,inventory_dict[line])
                record.sale_id._amount_all()
                #record.action_assign() 
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                    'title': 'Something is wrong.',
                    'message': 'Please reload your screen.',
                    'sticky': True,
                }
        }


    def action_reset_to_inscanning(self):
        if self.state in ['scanned']:
            for rec in self:
                if not rec.sale_id:
                    rec.sale_id = rec.get_order_id(self)
                rec.write({'state':'in_scanning'})
                rec.sale_id.write({'state':'in_scanning'})
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
    def button_open_quick_scan_spt(self):
        if self.state in ['in_scanning','confirmed']:
            for record in self:
                if not record.sale_id:
                    record.sale_id = record.get_order_id(self)
                for line in range(len(record.move_ids_without_package)):
                    line =  record.move_ids_without_package[line]
                    record.check_duplicate_move(line)
                if record.state not in ['done','cancel']:
                    wizard_id = self.env['stock.picking.barcode.spt'].create({
                        'picking_id':record.id,
                    })
                    return {
                        'name': 'Scan Order',
                        'view_mode': 'form',
                        'target': 'new',
                        'res_id':wizard_id.id,
                        'res_model': 'stock.picking.barcode.spt',
                        'type': 'ir.actions.act_window',
                    }
                else:
                    raise UserError(_("You can't change the stock when in %s."%(record.state)))
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }

    def open_remove_done_quantity_wizard(self):
        self.ensure_one()
        if self.state in ['in_scanning','confirmed']:
            if not self.sale_id:
                self.sale_id = self.get_order_id(self)
            for line in range(len(self.move_ids_without_package)):
                    line =  self.move_ids_without_package[line]
            wizard_id = self.env['remove.done.quantity.spt'].create({'picking_id' : self.id})
            return {
                'name': 'Remove Items',
                'view_mode': 'form',
                'target': 'new',
                'res_id':wizard_id.id,
                'res_model': 'remove.done.quantity.spt',
                'type': 'ir.actions.act_window',
                "context":{
                    "default_picking_id":self.id,
                },
                }  
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
            


    def action_delivery_restore(self):
        if self.state not in ['confirmed','in_scanning']:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else:
            return {
                "name":_("Recover Delivery"),
                "type":"ir.actions.act_window",
                "res_model":"delivery.recovery.selection.wizard",
                "view_mode":"form",
                "target":"new",
                "context":{'default_sale_id':self.sale_id.id}
            }
            
    def action_picking_to_order(self):
        list_view = self.env.ref('sale.view_order_tree')
        form_view = self.env.ref('sale.view_order_form')
        action =  {
            "name":_("Sale Order"),
            "type":"ir.actions.act_window",
            "res_model":"sale.order",
            "target":"current",
        }
        if not self.sale_id:
            self.sale_id = self.get_order_id(self)
        if self.sale_id:
            action.update( {
                "view_mode":"form",
                "res_id":self.sale_id.id,
            })
        else:
            action.update( {
                "view_mode":"tree,form",
                "views":[(list_view.id,"tree"),(form_view.id,"form")],
                "domain":[('picking_ids','in',self.ids)],
            })
        return action


    def get_order_id(self,picking_id):
        sale_obj = self.env['sale.order']
        order_id = False
        if picking_id and picking_id.origin:
            order_id = sale_obj.search([('name','=',picking_id.origin)],limit=1)
        return order_id


    @api.onchange('fright','shipping_id')
    def _check_order_weight(self):
        for rec in self:
            domain = []
            if rec.fright:
                carrier_ids = self.env['delivery.carrier'].search([('is_freight','=',True)])
                rec.carrier_id = carrier_ids[0].id if carrier_ids else False
                domain.append(('id','in',carrier_ids.ids))
            else:
                carrier_ids = self.env['delivery.carrier'].search([('delivery_type','=',rec.provider),('is_default','=',True)],limit=1)
                rec.carrier_id = carrier_ids.id if carrier_ids else False
                domain.append(('delivery_type','=',rec.provider))
                domain.append(('is_freight','=',False))

        return {'domain':{'carrier_id':domain}}

    # @api.onchange('weight_unit')
    # def onchange_check_weighty_unit(self):
    #     for rec in self:
    #         if rec.carrier_id and rec.carrier_id.delivery_type.lower() == 'ups' and rec.is_ups:
    #             if rec.carrier_id.ups_package_weight_unit != rec.weight_unit.upper()+'S':
    #                 raise UserError('Selected service type \'%s\' weight unit is %s & order weight unit is %s'%(rec.carrier_id.name,rec.carrier_id.ups_package_weight_unit,rec.weight_unit.upper()+'S'))
    #             else:
    #                 rec._onchange_weight()

    @api.onchange('tracking_number_spt')
    def _add_tracking_ref_in_order(self):
        for rec in self:
            if rec.sale_id:
                rec.sale_id.kits_carrier_tracking_ref = rec.tracking_number_spt

    @api.onchange('partner_id')
    def _get_recipent_address(self):
        if self.partner_id:
            self.recipient_id = self.partner_id.id
            self.country_id = self.partner_id.country_id.id
            self.state_id = self.partner_id.state_id.id
            self.street = self.partner_id.street
            self.street_2 = self.partner_id.street2
            self.zip_code = self.partner_id.zip
            self.city = self.partner_id.city
            self.phone = self.partner_id.phone
            
    @api.onchange('shipping_id')
    def _onchange_fedex_flag(self):
        for rec in self:
            rec.is_fedex = False
            rec.is_ups = False
            rec.carrier_id = False
            rec.is_provider_pick_up = False
            if rec.shipping_id:
                if rec.shipping_id and rec.shipping_id.name.lower() == 'pick up':
                    rec.is_provider_pick_up = True
                else:
                    shipping_id = self.env['delivery.carrier'].search([('delivery_type','=',rec.shipping_id.provider),('is_default','=',True)],limit=1)
                    if shipping_id:
                        if shipping_id.delivery_type.lower() == 'fedex':
                            rec.is_fedex = True
                        elif shipping_id.delivery_type.lower() == 'ups':
                            rec.is_ups = True
                        rec.carrier_id = shipping_id.id
                        rec.sale_id.carrier_id = shipping_id.id
            rec._get_recipent_address()
            rec._onchange_carrier_id()
            
    @api.onchange('carrier_id')
    def _onchange_carrier_id(self):
        for rec in self:
            if rec.carrier_id:
                rec.carrier_id = rec.carrier_id.id
                rec.sale_id.carrier_id = rec.carrier_id.id
                rec.package_type_id = getattr(rec.carrier_id,'%s_default_packaging_id'%rec.provider).id if rec.carrier_id else False
                # if rec.is_ups:
                #     rec.weight_unit = 'kg' if rec.carrier_id.ups_package_weight_unit == 'KGS' else 'lb'
            else:
                rec.carrier_id = False
                rec.package_type_id = False
            
            rec._onchange_dimention()

    @api.onchange('package_type_id')
    def _onchange_dimention(self):
        for rec in self:
            pass
            # rec.kits_length = rec.package_type_id.length
            # rec.width = rec.package_type_id.width
            # rec.height = rec.package_type_id.height


    def calculate_package(self):
        package = 1
        if self.is_ups and not self.fright:
            package_max_weight = self.package_type_id.max_weight * 2.2046 if self.weight_unit == 'lb' else self.package_type_id.max_weight
            package = math.ceil(self.weight / package_max_weight)

        return package
    
    @api.onchange('weight','fright','weight_unit')
    def _onchange_weight(self):
        for rec in self:
            rec.warning = False
            weight = rec.weight if rec.weight_unit != 'lb' else round(rec.weight / 2.2046,2)
            if rec.weight:
                if rec.weight_unit == 'lb' and round(weight,2) > 68.04:
                    rec.warning = 'The order weight is more than 150 lb, Add multiple package for ship order.'
                elif rec.weight_unit == 'kg' and rec.weight > 68.04:
                    rec.warning = 'The order weight is more than 68.04 kg, Add multiple package for ship order.'
                rec.actual_weight = weight
                rec.sale_id.actual_weight = weight
                rec.ups_no_of_package = self.calculate_package() if rec.warning else 1
            else:
                rec.actual_weight = rec.weight_total_kg
                rec.sale_id.actual_weight = rec.weight_total_kg

    def send_to_shipper(self):
        if not self._context.get('no_shipping_label'):
            self.ensure_one()
            res = self.carrier_id.send_shipping(self)[0]
            if self.carrier_id.free_over and self.sale_id and self.sale_id._compute_amount_total_without_delivery() >= self.carrier_id.amount:
            # if self.carrier_id.free_over and self.sale_id and self.sale_id._compute_amount_total_without_delivery() >= self.carrier_id.amount:
                res['exact_price'] = 0.0
            self.carrier_price = res['exact_price'] * (1.0 + (self.carrier_id.margin / 100.0))
            if res['tracking_number']:
                label = res.get('fedex_label') or res.get('ups_label')
                self.carrier_tracking_ref = res['tracking_number']
                self.tracking_number_spt = res['tracking_number']
                self.sale_id.kits_carrier_tracking_ref = res['tracking_number']
                self.sale_id.carrier_id = self.carrier_id.id
                if label:
                    base64_data = base64.b64encode(label[1])
                    byte_data = b64decode(base64_data,validate=True)
                    if byte_data[0:4] != b'%PDF':
                        raise ValueError('Missing the PDF file signature')
                    else:
                        self.shipping_label = base64.b64encode(byte_data)
                        self.file_name = label[0].split('.')[0] + '-' + res['tracking_number'] if self.carrier_id.delivery_type != 'fedex' else label[0].split('.')[0]
            order_currency = self.sale_id.currency_id or self.company_id.currency_id
            msg = _("Shipment sent to carrier %s for shipping with tracking number %s<br/>Cost: %.2f %s") % (self.carrier_id.name, self.carrier_tracking_ref, self.carrier_price, order_currency.name)
            self.message_post(body=msg)
            self._add_delivery_cost_to_so()
        else:
            pass

    def action_open_shipping_wizard(self):
        view_id = self.env.ref('delivery.choose_delivery_carrier_view_form').id
        delivery_method = self.env['delivery.carrier']
        shipping_method = False
        shipping_id = self.shipping_id
        if self.env.context.get('carrier_recompute'):
            name = _('Update shipping cost')
            carrier = self.sale_id.carrier_id
        else:
            name = _('Add a shipping method')
            carrier = (
                self.sale_id.partner_shipping_id.property_delivery_carrier_id
                or self.sale_id.partner_shipping_id.commercial_partner_id.property_delivery_carrier_id
            )
        domain = [('id','=',0)]
        if not carrier:
            if shipping_id and shipping_id.name.lower() == 'fedex':
                domain = [('delivery_type','=','fedex')]
                if self.partner_id.country_id.code == 'US' or self.partner_id.country_id.code == 'CA':
                    domain.append(('fedex_service_type','=','FEDEX_GROUND'))
                    # shipping_method = 'Fedex US'
                else:
                    domain.append(('fedex_service_type','=','INTERNATIONAL_ECONOMY'))
                    # shipping_method =  'Fedex International'
            if shipping_id and shipping_id.name.lower() == 'ups':
                if self.partner_id.country_id.code == 'US':
                    shipping_method = 'UPS US'
                else:
                    shipping_method = 'UPS BE'
        carrier = delivery_method.search(domain,limit=1)
        return {
            'name': name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'choose.delivery.carrier',
            'view_id': view_id,
            'views': [(view_id, 'form')],
            'target': 'new',
            'context': {
                'default_order_id': self.sale_id.id,
                'default_carrier_id': carrier.id,
                'default_weight_of_case': self.weight_of_cases,
                'default_weight_of_glasses': self.shipping_weight,
            }
        }

    def action_check_shipping_cost(self):
        view_id = self.env.ref('delivery.choose_delivery_carrier_view_form').id
        return {
            'name': _('Check Shipping Cost'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'choose.delivery.carrier',
            'view_id': view_id,
            'views': [(view_id, 'form')],
            'target': 'new',
            'context': {
                'default_order_id': self.sale_id.id,
                'default_weight_of_case': self.sale_id.case_weight_kg,
                'default_weight_of_glasses': self.sale_id.glass_weight_kg,
                'estimated_shipping_cost':False,
                'get_product_uom':True,
            }
        }

    @api.model
    def create(self,vals):
        if not 'currecnt_sender_name' in vals.keys():
            company_id = self.env.company
            vals.update({'currecnt_sender_name':company_id.id,
                         'kits_street':company_id.street,
                         'kits_street_1':company_id.street2,
                         'kits_country_id':company_id.country_id.id,
                         'kits_state_id':company_id.state_id.id,
                         'kits_zip_code':company_id.zip,
                         'kits_city':company_id.city})
        if not 'carrier_id' in vals.keys():
            service_type_id = self.env['delivery.carrier'].search([('delivery_type','=','fedex'),('is_default','=',True)],limit=1)
            vals.update({'carrier_id':service_type_id.id if vals.get('shipping_id') else False,
                         'package_type_id':service_type_id.fedex_default_packaging_id.id,
                         'kits_length':service_type_id.fedex_default_packaging_id.length,
                         'width':service_type_id.fedex_default_packaging_id.width,
                         'height':service_type_id.fedex_default_packaging_id.height})
        if not 'recipient_id' in vals.keys() and 'origin' in vals.keys():
            order_id = self.env['sale.order'].search([('name','=',vals.get('origin'))],limit=1)
            if order_id:
                vals.update({
                    'recipient_id':order_id.partner_id.id,
                    'country_id':order_id.partner_id.country_id.id,
                    'state_id':order_id.partner_id.state_id.id,
                    'street':order_id.partner_id.street,
                    'street_2':order_id.partner_id.street2,
                    'zip_code':order_id.partner_id.zip,
                    'city':order_id.partner_id.city,
                    'phone':order_id.partner_id.phone,
                })
        if "active_model" in self._context.keys() and self._context.get('active_model') == 'sale.order':
            if 'params' in self._context.keys():
                sale_id = self.env['sale.order'].browse(self._context['params'].get('id'))
            else:
                sale_id = self.env['sale.order'].search([('name','=',vals['origin'])]) if vals else False
            if sale_id and sale_id.picking_ids.filtered(lambda picking: picking.state != 'cancel' ):
                raise UserError("You can not create more then one delivery order of same sale order.")
        
        if "origin" in vals.keys():
            sale_order_id = self.env['sale.order'].search([('name','=',vals['origin'])])
            if sale_order_id and len(sale_order_id.picking_ids.filtered(lambda x:x.state != 'cancel' and x.picking_type_id.name != "Receipts")) >= 1:
                raise UserError(_('You can not add product in this order.\n%s has already one delivery order')%(vals.get('origin')))

        order = self.env['sale.order'].search([('name','=',vals.get('origin'))],limit=1)
        if order and vals.get('origin'):
            vals.update(include_cases=order.include_cases,no_of_cases=order.no_of_cases)
        return super(stock_picking,self).create(vals)
    
    def add_shipping_cost(self):
        if self.sale_id.state not in ['scan','shipped','draft_inv','open_inv','cancel','merged','done','assigned']:
            old_shipping_value = self.sale_id.amount_is_shipping_total
            shipping_product = self.env['product.product'].search([('is_shipping_product','=',True)],limit=1)
            shipping_line = self.sale_id.order_line.filtered(lambda x: x.product_id.is_shipping_product == True)
            vals = {
                'product_id':shipping_product.id or False,
                'product_uom_qty':1,
                'price_unit':round(self.calulate_shipping_cost,2) or 0.00,
                'unit_discount_price':round(self.calulate_shipping_cost,2) or 0.00,
                'fix_discount_price':0.0,
                'discount':0.0,
                'picked_qty_subtotal':0.0,
                'order_id':self.sale_id.id or False,
                'is_shipping_product':True,
            }
            if shipping_line:
                shipping_line.write(vals)
            else:
                self.env['sale.order.line'].create(vals)

            self.sale_id.message_post(body='Shipping Cost:%s --> %s'%(old_shipping_value,self.calulate_shipping_cost))

            return {
                'effect': {
                    'fadeout': 'slow',
                    'message': 'Shipping cost is added !!!',
                    'type': 'rainbow_man',
                }
            }
        else:
            raise UserError('You can not add shipping cost after order ready to ship.')

    def get_label(self):
        if self.shipping_id and self.shipping_id.name.lower() != 'pick up':
            self.send_to_shipper()

    def get_rate(self):
        vals = {}
        if self.carrier_id and self.sale_id and self.partner_id:
            vals.update({'carrier_id':self.carrier_id.id,'order_id':self.sale_id.id,'partner_id':self.partner_id})
            shipping_cost_wizard = self.env['choose.delivery.carrier'].create(vals)
            result = shipping_cost_wizard.update_price()
            if shipping_cost_wizard.delivery_price:
                self.write({'calulate_shipping_cost':shipping_cost_wizard.delivery_price})

    def action_download_pdf_report(self):
        for rec in self:
            if rec.tracking_number_spt and rec.exemption:
                return self.env.ref('tzc_sales_customization_spt.shipping_detail_receipt_report').report_action(rec)
            else:
                raise UserError('Make sure Tracking Number and Exemption # number is added in order.')

    @api.onchange('no_of_cases')
    def _onchange_no_of_cases(self):
        for rec in self:
            rec.sale_id.no_of_cases = rec.no_of_cases if rec.sale_id else None

    @api.onchange('include_cases')
    def _onchange_include_case(self):
        for rec in self:
            rec.sale_id.include_cases = rec.include_cases

    def _compute_order_qty_count_spt(self):
        for record in self:
            qty_total = 0
            for line in range(len(self.move_ids_without_package)):
                line = self.move_ids_without_package[line]
                if line.sale_line_id:
                    if line.quantity_done > line.sale_line_id.product_uom_qty:
                        qty_total = 1
                else:
                        qty_total = 1

            record.order_qty_count_spt = qty_total


    def button_validate(self):
        if self.state in ['assigned']:
            context = self._context.copy()
            context.update({'no_shipping_label':True,'ship':True})
            self.env.context = context
            error_message = ''
            # shipping_error = []
            website_error_msg = "This order comes from abandoned cart confirmed by %s, please do recount if needed.\n\n"%(self.sale_id.user_id.name)
            for record in self:
                if not record.tracking_number_spt and record.shipping_id and record.shipping_id.name.lower() != 'pick up':
                    raise UserError('Please add tracking number for ship order.')
                record.update_sale_order_spt()
                product_dict = {}
                # product record with qty
                for line in record.move_ids_without_package.sorted(lambda x: x.product_id.variant_name):
                    if line.product_id in product_dict.keys():
                        product_dict[line.product_id ] = line.quantity_done + product_dict[line.product_id ]
                    else:
                        product_dict[line.product_id ] = line.quantity_done

                for product_id in product_dict.keys():
                    if product_dict[product_id] > product_id.qty_available or product_id.qty_available < 0.00 :
                        error_message = error_message+("[%s]-%s you are trying to ship %s qty from the potential qty %s.\n")%(product_id.barcode,product_id.name,str(int(product_dict[product_id])),str(int(product_id.qty_available)))

            if error_message:
                if self.sale_id.is_confirm_by_saleperson:
                    error_message = website_error_msg + error_message
                raise UserError(_(error_message))

            quant_obj = self.env['stock.quant']
            if not self.shipping_id:
                raise ValidationError("Please add shipping details before order shipped.")
            ## check if picking have more then quantity then order
            custom_warning = False
            for line in range(len(self.move_ids_without_package)):
                line = self.move_ids_without_package[line]
                if line.sale_line_id:
                    if line.quantity_done > line.product_uom_qty:
                        custom_warning = True
                else:
                    custom_warning = True if line.quantity_done > 0 else False
                if line.product_id.virtual_available < line.reserved_availability:
                    line_ids = line.move_line_ids.mapped(lambda move_line: move_line if move_line.state not in ['done','cancel'] else None)
                    if line_ids:
                        available_quantity = quant_obj._get_available_quantity(line_ids[0].product_id, line_ids[0].location_id, None, None, None, False)
                        quants = quant_obj._gather(line_ids[0].product_id, line_ids[0].location_id, lot_id=None, package_id=None, owner_id=None, strict=False)
                        
                        quantity = -line_ids[0].product_qty
                        rounding = line_ids[0].product_id.uom_id.rounding
                        if float_compare(quantity, 0, precision_rounding=rounding) > 0:
                            # if we want to reserve
                            if float_compare(quantity, available_quantity, precision_rounding=rounding) > 0:
                                self._cr.execute('''
                                    Update stock_move_line set product_qty = 0.00 where id =%s
                                ''',[str(line_ids[0].id)])
                                self._cr.commit()
                        elif float_compare(quantity, 0, precision_rounding=rounding) < 0:
                            # if we want to unreserve
                            available_quantity = sum(quants.mapped('reserved_quantity'))
                            if float_compare(abs(quantity), available_quantity, precision_rounding=rounding) > 0:
                                self._cr.execute('''
                                        Update stock_move_line set product_qty = 0.00 where id =%s
                                    ''',[str(line_ids[0].id)])
                                self._cr.commit()
                                    
            if custom_warning:
                message = "You are shipping more than what was initially ordered, please click on UPDATE ORDER button before ship."
                raise ValidationError(message)
            res =  super(stock_picking, self).button_validate()

            template_id = self.env.ref('tzc_sales_customization_spt.tzc_order_shipped_notification_to_salesperson_spt')
            template_id.with_context({'salesperson_notify':True}).send_mail(self.id,force_send=True)
            # template_id.with_context({'default_attachment_ids':[(6,0,[attchment_id.id])]}).send_mail(self.id,force_send=True)
            # Check backorder should check for other barcodes
            immediate_transfer_view_id = self.env.ref('stock.view_immediate_transfer').id
            # if not self.sale_id:
            #     self.sale_id = self.get_order_id(self)
            if res and 'view_id' in res.keys() and self._check_backorder():
                self.env['stock.backorder.confirmation'].create({
                'pick_ids': [(4, self.id)]
                        }).process_cancel_backorder()
                self.mapped('sale_id').write({'state': 'shipped'})
                self.write({'state': 'done'})
                self._cr.commit()
                self.set_order_status()
                self._cr.commit()
                return 

            self.mapped('sale_id').write({'state': 'shipped'})
            self.write({'state': 'done'})
            self._cr.commit()
            self.set_order_status()
            self._cr.commit()
            return True
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }

    def _send_confirmation_email(self):
        for stock_pick in self.filtered(lambda p:p.picking_type_id.code == 'outgoing'):
            if stock_pick.partner_id and stock_pick.partner_id.user_ids and stock_pick.sale_id.partner_verification():
                delivery_template_id = stock_pick.company_id.stock_mail_confirmation_template_id.id
                stock_pick.with_context(force_send=True).message_post_with_template(delivery_template_id, email_layout_xmlid='mail.mail_notification_light')

    def get_scanned_product(self):
        scaned_product = {}
        for rec in self:
            scaned_product = ast.literal_eval(rec.preiviews_scanning_products_data) if rec.preiviews_scanning_products_data else {}
            scaned_product.update({rec.id:{}})
            for line in rec.move_ids_without_package:
                scaned_product[rec.id].update({line.product_id.id:line.quantity_done})
        
        return scaned_product
    


    def create_address_line_for_sale(self,source_id,take_name=False):
        address = ''
        if take_name == True:
            if source_id.name:
                address += str(source_id.name)+'\n'+'(A division of Tanzacan Tradelink Inc.)'
            if source_id.street:
                address += '\n'+str(source_id.street)
        else:
            if source_id.street:
                address += '\n'+'(A division of Tanzacan Tradelink Inc.)'
                address += '\n'+str(source_id.street) 
        if source_id.street2:
            address += '\n'+str(source_id.street2)
        if source_id.city:
            address+= '\n'+str(source_id.city)
        if source_id.state_id:
            address += '\n'+str(source_id.state_id.name)
        if source_id.zip:
            address += '\n'+source_id.zip
        if source_id.phone:
            address += '\nTel:-'+ source_id.phone
        if source_id.email:
            address += '\nEmail:-'+ source_id.email 
        
        return address

    def create_update_qty_log(self,product_id,product_data):
        self.env['update.qty.log'].create({
            'product_default_code':product_id.default_code,
            'created_date':datetime.now(),
            'user_id':self.env.user.id,
            'origin_order_id':self.sale_id.id,
            'before_qty_on_hand':product_data['before_qty_on_hand'],
            'before_available_qty':product_data['before_available_qty'],
            'before_reserved_qty':product_data['before_reserved_qty'],
            'after_qty_on_hand':product_id.qty_available,
            'after_available_qty':product_id.available_qty_spt,
            'after_reserved_qty':product_id.reversed_qty_spt,
        })

    def action_cancel(self):
        for record in self:
            for line in range(len(record.move_ids_without_package)):
                line = record.move_ids_without_package[line]
                line_ids = line.move_line_ids.mapped(lambda move_line: move_line if move_line.state not in ['done','cancel'] else None)
                if line_ids:
                    self._cr.execute('''
                        Update stock_move_line set product_qty = 0.00 where id =%s
                    ''',[str(line_ids[0].id)])
            # record.action_assign()
        return super(stock_picking,self).action_cancel() 

    def action_update_order_cases(self):
        for record in self:
            record.show_update_button = False
            sale_order = record.sale_id
            if not sale_order:
                sale_order = self.env['sale.order'].search([('picking_ids','in',record.ids)],limit=1)
            sale_order.include_cases = record.include_cases
            sale_order.no_of_cases = record.no_of_cases

    def check_duplicate_move(self,line):
        self.ensure_one()
        self._cr.execute("""select id from stock_move where picking_id = {} and product_id={}  order by id""".format(self.id,line.product_id.id))
        move_data = self._cr.fetchall()

        move_id = self.env['stock.move'].browse([id[0] for id in move_data])
        if move_id:
            if len(move_id)>1 :
                self.env['kits.double.move.log'].genarete_double_move_log(move_id[0].product_id,self,move_id)

                move_id_list = []
                for move in move_id:
                    if (move.quantity_done ==0 and move.product_uom_qty) or  (move.quantity_done ==0 and not move.product_uom_qty):
                        self._cr.execute("""delete from stock_move_line where move_id in ({})""".format(move.id))
                        self._cr.execute("""delete from stock_move where id={}""".format(move.id))
                    else:
                        if move.state == 'done':
                            self._cr.execute("""delete from stock_move_line where move_id in ({})""".format(move.id))
                            self._cr.execute("""delete from stock_move where id={}""".format(move.id))
                        else:
                            move_id_list.append(str(move.id))
                if len(move_id_list)>1:
                    move_id_list.sort()
                    self._cr.execute("""delete from stock_move_line where move_id in ({})""".format(','.join(move_id_list[1:])))
                    self._cr.execute("""delete from stock_move where id in ({})""".format(','.join(move_id_list[1:])))

                move_id= move_id.filtered(lambda move : move.exists())
        return move_id

    def get_picking_order_values(self):
        for rec in self:
            if rec.sale_id:
                rec.sale_id.write({
                    'include_cases':rec.include_cases,
                    'no_of_cases':rec.no_of_cases,
                    'shipping_id':rec.shipping_id.id,
                    'carrier_id':rec.carrier_id.id,
                    'kits_carrier_tracking_ref':rec.tracking_number_spt,
                    'glass_weight_kg':rec.shipping_weight,
                    'case_weight_kg':rec.weight_of_cases,
                    'weight_total_kg':rec.weight_total_kg,
                    'actual_weight':rec.actual_weight,
                    'estimate_shipping_cost':rec.calulate_shipping_cost
                })


    def get_wh_user(self):
        user_id = self.env['res.users'].search([('is_warehouse','=',True)],limit=1)
        return user_id.email


    @api.model
    def _fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(stock_picking, self)._fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar, submenu=submenu)
        if view_type == 'form':
            doc = etree.fromstring(res['arch'])
            is_admin = self.env.user.has_group('base.group_system')
            is_warehouse = self.env.user.is_warehouse
            if is_admin or is_warehouse:
                for button_reset in doc.xpath("//button[@name='action_reset_to_inscanning']"):
                    button_reset.attrib['invisible'] = '0'
            res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    # def is_accessible_to(self,user):
    #     self = self.sudo()
    #     self.ensure_one()
    #     result = False
    #     if user:
    #         if user.is_warehouse or user.is_salesperson or user.is_sales_manager or user.has_group('base.group_system '):
    #             result = True
    #     return result

    def write(self,vals):
        update = self.env['ir.model']._updated_data_validation(field_list,vals,self._name)
        if update:
            vals.update({'updated_by':self.env.user.id,'updated_on':datetime.now()})
        return super(stock_picking,self).write(vals)

    def name_get(self):
        name = []
        for rec in self:
            rec_name = '%s (%s)'%(rec.origin,rec.name)
            name.append((rec.id,rec_name))
        return name
    
    def action_delivery_restore(self):
        self.ensure_one()
        if self.state not in ['confirmed','in_scanning']:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }
        else:
            return {
                "name":_("Recover Delivery"),
                "type":"ir.actions.act_window",
                "res_model":"delivery.recovery.selection.wizard",
                "view_mode":"form",
                "target":"new",
                "context":{'default_sale_id':self.sale_id.id}
            }
    
    def button_open_package_scan(self):
        self.ensure_one()
        return {
            'name':_("Scan Packages"),
            'type':'ir.actions.act_window',
            'res_model':'kits.scan.package.products',
            'view_mode':'form',
            'view_id':self.env.ref('tzc_sales_customization_spt.kits_scan_package_products_form_view').id,
            'context':{'default_picking_id':self.id},
            'target':'new'
        }
    
    def button_open_remove_package(self):
        self.ensure_one()
        return {
            'name':_("Scan Packages"),
            'type':'ir.actions.act_window',
            'res_model':'kits.remove.package.products',
            'view_mode':'form',
            'view_id':self.env.ref('kits_package_product.kits_remove_pacakge_product_form_view').id,
            'context':{'default_picking_id':self.id},
            'target':'new'
        }