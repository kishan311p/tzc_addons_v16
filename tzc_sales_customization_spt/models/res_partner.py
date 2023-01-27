# -*- coding: utf-8 -*-
# Part of SnepTech See LICENSE file for full copyright and licensing details.##
##############################################################################
from odoo import models, fields, api, _
import re
import base64
import requests
import json
from lxml import etree
import os
from io import BytesIO
from odoo.modules import get_module_path
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from bs4 import BeautifulSoup
from operator import itemgetter

from odoo.addons.auth_signup.models.res_partner import now
from odoo.exceptions import UserError, ValidationError
from datetime import datetime

field_list = ['customer_type','company_type','name','parent_id','type','street','street2','city','state_id','country_id','zip','l10n_ca_pst',
              'customer_sales_rank','previous_total_sales','business_fax','business_type','function','phone','email','mobile','website','title',
              'category_id','internal_flag_id','eto','business_type_ids','is_customer','is_vendor','mail_notification','child_ids','user_id','property_delivery_carrier_id',
              'team_id','property_payment_term_id','property_account_position_id','property_stock_customer','property_stock_supplier','property_supplier_payment_term_id',
              'ref','webiste_id','bank_ids','property_account_receivable_id','property_account_payable_id','comment','image_1920','is_email_verified']

class res_partner(models.Model):
    _inherit = 'res.partner'

    internal_id = fields.Char('Internal ID ')
    previous_total_sales = fields.Float('Previous Total Sales')
    customer_sales_rank = fields.Float('Customer Sales Rank')
    internal_flag_id = fields.Many2one('internal.flag.spt',string='Internal Flag')
    eto = fields.Selection([('fs','FS'),('registered','REG')],string='ETO')
    business_type_ids = fields.Many2many('business.type.spt','business_type_partner_rel','partner_id','business_id',string='Business Type')
    customer_type = fields.Selection([('b2c','B2C'),('b2b_regular','B2B-Regular'),('b2b_fs','B2B-Fs')],default="b2c",tracking=True,string='Customer Type')
    sales_person_id= fields.Char('Sales Person ID',compute='_get_sales_person_id')
    business_fax = fields.Char('Business fax')
    business_type = fields.Char('Type of Business')
    contact_name_spt = fields.Char('Contact Name')
    last_logged_on = fields.Datetime('Last Logged On',compute="_get_partner_data")
    last_order_date = fields.Datetime('Last Order Date',compute="_compute_info_fields",store=True)
    last_order_id = fields.Many2one('sale.order','Last Order',compute="_compute_info_fields",store=True)
    last_order_value = fields.Char('Last Order Value',compute="_compute_info_fields",store=True)
    total_invoiced_count = fields.Integer('Invoices Count',compute="_compute_info_fields",store=True)
    user_state = fields.Selection([('new', 'Never Connected'), ('active', 'Confirmed')],'State ',compute="_compute_info_fields",store=True)
    is_customer = fields.Boolean('Is Customer')
    is_vendor = fields.Boolean('Is Vendor')
    is_email_verified = fields.Boolean(compute="_compute_info_fields",string='User Verified',store=True)
    sale_order_count = fields.Integer(compute='_compute_sale_order_count', string='Sale Order Count',store=True)
    email = fields.Char(track_visibility='onchange')
    country_id = fields.Many2one('res.country', string='Country', ondelete='restrict',track_visibility="onchange")
    internal_contacts_ids = fields.Many2many('res.users','internal_contact_res_users_rel','internal_contacts_id','res_users_id','Designated Salespersons',compute="_get_partner_data",compute_sudo=True)
    designeted_country_ids = fields.Many2many('res.country','res_partner_res_country_rel','res_partner_id','res_country_id','Sales Manager\'s Designated Countries',compute="_get_partner_data",compute_sudo=True)
    notify_salesperson_country_ids = fields.Many2many('res.country','notify_salesperson_country_res_country_rel','notify_salesperson_country_id','res_country_id','Notify Salesperson Country',compute="_get_partner_data",compute_sudo=True)
    notify_salesperson = fields.Boolean(compute="_notify_salesperson")
    is_visitor_connected = fields.Boolean('Visitor Connected')
    visiter_id = fields.Many2one('website.visitor','Website Visitor Partner')
    is_salesmanager = fields.Boolean(default=False)

    is_granted_portal_access = fields.Boolean('Is Granted Portal Access',compute="_compute_info_fields",store=True)
    is_salesperson = fields.Boolean(compute="_get_partner_data",default=False,compute_sudo=True)
    mail_notification = fields.Boolean('Email Notification',default=True)
    territory = fields.Many2one('res.country.group','Territory',compute="_get_partner_data")
    access_field_flag = fields.Boolean("Access Field Flag",compute="_get_partner_data",default=True,compute_sudo=True)
    updated_on = fields.Datetime('Updated On')
    updated_by = fields.Many2one('res.users','Updated By')
    website_id = fields.Many2one('website','Website')
    result = fields.Char("Result")
    fail_reason = fields.Char("Fail Reason")
    mail_risk = fields.Char("Mail Risk")
    catalog_ids = fields.Many2many('sale.catalog', string='Catalogs')
    catalog_count = fields.Integer('Number of catalog',compute="_catalog_count",store=True)
    name_get_partner = fields.Boolean(compute='_get_partner_name',string='Flag')
    is_user_internal = fields.Boolean("Is User Internal",store=True,compute="_compute_is_user_internal")
    is_internal_user = fields.Boolean(string="Is Internal User")
    signup_from_website = fields.Boolean(string='Sinup From Website')
    mailgun_verification_status = fields.Selection([('approved','Mg Approved'),('rejected','MG Rejected')],'MG Status')
    
    # wk_website_loyalty_points = fields.Float(
    #     string='Website Loyalty Points',
    #     help='The points are the points with which the user is awarded of being Loyal !',
    #     digits=dp.get_precision('Loyalty Points'),
    #     default=0
    # )
    # b2b_pricelist_id = fields.Many2one(comodel_name='product.pricelist',string="Pricelist")
    # preferred_currency= fields.Many2one(comodel_name='res.currency',string="Preferred Currency")
    # b2b_wishlist_count = fields.Integer('Wishlist Count',compute='_compute_b2b_wishlist_count')
    # b2b_recent_view_count = fields.Integer('Recent View Count',compute='_compute_b2b_wishlist_count')
    # promo_code_ids = fields.Many2many('sale.coupon.program','sale_coupon_program_partner_real','partner_id','sale_coupon_id','Coupon',compute="_compute_get_coupons")
    temp_pricelist_id = fields.Many2one('product.pricelist', 'Temp Pricelist')
    # catalog_ids = fields.Many2many('sale.catalog', string='Catalogs')
    # catalog_count = fields.Integer('Number of catalog',compute="_catalog_count",store=True,compute_sudo=True)

    def _get_default_lang(self):
        lang = self.env['res.lang'].search([('code','=','en_US')],limit=1)
        return lang.code

    def _get_all_data(self):
        return [[code, name] for code,_, name,active,image in self.env['res.lang'].get_available()]

    kits_lang = fields.Selection(_get_all_data,string='Language',default=_get_default_lang)

    def write(self,vals):
        user_obj = self.env['res.users']
        config_parameter_obj = self.env['ir.config_parameter'].sudo()
        if 'active' in vals.keys() and (vals['active'] == False or vals['active']):
            if not self.env.user.has_group('base.group_system') and not self.env.user.name == 'Public user':
                if self._context.get('params') and not self._context.get('params').get('model') == 'sale.catalog':
                    raise UserError(_('Due to security restrictions, you are not allowed to "%s" this record \n Contact your administrator to request access if necessary.'%('Unarchive' if vals['active'] else "Archive")))

        if 'country_id' in vals.keys():
            canada_country_id = self.env.ref('base.ca').id
            usd_currency_id = self.env.ref('base.USD').id
            cad_currency_id = self.env.ref('base.CAD').id
            cad_public_pricelist = self.env.ref('product.list0')
            usd_public_pricelist = self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt')
            for partner in self:
                if vals['country_id']== canada_country_id:
                    if partner.property_product_pricelist.currency_id.id != cad_currency_id:
                        vals.update({'property_product_pricelist':cad_public_pricelist.id})
                else:
                    if partner.property_product_pricelist.currency_id.id != usd_currency_id:
                        vals.update({'property_product_pricelist':usd_public_pricelist.id})
        if 'email' in vals.keys() and vals['email']:
            vals.update({'email':vals['email'].lower()})        
        
        update = self.env['ir.model']._updated_data_validation(field_list,vals,self._name)
        if update:
            vals.update({'updated_by':self.env.user.id,'updated_on':datetime.now()})
        # return super(res_partner, self).write(vals)
        if self.ids:
            customers = self.filtered(lambda record: record.customer_rank)
            if len(customers) > 0:
                self.env['customer.index'].create({'updated': ','.join([str(x) for x in customers.ids])})
        
        for partner in self:  
            if 'customer_type' in vals.keys() and partner.customer_type == 'b2c' and vals['customer_type'] in ['b2b_regular']:
                config_parameter = config_parameter_obj.sudo().get_param('user_ids_spt', False)
                user_ids =user_obj.search([('id','in',eval(config_parameter)+partner.user_id.ids)])
                # approve email for customer
                self.env.ref('tzc_sales_customization_spt.tzc_mail_template_customer_approve_notify_spt').sudo().send_mail(partner.id,force_send=True,email_values={'partner_ids':[(6,0,partner.ids)]})
                # approve email for salesperson and admin
                self.env.ref('tzc_sales_customization_spt.tzc_mail_template_customer_approve_notify_salesperson_spt').sudo().send_mail(partner.id,force_send=True,email_values={'partner_ids':[(6,0,user_ids.mapped('partner_id').ids)]})
            if self._context.get('set_user_spt'):
                vals['user_id'] = partner.user_id.id
            if partner.customer_rank or partner.is_customer:
                vals['is_customer'] = True
                if  not partner.customer_rank:
                    vals['customer_rank'] = 1
            if 'email' in vals:
                if partner.user_ids:
                    partner.user_ids[0].login = vals['email']
            
            if 'user_id' in vals.keys() and not vals['user_id']:
                    del vals['user_id']
                    if not partner.user_id: 
                        raise UserError('Salespreson is required.')        
        res = super(res_partner,self).write(vals)
        for rec in self:
            if 'email' in vals.keys() and vals['email']:
                mailing_contact_id = self.env['mailing.contact'].search([('odoo_contact_id','=',self.id)])
                # mailing_contact_id.mailgun_verification = False
                mailing_contact_id.mailgun_verification_status = 'rejected'
                # rec.mailgun_verification = False
                rec.mailgun_verification_status = 'rejected'
                
        return res


    @api.depends('user_ids','user_ids.is_email_verified','user_ids.login_date','user_ids.state','sale_order_ids','sale_order_ids.date_order','sale_order_ids.amount_total')
    def _compute_info_fields(self):
        for record in self:
            # record.last_logged_on = record.user_ids[0].login_date if record.user_ids else False
            record.last_order_date = record.sale_order_ids[0].date_order if record.sale_order_ids else False
            record.last_order_id = record.sale_order_ids[0].id if record.sale_order_ids else False
            record.last_order_value = record.sale_order_ids[0].amount_total if record.sale_order_ids else False
            record.total_invoiced_count = len(record.invoice_ids)
            record.user_state = record.user_ids[0].state if record.user_ids else False
            record.is_email_verified = record.user_ids[0].is_email_verified if record.user_ids else False
            record.is_granted_portal_access = True if len(record.user_ids) > 0 else False

    
    def _get_sales_person_id(self):
        for record in self:
            record.sales_person_id = record.user_id.partner_id.internal_id

    def _get_partner_data(self):
        for rec in self:
            rec.internal_contacts_ids = self.env['res.users'].browse(rec.user_ids.allow_user_ids.ids)
            rec.designeted_country_ids = self.env['res.country'].browse(rec.user_ids.contact_allowed_countries.ids)
            rec.notify_salesperson_country_ids = self.env['res.country'].browse(rec.user_ids.country_ids.ids)
            rec.is_salesperson = rec.user_ids.is_salesperson
            rec.territory = rec.user_ids.territory.id
            # visitor_id = self.env['website.visitor'].search([('partner_id','=',rec.id)],limit=1)
            # rec.last_logged_on = visitor_id.last_connection_datetime
            if rec.check_partner_access_right():
                # rec.catalog_count = len(rec.catalog_ids)
                # rec.sale_order_count = len(rec.sale_order_ids)
                rec.access_field_flag = True
            else:
                rec.sale_order_count = 0.0
                rec.catalog_count = 0.0
                rec.access_field_flag = False

        self.env['res.partner'].search([('is_salesmanager','=',True)]).sudo().write({'is_salesmanager':False})
        self.env.ref('tzc_sales_customization_spt.group_sales_manager_spt').users.mapped('partner_id').sudo().write({'is_salesmanager':True})
    
    def check_partner_access_right(self):
        is_accessible = False
        if self.user_ids:
            customer_ids = self.search([('user_id','=', self.env.user.id)])
            if self.user_ids and (self.user_ids[0].has_group('base.group_portal') or self.user_ids[0].has_group('base.group_public')):
                is_accessible = True
            else:
                if self.env.user.has_group('base.group_system') or self.env.user.has_group('tzc_sales_customization_spt.group_marketing_user'):
                    is_accessible = True
                elif self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt'):
                    if self.id in self.env.user.allow_user_ids.mapped('partner_id').ids or self.id == self.env.user.partner_id.id or self.id in customer_ids.ids:
                        is_accessible = True
                elif self.env.user.has_group('sales_team.group_sale_salesman'):
                    if self.id in customer_ids.ids or self.id == self.env.user.partner_id.id:
                        is_accessible = True
        else:
            is_accessible = True

        return is_accessible

    @api.depends('sale_order_ids')
    def _compute_sale_order_count(self):
        # super(res_partner,self)._compute_sale_order_count()
        for rec in self:
            # if rec.check_partner_access_right():
            rec.sale_order_count = len(rec.sale_order_ids)
            # else:
            #     rec.sale_order_count = 0.0


    
    @api.depends('catalog_ids')
    def _catalog_count(self):
        for record in self:
            # if record.check_partner_access_right():
            record.catalog_count = len(record.catalog_ids)
            # else:
            #     record.catalog_count = 0.0

    @api.onchange('is_customer')
    def _onchange_customer_rank(self):
        for record in self:
            if record.is_customer:
                record.customer_rank = 1

    @api.onchange('customer_rank')
    def _onchange_is_customer(self):
        for record in self:
            if record.customer_rank > 0:
                record.is_customer = True

    @api.onchange('is_vendor')
    def _onchange_supplier_rank(self):
        for record in self:
            if record.is_vendor:
                record.supplier_rank = 1

    @api.onchange('supplier_rank')
    def _onchange_is_vendor(self):
        for record in self:
            if record.supplier_rank > 0:
                record.is_vendor = True
                

    @api.onchange('country_id')
    def _onchange_country_spt(self):
        canada_country_id = self.env.ref('base.ca').id
        cad_public_pricelist = self.env.ref('product.list0')
        usd_public_pricelist = self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt')
        for record in self:
            if record.country_id.id == canada_country_id:
                record.property_product_pricelist = cad_public_pricelist.id
            else:
                record.property_product_pricelist = usd_public_pricelist.id
        # Country Change Warning
        if self._origin.id:
            orders = self.sale_order_ids.filtered(lambda x: x.state not in ('cancel','merged','draft_inv','open_inv','paid'))
            message = 'You have modified the country to "%s" for client "%s".%s'%(self.country_id.name,self.name,'\nCurrency, Pricelist and Fiscal position of following orders will be changed.   Please Review following #orders: %s'%(','.join(orders.mapped('name'))) if orders else '')
            if not self.country_id:
                message = 'You removed the country of client "%s"'%(self.name)
            return {
                'warning':{
                    "title":'Warning',
                    "message":message
                    }
                }
    
    
    @api.onchange('email')
    def _onchange_email(self):
        if self._origin.id:
            message = 'You have changed email to "%s" of customer "%s".\nSo the old email address will not be valid for login, Please notify customer for the same.'%(self.email,self.name)
            if not self.email:
                message = 'You removed email of customer "%s".'%(self.name)
            return {
                "warning":{
                    "title":"Warning",
                    "message":message,
                }
            }

    def update_salesperson(self):
        if self.check_partner_access_right():
            orders = self.sale_order_ids.filtered(lambda x: x.state not in ('cancel','paid','open_inv','draft_inv'))
            message = 'Are you sure? You want to change the salesperson of customer %s.%s'%(self.name,'\nPlease review following orders if you want to change salesperson or any other details\n%s'%(','.join(orders.mapped('name'))) if orders else '')
            return {
                'name':_("Assign Salesperson"),
                'view_mode': 'form',
                'view_type': 'form',
                'res_model': 'kits.assign.salesperson.wizard',
                'type': 'ir.actions.act_window',
                'target': 'new',
                'context' : {
                    'default_message':message,
                    'default_partner_id' : self.id,
                }
            }
        else:
            raise UserError('You can\'t change your superior salesperson.')


    def action_change_contact_country(self):
        self.ensure_one()
        if self.check_partner_access_right():
            return {
                'name':_('Chagne Country'),
                'type':'ir.actions.act_window',
                'res_model':'kits.change.contact.country',
                'view_mode':'form',
                'context':{'default_partner_id':self.id},
                'target':'new',
            }
        else:
            raise UserError('You can\'t change your superior country.')

    # def action_website_activity(self):
    #     if self.check_partner_access_right():
    #         web_visitor_id = self.env['website.visitor'].search([('partner_id','=',self.id)])
    #         if web_visitor_id:
    #             return {
    #                 "name":_("Website Activity"),
    #                 "type":"ir.actions.act_window",
    #                 "res_model":"website.visitor",
    #                 "view_mode":"form",
    #                 "res_id":web_visitor_id.id,
    #             }
    #         else:
    #             raise UserError(_('Customer %s has not visited any website pages.') %(self.name))
    #     else:
    #         raise UserError(_('You can\'t check your superior activities.'))

    def action_customer_order(self):
        tree_view_id = self.env.ref('sale.view_order_tree').id
        form_view_id = self.env.ref('sale.view_order_form').id

        sale_order_ids = self.env['sale.order'].search([('user_id','=',self.user_ids.id)])

        return {
            'name' : _('Order'),
            'view_mode': 'tree,form',
            'view_type': 'form',
            'views':[(tree_view_id,'tree'),(form_view_id,'form')],
            'domain':[('id','in',sale_order_ids.ids)],
            'res_model':'sale.order',
            'type':'ir.actions.act_window'
        }

    def action_customer(self):
        # pass
        tree_view_id = self.env.ref('tzc_sales_customization_spt.tzc_res_partner_inherit_tree_view_spt').id
        form_view_id = self.env.ref('base.view_partner_form').id

        partner_ids = self.env['res.partner'].search([('user_id','in',self.user_ids.ids)])

        return {
            'name': _('Customers'),
            'view_mode': 'tree,form',
            'view_type':'form',
            'views':[[tree_view_id,'tree'],[form_view_id,'form']],
            'domain': [('id','in',partner_ids.ids)],
            'res_model': 'res.partner',
            'type': 'ir.actions.act_window'
        }


    def partner_approved_spt(self):
        # return self.env.ref('tzc_sales_customization_spt.tzc_mail_template_customer_approve_spt').send_mail(self.id,force_send=True)
        self.ensure_one()
        ir_model_data = self.env['ir.model.data']
        if self.email:
            template_id = ir_model_data._xmlid_lookup('tzc_sales_customization_spt.tzc_mail_template_customer_approve_spt')[2]
            ctx = ({
                    'default_model': 'res.partner',
                    'default_res_id': self.id,
                    'default_use_template': bool(template_id),
                    'default_template_id': template_id,
                    'custom_layout': 'mail.mail_notification_light',
                    'force_email': True,
                })

            return {
                'name': _('Send Mail'),
                'type': 'ir.actions.act_window',
                'res_model': 'mail.compose.message',
                'binding_model':"res.partner",
                'view_mode' : 'form',
                'target': 'new',
                'context':ctx,
            }
        else:
            raise UserError('Please check customer email.')

    
    def action_reset_password(self):
        for record in self:
            if record.check_partner_access_right():
                if record.user_ids:
                    record.user_ids[0].sudo().action_reset_password()
                else:
                    raise UserError(_("%s's no user found.")%(record.name) )
            else:
                raise UserError('You can\'t change your superior password.')


    def action_verify_email(self):
        success,failed = [],[]
        for rec in self:
            res = rec.user_ids.sudo().write({'is_email_verified':True}) if rec.user_ids and rec.user_ids.ids else False
            if res:
                success.append(rec.id)
            else:
                failed.append(rec.id)
        success_message = "From %s contacts, %s contact's email has been successfully verified."%(len(self),len(success))
        failed_messsage = "%s contact's email failed to verify due to not having Portal Access."%(len(failed))
        return {
            'name':_("Verify Email"),
            "type":"ir.actions.act_window",
            "view_mode":"form",
            "res_model":"warning.spt.wizard",
            "view_id":self.env.ref('tzc_sales_customization_spt.res_partner_email_verify_warnig_wizard_spt_form_view').id,
            'context':{'default_success_partner_ids':[(6,0,success)],'default_failed_partner_ids':[(6,0,failed)],'default_verify_mail_success':success_message,'default_verify_mail_failed':failed_messsage},
            "target":"new",
        }

    @api.model
    def _get_view(self, view_id=None, view_type='form',**options):
        fields_list = ['name','email','phone','mobile','same_vat_partner_id']
        button_list = ['action_view_opportunity','schedule_meeting',str(self.env.ref('sale.act_res_partner_2_sale_order').id),'action_view_partner_invoices','action_catalogs']
        arch,view = super(res_partner, self)._get_view(view_id=view_id, view_type=view_type,**options)
        if view_type == 'form':
            doc = arch
            is_admin = not self.env.user.has_group('base.group_system')
            is_manager = self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt')
            if is_admin:
                for email_node in  doc.xpath('//field[@name="email"]'):
                    email_node.attrib['readonly'] = '1' if not is_manager else '0'
                for country_id_node in  doc.xpath('//field[@name="country_id"]'):
                    country_id_node.attrib['readonly'] = '1' if not is_manager else '0'
                for company_type_node in  doc.xpath('//field[@name="company_type"]'):
                    company_type_node.attrib['readonly'] = '1'
                for is_vendor_node in  doc.xpath('//field[@name="is_vendor"]'):
                    is_vendor_node.attrib['readonly'] = '1'
                for is_customer_node in  doc.xpath('//field[@name="is_customer"]'):
                    is_customer_node.attrib['readonly'] = '1'
                for user_id in doc.xpath('//field[@name="user_id"]'):
                    user_id.attrib['readonly'] = '1' if not is_admin or not is_manager else '0'
            for country_id in doc.xpath('//field[@name="country_id"]'):
                country_id.attrib['readonly'] = '1'
            for state_id in doc.xpath('//field[@name="state_id"]'):
                state_id.attrib['readonly'] = '1'
            str_xml = etree.tostring(doc, encoding='unicode')
            if not self._context.get('resend') or not self._context.get('campaign') or not self._context.get('raise_campaign'):
                soup = BeautifulSoup(str_xml, "html.parser")
                fields = soup.find_all('field')
                notebook = soup.find_all('notebook')
                buttons = soup.find_all('button')
                for field in fields:
                    if field.attrs.get('name') not in fields_list:
                        field['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for book in notebook:
                    book['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for btn in buttons:
                    if btn.attrs.get('name') in button_list:
                        btn['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for div in soup.find_all('div'):
                    if div.attrs.get('role') == 'status':
                        div['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for label in soup.find_all('label'):
                    if label.attrs.get('name') == 'address_name':
                        label['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                soup.find_all('group')[1]['attrs'] = "{'invisible':[('access_field_flag','=',False)]}"
                arch = etree.fromstring(str(soup))
        return arch,view
    
    def _get_partner_name(self):
        for rec in self:
            partner_name = []
            if rec.name:
                partner_name.append(rec.name)
            if rec.city:
                partner_name.append(rec.city)
            if rec.country_id:
                partner_name.append(rec.country_id.name)

            name = ', '.join(partner_name)
            rec.name_get_partner = True
            rec.display_name = name if name else None

    def _compute_signup_url(self):
        """ proxy for function field towards actual implementation """
        result = self.with_context(signup_force_type_in_url='reset').sudo()._get_signup_url_for_action()
        for partner in self:
            if any(u.has_group('base.group_user') for u in partner.user_ids if u != self.env.user):
                self.env['res.users'].sudo().check_access_rights('write')
            partner.signup_url = result.get(partner.id, False)

    def _notify_salesperson(self):
        for rec in self:
            rec.notify_salesperson = False
            if self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt') or self.env.user.is_salesperson:
                rec.notify_salesperson = True

    @api.constrains('email')
    def _check_email(self):
        partner_obj = self.env['res.partner']
        for partner in self:
            if partner.email:
                partners = partner_obj.search([('email','=',partner.email.lower()),('id','!=',partner.id)])
                if len(partners) >0:
                    message = "This email is already assigned to "+ partners[0].display_name
                    raise UserError(_(message))

    @api.constrains('email')
    def _check_email_validation(self):
        for record in self:
            single_email_re = re.compile(r"""^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\D{2,4})+$""", re.VERBOSE)
            if record.email and not single_email_re.match(record.email):
                raise UserError(_('Invalid Email! Please enter a valid email address.'))

    @api.constrains('internal_id')
    def _check_internal_id(self):
        partner_obj = self.env['res.partner']
        for partner in self:
            if partner.internal_id:
                partner_count = partner_obj.search_count([('internal_id','=',partner.internal_id)])
                if partner_count >1:
                    raise UserError(_('Internal id must be uniue'))
        
    @api.constrains('property_product_pricelist', 'country_id')
    def _check_pricelist(self):
        canada_country_id = self.env.ref('base.ca').id
        usd_currency_id = self.env.ref('base.USD').id
        cad_currency_id = self.env.ref('base.CAD').id
        for partner in self:
            if 'bypass_validation_spt' in self.env.context.keys() and self.env.context['bypass_validation_spt'] or partner.parent_id:
                pass
            else:
                if partner.country_id:
                    if partner.country_id.id == canada_country_id and partner.property_product_pricelist.currency_id.id != cad_currency_id:
                        raise UserError(_('Canadian customer must have pricelist having CAD as currency'))
                    if partner.country_id.id != canada_country_id and partner.property_product_pricelist.currency_id.id != usd_currency_id:
                        raise UserError(_('Non-Canadian customer must have pricelist having USD as currency'))


    @api.model
    def default_get(self, fields):
        user_obj = self.env['res.users']
        config_parameter_obj = self.env['ir.config_parameter']
        vals = super(res_partner, self).default_get(fields)
        vals.update({
            'user_id': self.env.user.id
            })        
        if 'user_id' in vals.keys():
            user_id = user_obj.browse(vals['user_id'])
            if user_id and not user_id.is_salesperson and not self:
                config_parameter = config_parameter_obj.sudo().get_param('default_sales_person_id', False)
                if config_parameter:
                    vals['user_id'] = eval(config_parameter)

        usd_public_pricelist = self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt')
        vals['property_product_pricelist'] = usd_public_pricelist.id
        
        return vals

    @api.model
    def create(self,vals):
        if self.env.user.has_group('base.group_user') and not self.env.user.has_group('tzc_sales_customization_spt.group_partner_access_salesperson'):
            raise ValidationError("You can not create partner !!!")
        user_obj = self.env['res.users']
        config_parameter_obj = self.env['ir.config_parameter'].sudo()
        self = self.with_context(bypass_validation_spt=True)
        canada_country_id = self.env.ref('base.ca').id
        if 'internal_id' not in vals.keys() or 'internal_id' in vals.keys() and not vals['internal_id']:
            internal_id = self.env['ir.sequence'].next_by_code('tzc.partner.internal.id.seq.spt')
            vals.update({'internal_id':internal_id})
        if 'email' in vals.keys() and vals['email']:
            vals.update({
                'email':vals['email'].lower()
            })
        if 'user_id' in vals.keys() and not vals['user_id']:
            vals.update({
                'user_id': eval(config_parameter_obj.sudo().get_param('default_sales_person_id', 'False'))
            })
            self = self.with_context(stop_assign_mail=True)
        
        if 'type' in vals and vals['type'] != 'contact':
            self = self.with_context({'mail_create_nosubscribe':True,'tracking_disable':True})
        
        res = super(res_partner,self).create(vals)
        if not self.env.user.has_group('base.group_user'):
            res.company_type = 'company'
        if res.country_id and res.country_id.id != canada_country_id:
            usd_public_pricelist = self.env.ref('tzc_sales_customization_spt.usd_public_pricelist_spt')
            res.property_product_pricelist = usd_public_pricelist.id
        else:
            cad_public_pricelist = self.env.ref('product.list0')
            res.property_product_pricelist = cad_public_pricelist.id
        
        for rec in res:
            if rec.email:
                    api_key = self.env['ir.config_parameter'].sudo().get_param('mailgun.webhook.signin.key',False)
                    params = {"address":rec.email}
                    request = requests.get("https://api.mailgun.net/v4/address/validate",auth=("api", api_key),params=params).json()
                    if request.get('result') == 'undeliverable':
                        rec.fail_reason = request.get('reason')[0]
                    else:
                        # rec.mailgun_verification = True
                        rec.mailgun_verification_status = 'approved'
                    rec.result = request.get('result')
                    rec.mail_risk = request.get('risk')

        if 'customer_type' in vals.keys() and vals['customer_type'] in ['b2b_regular','b2b_regular']:
                config_parameter = config_parameter_obj.sudo().get_param('user_ids_spt', False)
                user_ids =user_obj.search([('id','=',eval(config_parameter)+res.user_id.ids)])
                self.env.ref('tzc_sales_customization_spt.tzc_mail_template_customer_approve_notify_spt').sudo().send_mail(res.id,force_send=True,email_values={'partner_ids':[(6,0,res.ids)]})
                # self.env.ref('tzc_sales_customization_spt.tzc_mail_template_customer_approve_notify_salesperson_spt').sudo().send_mail(res.id,force_send=True,email_values={'partner_ids':[(6,0,user_ids.mapped('partner_id').ids)]})
        if res.customer_rank or res.is_customer:
                res._onchange_is_customer()
                res._onchange_customer_rank()

        if res.customer_rank:
            self.env['customer.index'].create({'created': str(res.id)})

        return res

    def unlink(self):
        res = super(res_partner,self).unlink()
        if not self.env.context.get('confirm_delete') and (not self.env.user.has_group('base.group_system') or not self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt')):
            raise UserError(_('Due to security restrictions, you are not allowed to access "Contact" (res.partner) records \n Contact your administrator to request access if necessary.'))
        if self.ids:
            customers = self.filtered(lambda record: record.customer_rank)
            if len(customers) > 0:
                self.env['customer.index'].create({'deleted': ','.join([str(x) for x in customers.ids])})
        return res
    
    def action_grant_portal_access_spt(self):
        # catalog_obj = self.env['sale.catalog']
        # catalog_obj.connect_server()
        # method = catalog_obj.get_method('action_grant_portal_access_spt')
        # if method['method']:
        #     localdict = {'self':self}
        #     exec(method['method'], localdict)
        # portal_users = localdict['portal_users']
        # partner_ids  = localdict['partner_ids']
        # temp  = localdict['temp']
        # user_changes = localdict['user_changes']

        user_changes = []
        partner_ids = []
        contact_ids = set()
        portal_users = 'These contacts are alredy user: <br/>'
        temp = False
        for partner in self:
            contact_partners = partner.child_ids.filtered(lambda p: p.type in ('contact', 'other')) | partner
            for contact in contact_partners:
                # make sure that each contact appears at most once in the list
                if contact.id not in contact_ids:
                    contact_ids.add(contact.id)
                    is_portal = False
                    if contact.user_ids:
                        is_portal = self.env.ref('base.group_portal') in contact.user_ids[0].groups_id
                    user_changes.append((0, 0, {
                        'partner_id': contact.id,
                        'email': contact.email,
                        'is_portal': is_portal,
                    }))
                    if contact.user_ids:
                        temp = True
                        portal_users = portal_users + contact.name +'<br/>'
                        partner_ids.append(contact.id) 
        if temp:
            return {
                'name': '',
                'view_mode': 'form',
                'target': 'new',
                'context':{'default_name': portal_users,'default_partner_ids': partner_ids,'default_user_details':user_changes},
                'res_model': 'portal.users.message.wizard.spt',
                'type': 'ir.actions.act_window',
            }
        return {
                'name': 'Grant portal access',
                'view_mode': 'form',
                'target': 'new',
                'context':{'default_user_ids':user_changes},
                'res_model': 'portal.wizard',
                'type': 'ir.actions.act_window',
            }

    def action_open_pricelist_wizard(self):
        wizard_id = self.env['pricelist.partner.wizard.spt'].create({
            'partner_ids' : [(6,0,self.ids)]
        })
        return {
            'name': _('Set Pricelist'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'pricelist.partner.wizard.spt',
            'res_id': wizard_id.id,
            'target': 'new',
        }

    @api.model
    def _fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        fields_list = ['name','email','phone','mobile']
        button_list = ['action_view_opportunity','schedule_meeting',str(self.env.ref('sale.act_res_partner_2_sale_order').id),'action_view_partner_invoices','action_catalogs']
        res = super(res_partner, self)._fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar, submenu=submenu)
        if view_type == 'form':
            doc = etree.fromstring(res['arch'])
            is_admin = not self.env.user.has_group('base.group_system')
            is_manager = self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt')
            if is_admin:
                for email_node in  doc.xpath('//field[@name="email"]'):
                    email_node.attrib['readonly'] = '1' if not is_manager else '0'
                for country_id_node in  doc.xpath('//field[@name="country_id"]'):
                    country_id_node.attrib['readonly'] = '1' if not is_manager else '0'
                for company_type_node in  doc.xpath('//field[@name="company_type"]'):
                    company_type_node.attrib['readonly'] = '1'
                for is_vendor_node in  doc.xpath('//field[@name="is_vendor"]'):
                    is_vendor_node.attrib['readonly'] = '1'
                for is_customer_node in  doc.xpath('//field[@name="is_customer"]'):
                    is_customer_node.attrib['readonly'] = '1'
                for user_id in doc.xpath('//field[@name="user_id"]'):
                    user_id.attrib['readonly'] = '1' if not is_admin or not is_manager else '0'
            for country_id in doc.xpath('//field[@name="country_id"]'):
                country_id.attrib['readonly'] = '1'
            for state_id in doc.xpath('//field[@name="state_id"]'):
                state_id.attrib['readonly'] = '1'
            res['arch'] = etree.tostring(doc, encoding='unicode')
            if not self._context.get('resend') or not self._context.get('campaign') or not self._context.get('raise_campaign'):
                soup = BeautifulSoup(res['arch'], "html.parser")
                fields = soup.find_all('field')
                notebook = soup.find_all('notebook')
                buttons = soup.find_all('button')
                for field in fields:
                    if field.attrs.get('name') not in fields_list:
                        field['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for book in notebook:
                    book['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for btn in buttons:
                    if btn.attrs.get('name') in button_list:
                        btn['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for div in soup.find_all('div'):
                    if div.attrs.get('role') == 'status':
                        div['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                for label in soup.find_all('label'):
                    if label.attrs.get('name') == 'address_name':
                        label['attrs'] =  "{'invisible':[('access_field_flag','=',False)]}"
                soup.find_all('group')[1]['attrs'] = "{'invisible':[('access_field_flag','=',False)]}"
                res['arch'] = str(soup)
        return res
    
    def action_open_eto_wizard(self):
        wizard_id = self.env['eto.partner.wizard.spt'].create({
            'partner_ids' : [(6,0,self.ids)]
        })
        return {
            'name': _('Set ETO'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'eto.partner.wizard.spt',
            'res_id': wizard_id.id,
            'target': 'new',
        }
    @api.depends('user_ids','user_ids.is_internal_user')
    def _compute_is_user_internal(self):
        for record in self:
            record.is_user_internal = record.user_ids.has_group('base.group_user') if len(record.user_ids) == 1 else record.user_ids[0].has_group('base.group_user') if len(record.user_ids) > 1 else False

    @api.depends('signup_token', 'signup_expiration')
    def _compute_signup_valid(self):
        for partner, partner_sudo in zip(self, self.sudo()):
            partner.signup_valid = bool(partner_sudo.signup_token)

    def give_required_sale_orders_names(self,orders,states):
        success_orders = []
        for order in orders:
            if order.state not in states:
                success_orders.append(order.name)
        return success_orders

    @api.constrains('user_id')
    def _check_salesperson(self):
        for record in self:
            if not self.user_id:
                path = get_module_path('tzc_sales_customization_spt')
                directory = 'Logs'
                dir_path = path+"/"+directory
                f_name = 'salesperson_change.log'
                final_path = '/'.join([dir_path,f_name])
                dir_available = os.path.exists(dir_path)
                mode = 'w'
                try:
                    if not dir_available:
                        try:
                            os.system('mkdir %s'%(dir_path))
                        except PermissionError:
                            os.system('chmod 777 {}'.format(dir_path))
                            os.system('mkdir %s'%(dir_path))
                    if os.path.exists(final_path):
                        mode = 'a+'
                    message = "{} {} Contact {}'s salesperson has been removed. {}".format(datetime.now(),self.env.user.name,record.name,record._context)
                    with open(final_path,mode) as f:
                        f.write("\n"+message)
                except Exception as e:
                    message = "{} {} {}".format(datetime.now(),self.env.user.name,str(e))
                    with open(final_path,mode) as f:
                        f.write("\n"+message)

    def action_contact_mailgun_verification(self):
        mailing_contact_obj = self.env['mailing.contact']
        # for rec in self.filtered(lambda x:x.email and not x.mailgun_verification):
        for rec in self.filtered(lambda x:x.email and x.mailgun_verification_status == 'rejected'):
            rec.fail_reason = False
            rec.result = False
            rec.mail_risk = False
            mailing_contact_id = mailing_contact_obj.search([('email','=',rec.email)],limit=1)
            if mailing_contact_id:
                # if mailing_contact_id.mailgun_verification:
                if mailing_contact_id.mailgun_verification_status == 'approved':
                    # rec.mailgun_verification = True
                    rec.mailgun_verification_status = 'approved'
                    rec._cr.commit()
                else:
                    mail_verify = mailing_contact_obj.email_verification(mailing_contact_id.email)
                    if mail_verify.get('result') == 'deliverable':
                        # mailing_contact_id.mailgun_verification = True
                        mailing_contact_id.mailgun_verification_status = 'approved'
                        # rec.mailgun_verification = True
                        rec.mailgun_verification_status = 'approved'
                    else:
                        rec.fail_reason = mail_verify.get('reason')[0]
                    rec.result = mail_verify.get('result')
                    rec.mail_risk = mail_verify.get('risk')
                    rec._cr.commit()
            else:
                rec.action_partner_assign_tp_mailing_cantact()
                mail_con_id = mailing_contact_obj.search([('email','=',rec.email)],limit=1)
                email_verify = mailing_contact_obj.email_verification(mail_con_id.email)
                if email_verify.get('result') == 'deliverable':
                    # mail_con_id.mailgun_verification = True
                    mail_con_id.mailgun_verification_status = 'approved'
                    # rec.mailgun_verification = True
                    rec.mailgun_verification_status = 'approved'
                else:
                    rec.fail_reason = email_verify.get('reason')[0]
                rec.result = email_verify.get('result')
                rec.mail_risk = email_verify.get('risk')
                rec._cr.commit()

    def action_partner_assign_tp_mailing_cantact(self):
        mailing_contact_obj = self.env['mailing.contact']
        sale_order_obj = self.env['sale.order']
        success_partner_ids = []
        mailing_contact_list = []
        failed_partner_ids=[]
        tags_obj = self.env['res.partner.category']
        for partner in self:
            if partner.email:
                try:
                    mailing_contact_id = mailing_contact_obj.search([('email','=',partner.email)])
                    tag_ids = []
                    for tag in partner.category_id:
                        tag_id = tags_obj.sudo().search(
                            [('name', '=', tag.name)])
                        if not tag_id:
                            tag_id = tags_obj.create(
                                {'name': tag.name})
                        tag_ids.append(tag_id.id)
                    orders = len(sale_order_obj.search([('partner_id','=',partner.id)]))

                    mailing_contact_dict = {
                            'email': partner.email,
                            'odoo_contact_id' : partner.id,
                            'status_type' : partner.customer_type,
                            'action_type': 'confirmed' if partner.user_state == 'active' else 'not_connected',
                            'source' : 'odoo_contact',
                            'country_id': partner.country_id.id if partner.country_id else None,
                            'orders': orders
                        }
                    if not mailing_contact_id:
                        mailing_contact_dict['name'] = partner.name
                        mailing_contact_dict['company_name']= partner.company_name
                        mailing_contact_dict['internal_id']= partner.internal_id
                        mailing_contact_dict['phone']= partner.phone
                        mailing_contact_dict['street']= partner.street
                        mailing_contact_dict['street2']= partner.street2
                        mailing_contact_dict['city']= partner.city
                        mailing_contact_dict['zip']= partner.zip
                        mailing_contact_dict['website']= partner.website
                        mailing_contact_dict['state_id']= partner.state_id.id if partner.state_id else None,
                        mail_contact = mailing_contact_obj.create(mailing_contact_dict)
                        mailing_contact_list.append(mail_contact.id)
                    else:
                        mailing_contact_id.write(mailing_contact_dict)
                        mailing_contact_list.append(mailing_contact_id.id)

                    
                    success_partner_ids.append(partner.id)
                    
                except Exception as e:
                    failed_partner_ids.append(partner.id)
            else:
                failed_partner_ids.append(partner.id)

        success_message = "From the %s contacts %s contacts are sucessfully exported to mailing contacts"%(len(self),len(success_partner_ids))
        failed_message = '%s contacts are failed due to the wrong email.'%(len(failed_partner_ids))
        return {
                'name': _('Contacts Messages'),
                'type': 'ir.actions.act_window',
                'res_model': "error.message.wizard.spt",
                'view_type': 'form',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_message':success_message,
                    "default_failed_message":failed_message,
                    'default_failed_partner_ids':[(6,0,failed_partner_ids)],
                    'default_contact_ids': [(6,0,mailing_contact_list)],
                }
            }

    def action_change_password_of_related_user(self):
        if self.check_partner_access_right():
            users = self.env['res.users'].search([('partner_id','in',self.ids)])
            change_password_wiz = self.env['change.password.wizard'].create({
                'user_ids':[(0,0,{'user_id':user.id,'user_login':user.login}) for user in users]
                })
            return {
                "name":_('Change Password'),
                "type":"ir.actions.act_window",
                "res_model":'change.password.wizard',
                "view_mode":"form",
                'res_id':change_password_wiz.id,
                "target":"new"
            }
        else:
            raise UserError('You can\'t change your superior password.')

    def action_contact_mass_mailing(self):
        # rejected_partner_ids = self.filtered(lambda x: x.mailgun_verification_status == 'rejected').ids
        # if rejected_partner_ids:
        #     return self.mailgun_varified()
        # partner_ids = self.filtered(lambda x: not x.email).ids
        # mail_context = self._context.copy()
        # message = 'From the %s contacts %s contact%s have no email.'%(len(self),len(partner_ids),'s' if len(partner_ids) > 1 else '')
        # if partner_ids:
        #     mail_context.update({
        #                 'default_partner_ids':partner_ids,
        #                 'default_email_partner_ids':self.ids,
        #                 'default_message':message,
        #             })
        #     return {
        #             'name': _('Mass Mailing Message Wizard'),
        #             'type': 'ir.actions.act_window',
        #             'res_model': "mass.mailing.message.wizard",
        #             'view_type': 'form',
        #             'view_mode': 'form',
        #             'target': 'new',
        #             'context':mail_context,
        #         }
        # else:
        #     mail_context.update({
        #                     'default_composition_mode': 'mass_mail',
        #                     'default_partner_to': ','.join(str(id.id) for id in self),
        #                     'default_use_template': True,
        #                     'default_template_id': self.env.ref('mail.email_template_partner').id,
        #                     'default_no_auto_thread':False,
        #                     'campaign' : True,
        #                     'default_mail_server_id':eval(self.env['ir.config_parameter'].sudo().get_param('mass_mailing.mail_server_id')),
        #                     'raise_campaign':True
        #                 })
        #     return {
        #         'name': _('Send Mail'),
        #         'type': 'ir.actions.act_window',
        #         'res_model': 'mail.compose.message',
        #         'binding_model':"res.partner",
        #         'view_mode' : 'form',
        #         'target': 'new',
        #         'context':mail_context,
        #     }
        partner_ids = self.filtered(lambda x: x.mailgun_verification_status == 'approved' and x.email).ids
        none_mails_partner_ids = self.filtered(lambda x: not x.email).ids
        # mail_context = self._context.copy()
        # message = 'Out of %s customer %s customer is eligible for mail.'%(len(self),len(partner_ids))
        message = 'Out of %s customers %s customer has been sent an email successfully.'%(len(self),len(partner_ids))
        return {
            'name': _('Send Mail'),
            'type': 'ir.actions.act_window',
            'res_model': 'mail.compose.message',
            'binding_model':"res.partner",
            'view_mode' : 'form',
            'target': 'new',
            'context':{
                        'default_model': 'res.partner',
                        'default_partner_to': ','.join(str(id) for id in partner_ids),
                        'default_template_id': self.env.ref('tzc_sales_customization_spt.email_template_partner').id,
                        'default_composition_mode': 'mass_mail',
                        'campaign':True,
                        'raise_campaign':True,
                        'none_mails_partner_ids':none_mails_partner_ids,
                        'email_partner_ids':self.ids,
                        'wiz_message':message,
                        'verify_partner_ids':partner_ids,
                        'default_mail_server_id':eval(self.env['ir.config_parameter'].sudo().get_param('mass_mailing.mail_server_id')),
                        'active_model':'mass.mailing.message.wizard',
                    },
        }

    def action_delete_partner(self):
        restricted_contact_ids = []
        not_allow_for_delete = []
        allow_contact_for_delete = []
        # deleted_data= False
        error_message = 'Due to security restrictions, you are not allowed to delete this contact.'
        if len(self) > 1:
            # list_view_data = []
            # deleted = 0
            # archived = 0
            for rec in self:
                data = False
                if rec.user_ids and rec.user_ids[0].has_group('base.group_user'):
                    contacts = self.search([('user_id','=',rec.user_ids[0].id)])
                    if contacts:
                        restricted_contact_ids.append(rec.id)
                    else:
                        if self.env.user.has_group('bsae.group_system'):
                            # data = rec.delete_contact()
                            allow_contact_for_delete.append(rec.id)
                        else:
                            not_allow_for_delete.append(rec.id)
                else:
                    # data = rec.delete_contact()
                    allow_contact_for_delete.append(rec.id)
                
                # if data:
                #     list_view_data.append(data[0][0])
                #     deleted += data[2]
                #     archived += data[1]

            # deleted_data = list_view_data,archived,deleted
        elif len(self) == 1:
            user_id = self.user_ids[0] if self.user_ids else False
            if user_id and user_id.has_group('base.group_user'):
                if self.env.user.has_group('base.group_system'):
                    contact_ids = self.search([('user_id','=',user_id.id)])
                    if user_id.is_warehouse:
                        return {
                                    'name':_("Contact Delete Confirmation"),
                                    'view_mode': 'form',
                                    'view_type': 'form',
                                    'res_model': 'kits.warning.message.wizard',
                                    'type': 'ir.actions.act_window',
                                    'target': 'new',
                                    'context' : {
                                        'default_message' : 'Are you sure you want to delete "%s" contact ?'%(self.name),
                                        'default_partner_ids' : [(6,0,contact_ids.ids)],
                                        'default_selected_id' : self.id,
                                        'total_count':len(self)
                                    }
                                }
                    else:
                        if contact_ids:
                            return {
                                    'name':_("Contact Delete"),
                                    'view_mode': 'form',
                                    'view_type': 'form',
                                    'res_model': 'kits.assign.salesperson.wizard',
                                    'type': 'ir.actions.act_window',
                                    'target': 'new',
                                    'context' : {
                                        'default_partner_id' : self.id,
                                        'default_partner_ids' : [(6,0,contact_ids.ids)],
                                        'default_hide_button':True,
                                        'default_message':'The selected contact is assigned in contact as a salesperson, Please change below contacts salesperson.',
                                        'total_count':len(self),
                                    }
                                }
                        else:
                            allow_contact_for_delete.append(self.id)
                            # deleted_data = self.delete_contact()
                else:
                    raise UserError(error_message)
            else:
                allow_contact_for_delete.append(self.id)
                # deleted_data = self.delete_contact()
            
        if allow_contact_for_delete or not_allow_for_delete or restricted_contact_ids:
        # if deleted_data:
            message = (f'Out of {len(self)} contacts {len(allow_contact_for_delete)} will be deleted.')
            # message = (f'Out of {len(self)} contacts {deleted_data[2]} is deleted and following {deleted_data[1]} is archived')
            return {
                    'name':_('Delete Contact'),
                    'type':'ir.actions.act_window',
                    'res_model':'kits.confirm.contact.delete.wizard',
                    'view_mode':'form',
                    'context':{'default_partner_ids':[(6,0,allow_contact_for_delete)],'default_message':message,'default_internal_contacts':[(6,0,restricted_contact_ids+not_allow_for_delete)]},
                    # 'context':{'default_partner_ids':[(6,0,deleted_data[0])],'default_message':message,'default_internal_contacts':[(6,0,restricted_contact_ids+not_allow_for_delete)]},
                    # 'context':{'default_partner_ids':[(6,0,deleted_data[0])],'default_message':message,'default_internal_contacts':[(6,0,restricted_contact_ids)],'default_error_partners':[(6,0,not_allow_for_delete)]},
                    'target':'new',
                }


        # user_obj = self.env['res.users'].sudo()
        # order_contacts = self.env['res.partner']
        # action = {
        #     'name':_('Confirm Delete'),
        #     'type':'ir.actions.act_window',
        #     'res_model':'kits.confirm.contact.delete.wizard',
        #     'view_mode':'form',
        #     'context':{'default_partner_ids':[(6,0,self.ids)]},
        #     'target':'new',
        # }
        # if not self._context.get('restrict_delete_rules'):
        #     is_manager = self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt')
        #     is_admin = self.env.user.has_group('base.group_system')
        #     designated_country_ids = self.env.user.contact_allowed_countries
        #     salespersons = self.env.user.allow_user_ids
        #     internal_contacts = False
        #     if is_manager or is_admin:
        #         internal_contacts = self.filtered(lambda x: x.is_user_internal)
        #         self = self - internal_contacts
        #         for order_partner in self.env['sale.order'].search([('partner_id','in',self.ids)]).mapped('partner_id'):
        #             order_contacts |= order_partner
        #         for picking_partner in self.env['stock.picking'].search([('partner_id','in',self.ids)]).mapped('partner_id'):
        #             order_contacts |= picking_partner
        #         for invoice_partner in self.env['account.move'].search([('partner_id','in',self.ids)]).mapped('partner_id'):
        #             order_contacts |= invoice_partner
        #         self = self - order_contacts
        #     if is_manager or is_admin:
        #         ctx = {'default_partner_ids':self.ids}
        #         if internal_contacts:
        #             ctx.update({'default_internal_contacts':[(6,0,internal_contacts.ids)]})
        #         if order_contacts:
        #             ctx.update({'default_order_partners':[(6,0,order_contacts.ids)]})
        #         action.update({'context':ctx})
        #     return action
        # else:
        #     users = user_obj.search([('login','in',self.mapped('email'))])
        #     users.unlink() if users else None
        #     mailing_contacts = self.env['mailing.contact'].search(['|',('email','in',self.mapped('email')),('odoo_contact_id','in',self.ids)])
        #     mailing_contacts.write({'source':'imported'})
        # return self.sudo().with_context(confirm_delete=True).unlink()

    def delete_contact(self):
        # deleted_ids = []
        # archived_count = 0
        # deleted_count = 0
        for rec in self:
            try:
                if rec.user_ids:
                    for user in rec.user_ids:
                        try:
                            user.active = False
                            user._cr.commit()
                            user.unlink()
                        except:
                            rec._cr.rollback()
                            pass
                rec.active = False
                rec._cr.commit()
                rec.unlink()
                # deleted_ids.append(self.id)
                # deleted_count += 1
            except:
                rec._cr.rollback()
                partner_id = rec.exists()
                # deleted_ids.append(rec.id)
                # archived_count += 1
                pass

    
    def export_customer_action(self):
        try:
            res_company_id = self.env.ref("base.main_company")
            # self.env['kits.quickbooks.backend'].with_context(backend_id=res_company_id.kits_quickbooks_backend_id).action_test_connection()
            self.env['kits.quickbooks.backend'].with_context(backend_id=res_company_id.kits_quickbooks_backend_id, product_ids=self).export_customer_action()
        except Exception as e:
            raise UserError(_(str(e)))

    # def _compute_get_coupons(self):
    #     sale_order_obj = self.env['sale.order']
    #     coupon_obj = self.env['sale.coupon.program']
    #     for record in self:
    #         promotion_name_list = []
    #         order_ids = sale_order_obj.search([('partner_id','=',record.id),('applied_coupon_ids','!=',False)]).filtered(lambda x: x if x.applied_coupon_ids else None)
    #         if order_ids:
    #             promotion_name_list = order_ids.mapped('promo_code')
    #         promotion_ids = coupon_obj.search([('promo_code','in',promotion_name_list)])
    #         record.promo_code_ids = [] if not promotion_ids else [(6,0,promotion_ids.ids)]

    def action_contact_campaign(self):
        campaign = False
        campaign = self.env['marketing.participant'].search([('model_name','=','res.partner'),('res_id','=',self.id)]).mapped('campaign_id')
        return {
            "name":_("Campaign"),
            "type":"ir.actions.act_window",
            "res_model":"marketing.campaign",
            "view_mode":"tree,form",
            'domain': [('id','in',campaign.ids)]
        }


    def action_catalogs(self):
        return {
                "name":_("Catalog"),
                "type":"ir.actions.act_window",
                "res_model":"sale.catalog",
                "view_mode":"tree",
                'domain': [('partner_ids','=',self.id)]
            }

    def signup_prepare(self, signup_type="signup", expiration=False):
        """ generate a new token for the partners with the given validity
        """
        if signup_type == 'reset' and expiration:            
            expiration = now(hours=int(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.reset_pass_expire_hours')))
        return super(res_partner,self).signup_prepare(signup_type=signup_type,expiration=expiration)

    def get_product_price(self,product,partner):
        if partner and type(partner) == int:
            partner = self.browse(partner)
            pricelist_id = partner.property_product_pricelist
            query = f"""select
                        ppi.product_id as "product_id",
                        ppi.id as "ID",
                        (select case when pro.sale_type is Null then 'False' else pro.sale_type end) as "sale_type",
                        (select case when pro.lst_price_usd is Null then 0.00 else pro.lst_price_usd end) as "lst_price_usd",
                        (select case when pro.price_wholesale is Null then 0.00 else pro.price_wholesale end) as "price_wholesale",
                        (select case when pro.price_wholesale_usd is Null then 0.00 else pro.price_wholesale_usd end) as "price_wholesale_usd",
                        (select case when pro.clearance_cad is Null then 0.00 else pro.clearance_cad end) as "clearance_cad",
                        (select case when pro.clearance_usd is Null then 0.00 else pro.clearance_usd end) as "clearance_usd",
                        (select case when pro.on_sale_cad is Null then 0.00 else pro.on_sale_cad end) as "on_sale_cad",
                        (select case when pro.on_sale_usd is Null then 0.00 else pro.on_sale_usd end) as "on_sale_usd"
                    from product_pricelist pp
                        LEFT JOIN product_pricelist_item ppi ON ppi.pricelist_id = {pricelist_id.id}
                        LEFT JOIN product_product pro ON ppi.product_id = pro.id
                    where pro.active = True and pro.id in ({','.join([str(pro) for pro in product.ids])}) and pp.id = {pricelist_id.id}"""
            print(query)
            self.env.cr.execute(query)
            data = self._cr.fetchall()
            data_dict = {}
            for i in data:
                i = list(i)
                product_id = self.env['product.product'].browse(i[0])
                pricelist_price = self.env['product.pricelist.item'].browse(i[1])
                data_dict.update({i[0]:{'sale_type':i[2],
                                        'pricelist_price':'{:,.2f}'.format(float(pricelist_price.price.split(pricelist_id.currency_id.symbol)[1].strip()))}})
                if pricelist_price.currency_id.name == 'USD':
                    if data_dict.get(i[0]):
                        data_dict.get(i[0]).update({'lst_price':'{:,.2f}'.format(i[3]),
                                                    'wholesale_price':'{:,.2f}'.format(i[5]),
                                                    'clearance_price':'{:,.2f}'.format(i[7]),
                                                    'on_sale_price':'{:,.2f}'.format(i[9])})
                if pricelist_price.currency_id.name == 'CAD':
                    if data_dict.get(i[0]):
                        data_dict.get(i[0]).update({'lst_price':round(product_id.lst_price,2),
                                                    'wholesale_price':'{:,.2f}'.format(i[4]),
                                                    'clearance_price':'{:,.2f}'.format(i[6]),
                                                    'on_sale_price':'{:,.2f}'.format(i[8])})
            return data_dict
        else:
            return {'error':'Please check passing data.'}

    def _get_contact_name(self,partner,name):
        return name
    
    def name_get(self):
        contact_name = []
        res = super(res_partner,self).name_get()
        for partner in res:
            partner_name = []
            partner_id = self.browse(partner[0])
            if partner_id:
                if partner_id.name:
                    partner_name.append(partner_id.name)
                if partner_id.city:
                    partner_name.append(partner_id.city)
                if partner_id.country_id:
                    partner_name.append(partner_id.country_id.name)

            name = ', '.join(partner_name)
            contact_name.append((partner_id.id,name))

        return contact_name
    
    def action_create_contact_report(self):
        partner_ids = self.search([]).filtered(lambda x:x.company_type == 'company')
        wizard_obj = self.env['product.info.wizard.spt']

        f_name = 'Contact Detail'
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Contacts", index=0)
        header_font = Font(name='Calibri',size='11',bold=True)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd,bottom=bd)
        left_alignment = Alignment(
            vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
        right_alignment = Alignment(
            vertical='center', horizontal='right', text_rotation=0, wrap_text=True)

        header_row = 1
        sheet.cell(row=header_row, column=1).value = 'ID'
        sheet.cell(row=header_row, column=2).value = 'Internal ID'
        sheet.cell(row=header_row, column=3).value = 'Name'
        sheet.cell(row=header_row, column=4).value = 'Is a Company'
        sheet.cell(row=header_row, column=5).value = 'Salesperson'
        sheet.cell(row=header_row, column=6).value = 'Sales Person ID'
        sheet.cell(row=header_row, column=7).value = 'Active'
        sheet.cell(row=header_row, column=8).value = 'Phone'
        sheet.cell(row=header_row, column=9).value = 'Mobile'
        sheet.cell(row=header_row, column=10).value = 'Email'
        sheet.cell(row=header_row, column=11).value = 'Street'
        sheet.cell(row=header_row, column=12).value = 'Zip'
        sheet.cell(row=header_row, column=13).value = 'State'
        sheet.cell(row=header_row, column=14).value = 'Country'
        sheet.cell(row=header_row, column=15).value = 'Business Type'
        sheet.cell(row=header_row, column=16).value = 'Customer Type'
        sheet.cell(row=header_row, column=17).value = 'ID'
        sheet.cell(row=header_row, column=18).value = 'Created on'
        sheet.cell(row=header_row, column=19).value = 'Last Modified on'
        sheet.cell(row=header_row, column=20).value = 'Last Order Value'
        sheet.cell(row=header_row, column=21).value = 'Contact Name'
        sheet.cell(row=header_row, column=22).value = 'Shipping addresses'

        sheet.cell(row=header_row, column=1).font = header_font
        sheet.cell(row=header_row, column=1).border = top_bottom_border
        sheet.cell(row=header_row, column=2).font = header_font
        sheet.cell(row=header_row, column=2).border = top_bottom_border
        sheet.cell(row=header_row, column=3).font = header_font
        sheet.cell(row=header_row, column=3).border = top_bottom_border
        sheet.cell(row=header_row, column=4).font = header_font
        sheet.cell(row=header_row, column=4).border = top_bottom_border
        sheet.cell(row=header_row, column=5).font = header_font
        sheet.cell(row=header_row, column=5).border = top_bottom_border
        sheet.cell(row=header_row, column=6).font = header_font
        sheet.cell(row=header_row, column=6).border = top_bottom_border
        sheet.cell(row=header_row, column=7).font = header_font
        sheet.cell(row=header_row, column=7).border = top_bottom_border
        sheet.cell(row=header_row, column=8).font = header_font
        sheet.cell(row=header_row, column=8).border = top_bottom_border
        sheet.cell(row=header_row, column=9).font = header_font
        sheet.cell(row=header_row, column=9).border = top_bottom_border
        sheet.cell(row=header_row, column=10).font = header_font
        sheet.cell(row=header_row, column=10).border = top_bottom_border
        sheet.cell(row=header_row, column=11).font = header_font
        sheet.cell(row=header_row, column=11).border = top_bottom_border
        sheet.cell(row=header_row, column=12).font = header_font
        sheet.cell(row=header_row, column=12).border = top_bottom_border
        sheet.cell(row=header_row, column=13).font = header_font
        sheet.cell(row=header_row, column=13).border = top_bottom_border
        sheet.cell(row=header_row, column=14).font = header_font
        sheet.cell(row=header_row, column=14).border = top_bottom_border
        sheet.cell(row=header_row, column=15).font = header_font
        sheet.cell(row=header_row, column=15).border = top_bottom_border
        sheet.cell(row=header_row, column=16).font = header_font
        sheet.cell(row=header_row, column=16).border = top_bottom_border
        sheet.cell(row=header_row, column=17).font = header_font
        sheet.cell(row=header_row, column=17).border = top_bottom_border
        sheet.cell(row=header_row, column=18).font = header_font
        sheet.cell(row=header_row, column=18).border = top_bottom_border
        sheet.cell(row=header_row, column=19).font = header_font
        sheet.cell(row=header_row, column=19).border = top_bottom_border
        sheet.cell(row=header_row, column=20).font = header_font
        sheet.cell(row=header_row, column=20).border = top_bottom_border
        sheet.cell(row=header_row, column=21).font = header_font
        sheet.cell(row=header_row, column=21).border = top_bottom_border
        sheet.cell(row=header_row, column=22).font = header_font
        sheet.cell(row=header_row, column=22).border = top_bottom_border

        row_index=header_row+1
        for partner in partner_ids:

            sale_order_ids = partner.sale_order_ids.filtered(lambda x:x.state != 'cancel')
            last_order_date = max(sale_order_ids.mapped('create_date')) if sale_order_ids else None
            order_id = partner.sale_order_ids.search([('create_date','=',last_order_date)],limit=1) if last_order_date else None
            shipping_address = []
            for address in partner.child_ids.filtered(lambda x:x.type == 'delivery'):
                delivery_address = []
                if address.street:
                    delivery_address.append(address.street)
                if address.street2:
                    delivery_address.append(address.street2)
                if address.city:
                    delivery_address.append(address.city)
                if address.state_id:
                    delivery_address.append(address.state_id.name)
                if address.country_id:
                    delivery_address.append(address.country_id.name)
                if address.zip:
                    delivery_address.append(address.zip)
                shipping_address.append(','.join(delivery_address))

            sheet.cell(row=row_index, column=1).value = partner.id
            sheet.cell(row=row_index, column=1).alignment = left_alignment
            sheet.cell(row=row_index, column=2).value = partner.internal_id or ''
            sheet.cell(row=row_index, column=3).value = partner.name or ''
            sheet.cell(row=row_index, column=4).value = 'Yes' if partner.company_type == 'company' else 'No'
            sheet.cell(row=row_index, column=5).value = partner.user_id.name or ''
            sheet.cell(row=row_index, column=6).value = partner.user_id.partner_id.internal_id or ''
            sheet.cell(row=row_index, column=7).value = 'Yes' if partner.active else 'No'
            sheet.cell(row=row_index, column=8).value = partner.phone or ''
            sheet.cell(row=row_index, column=9).value = partner.mobile or ''
            sheet.cell(row=row_index, column=10).value = partner.email or ''
            sheet.cell(row=row_index, column=11).value = ','.join([l for l in [partner.street or '', partner.street2 or ''] if l])
            sheet.cell(row=row_index, column=12).value = partner.zip or ''
            sheet.cell(row=row_index, column=13).value = partner.state_id.name or ''
            sheet.cell(row=row_index, column=14).value = partner.country_id.name or ''
            sheet.cell(row=row_index, column=15).value = ','.join(partner.business_type_ids.mapped('name')) or ''
            sheet.cell(row=row_index, column=16).value = dict(self._fields['customer_type'].selection).get(partner.customer_type) or ''
            sheet.cell(row=row_index, column=17).value = partner._get_external_ids()[partner.id][0] if partner._get_external_ids()[partner.id] else ''
            sheet.cell(row=row_index, column=18).value = partner.create_date.strftime('%d-%m-%Y') or  ''
            sheet.cell(row=row_index, column=18).alignment = left_alignment
            sheet.cell(row=row_index, column=19).value = partner.write_date.strftime('%d-%m-%Y') or ''
            sheet.cell(row=row_index, column=19).alignment = left_alignment
            sheet.cell(row=row_index, column=20).value = order_id.picked_qty_order_total if order_id and order_id.picked_qty_order_total else order_id.amount_total if order_id else 0.0
            sheet.cell(row=row_index, column=20).alignment = right_alignment
            sheet.cell(row=row_index, column=21).value = partner.contact_name_spt or ''
            sheet.cell(row=row_index, column=22).value = ' | '.join(shipping_address)
            row_index += 1
        
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 25
        sheet.column_dimensions['K'].width = 30
        sheet.column_dimensions['L'].width = 25
        sheet.column_dimensions['M'].width = 15
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['P'].width = 15
        sheet.column_dimensions['Q'].width = 20
        sheet.column_dimensions['R'].width = 17
        sheet.column_dimensions['S'].width = 17
        sheet.column_dimensions['T'].width = 15
        sheet.column_dimensions['U'].width = 20
        sheet.column_dimensions['V'].width = 40

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        wizard_id = wizard_obj.create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=product.info.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (wizard_id.id,f_name),
            'target': 'self',
        }

    def action_contact_mail_template(self):
            templates_id = self.env['mail.template'].search([('model_id','=',self.env.ref('base.model_res_partner').id)]).filtered(lambda x:list(x.get_external_id().values()) == [''])
            list_view = self.env.ref('tzc_sales_customization_spt.manage_template_tree_view')
            form_view = self.env.ref('tzc_sales_customization_spt.manage_template_form_view')
            action = {
                'name' : _('Email Templates'),
                'type' : 'ir.actions.act_window',
                'res_model' : 'mail.template',
                'view_mode' : 'tree,form',
                'views' : [(list_view.id,'tree'),(form_view.id,'form')],
                'target' : 'current',
                'domain':[('id','in',templates_id.ids)]
                # 'domain':[('model_id','=',self.env.ref('base.model_res_partner').id)]
            }
            return action
    def action_internal_contact(self):
        list_view = self.env.ref('tzc_sales_customization_spt.tzc_internal_contacts_tree_view_spt')
        form_view = self.env.ref('base.view_partner_form')
        action = {
                'name': _('Internal Contacts'),
                'type': 'ir.actions.act_window',
                'res_model': 'res.partner',
                'view_mode' : 'tree,form',
                'views': [(list_view.id,'tree'),(form_view.id,'form')],
                'target': 'current',
            }
        if self.env.user.has_group('base.group_system') or self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt') or self.env.user.has_group('tzc_sales_customization_spt.group_marketing_user') or self.env.user.is_salesperson:
            if self.env.user.has_group('base.group_system') or self.env.user.has_group('tzc_sales_customization_spt.group_marketing_user'):
                action.update({'domain': [('is_user_internal','=',True)]})
            elif self.env.user.has_group('tzc_sales_customization_spt.group_sales_manager_spt'):
                users = self.env.user.allow_user_ids
                action.update({'domain': ['|',('id','in',users.mapped('partner_id').ids+[self.env.user.partner_id.id]),('user_id','in',users.ids),('is_user_internal','=',True)]})
            elif self.env.user.is_salesperson:
                action.update({'domain':[('id','=',self.env.user.partner_id.id),('is_user_internal','=',True)]})
            
            return action
        else:
            action.update({'domain': [('id','in',[])]})
        return action
