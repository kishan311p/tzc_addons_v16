from odoo import models,fields,api,_
from odoo.exceptions import UserError
import random
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO
from odoo.tools.misc import formatLang, format_date, get_lang
from werkzeug.urls import url_encode
from lxml import etree
import requests
import json

field_list = ['sequence_name','sale_manager_id','state']

class account_move(models.Model):
    _inherit = 'account.move'

    def _get_sales_manager_domain(self):
        managers = self.env.ref('tzc_sales_customization_spt.group_sales_manager_spt').sudo().users
        return [('id','in',managers.ids)]

    def _get_default_sale_manager_id(self):
        order = self.env['sale.order'].search([('invoice_ids','in',self.ids)],limit=1)
        return order.sale_manager_id.id
    
    # quickbooks_backend_id = fields.Many2one("kits.quickbooks.backend", "Quickbooks Backend")
    # quickbooks_invoice_id = fields.Char("Quickbooks Invoice Id")

    commission_line_ids = fields.One2many('kits.commission.lines','invoice_id','Commissions')
    is_commission_paid = fields.Boolean('Paid ?')
    sale_manager_id = fields.Many2one('res.users','Sales Manager',domain=_get_sales_manager_domain,default=_get_default_sale_manager_id)

    sequence_name = fields.Char('Name Sequence')
    amount_is_admin = fields.Monetary(string='Admin Fee',store=True,compute_sudo=True,compute='_compute_amount')
    amount_is_shipping_total = fields.Monetary(string='Shipping Cost',store=True,compute_sudo=True,compute='_compute_amount')
    ordered_qty = fields.Integer('Ordered Quantity',compute="_compute_qty")
    delivered_qty = fields.Integer('Delivered Quantity',compute="_compute_qty")
    picked_qty = fields.Integer('Picked Quantity',compute="_compute_qty")
    amount_without_discount = fields.Monetary(string='Subtotal',compute_sudo=True,compute='_compute_amount', store=True, tracking=4)
    amount_discount = fields.Monetary(string='Discount',compute_sudo=True,compute='_compute_amount', store=True, tracking=4)
    global_discount = fields.Float('Additional Discount',compute_sudo=True,store=True,compute='_compute_amount')
    order_id = fields.Many2one('sale.order','Sale Order',compute_sudo=True,compute='_compute_order_id',store=True)
    report_file = fields.Binary()
    applied_promo_code = fields.Char("Applied Promo Code",related="invoice_line_ids.sale_line_ids.order_id.applied_promo_code")
    country_id = fields.Many2one('res.country','Country',related="partner_id.country_id")
    sale_manager_id = fields.Many2one('res.users','Sales Manager',domain=_get_sales_manager_domain,default=_get_default_sale_manager_id)
    street = fields.Char('Street',related='partner_id.street')
    street2 = fields.Char('Street2',related='partner_id.street2')
    city = fields.Char(related="partner_id.city")
    postal_code = fields.Char(related="partner_id.zip")
    state_id = fields.Many2one('res.country.state','State',related="partner_id.state_id")
    country_id = fields.Many2one('res.country','Country',related="partner_id.country_id")
    updated_on = fields.Datetime('Updated On  ')
    updated_by = fields.Many2one('res.users','Updated By  ')
    count_return_credit_notes = fields.Integer(compute="_count_return_credit_notes")
    is_commission_paid = fields.Boolean('Paid ?',track_visibility='onchange')
    inv_payment_status = fields.Selection([('full','Fully Paid'),('partial','Partial Paid'),('over','Over Paid')],'Payment Status',compute="_compute_inv_payment_status",copy=False)
    filtere_state = fields.Char(compute="_compute_payment_status",copy=False,store=True)
    is_admin = fields.Char(compute='_compute_is_admin', string='is_admin')
    report_token = fields.Char('Report Access Token')
    sale_order_number = fields.Char('Order Number',compute="_compute_sale_order_number")
    kits_amount_tax = fields.Float('Tax')
    kits_amount_total = fields.Float('Total')
    kits_amount_residual = fields.Float('Amount Due')

    def _compute_is_admin(self):
        for record in self:
            record.is_admin = True if self.env.user.has_group('base.group_system') else False
    
    @api.depends('invoice_line_ids','invoice_line_ids.quantity','order_id','order_id.ordered_qty','order_id.picked_qty','order_id.delivered_qty')
    def _compute_qty(self):
        for record in self:
            sale_id = record.order_id
            if not sale_id:
                sale_id = self.env['sale.order'].search([('invoice_ids','in',record.ids)])
            record.ordered_qty = sale_id.ordered_qty
            record.delivered_qty = sale_id.picked_qty
            record.picked_qty = sale_id.delivered_qty
    
    @api.depends('line_ids','line_ids.sale_line_ids')
    def _compute_order_id(self):
        for record in self:
            record.order_id = False
            if record.line_ids and record.line_ids[0].sale_line_ids:
                record.order_id = record.line_ids[0].sale_line_ids[0].order_id.id

    @api.onchange('invoice_user_id','order_id.sale_manager_id')
    def _onchange_user_id(self):
        for record in self:
            order = self.env['sale.order'].search([('invoice_ids','in',record.ids)],limit=1)
            record.sale_manager_id = order.sale_manager_id.id

    def update_name(self):
        for record in self:
            name = ''
            if record.sequence_name:
                name = record.name[0:4]+record.sequence_name+record.name[-2:]
                update_name = self.search([('name','=',name),('id','!=',record.id)])
                if update_name:
                    message = "This Sequence is already assigned to "+ update_name[0].display_name
                    raise UserError(_(message))
                else:
                    record.name = name

    def button_cancel(self):
        if self.state in ['draft','cancel']:
            self.commission_line_ids.action_cancel()
            # self.with_context(from_cancel=True).write({'is_commission_paid':False})
            sale_id = self.env['sale.order'].search([('invoice_ids','in',self.ids)],limit=1)
            if sale_id.picking_ids.filtered(lambda x: x.state != 'cancel' and 'WH/OUT' in x.name).state =='done':
                state = 'shipped'
            else:
                if sale_id.source_spt != 'Manually':
                    state = 'received'
                elif sale_id.state == 'cancel':
                    state = 'cancel'
                else:
                    state = 'draft'
            sale_id.write({'state': state})  
            return super(account_move,self).button_cancel()
        else:
            raise UserError("You cannot delete an item linked to a posted entry")
            # return {
            #     'type': 'ir.actions.client',
            #     'tag': 'display_notification',
            #     'params': {
            #             'title': 'Something is wrong.',
            #             'message': 'Please reload your screen.',
            #             'sticky': True,
            #         }
            #     }
    
    def action_post(self):
        if self.state in ['draft']:
            sale_obj = self.env['sale.order']
            for move in self:
                sale_obj.search([('invoice_ids','=',move.id),('state','!=','cancel')]).write({'state':'open_inv'})
                error = []
                # if move.order_id:
                #     if move.order_id.picked_qty_order_subtotal != move.amount_discount:
                #         error.append('Order Subtotal')   
                #     if move.order_id.picked_qty_order_total != move.amount_total:
                #         error.append('Order Total')
                #     if move.order_id.picked_qty_order_tax != move.amount_tax:
                #         error.append('Order Tax ')
                #     if move.order_id.picked_qty_order_discount != move.amount_discount:
                #         error.append('Order Discount')           
                if error:
                    raise UserError(','.join(error)+' not match with invoice,please correct it then proceed.')
            res = super(account_move,self).action_post()
            res_company_id = self.env.ref("base.main_company")         
            # Comment below to test
            # if res_company_id:
            #     try:
            #         res_company_id.kits_quickbooks_backend_id.action_test_connection()
            #         self.create_invoice(res_company_id)
            #     except Exception as e:
            #         raise UserError(e)
            self.commission_line_ids.filtered(lambda x: x.state == 'cancel').sudo().unlink()
            if not res:
                commission_line_obj = self.env['kits.commission.lines']
                user_obj = self.env['res.users']
                for record in self:
                    users = dict(saleperson=user_obj,manager=user_obj)
                    if record.invoice_user_id:
                        if record.partner_id.user_ids and record.invoice_user_id.ids != record.partner_id.user_ids.ids:
                            users['saleperson'] = record.invoice_user_id
                    if record.sale_manager_id:
                        users['manager'] = record.sale_manager_id
                    for user in users:
                        values = self._get_commission_line_detail(user,users[user])
                        if record.state == 'posted' and values and values['amount']:
                            commission_id = commission_line_obj.search([('user_id','=',users[user].id),('is_product_brand_commissison','=',False),('state','=','draft'),('commission_for','=',user),('invoice_id','=',record.id)],limit=1)
                            if not commission_id:
                                commission_line_obj.create(values)
                            else:
                                commission_id.write(values)
                    if record.state == 'posted' and any(record.line_ids.mapped('product_id.product_brand_commission')):
                        if record.partner_id.user_ids and record.invoice_user_id.ids != record.partner_id.user_ids.ids:
                            brand_commission_line_id = commission_line_obj.search([('user_id','=',record.invoice_user_id.id),('is_product_brand_commissison','=',True),('state','=','draft'),('commission_for','=','saleperson'),('invoice_id','=',record.id)],limit=1)
                            commission = 0.0
                            for line in record.line_ids.filtered(lambda x:x.product_id.product_brand_commission):
                                commission += round(round(line.discount_unit_price,2) * line.quantity * line.product_id.product_brand_commission * 0.01,2)
                            vals = self.get_product_brand_commission_vals(record,commission)
                            if brand_commission_line_id and commission:
                                brand_commission_line_id.write(vals)
                            else:
                                if commission:
                                    brand_commission_line_id.create(vals)

                    if record.commission_line_ids and record.inv_payment_status in ['full','over']:
                        record.commission_line_ids.write({'state':'paid'})
                    elif record.commission_line_ids and record.inv_payment_status == 'partial':
                        record.commission_line_ids.write({'state':'draft'})
                    else:
                        record.commission_line_ids.write({'state':'draft'})
            return res
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }

    def _get_commission_line_detail(self,type,user):
        rule = user.manager_commission_rule_id if type == 'manager' else user.commission_rule_id
        vals = {
            # 'name':'Commission of invoice {} for {} {}. '.format(self.name,type,user.name),
            'invoice_id':self.id,
            'user_id':user.id,
            'amount':round(rule.get_commission(self),2) if rule else 0.0,
            'rule_id':rule.id,
            'commission_for':type,
            'create_type':'by_system',
        }
        return vals

    def get_product_brand_commission_vals(self,record,commission):
        vals = {'create_type':'by_system',
                'invoice_id':record.id,
                'amount':commission,
                'user_id':record.invoice_user_id.id,
                'commission_for':'saleperson',
                'is_product_brand_commissison':True}
        return vals

    def write(self,vals):
        update = self.env['ir.model']._updated_data_validation(field_list,vals,self._name)
        if update:
            vals.update({'updated_by':self.env.user.id,'updated_on':datetime.now()})
        res = super(account_move,self).write(vals)
        if 'inv_payment_status' in vals and not self._context.get('from_cancel'):
            if vals['inv_payment_status'] in ['over','full']:
                self.commission_line_ids.write({'state':'paid'})
            elif vals['inv_payment_status'] == 'partial':
                self.commission_line_ids.write({'state':'draft'})
        return res

    @api.depends('inv_payment_status')
    def _compute_payment_status(self):
        for rec in self:
            rec.filtere_state = False
            if rec.inv_payment_status:
                rec.filtere_state = rec.inv_payment_status

    def _compute_inv_payment_status(self):
        for rec in self:
            order_id = self.env['sale.order'].search([('invoice_ids','in',rec.ids)],limit=1)
            rec.inv_payment_status = False
            if order_id and order_id.payment_status:
                rec.inv_payment_status = order_id.payment_status
    
    def action_cancel(self):
        for record in self:
            payment_ids = self.env['account.payment'].search([]).filtered(lambda pay: record.id in pay.reconciled_invoice_ids.ids)
            # payment_ids = self.env['account.payment'].search([('invoice_ids','in',record.id)])

            for inv in record:
                if inv.invoice_line_ids:
                    if payment_ids:
                        for payment in payment_ids:
                            payment.state='cancel'
                    inv.button_cancel()
    
    def _count_return_credit_notes(self):
        for record in self:
            order = self.env['sale.order'].search([('invoice_ids','in',record.ids)],limit=1)
            record.count_return_credit_notes = len(order.kits_credit_payment_ids)

    def action_get_return_credit_notes(self):
        notes = self.env['sale.order'].search([('invoice_ids','in',self.ids)],limit=1).kits_credit_payment_ids
        action =  {
            'name':_('Credit Payments'),
            'type':'ir.actions.act_window',
            'res_model':'account.payment',
            'view_mode':'form',
            'context':{'default_is_return_credit':True},
            'target':'self',
        }
        if len(notes) == 1:
            action['res_id'] = notes.id
        else:
            action['view_mode'] = 'tree,form'
            action['domain'] = [('id','in',notes.ids)]
        return action

    @api.depends(
        'invoice_line_ids',
        'line_ids.debit',
        'line_ids.credit',
        'line_ids.matched_debit_ids.debit_move_id.move_id.payment_id.is_matched',
        'line_ids.matched_debit_ids.debit_move_id.move_id.line_ids.amount_residual',
        'line_ids.matched_debit_ids.debit_move_id.move_id.line_ids.amount_residual_currency',
        'line_ids.matched_credit_ids.credit_move_id.move_id.payment_id.is_matched',
        'line_ids.matched_credit_ids.credit_move_id.move_id.line_ids.amount_residual',
        'line_ids.matched_credit_ids.credit_move_id.move_id.line_ids.amount_residual_currency',
        'line_ids.balance',
        'line_ids.currency_id',
        'line_ids.amount_currency',
        'line_ids.amount_residual',
        'line_ids.amount_residual_currency',
        'line_ids.payment_id.state',
        'line_ids.full_reconcile_id')
    def _compute_amount(self):
        for move in self:
            total_untaxed, total_untaxed_currency = 0.0, 0.0
            total_tax, total_tax_currency = 0.0, 0.0
            total_residual, total_residual_currency = 0.0, 0.0
            total, total_currency = 0.0, 0.0
            total_amount_is_shipping_total = 0.0
            total_amount_is_admin = 0.0
            global_discount = 0.0
            amount_without_discount = 0.0
            amount_discount = 0.0
            for line in move.line_ids:
                if move.is_invoice(True):
                    # === Invoices ===
                    if line.display_type == 'tax' or (line.display_type == 'rounding' and line.tax_repartition_line_id):
                        # Tax amount.
                        total_tax += line.balance
                        total_tax_currency += line.amount_currency
                        total += line.balance
                        total_currency += line.amount_currency
                    elif line.display_type in ('product', 'rounding'):
                        # Untaxed amount.
                        total_untaxed += line.balance
                        total_untaxed_currency += line.amount_currency
                        total += line.balance
                        total_currency += line.amount_currency
                    elif line.display_type == 'payment_term':
                        # Residual amount.
                        total_residual += line.amount_residual
                        total_residual_currency += line.amount_residual_currency
                else:
                    # === Miscellaneous journal entry ===
                    if line.debit:
                        total += line.balance
                        total_currency += line.amount_currency
            
            for line in range(len(move.invoice_line_ids)):
                line = move.invoice_line_ids[line]
                if line.product_id.is_shipping_product:
                    total_amount_is_shipping_total += line.price_subtotal
                if line.product_id.is_admin:
                    total_amount_is_admin += line.price_subtotal
                if line.product_id.is_global_discount:
                    global_discount += line.price_subtotal
                if line.product_id.type != 'service':
                    amount_discount +=  round(((line.quantity * line.price_unit) - line.price_subtotal),2)
                    amount_without_discount = round(amount_without_discount +( line.quantity * line.price_unit),2)
                    
            move.amount_is_admin = total_amount_is_admin
            move.amount_is_shipping_total = total_amount_is_shipping_total
            move.amount_without_discount = amount_without_discount
            move.global_discount = - global_discount
            move.amount_discount = move.order_id.picked_qty_order_discount
            # move.amount_discount = amount_discount + abs(global_discount)
            sign = move.direction_sign
            move.amount_untaxed = sign * total_untaxed_currency
            move.amount_tax = sign * total_tax_currency
            move.kits_amount_tax = move.order_id.picked_qty_order_tax
            move.kits_amount_total = move.order_id.picked_qty_order_total
            move.kits_amount_residual = move.order_id.picked_qty_order_total
            move.amount_total = sign * total_currency
            move.amount_residual = -sign * total_residual_currency
            move.amount_untaxed_signed = -total_untaxed
            move.amount_tax_signed = -total_tax
            move.amount_total_signed = abs(total) if move.move_type == 'entry' else -total
            move.amount_residual_signed = total_residual
            move.amount_total_in_currency_signed = abs(move.amount_total) if move.move_type == 'entry' else -(sign * move.amount_total)
            is_paid = self.env['account.payment'].search([('move_id','=',move.id)])
            in_payment_set = move.amount_residual -sum(is_paid.mapped('amount'))
            # Compute 'invoice_payment_state'.
            if move.move_type == 'entry':
                move.payment_state = False
            elif move.state == 'posted' and is_paid:
                if move.id in in_payment_set:
                    move.payment_state = 'in_payment'
                else:
                    move.payment_state = 'paid'
            else:
                move.payment_state = 'not_paid'


    @api.model
    def create(self,vals):
        res = super(account_move,self).create(vals)
        for record in range(len(res)):
            record = res[record]
            random_letter = record.random_string()
            record.name = random_letter
            if record.order_id:
                invoice_ids = self.search([('state','!=','cancel'),('id','in',record.order_id.invoice_ids.ids)])
                if len(invoice_ids) >1:
                    raise UserError(_("Once Create Invoice You Can't Create Other."))
                record.global_discount = record.order_id.global_discount
            
                if record.order_id.partner_id and record.order_id.partner_id.user_id:
                    record.user_id = record.order_id.partner_id.user_id.id
                
            if not record.sale_manager_id:
                order = self.env['sale.order'].search(['|',('name','=',record.invoice_origin),('invoice_ids','in',record.ids)],limit=1)
                record.sale_manager_id = order.sale_manager_id.id
        return res
  
    def random_string(self):
        config_parameter_obj = self.env['ir.config_parameter']
        string = ''
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        while len(string)!= 4:
            random_letter = random.choices(letters)
            if random_letter[0] not in string:
                string = string + random_letter[0]
        sequence = 0
        config_parameter_id = config_parameter_obj.search([('key','=','account_move_random_letter_spt')])
        if not config_parameter_id:
            config_parameter_id = config_parameter_obj.create({'key': 'account_move_random_letter_spt','value': sequence})
        if datetime.now().year != (datetime.now() + timedelta(days=-1)).year:
            config_parameter_id.write({'value':0})
        sequence = eval(config_parameter_id.value)
        sequence += 1 
        config_parameter_id.write({'value':sequence})
        for i in range(4-len(str(sequence))):
            string += '0'
        string += str(sequence)

        new_string = string[0:1]+str(datetime.now().year)[-2:]+string[1:2]+string[-4:]+string[2:4] 

        return new_string
    

    @api.model
    def default_get(self, default_fields):
        partner_obj = self.env['res.partner']
        rec = super(account_move, self).default_get(default_fields)
        if 'partner_id' in rec.keys():
            partner_id = False
            partner_id = partner_obj.browse(rec['partner_id'])
            if partner_id and partner_id.user_id:
                rec['user_id'] = partner_id.user_id.id
        return rec

    @api.onchange('partner_id')
    def onchange_partner_id_user_id_spt(self):
        for record in self:
            if record.partner_id and record.partner_id.user_id:
                record.user_id = record.partner_id.user_id.id
    
    
    # Not used in 13.0
    def excel_order_report(self):
        return {
                'name': 'Invoice Report',
                'view_mode': 'form',
                'target': 'new',
                'context':{'default_invoice_id':self.id,},
                'res_model': 'sale.order.report.wizard.spt',
                'type': 'ir.actions.act_window',
            }
    

    # def excel_report_line(self):
    #     workbook = Workbook()
    #     sheet = workbook.create_sheet(title='Product',index=0)

    #     bd = Side(style='thin', color="000000")
    #     all_border = Border(left=bd, top=bd, right=bd, bottom=bd)
    #     right_border = Border(right=bd)
    #     bottom_border = Border(top=bd)
    #     bottom_border = Border(top=bd)
    #     all_font = Font(size=12, bold=True)
    #     dict_data = self.calculat_data_for_invoice()
    #     row_index = 2
    #     sub_total = 0.00
    #     if dict_data:
    #         for data in dict_data:
    #             sheet.cell(row=row_index, column=1).value = dict_data[data]['categ_id']
    #             sheet.cell(row=row_index, column=2).value = dict_data[data]['qty']
    #             sheet.cell(row=row_index, column=3).value = dict_data[data]['total_discount']
    #             sheet.cell(row=row_index, column=4).value = dict_data[data]['total_price']
    #             sheet.cell(row=row_index, column=5).value = round(dict_data[data]['total_price']*dict_data[data]['qty'],2)
    #             sub_total = round(dict_data[data]['total_price']*dict_data[data]['qty'],2) + sub_total
    #             sheet.cell(row=row_index, column=5).border = right_border
    #             row_index += 1

    #         sheet.cell(row=row_index, column=1).border = bottom_border
    #         sheet.cell(row=row_index, column=2).border = bottom_border
    #         sheet.cell(row=row_index, column=3).border = bottom_border
    #         sheet.cell(row=row_index, column=4).border = bottom_border
    #         sheet.cell(row=row_index, column=5).border = bottom_border

    #         sheet.cell(row=1, column=1).value = 'Product'
    #         sheet.cell(row=1, column=2).value = 'Qty'
    #         sheet.cell(row=1, column=3).value = 'Disc.%'
    #         sheet.cell(row=1, column=4).value = 'Price'
    #         sheet.cell(row=1, column=5).value = 'Subtotal'

    #         sheet.cell(row=1, column=1).border = all_border
    #         sheet.cell(row=1, column=2).border = all_border
    #         sheet.cell(row=1, column=3).border = all_border
    #         sheet.cell(row=1, column=4).border = all_border
    #         sheet.cell(row=1, column=5).border = all_border

    #         sheet.cell(row=1, column=1).font = all_font
    #         sheet.cell(row=1, column=2).font = all_font
    #         sheet.cell(row=1, column=3).font = all_font
    #         sheet.cell(row=1, column=4).font = all_font
    #         sheet.cell(row=1, column=5).font = all_font

    #         sheet.column_dimensions['A'].width = 40
    #         sheet.column_dimensions['B'].width = 10
    #         sheet.column_dimensions['C'].width = 12
    #         sheet.column_dimensions['D'].width = 12
    #         sheet.column_dimensions['E'].width = 12

    #         sheet.column_dimensions['A'].hight = 20
    #         sheet.column_dimensions['B'].hight = 20
    #         sheet.column_dimensions['C'].hight = 20
    #         sheet.column_dimensions['D'].hight = 20
    #         sheet.column_dimensions['E'].hight = 20

    #         sheet.cell(row=row_index, column=4).border =   all_border
    #         sheet.cell(row=row_index, column=5).border =   all_border
    #         sheet.cell(row=row_index+1, column=4).border = all_border
    #         sheet.cell(row=row_index+1, column=5).border = all_border
    #         sheet.cell(row=row_index+2, column=4).border = all_border
    #         sheet.cell(row=row_index+2, column=5).border = all_border
    #         sheet.cell(row=row_index+3, column=4).border = all_border
    #         sheet.cell(row=row_index+3, column=5).border = all_border
    #         sheet.cell(row=row_index+4, column=4).border = all_border
    #         sheet.cell(row=row_index+4, column=5).border = all_border
    #         sheet.cell(row=row_index+5, column=4).border = all_border
    #         sheet.cell(row=row_index+5, column=5).border = all_border


    #         sheet.cell(row=row_index, column=4).value = 'Subtotal'
    #         sheet.cell(row=row_index, column=5).value = round(sub_total,2)
    #         sheet.cell(row=row_index+1, column=4).value = 'Shipping'
    #         sheet.cell(row=row_index+1, column=5).value = round(self.amount_is_shipping_total,2)
    #         sheet.cell(row=row_index+2, column=4).value = 'Admin Fee'
    #         sheet.cell(row=row_index+2, column=5).value = round(self.amount_is_admin,2)
    #         sheet.cell(row=row_index+3, column=4).value = 'Additional Discount'
    #         sheet.cell(row=row_index+3, column=5).value = round(self.amount_discount,2)
    #         sheet.cell(row=row_index+4, column=4).value = 'Taxes'
    #         sheet.cell(row=row_index+4, column=5).value = round(self.amount_tax,2)
    #         sheet.cell(row=row_index+5, column=4).value = 'Total'
    #         sheet.cell(row=row_index+5, column=5).value = round(sub_total + self.amount_is_shipping_total + self.amount_is_admin + self.amount_tax - self.amount_discount,2) 
            
    #         fp = BytesIO()
    #         workbook.save(fp)
    #         fp.seek(0)
    #         data = fp.read()
    #         fp.close()
    #     self.report_file = base64.b64encode(data)

    # def calculat_data_for_invoice(self):
    #     data_dict = {}
    #     total_amount = 0.0
    #     for line in range(len(self.invoice_line_ids)):
    #         line =  self.invoice_line_ids[line]
    #         total_amount =line.price_unit
    #         total_line = 0
    #         if line.product_id.type != 'service': 
    #             if line.product_id.categ_id.name in data_dict.keys():
    #                 data_dict[line.product_id.categ_id.name]['qty'] = data_dict[line.product_id.categ_id.name]['qty'] + line.quantity
    #                 data_dict[line.product_id.categ_id.name]['totat_amount'] = data_dict[line.product_id.categ_id.name]['totat_amount'] + total_amount
    #                 data_dict[line.product_id.categ_id.name]['total_discount'] = data_dict[line.product_id.categ_id.name]['total_discount'] + line.discount 
    #                 data_dict[line.product_id.categ_id.name]['total_price'] = data_dict[line.product_id.categ_id.name]['total_price'] + line.discount_unit_price
    #             else:
    #                 name ='Assorted Eyeglasses' if line.product_id.categ_id.name == 'E' else 'Assorted Sunglasses' if line.product_id.categ_id.name == 'S' else line.product_id.categ_id.name
    #                 data_dict[line.product_id.categ_id.name] = {'categ_id':name,'qty': line.quantity ,'totat_amount': total_amount,'discount_on_line': 0.0,'total_line': 0.0,'total_discount':line.discount , 'total_price': line.discount_unit_price }
    #             data_dict[line.product_id.categ_id.name]['total_line'] = data_dict[line.product_id.categ_id.name]['total_line'] + 1
    #         if line.discount:
    #             data_dict[line.product_id.categ_id.name]['discount_on_line'] = data_dict[line.product_id.categ_id.name]['discount_on_line'] + 1
    #     for data in data_dict:
    #         data_dict[data]['total_price'] = round(data_dict[data]['total_price']/data_dict[data]['total_line'] if data_dict[data]['total_line'] else 1,2)
    #         data_dict[data]['total_discount'] = round(data_dict[data]['total_discount']/data_dict[data]['discount_on_line'] if data_dict[data]['discount_on_line'] else 1,2) if round(data_dict[data]['total_discount']/data_dict[data]['discount_on_line'] if data_dict[data]['discount_on_line'] else 1,2) > 1 else 0
    #         data_dict[data]['totat_amount'] = round(data_dict[data]['totat_amount']/data_dict[data]['total_line'],2) 
        
    #     return data_dict
    

    def action_invoice_sent(self):
        """ Open a window to compose an email, with the edi invoice template
            message loaded by default
        """
        self.ensure_one()
        template = self.env.ref('account.email_template_edi_invoice', raise_if_not_found=False)
        lang = get_lang(self.env)
        if template and template.lang:
            lang = template._render_lang(self.ids)[self.id]
        else:
            lang = lang.code
        compose_form = self.env.ref('account.account_invoice_send_wizard_form', raise_if_not_found=False)
        ctx = dict(
            default_model='account.move',
            default_res_id=self.id,
            # For the sake of consistency we need a default_res_model if
            # default_res_id is set. Not renaming default_model as it can
            # create many side-effects.
            default_res_model='account.move',
            default_use_template=bool(template),
            default_template_id=template and template.id or False,
            default_composition_mode='comment',
            mark_invoice_as_sent=True,
            default_email_layout_xmlid="mail.mail_notification_light",
            model_description=self.with_context(lang=lang).type_name,
            force_email=True,
            default_is_print=False
        )
        return {
            'name': _('Send Invoice'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'account.invoice.send',
            'views': [(compose_form.id, 'form')],
            'view_id': compose_form.id,
            'target': 'new',
            'context': ctx,
        }
    

    def get_access_token_spt(self):
        self.ensure_one()
        auth_param = url_encode(self.partner_id.signup_get_auth_param()[self.partner_id.id])
        return auth_param

    # Not used in 13
    def line_ordering_by_product(self):
        product_list = []
        product_list = self.invoice_line_ids.mapped(lambda line: line.product_id.name_get()[0][1].strip()) if self.invoice_line_ids else []
        # for line in self.invoice_line_ids:
        #     product_name = line.product_id.name_get()[0][1].split('(')
        #     product_list.append(product_name[0])
        # product_list = list(set(product_list))
        product_list.sort()
        return product_list
    
    def line_product_dict(self,product_name):
        product_dict = {}
        product_dict[product_name] = {'line_ids': self.invoice_line_ids.filtered(lambda x:x.product_id.name_get()[0][1].strip() == product_name)}
        # for line in self.invoice_line_ids:
            # line_dict = {}
            # product_name = line.product_id.name_get()[0][1].split('(')
            # if line.product_id.name in product_dict.keys():
            #     product_dict[product_name[0]]['line_ids'].append(line) 
            # else:
            #     line_dict['line_ids'] = [line]
            #     product_dict[product_name[0]] = line_dict
        return product_dict

    def get_invoice_courier(self):
        self.ensure_one()
        sale_id = self.order_id
        if not sale_id:
            sale_id = self.env['sale.order'].search([('invoice_ids','in',self.ids)])
        picking = sale_id.picking_ids.filtered(lambda x: 'wh/out' in x.name.lower() and x.state == 'done')
        courier = ''
        tracking = ''
        if len(picking):
            try:
                courier = picking.shipping_id.name
                tracking = picking.tracking_number_spt
            except:
                courier = picking[0].shipping_id.name
                tracking = picking[0].tracking_number_spt
        payment_term_id = self.invoice_payment_term_id.name
        if sale_id:
            if not payment_term_id:
                payment_term_id = sale_id.payment_term_id.name
            if not payment_term_id:
                payment_term_id = self.invoice_date or sale_id.date_order.date()
        return courier,tracking,payment_term_id

    def action_invoice_to_order(self):
        list_view = self.env.ref('sale.view_order_tree')
        form_view = self.env.ref('sale.view_order_form')
        sale_order_id = self.env['sale.order'].search([('invoice_ids','in',self.ids)],limit=1)
        action = {
            "name":_("Sale Order"),
            "type":"ir.actions.act_window",
            "res_model":"sale.order",
            "target":"current",
            }
        if sale_order_id:
            action.update({
                "view_mode":"form",
                "res_id":sale_order_id.id,
            })
        else:
            action.update({
                "view_mode":"tree,form",
                "views":[(list_view.id,"tree"),(form_view.id,"form")],
                "domain":[('id','=',sale_order_id.id)],
            })
        return action

    @api.model
    def _get_view(self, view_id=None, view_type='form',**options):
        arch,view = super(account_move, self)._get_view(view_id=view_id, view_type=view_type,**options)
        if view_type == 'form':
            doc = arch
            for manager_id in doc.xpath("//field[@name='sale_manager_id']"):
                manager_id.attrib['readonly'] = '0' if self.env.user.has_group('base.group_system') else '1'
            # arch = etree.tostring(doc, encoding='unicode')
        return arch,view

    # # def is_accessible_to(self,user):
    # #     self = self.sudo()
    # #     self.ensure_one()
    # #     result = False
    # #     if user:
    # #         if user in self.partner_id.user_ids or user == self.user_id or user == self.sale_manager_id:
    # #             result = True
    # #     return result


    def button_draft(self):
        if self.state in ['posted','cancel']:
            res = super(account_move, self).button_draft()
            self.env['sale.order'].sudo().search([('invoice_ids','in',self.ids)]).write({'state': 'draft_inv'})
            res_company_id = self.env.ref("base.main_company")         
            
            # Commented below for testing. Quickbooks backend.
            # if res_company_id:
            #     try:
            #         res_company_id.kits_quickbooks_backend_id.action_test_connection()
            #         self.delete_invoice(res_company_id)
            #     except Exception as e:
            #         raise UserError(e)
            return res    
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                        'title': 'Something is wrong.',
                        'message': 'Please reload your screen.',
                        'sticky': True,
                    }
                }


    def get_html_field_val(self,value):
        val = False
        if value != '<p><br></p>':
            val = True
        
        return val


    def format_name(self,name):
        if '%' in name:
            name = name.replace('%', '%25')
        elif "'" in name:
            name = name.replace("'", "\\'")
        elif '=' in name:
            name = name.replace('=', '%3D')
        elif '<' in name:
            name = name.replace('<', '%3C')
        elif '>' in name:
            name = name.replace('>', '%3E')
        elif '&' in name:
            name = name.replace('&', '%26')
        elif '#' in name:
            name = name.replace('#', '%23')
        return name

    def _compute_sale_order_number(self):
        for rec in self:
            try:
                sale_order_id = self.env['sale.order'].search([('invoice_ids','in',self.ids)],limit=1)
                rec.sale_order_number = sale_order_id.name
            except:
                rec.sale_order_number = ""

    def ordered_qty_button(self):
        pass

    # QuickBook Backend Don't Remove
    
    # def get_customer(self,res_company_id,headers,customer_name=False):
    #     customer_id = False
    #     customer_name = self.format_name(self.partner_id.display_name if self.partner_id.display_name else customer_name)
    #     customer_url_query = "Select * from Customer where FullyQualifiedName='%s'" % (customer_name)
    #     customer_url = res_company_id.kits_quickbooks_backend_id.base_url + '/v3/company/' + res_company_id.kits_quickbooks_backend_id.company_id + '/query?query=' + customer_url_query
    #     response = requests.request("GET", customer_url, headers=headers)
    #     response_dict = json.loads(response.text)
    #     if "QueryResponse" in response_dict.keys() and response_dict.get("QueryResponse") == {}:
    #         customer = self.with_context(res_company_id=res_company_id).create_customer(headers, res_company_id)
    #         customer_id = json.loads(customer.text).get("Customer").get("Id") if json.loads(customer.text).get("Customer") else ""
    #     elif "Fault" in response_dict.keys() and response_dict.get("Fault"):
    #         error_msg = response_dict.get('Fault').get("Error")[0].get("Message") if response_dict.get('Fault').get("Error")[0] else ""
    #         raise UserError(_(f"{error_msg}"))
    #     else:
    #         if "QueryResponse" in response_dict.keys() and response_dict.get("QueryResponse"):
    #             customer_id = response_dict.get("QueryResponse").get("Customer")[0].get("Id") if response_dict.get("QueryResponse").get("Customer")[0] else ""        
    #     return customer_id
        
    # def create_invoice(self,res_company_id):
    #     try:
    #         inv_response = {}
    #         headers = {
    #             'Accept': 'application/json',
    #             'Authorization': 'Bearer '+ res_company_id.kits_quickbooks_backend_id.access_token,
    #             'Content-Type': 'application/json'
    #             }
    #         customer_id = self.get_customer(res_company_id,headers)
    #         url = res_company_id.kits_quickbooks_backend_id.base_url + '/v3/company/' + res_company_id.kits_quickbooks_backend_id.company_id + '/invoice'
    #         product_list = []
    #         for invoice_line in self.invoice_line_ids:
    #             product_name = self.format_name(invoice_line.product_id.display_name)
    #             product_url_query = "Select * from Item where Name='%s'"%(product_name)
    #             product_url = res_company_id.kits_quickbooks_backend_id.base_url + '/v3/company/' + res_company_id.kits_quickbooks_backend_id.company_id + '/query?query=' + product_url_query
    #             res = requests.request("GET", url=product_url, headers=headers)
    #             res.raise_for_status()
    #             res_dict = json.loads(res.text)
    #             new_product = False
    #             product_id = False
    #             product_name = False
    #             if "QueryResponse" in res_dict.keys() and res_dict.get("QueryResponse") == {}:
    #                 new_product = self.create_product(headers, invoice_line.product_id, res_company_id)                       
    #                 product = new_product if new_product else product
    #                 product_id =  json.loads(product.text).get("Item").get("Id") if json.loads(product.text).get("Item") else ""
    #                 product_name = json.loads(product.text).get("Item").get("Name")
    #             elif "Fault" in res_dict.keys() and res_dict.get("Fault"):
    #                 error_msg = res_dict.get('Fault').get("Error")[0].get("Message") if res_dict.get('Fault').get("Error")[0] else ""
    #                 raise UserError(_(f"{error_msg}"))
    #             else:
    #                 if "QueryResponse" in res_dict.keys() and res_dict.get("QueryResponse"): 
    #                     product_id = res_dict.get("QueryResponse").get("Item")[0].get("Id") if res_dict.get("QueryResponse").get("Item")[0] else ""
    #                     product_name = res_dict.get("QueryResponse").get("Item")[0].get("Name")
    #             sales_item_line_dict = {
    #                 "Qty": invoice_line.quantity,
    #                 "UnitPrice": invoice_line.price_unit,
    #                 "ItemRef": {
    #                     "name": product_name if product_name else invoice_line.product_id.display_name,
    #                     "value": str(product_id),
    #                     }
    #                 }
    #             if invoice_line.tax_ids and self.invoice_line_ids.mapped('tax_ids')[0].quickbooks_taxes_id:
    #                 sales_item_line_dict.update({
    #                     "Qty": invoice_line.quantity,
    #                     "UnitPrice": invoice_line.price_unit,
    #                     "ItemRef": {
    #                         "name": invoice_line.product_id.display_name,
    #                         "value": str(product_id),
    #                     },
    #                     'TaxCodeRef':{
    #                         'value': 'TAX'
    #                         },
    #                 })
    #             product_dict = {"Description": invoice_line.product_id.description, 
    #                             "DetailType": "SalesItemLineDetail", 
    #                             "SalesItemLineDetail": sales_item_line_dict,
    #                             "Amount": invoice_line.quantity * invoice_line.price_unit, 
    #                         }
    #             product_list.append(product_dict)
                                
    #         if self.amount_discount != 0:
    #             product_list.append({
    #                 'Amount': self.amount_discount, 
    #                 'DetailType': 'DiscountLineDetail', 
    #                 'DiscountLineDetail': {
    #                     'PercentBased': False, 
    #                     'DiscountAccountRef':{
    #                         "name": "Discounts given", 
    #                         "value": "86"
    #                     } ,
    #                 }
    #             })
    #         payload = {
    #                     "domain": "QBO",
    #                     "TotalAmt": self.amount_total,
    #                     "Line": product_list,
    #                     "DueDate": str(self.invoice_date_due),
    #                     "CustomField":[
    #                         {
    #                             'DefinitionId': '2', 
    #                             'Name': 'Odoo Invoice No', 
    #                             'Type': 'StringType', 
    #                             'StringValue': self.name,
    #                         }
    #                     ],
    #                     "CustomerMemo": {
    #                         "value": "Thank you for your business and have a great day!",
    #                         },
    #                     "CustomerRef": {
    #                         "name": self.partner_id.display_name, 
    #                         "value": str(customer_id),
    #                         },
    #                     "ShipAddr": {
    #                         "City": self.partner_id.city, 
    #                         "Line1": self.partner_id.street, 
    #                         "PostalCode": self.partner_id.zip, 
    #                         "Lat": "37.4238562", 
    #                         "Long": "-122.1141681", 
    #                         "CountrySubDivisionCode": self.partner_id.country_id.code  if self.partner_id.country_id.code else "US", 
    #                         }, 
    #                     "ApplyTaxAfterDiscount": False, 
    #                     } 
    #         if self.amount_tax != 0:
    #             tax_id = self.invoice_line_ids.mapped('tax_ids')[0]
    #             tax_code_id = tax_id.quickbooks_taxes_id.quickbooks_tax_code_id
    #             tax_code_list = self.find_tax_code(res_company_id,headers,tax_code_id)
    #             if tax_code_list:
    #                 tax_line_list = []
    #                 for tax_code_rate in tax_code_list[0].get("SalesTaxRateList").get("TaxRateDetail"):
    #                     tax_rate_id = tax_code_rate.get("TaxRateRef").get("value")
    #                     tax_rate = self.env['account.tax'].get_tax_rate(res_company_id,headers,tax_rate_id)
    #                     tax_line_list.append({
    #                         "DetailType": "TaxLineDetail", 
    #                         "Amount": self.amount_tax, 
    #                         "TaxLineDetail": {
    #                             "NetAmountTaxable": self.amount_total, 
    #                             "TaxPercent": tax_rate, 
    #                             "TaxRateRef": {
    #                                 "value": tax_rate_id,
    #                             }, 
    #                             "PercentBased": True
    #                         }
    #                     })
    #                 payload.update({
    #                     "TxnTaxDetail": {
    #                         "TxnTaxCodeRef": {
    #                             "value": str(tax_code_id),
    #                             "name": tax_code_list[0].get("Name"),
    #                         }, 
    #                         "TotalTax": self.amount_tax, 
    #                         "TaxLine": tax_line_list,
    #                     },
    #                 })
    #             else:
    #                 raise UserError("Quickbooks Tax Not Found!")
                
    #         if self.invoice_payment_term_id:
    #             if not self.invoice_payment_term_id.quickbooks_payment_term_id.payment_term_id:
    #                 raise UserError("Quickbooks Payment Term Not Found!")
    #             term_object = self.get_quickbooks_term(res_company_id,headers,self.invoice_payment_term_id.quickbooks_payment_term_id.payment_term_id) 
    #             payload.update({
    #                 'SalesTermRef':{
    #                     'value': term_object.get('term_id'), 
    #                     'name': term_object.get('term_name')
    #                 }
    #             })
    #         currency_object = self.get_quickbooks_currency(res_company_id,headers)
    #         if currency_object and currency_object.get("currency_code"):
    #             payload.update({"CurrencyRef": {
    #                 "name": currency_object.get('currency_name'), 
    #                 "value": currency_object.get('currency_code')
    #                 }})
    #         inv_response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
    #         inv_response.raise_for_status()
    #         self.write({
    #             'quickbooks_invoice_id':json.loads(inv_response.text).get("Invoice").get("Id"),
    #         })
    #     except Exception as e:
    #         error_msg = json.loads(inv_response.text).get('Fault').get("Error")[0].get("Detail") if (inv_response != {} and "Fault" in json.loads(inv_response.text).keys() and json.loads(inv_response.text).get('Fault')) else e
    #         raise UserError(_(f"{error_msg}"))
    #     return{
    #         'type': 'ir.actions.client',
    #         'tag': 'display_notification',
    #         'params': {
    #             'title': "Everything seems properly set up!",
    #             'message': "Invoice Created Successfully",
    #             'sticky': False,
    #             }
    #         }

    # def find_tax_code(self,res_company_id,headers,tax_code_id):
    #     tax_code_query = "Select * from TaxCode where Id = '{}'".format(tax_code_id)
    #     tax_code_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/query?query=" + tax_code_query
    #     tax_code_response = requests.request("GET",tax_code_url,headers=headers)
    #     tax_code_dict = json.loads(tax_code_response.text)
    #     if "QueryResponse" in tax_code_dict.keys() and tax_code_dict.get("QueryResponse"):
    #         return tax_code_dict.get("QueryResponse").get("TaxCode")
    #     else:
    #         return False 

    # def get_quickbooks_term(self,res_company_id,headers,term_id):
    #     term_query = "Select * from Term where Id = '{}'".format(term_id)
    #     term_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/query?query=" + term_query
    #     term_response = requests.request("GET",term_url,headers=headers)
    #     term_response.raise_for_status()
    #     term_dict = json.loads(term_response.text)
    #     if "QueryResponse" in term_dict.keys() and term_dict.get("QueryResponse") == {}:
    #         raise UserError("Payment Term Not Found in Quickbooks!")
    #     elif "Fault" in term_dict.keys() and term_dict.get("Fault"):
    #         error_msg = term_dict.get('Fault').get("Error")[0].get("Message") if term_dict.get('Fault').get("Error")[0] else ""
    #         raise UserError(_(f"{error_msg}"))
    #     else:
    #         if "QueryResponse" in term_dict.keys() and term_dict.get("QueryResponse"):
    #             term_id,term_name = term_dict.get("QueryResponse").get("Term")[0].get("Id") if term_dict.get("QueryResponse").get("Term")[0] else "",term_dict.get("QueryResponse").get("Term")[0].get("Name")    
    #     return {"term_id":term_id,"term_name":term_name}

    # def get_quickbooks_currency(self,res_company_id,headers):
    #     currency_query = "Select * from CompanyCurrency where code = '{}'".format(self.currency_id.name)
    #     currency_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/query?query=" + currency_query
    #     currency_response = requests.request("GET", currency_url, headers=headers)
    #     currency_dict = json.loads(currency_response.text)
    #     if "QueryResponse" in currency_dict.keys() and currency_dict.get("QueryResponse"):
    #         currency_code,currency_name = currency_dict.get("QueryResponse").get("CompanyCurrency")[0].get("Code"),currency_dict.get("QueryResponse").get("CompanyCurrency")[0].get("Name")
    #         return {"currency_code":currency_code,"currency_name":currency_name}
    #     elif "Fault" in currency_dict.keys() and currency_dict.get("Fault"):
    #         error_msg = currency_dict.get('Fault').get("Error")[0].get("Message") if currency_dict.get('Fault').get("Error")[0] else ""
    #         raise UserError(_(f"{error_msg}"))


    # def customer_payload(self, partner_id):
    #     fpos_id = self.env['account.fiscal.position']._get_fpos_by_region(country_id=partner_id.country_id.id, state_id=partner_id.state_id.id, zipcode=False, vat_required=False)
    #     if fpos_id:
    #         tax_id = fpos_id.tax_ids[0].tax_dest_id
    #     payload = {
    #             "FullyQualifiedName": partner_id.display_name, 
    #             "PrimaryEmailAddr": {
    #                 "Address": partner_id.email,
    #             }, 
    #             "DisplayName": partner_id.display_name, 
    #             "Suffix": "", 
    #             "Title": "Mr", 
    #             "MiddleName": "",   
    #             "Notes": "Here are other details.", 
    #             "FamilyName": "", 
    #             "PrimaryPhone": {
    #                 "FreeFormNumber": partner_id.mobile
    #             }, 
    #             "CompanyName": partner_id.company_id.display_name, 
    #             "BillAddr": {   
    #                 "CountrySubDivisionCode": "CA", 
    #                 "City": partner_id.city,  
    #                 "PostalCode": partner_id.zip[:31] if (partner_id.zip and len(partner_id.zip) > 30) else partner_id.zip, 
    #                 "Line1": partner_id.street,
    #                 "Country": partner_id.country_id.display_name,
    #             }, 
    #             "GivenName": partner_id.name,
    #         }
    #     if tax_id:
    #         if tax_id.quickbooks_taxes_id:
    #             payload.update({
    #                 'DefaultTaxCodeRef':{
    #                     'value': tax_id.quickbooks_taxes_id.quickbooks_tax_code_id,
    #                 }
    #             })
    #         else:
    #             raise UserError("Quickbooks tax Not Found!")
    #     return payload

    # def create_customer(self, headers, res_company_id):
    #     try:
    #         res = {}
    #         url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/customer"
    #         payload = self.customer_payload(self.partner_id)
    #         currency_object = self.get_quickbooks_currency(res_company_id,headers)
    #         if currency_object:
    #             payload.update({"CurrencyRef": {
    #                                     "name": currency_object.get('currency_name'), 
    #                                     "value": currency_object.get('currency_code')
    #                                     }})
    #         res = requests.request("POST", url, headers=headers, data=json.dumps(payload)) 
    #         if res.status_code == 200:
    #             return res
    #         res.raise_for_status() 
    #     except Exception as e:
    #         error_msg = json.loads(res.text).get('Fault').get("Error")[0].get("Detail") if (res != {} and 'Fault' in json.loads(res.text).keys() and json.loads(res.text).get('Fault')) else e
    #         raise UserError(f"Customer cannot be created because of {str(error_msg)}!")
        
    # def product_payload(self, product_id):
    #     payload = json.dumps({
    #             "FullyQualifiedName": product_id.display_name, 
    #             "domain": "QBO", 
    #             "Name": product_id.display_name, 
    #             "TrackQtyOnHand": False if (product_id.is_admin or product_id.is_shipping_product or product_id.is_global_discount) else True, 
    #             # "Type": dict(self.env['product_id.product_id']._fields['type'].selection).get(product_id.type), 
    #             "PurchaseCost": 0, 
    #             "QtyOnHand": 0 if (product_id.is_admin or product_id.is_shipping_product or product_id.is_global_discount) else product_id.qty_available,   
    #             "Active": True, 
    #             "UnitPrice": product_id.list_price, 
    #             "Description": product_id.description if product_id.description else '',  
    #             "IncomeAccountRef": {
    #                 "name": "Sales of Product Income", 
    #                 "value": "79"
    #                 }, 
    #             "AssetAccountRef": {
    #                 "name": "Inventory Asset", 
    #                 "value": "81"
    #                 }, 
    #             "InvStartDate": str(fields.date.today() - relativedelta(months=6)), 
    #             "Type": "Service" if (product_id.is_admin or product_id.is_shipping_product or product_id.is_global_discount) else "Inventory", 
    #             "ExpenseAccountRef": {
    #                 "name": "Cost of Goods Sold", 
    #                 "value": "80"
    #                 },
    #             })
    #     return payload

    # def create_product(self, headers, product, res_company_id):
    #     try:
    #         url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/item?minoversion=62"
    #         payload = self.product_payload(product)
    #         response = requests.request("POST", url, headers=headers, data=payload)
    #         if response.status_code == 200:
    #             return response
    #         response.raise_for_status()
    #     except Exception as e:
    #         error_msg = json.loads(response.text).get('Fault').get("Error")[0].get("Detail") if json.loads(response.text).get('Fault').get("Error")[0] else e
    #         raise UserError(_(str(error_msg)))
        
    # def delete_invoice(self,res_company_id):
    #     try:
    #         del_inv_response = {}
    #         inv_query = "Select * from Invoice where Id = '%s'"%(self.quickbooks_invoice_id)
    #         inv_response = self.find_invoice(res_company_id,inv_query)
    #         if inv_response:
    #             if "QueryResponse" in json.loads(inv_response.text).keys() and json.loads(inv_response.text).get("QueryResponse") != {}:
    #                 try:
    #                     payment_response = {}
    #                     headers = {
    #                         'Accept': 'application/json',
    #                         'Authorization': 'Bearer '+ res_company_id.kits_quickbooks_backend_id.access_token,
    #                         'Content-Type': 'application/json'
    #                     }
    #                     if "LinkedTxn" in json.loads(inv_response.text).get("QueryResponse").get("Invoice")[0] and json.loads(inv_response.text).get("QueryResponse").get("Invoice")[0].get("LinkedTxn") != []: 
    #                         payment_id = json.loads(inv_response.text).get("QueryResponse").get("Invoice")[0].get("LinkedTxn")[0].get("TxnId")
    #                         if payment_id:
    #                             del_payment_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/payment?operation=delete&minorversion=65"
    #                             payload = {
    #                                 "SyncToken": "2", 
    #                                 "Id": payment_id,
    #                             }
    #                             payment_response = requests.request("POST",del_payment_url,headers=headers,data=json.dumps(payload))
    #                             payment_response.raise_for_status()
    #                 except Exception as e:
    #                     error_msg = json.loads(payment_response.text).get('Fault').get("Error")[0].get("Detail") if (payment_response != {} and "Fault" in json.loads(payment_response.text).keys() and json.loads(payment_response.text).get('Fault')) else e
    #                     raise UserError(_(str(error_msg)))
    #                 del_inv_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/invoice?operation=delete&minorversion=65" 
    #                 payload = json.dumps({ 
    #                     "SyncToken": "3",
    #                     "Id": self.quickbooks_invoice_id,
    #                 })
    #                 del_inv_response = requests.request("POST",del_inv_url,headers=headers,data=payload)
    #                 del_inv_response.raise_for_status()
                    
    #     except Exception as e:
    #         error_msg = json.loads(del_inv_response.text).get('Fault').get("Error")[0].get("Detail") if (del_inv_response != {} and "Fault" in json.loads(del_inv_response.text).keys() and json.loads(del_inv_response.text).get('Fault')) else e
    #         raise UserError(_(str(error_msg)))

    # def find_invoice(self,res_company_id,inv_query):
    #     if self.quickbooks_invoice_id:
    #         inv_query = "Select * from Invoice where Id = '%s'"%(self.quickbooks_invoice_id)
    #         inv_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/query?query=" + inv_query + "&minorversion=65"
    #         headers = {
    #         'Accept': 'application/json',
    #         'Authorization': 'Bearer '+ res_company_id.kits_quickbooks_backend_id.access_token,
    #         'Content-Type': 'application/json'
    #         }
    #         inv_response = requests.request("GET",inv_url,headers=headers)
    #         inv_response.raise_for_status()
    #         return inv_response

    # def create_term(self,res_company_id,headers,term_id):
    #     term_url = res_company_id.kits_quickbooks_backend_id.base_url + "/v3/company/" + res_company_id.kits_quickbooks_backend_id.company_id + "/term?minorversion=65"
    #     payload ={
    #         "DueDays": str(term_id.line_ids[0].days),
    #         "Name": term_id.name,
    #     }
    #     term_response = requests.request("POST",term_url,headers=headers,data=json.dumps(payload))
    #     term_response.raise_for_status()
    #     return term_response
