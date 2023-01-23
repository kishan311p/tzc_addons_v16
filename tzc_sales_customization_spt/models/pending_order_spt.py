# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from io import BytesIO
import base64
import openpyxl

class pending_order_spt(models.Model):
    _name = 'pending.order.spt'
    _inherit = ['portal.mixin']
    _rec_name = 'catalog_id'
    _order = 'execution_time desc'
    _description = "Pending Order"

    catalog_id = fields.Many2one('sale.catalog', 'Catalog')
    customer_id = fields.Many2one('res.partner', 'Customer')
    pending_order_line_ids = fields.One2many('pending.order.line.spt', 'pending_order_id', 'pending Order Line')
    execution_time = fields.Datetime('Execution Time')
    state = fields.Selection([('draft','Draft'),('pending','Pending'),('done','Done'),('decline','Decline')], string='Status')
    sale_order_id = fields.Many2one('sale.order', 'Sale Order')
    reject_reason = fields.Text(string="Reject Reason")
    user_id = fields.Many2one('res.users', string='Responsible',compute="_compute_user_id",store=True)

    @api.depends('catalog_id')
    def _compute_user_id(self):
        for record in self:
            user_id = False
            if record.catalog_id:
                user_id = record.catalog_id.user_id.id
            record.user_id = user_id

    def p_ord_line_ordering_by_product(self):
        product_list = []
        product_list = self.pending_order_line_ids.mapped(lambda line:line.product_id.name_get()[0][1].strip()) if self.pending_order_line_ids else []
        # for line in self.pending_order_line_ids:
        #     product_name = line.product_id.name_get()[0][1].split('(')
        #     product_list.append(product_name[0])
        product_list = list(set(product_list))
        product_list.sort()
        return product_list

    def create_sale_order(self):
        sale_order_obj = self.env['sale.order']
        sale_order_line_obj = self.env['sale.order.line']
        sale_catalog_line_obj = self.env['sale.catalog.line']
        for record in self:
            fiscal_position_id = self.env['account.fiscal.position'].sudo().get_fiscal_position(record.customer_id.id)
            so_id = sale_order_obj.create({
                'partner_id': record.customer_id.id,
                'catalog_id': record.catalog_id.id,
                'user_id':record.catalog_id.user_id.id,
                'fiscal_position_id':fiscal_position_id,
            })
            for line_id in record.pending_order_line_ids.filtered(lambda x:x.qty != 0):
                if record.customer_id.property_product_pricelist.currency_id.name == 'USD':
                    price_unit = line_id.product_id.lst_price_usd
                elif record.customer_id.property_product_pricelist.currency_id.name == 'CAD':
                    price_unit = line_id.product_id.lst_price
                else:
                    price_unit = line_id.product_price
                # price_unit = self.env.user.partner_id.property_product_pricelist.get_product_price(line_id.product_id,line_id.qty,self.env.user.partner_id)
                product_price = price_unit
                if line_id.cataog_line_id.sale_type == 'on_sale' and so_id.partner_id and so_id.partner_id.property_product_pricelist :
                    if so_id.partner_id.property_product_pricelist.currency_id.name == 'CAD':
                        price_unit = line_id.cataog_line_id.product_pro_id.on_sale_cad
                    else:
                        price_unit = line_id.cataog_line_id.product_pro_id.on_sale_usd
                
                if line_id.cataog_line_id.sale_type == 'clearance' and so_id.partner_id and so_id.partner_id.property_product_pricelist :
                    if so_id.partner_id.property_product_pricelist.currency_id.name == 'CAD':
                        price_unit = line_id.cataog_line_id.product_pro_id.clearance_cad
                    else:
                        price_unit = line_id.cataog_line_id.product_pro_id.clearance_usd

                unit_discount_price = price_unit
                if line_id.cataog_line_id.discount:
                    unit_discount_price = round(product_price - round((product_price *(line_id.cataog_line_id.discount * 0.01)),2),2)
                disc = sale_catalog_line_obj.search([('catalog_id','=',record.catalog_id.id),('product_pro_id','=',line_id.product_id.id)],limit=1).discount or 0.0
                order_line_id = sale_order_line_obj.create({
                    'order_id': so_id.id,
                    'name' : line_id.product_id.display_name,
                    'product_id': line_id.product_id.id,
                    'discount': disc,
                    'product_uom_qty': line_id.qty,
                    'product_uom': line_id.product_id.uom_id.id,
                    'sale_type' : line_id.cataog_line_id.sale_type,
                    'price_unit' : product_price,
                    'unit_discount_price': unit_discount_price,
                })

            record.sale_order_id = so_id.id
            for line in so_id.order_line:
                line._onchange_discount_spt()
            verified = so_id.partner_verification()
            quotation_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_catalog_quotation_spt')
            quotation_template_id.send_mail(so_id.id,force_send=True) if verified else None
            confirmation_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_saleperson_quotation_spt')
            confirmation_template_id.send_mail(so_id.id,email_values={'email_to': so_id.user_id.partner_id.email},force_send=True)
            record.state = 'done'
            so_id.write({'state':'received'})
        
        # catalog_obj = self.env['sale.catalog']
        # catalog_obj.connect_server()
        # method = catalog_obj.get_method('create_sale_order')
        # if method['method']:
        #     localdict = {'self': self,'_':_,}
        #     exec(method['method'], localdict)
            # self._cr.commit()
            
    @api.model
    def create(self,vals):
        last_record = self.search([('state','=','pending')],order='id desc',limit=1)
        if last_record and last_record.execution_time:
            execution_time = last_record.execution_time + timedelta(minutes=int(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.order_delay', default=0)))
            vals.update({'execution_time':execution_time})
        else:
            vals.update({'execution_time':datetime.now()})
        return super(pending_order_spt,self).create(vals)

    def _get_report_base_filename(self):
        self.ensure_one()
        return '%s' % (self.catalog_id.name)

    def print_pending_catalog_order_excel(self):
        for record in self:
            workbook = Workbook()
            sheet = workbook.create_sheet(title='Pending Catalog Order', index=0)  # sheet name
            alignment = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
            alignment_right = Alignment(vertical='center', horizontal='right', text_rotation=0, wrap_text=True)
            alignment_left = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
            header_font = Font(name="Garamond",size="10",bold=False)
            table_header_font = Font(name="Garamond",size="10",bold=True)
            currency_symbol = self.env.user.partner_id.property_product_pricelist.currency_id.symbol
            currency_name = self.env.user.partner_id.property_product_pricelist.currency_id.name
            # ========================= HEADER =========================
            # Company LOGO
            img = BytesIO()
            img.flush()
            img.write(base64.b64decode(self.env.companies[0].logo))
            try:
                image = openpyxl.drawing.image.Image(img)
                image.width = 450
                image.height = 55
                sheet.merge_cells('A1:H5')
                sheet.add_image(image, 'A2')
            except:
                pass
            # LOGO END
            # ========================= END HEADER =========================
            sheet.merge_cells('I1:J5')
            address = []
            company = self.env.companies[0]
            if company.name:
                address.append(company.name)
                address.append('(A division of Tanzacan Tradelink Inc.)')
            if company.street:
                address.append(company.street)
            if company.street2:
                address.append(company.street2)
            if company.city:
                address.append(company.city)
            if company.state_id:
                address.append(company.state_id.name)
            if company.zip:
                address.append(company.zip)
            if company.country_id:
                address.append(company.country_id.name)
            address_line = '\n'.join(address)
            sheet.cell(row=1,column=9).value = address_line
            sheet.cell(row=1,column=9).font = header_font

            table_header_index = 7
            
            sheet.cell(row=table_header_index, column=1).value = 'Image'
            sheet.cell(row=table_header_index, column=2).value = 'Secondary Image'
            sheet.cell(row=table_header_index, column=3).value = 'Product'
            sheet.cell(row=table_header_index, column=4).value = 'CAT'
            sheet.cell(row=table_header_index, column=5).value = 'QTY'
            sheet.cell(row=table_header_index, column=6).value = 'WHOLESALE'
            sheet.cell(row=table_header_index, column=7).value = 'PRICE (%s)'%(currency_name or '')
            sheet.cell(row=table_header_index, column=8).value = 'DISC %'
            sheet.cell(row=table_header_index, column=9).value = 'Discounted Price'
            sheet.cell(row=table_header_index, column=10).value = 'SUBTOTAL'
            sheet.row_dimensions[table_header_index].height = 30

            sheet.cell(row=table_header_index, column=1).alignment = alignment
            sheet.cell(row=table_header_index, column=2).alignment = alignment
            sheet.cell(row=table_header_index, column=3).alignment = alignment
            sheet.cell(row=table_header_index, column=4).alignment = alignment
            sheet.cell(row=table_header_index, column=5).alignment = alignment
            sheet.cell(row=table_header_index, column=6).alignment = alignment
            sheet.cell(row=table_header_index, column=7).alignment = alignment
            sheet.cell(row=table_header_index, column=8).alignment = alignment
            sheet.cell(row=table_header_index, column=9).alignment = alignment
            sheet.cell(row=table_header_index, column=10).alignment = alignment

            sheet.cell(row=table_header_index, column=1).font = table_header_font
            sheet.cell(row=table_header_index, column=2).font = table_header_font
            sheet.cell(row=table_header_index, column=3).font = table_header_font
            sheet.cell(row=table_header_index, column=4).font = table_header_font
            sheet.cell(row=table_header_index, column=5).font = table_header_font
            sheet.cell(row=table_header_index, column=6).font = table_header_font
            sheet.cell(row=table_header_index, column=7).font = table_header_font
            sheet.cell(row=table_header_index, column=8).font = table_header_font
            sheet.cell(row=table_header_index, column=9).font = table_header_font
            sheet.cell(row=table_header_index, column=10).font = table_header_font

            total_products=0
            total_discount = 0.00
            total_subtotal = 0.00
            restricted_products = self.env['product.product'].search([('geo_restriction','in',self.env.user.country_id.ids)])
            row_index = table_header_index+1
            img1,img2=False,False
            for line in record.catalog_id.line_ids.sorted(lambda x: x.product_pro_id.variant_name):
                if line.product_pro_id not in restricted_products:
                    if line.image_variant_1920:
                        img1 = BytesIO()
                        img1.flush()
                        img1.write(base64.b64decode(line.image_variant_1920))
                        image1 = openpyxl.drawing.image.Image(img1)
                        image1.width = 100
                        image1.height = 50
                        sheet.add_image(image1,'A%s'%(row_index))
                    
                    if line.image_variant_1920:
                        img2 = BytesIO()
                        img2.flush()
                        img2.write(base64.b64decode(line.image_secondary_1920))
                        image2 = openpyxl.drawing.image.Image(img2)
                        image2.width = 100
                        image2.height = 50
                        sheet.add_image(image2,'B%s'%(row_index))



                    sheet.cell(row=row_index, column=3).value = line.product_pro_id.name_get()[0][1] or ''
                    sheet.cell(row=row_index, column=4).value = line.product_pro_id.categ_id.name or ''
                    sheet.cell(row=row_index, column=5).value = line.product_qty or 0
                    total_products += line.product_qty
                    sheet.cell(row=row_index, column=6).value = currency_symbol+str(round(line.product_price_wholesale,2) or 0.00)
                    sheet.cell(row=row_index, column=8).value = round(line.discount,2) or 0.00

                    price = self.env.user.partner_id.property_product_pricelist.get_product_price(line.product_pro_id,line.product_qty,self.env.user.partner_id)
                    product_price = price
                    if line.sale_type == 'on_sale':
                        if currency_name == 'USD':
                            price = line.product_pro_id.on_sale_usd
                            product_price = line.product_pro_id.lst_price_usd

                        else:
                            product_price = line.product_pro_id.lst_price
                            price = line.product_pro_id.on_sale_cad
                    if line.sale_type == 'clearance':
                        if currency_name == 'USD':
                            price = line.product_pro_id.clearance_usd
                            product_price = line.product_pro_id.lst_price_usd

                        else:
                            product_price = line.product_pro_id.lst_price
                            price = line.product_pro_id.clearance_cad
                    if not price:
                        if currency_name == 'USD':
                            price = line.product_pro_id.lst_price_usd
                            product_price = line.product_pro_id.lst_price_usd
                        else:
                            price = line.product_pro_id.lst_price
                            product_price = line.product_pro_id.lst_price

                    price = product_price

                    discount_amount = price * line.discount * 0.01
                    subtotal = (price-discount_amount) * line.product_qty
                    total_discount += round(discount_amount*line.product_qty,2)
                    total_subtotal += round(subtotal,2)

                    sheet.cell(row=row_index, column=7).value = '{} {:,.2f}'.format(currency_symbol,price or 0.00)
                    sheet.cell(row=row_index, column=9).value = '{} {:,.2f}'.format(currency_symbol,round(price-discount_amount,2) or 0.00)
                    sheet.cell(row=row_index, column=10).value = '{} {:,.2f}'.format(currency_symbol,subtotal or 0.00)
                
                    sheet.cell(row=row_index, column=3).alignment = alignment_left
                    sheet.cell(row=row_index, column=4).alignment = alignment_left
                    sheet.cell(row=row_index, column=5).alignment = alignment_right
                    sheet.cell(row=row_index, column=6).alignment = alignment_right
                    sheet.cell(row=row_index, column=7).alignment = alignment_right
                    sheet.cell(row=row_index, column=8).alignment = alignment_right
                    sheet.cell(row=row_index, column=9).alignment = alignment_right
                    sheet.cell(row=row_index, column=10).alignment = alignment_right

                    sheet.row_dimensions[row_index].height = 50
                    row_index+=1
            
            footer_row = row_index+1
            sheet.cell(row=footer_row, column=9).value = 'Subtotal'
            sheet.cell(row=footer_row, column=10).value = '{} {:,.2f}'.format(currency_symbol,total_subtotal+total_discount or 0.00)
            sheet.cell(row=footer_row, column=10).alignment = alignment_right
            footer_row += 1
            sheet.cell(row=footer_row, column=9).value = 'Discount'
            sheet.cell(row=footer_row, column=10).value = '{} {:,.2f}'.format(currency_symbol,total_discount or 0.00)
            sheet.cell(row=footer_row, column=10).alignment = alignment_right
            footer_row += 1
            sheet.cell(row=footer_row, column=9).value = 'Total (%s)'%(currency_name)
            sheet.cell(row=footer_row, column=10).value = '{} {:,.2f}'.format(currency_symbol,total_subtotal or 0.00)
            sheet.cell(row=footer_row, column=10).alignment = alignment_right
            footer_row += 1
            sheet.cell(row=footer_row, column=9).value = 'Total Quantity'
            sheet.cell(row=footer_row, column=10).value = total_products
            sheet.cell(row=footer_row, column=10).alignment = alignment_right
            footer_row += 1

            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 5
            sheet.column_dimensions['E'].width = 5
            sheet.column_dimensions['F'].width = 13
            sheet.column_dimensions['G'].width = 8
            sheet.column_dimensions['H'].width = 15
            sheet.column_dimensions['I'].width = 12
            sheet.column_dimensions['J'].width = 12

            fp = BytesIO()
            workbook.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()
            img.close()
            img1.close() if img1 else None
            img2.close() if img2 else None

            return data
