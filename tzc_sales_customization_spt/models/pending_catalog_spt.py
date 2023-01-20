# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from datetime import datetime, timedelta
from werkzeug.urls import url_encode
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from io import BytesIO
import base64
import openpyxl
import os
import re
from bs4 import BeautifulSoup

class pending_catalog_spt(models.Model):
    _name = 'pending.catalog.spt'
    _inherit = ['portal.mixin']
    _description = "Pending Catalog"
    _rec_name = 'catalog_id'
    _order = 'execution_time desc'

    catalog_id = fields.Many2one('sale.catalog', 'Catalog')
    customer_id = fields.Many2one('res.partner', 'Customer')
    execution_time = fields.Datetime('Execution Time')
    user_id = fields.Many2one('res.users', string='Responsible',compute="_compute_user_id",store= True)
    state = fields.Selection([('draft','Draft'),('sent','Sent')],default='draft', string='Status')
    type_name = fields.Char('Type Name', compute='_compute_type_name')

    @api.depends('catalog_id')
    def _compute_user_id(self):
        for record in self:
            user_id = False
            if record.catalog_id:
                user_id = record.catalog_id.user_id.id
            record.user_id = user_id

    def get_access_token_spt(self):
        self.ensure_one()
        auth_param = url_encode(self.customer_id.signup_get_auth_param()[self.customer_id.id])
        return auth_param
    
    def send_pending_catalog_spt(self):
        # catalog_obj = self.env['sale.catalog']
        # catalog_visitors_obj = self.env['catalog.visitors.spt']
        # catalog_obj.connect_server()
        # method = catalog_obj.get_method('send_pending_catalog_spt')
        # if method['method']:
        #     localdict = {'self': self,'_':_,}
        #     exec(method['method'], localdict)

        catalog_visitors_obj = self.env['catalog.visitors.spt']
        for record in self:
            if record.state != 'sent':
                customer_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_catalog_spt')
                customer_template_id.send_mail(record.id,email_values={'email_to': record.customer_id.email},force_send=True)

                sales_person_template_id = self.env.ref('tzc_sales_customization_spt.tzc_email_template_catalog_confirmation_spt')
                sales_person_template_id.send_mail(record.id,force_send=True)

                record.write({'state':'sent'})
                catalog_visitors_obj.create({
                    'catalog_id':record.catalog_id.id,
                    'customer_id':record.customer_id.id,
                })
            record.catalog_state_done()
            record.catalog_id._get_pending_catalog_count()

        # catalog_visitors_obj.create({
        #     'catalog_id':self.catalog_id,
        #     'customer_id':self.customer_id,
        # })

    def send_pending_catalog(self):
        catalog_todo = self.search([('state','=','draft'),('execution_time','<',datetime.now())])
        catalog_todo.send_pending_catalog_spt()
            
    @api.model
    def create(self,vals):
        last_record = self.search([],order='id desc',limit=1)
        if last_record and last_record.execution_time:
            execution_time = last_record.execution_time + timedelta(minutes=int(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.catalog_delay', default=0)))
            vals.update({'execution_time':execution_time})
        else:
            vals.update({'execution_time':datetime.now()})
        return super(pending_catalog_spt,self).create(vals)

    def get_base_url(self):
        self.ensure_one()
        return self.env['ir.config_parameter'].sudo().search([('key','=','web.base.url')]).value

    def line_ordering_by_product(self):
        product_list = []
        # product_list = self.catalog_id.line_ids.mapped('product_pro_id.name')
        user = self.env.user
        # product_list =  self.catalog_id.line_ids.mapped(lambda line:line.product_pro_id.name if user.country_id not in line.product_pro_id.geo_restriction else '')
        for line in self.catalog_id.line_ids:
            if user.country_id.id not in line.product_pro_id.geo_restriction.ids:
                product_name = line.product_pro_id.name_get()[0][1]
                product_list.append(product_name.strip())
            else:
                product_list.append('')
        #to filter blank values comes where geo ristricted product comes
        product_list = list(filter(None, product_list))
        product_list = list(set(product_list))
        product_list.sort()
        return product_list

    def line_product_dict(self,product_name):
        product_dict = {}
        user = self.env.user
        restricted_products = self.env['product.product'].search([('geo_restriction','in',user.country_id.ids)])
        catalog_id = self.catalog_id.line_ids.filtered(lambda line: line.product_pro_id.name_get()[0][1] == product_name)
        if catalog_id[0].product_pro_id.id not in restricted_products.ids:
            product_dict[product_name] = {'line_ids': [catalog_id[0]]}
        # for line in self.catalog_id.line_ids:
        #     line_dict = {}
        #     # if user.country_id not in line.product_pro_id.geo_restriction:
        #     if line.product_pro_id not in restricted_products:
        #         product_name = line.product_pro_id.name_get()[0][1].split('(')
        #         if line.product_pro_id.name in product_dict.keys():
        #             product_dict[product_name[0]]['line_ids'].append(line) 
        #         else:
        #             line_dict['line_ids'] = [line]
        #             product_dict[product_name[0]] = line_dict
            return product_dict
        else:
            product_dict[product_name] = {'line_ids': []}
            return product_dict

    #for print and download pending catalog from website

    def _compute_access_url(self):
        super(pending_catalog_spt, self)._compute_access_url()
        for catalog in self:
            catalog.access_url = '/my/catalog/%s' % (catalog.catalog_id.id)
    
    def _get_report_base_filename(self):
        self.ensure_one()
        return '%s' % (self.type_name)

    @api.depends('state')
    def _compute_type_name(self):
        for record in self:
            record.type_name = record.catalog_id.name



    def catalog_state_done(self):
        for record in self:
            order_ids = self.search([('catalog_id','=',record.catalog_id.id)])
            if not any(order_ids.filtered(lambda x: x.state != 'sent')):
                record.catalog_id.state = 'done'

    def print_pending_catalog_excel(self):
        for record in self:
            base_sample_file = '/'.join(os.path.dirname(__file__).split('/')[:-1])+'/sample/_base_template_catalog.xlsm'
            wb = load_workbook(base_sample_file,read_only=False, keep_vba=True)
            wrksht = wb.active

            alignment_left = Alignment(vertical='center', horizontal='left', text_rotation=0, wrap_text=True)
            alignment_center = Alignment(vertical='center', horizontal='center', text_rotation=0, wrap_text=True)
            currency_symbol = self.env.user.partner_id.property_product_pricelist.currency_id.symbol
            currency_name = self.env.user.partner_id.property_product_pricelist.currency_id.name
            main_font = Font(name='Calibri',size='16',bold=True)
            data_font = Font(name='Calibri',size='12')
            all_border = Border(left=Side(style='thin', color="000000"), 
                            right=Side(style='thin', color="000000"), 
                            top=Side(style='thin', color="000000"), 
                            bottom=Side(style='thin', color="000000"))

            partner_line = 7
            wrksht.cell(row=partner_line, column=1).value = self.env['sale.catalog'].create_address_line_for_sale(self.customer_id, take_name=True)

            wrksht.cell(row=partner_line+2, column=1).value = self.catalog_id.name
            wrksht.cell(row=partner_line+2, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+2, column=1).font = main_font

            wrksht.cell(row=partner_line+3, column=1).value = 'Date: ' + self.catalog_id.create_date.strftime('%d/%m/%Y')
            wrksht.cell(row=partner_line+3, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+3, column=1).font = data_font

            wrksht.cell(row=partner_line+4, column=1).value = 'Salesperosn: ' + self.catalog_id.user_id.name
            wrksht.cell(row=partner_line+4, column=1).alignment = alignment_left
            wrksht.cell(row=partner_line+4, column=1).font = data_font

            wrksht.cell(row=16, column=5).value = 'PRICE (%s)'%(currency_name or '')

            total_products=0
            total_discount = 0.00
            total_subtotal = 0.00
            restricted_products = self.env['product.product'].search([('geo_restriction','in',self.env.user.country_id.ids)])
            row_index = 17
            for line in record.catalog_id.line_ids.sorted(lambda x: x.product_pro_id.variant_name):
                if line.product_pro_id not in restricted_products:
                    wrksht.cell(row=row_index, column=1).border = all_border
                    wrksht.cell(row=row_index, column=2).border = all_border
                    wrksht.cell(row=row_index, column=3).value = line.product_pro_id.catalog_report_product_name() or ''
                    wrksht.cell(row=row_index, column=3).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=row_index, column=3).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,wrap_text=True)
                    wrksht.cell(row=row_index, column=3).border = all_border
                    wrksht.cell(row=row_index, column=4).value = line.product_qty or 0
                    wrksht.cell(row=row_index, column=4).alignment = alignment_center
                    wrksht.cell(row=row_index, column=4).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=row_index, column=4).border = all_border
                    total_products += line.product_qty
                    wrksht.cell(row=row_index, column=5).value = line.unit_discount_price or 0.00
                    wrksht.cell(row=row_index, column=5).alignment = alignment_center
                    wrksht.cell(row=row_index, column=5).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=row_index, column=5).number_format = '"$"#,##0.00'
                    wrksht.cell(row=row_index, column=5).border = all_border
                    
                    price = self.env.user.partner_id.property_product_pricelist.get_product_price(line.product_pro_id,line.product_qty,self.env.user.partner_id)
                    product_price = price
                    if line.sale_type == 'on_sale':
                        if currency_name == 'USD':
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.on_sale_usd
                        else:
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.on_sale_cad
                    if line.sale_type == 'clearance':
                        if currency_name == 'USD':
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.clearance_usd
                        else:
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.clearance_cad
                    if not price:
                        if currency_name == 'USD':
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.lst_price_usd
                        else:
                            product_price = line.product_pro_id.lst_price_usd
                            price = line.product_pro_id.lst_price
                    price = product_price
                    discount_amount = price * line.discount * 0.01
                    subtotal = (price-discount_amount) * line.product_qty
                    total_discount += round(discount_amount*line.product_qty,2)
                    total_subtotal += round(subtotal,2)
                    
                    wrksht.cell(row=row_index, column=6).value = '=D%s*E%s'%(str(row_index),str(row_index))
                    wrksht.cell(row=row_index, column=6).alignment = alignment_center
                    wrksht.cell(row=row_index, column=6).font = Font(name='Calibri',size='16',bold=False)
                    wrksht.cell(row=row_index, column=6).number_format = '"$"#,##0.00'
                    wrksht.cell(row=row_index, column=6).border = all_border

                    wrksht.cell(row=row_index, column=7).hyperlink = line.product_pro_id.image_url
                    wrksht.cell(row=row_index, column=7).alignment = alignment_center
                    wrksht.cell(row=row_index, column=7).font = Font(color="FF0000FF",underline='single',name='Calibri',size='16')
                    wrksht.cell(row=row_index, column=7).border = all_border

                    wrksht.cell(row=row_index, column=8).hyperlink = line.product_pro_id.image_secondary_url
                    wrksht.cell(row=row_index, column=8).alignment = alignment_center
                    wrksht.cell(row=row_index, column=8).font = Font(color="FF0000FF",underline='single',name='Calibri',size='16')
                    wrksht.cell(row=row_index, column=8).border = all_border
                    
                    
                    wrksht.row_dimensions[row_index].height = 240
                    row_index+=1

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            data = fp.read()
            fp.close()

            return data

    # def is_accessible_to(self,user):
    #     self = self.sudo()
    #     self.ensure_one()
    #     result = False
    #     if user:
    #         if user in self.customer_id.user_ids or user == self.user_id:
    #             result = True
    #     return result

    def get_description(self,description):
        soup = BeautifulSoup(description, "html.parser")
        data = soup(['p'])
        description_data = []
        for i in data:
            cleanr = re.compile('<.*?>')
            cleantext = re.sub(cleanr,'',str(i))
            description_data.append(cleantext)
        if not description_data or description_data == ['']:
            description_data = ['An exclusive eyewear catalog has been created for you. Click on the View Catalog button below to view the entire catalog.']
        return description_data
