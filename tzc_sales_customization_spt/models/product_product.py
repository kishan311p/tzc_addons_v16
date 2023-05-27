from odoo import _, api, fields, models, tools
from odoo.addons.http_routing.models.ir_http import slug
from odoo.exceptions import RedirectWarning, UserError, ValidationError, AccessError
from datetime import datetime,timedelta
from urllib.request import urlopen
import pandas as pd
import math
import re
import os
import random
import requests
import base64
import urllib
from bs4 import BeautifulSoup
from lxml import etree
from io import BytesIO
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import xlsxwriter
import csv

class product_template(models.Model):
    _inherit = 'product.product'
    
    def action_open_quants(self):
        res = super(product_template,self).action_open_quants()
        if isinstance(res, dict) and res.get('context',False):
            ctx = dict()
            ctx = res.get('context').copy()
            ctx.update({'kits_update_product_date': True})
            res['context'] = ctx
        return res
    
class ProductProduct(models.Model):
    _inherit = 'product.product'

    def _get_default_category_id(self):
        if self._context.get('categ_id') or self._context.get('default_categ_id'):
            return self._context.get('categ_id') or self._context.get('default_categ_id')
        category = self.env.ref('product.product_category_all', raise_if_not_found=False)
        if not category:
            category = self.env['product.category'].search([], limit=1)
        if category:
            return category.id
        else:
            err_msg = _('You must define at least one product category in order to be able to create products.')
            redir_msg = _('Go to Internal Categories')
            raise RedirectWarning(err_msg, self.env.ref('product.product_category_action_form').id, redir_msg)

    flex_hinges = fields.Selection([('yes','Yes'),('no','No')],'Flex Hinges')
    product_tmpl_id = fields.Many2one('product.template', 'Product Template',index=True)
    active = fields.Boolean(
        'Active', default=True,
        help="If unchecked, it will allow you to hide the product without removing it.",tracking=True)
    is_published_spt = fields.Boolean('Is Published ')
    product_color_name =  fields.Many2one('product.color.spt','Color Name')
    secondary_color_name =  fields.Many2one('product.color.spt','Secondary Color Name')
    variant_name = fields.Char('Variant Description')
    image_url = fields.Char('Image URL')
    image_secondary_url = fields.Char('Image Secondary URL')
    brand = fields.Many2one('product.brand.spt','Brand', index=True)
    model = fields.Many2one('product.model.spt','Model', index=True)
    color_code = fields.Many2one('kits.product.color.code','Manufacturing Color Code')
    eye_size = fields.Many2one('product.size.spt','Eye Size ', index=True)
    categ_id = fields.Many2one(
        'product.category', 'Product Category',
        change_default=True, default=_get_default_category_id, group_expand='_read_group_categ_id',
        required=True, help="Select category for the current product", index=True)

    write_date = fields.Datetime('Last Update')

    image_secondary = fields.Binary("Secondary Image")
    image_secondary_1920 = fields.Image("Secondary Image 1920", related="image_secondary", max_width=1920, max_height=1920, store=True)
    image_secondary_1024 = fields.Image("Secondary Image 1024", related="image_secondary", max_width=1024, max_height=1024, store=True)
    image_secondary_512 = fields.Image("Secondary Image 512", related="image_secondary", max_width=512, max_height=512, store=True)
    image_secondary_256 = fields.Image("Secondary Image 256", related="image_secondary", max_width=256, max_height=256, store=True)
    image_secondary_128 = fields.Image("Secondary Image 128", related="image_secondary", max_width=128, max_height=128, store=True)
    
    replenishable = fields.Boolean('Replenishable?')
    # available_in_slider_spt = fields.Boolean(string="Available in slider?")
    
    price_wholesale = fields.Float('Wholesale Price',digits='Product Price',help="Wholesale Price")
    price_msrp = fields.Float('MSRP',digits='Product Price',help="MSRP Price")
    price_wholesale_usd = fields.Float('Wholesale Price USD',digits='Product Price',help="Wholesale Price")
    price_msrp_usd = fields.Float('MSRP USD',digits='Product Price',help="MSRP Price")
    lst_price_usd = fields.Float('Our Price USD',digits='Product Price',help="Public Price In USD")
    # on_sale_cad = fields.Float('On Sale CAD',compute='_compute_onsale_price',store=True)
    on_sale_usd = fields.Float('On Sale Price')
    # on_sale_cad_in_percentage = fields.Float('On Sale CAD In Percentage',compute='_compute_onsale_price',store=True)
    on_sale_usd_in_percentage = fields.Float('On Sale Price In Percentage')
    eye_size_compute = fields.Integer('Eye Size Compute (Flag)',compute='_compute_eye_size',store=True)
    bridge_size_compute = fields.Integer('Bridge Size Compute',compute='_compute_bridge_size',store=True)
    temple_size_compute = fields.Integer('Temple Size Compute',compute='_compute_temple_size',store=True)
    sale_type = fields.Selection([('clearance','Clearance'),('on_sale','On Sale')],'Sale Type')
    # clearance_cad = fields.Float('Clearance CAD',compute='_compute_clearance_price',store=True)
    clearance_usd = fields.Float('Clearance Price')
    # clearance_cad_in_percentage = fields.Float('Clearance CAD In Percentage',compute='_compute_clearance_price',store=True)
    clearance_usd_in_percentage = fields.Float('Clearance Price In Percentage')
    temporary_out_of_stock = fields.Boolean('Temporary Out Of Stock')
    geo_restriction = fields.Many2many('res.country','product_with_country_real','product_id','country_id','Geo Restriction', index=True)
    new_arrivals = fields.Boolean('New Arrivals (Flag)')
    new_arrival_update = fields.Datetime('New Arrival Update',compute='_onchange_new_arrivals',store=True)
    length = fields.Float('Length (cm)')
    width = fields.Float('Width (cm)')
    height = fields.Float('Height (cm)')
    # material_ids = fields.Many2many('product.material.spt','product_with_material_real','product_id','material_id','Materials')
    material_id = fields.Many2one('product.material.spt','Materials')
    # shape_ids = fields.Many2many('product.shape.spt','product_with_shape_real','product_id','shape_id','Shapes')
    shape_id = fields.Many2one('product.shape.spt','Shapes')
    # website_id = fields.Many2one(compute="_compute_fields_stored", readonly=False, store=True)
    qty_available = fields.Float(
        'Quantity On Hand', compute='_compute_quantities', search='_search_qty_available',
        digits='Product Unit of Measure', compute_sudo=False,
        help="Current quantity of products.\n"
             "In a context with a single Stock Location, this includes "
             "goods stored at this Location, or any of its children.\n"
             "In a context with a single Warehouse, this includes "
             "goods stored in the Stock Location of this Warehouse, or any "
             "of its children.\n"
             "stored in the Stock Location of the Warehouse of this Shop, "
             "or any of its children.\n"
             "Otherwise, this includes goods stored in any Stock Location "
             "with 'internal' type.",store=True)

    manufacture_color_code = fields.Char('Manufacture Color Code',compute="_compute_eye_size",store=True)
    in_future_archive = fields.Boolean('In Future Archive (Flag)')
    case_type = fields.Selection([('original', 'Original'),('generic', 'Generic')],"Case Type",store=True)
    case_image_url = fields.Char('Case Image Url',store=True)
    case_image_url_1920 = fields.Image("Case Image 1920", max_width=1920, max_height=1920, store=True)
    case_image_url_1024 = fields.Image("Case Image 1024", max_width=1024, max_height=1024, store=True)
    case_image_url_512 = fields.Image("Case Image 512", max_width=512, max_height=512, store=True)
    case_image_url_256 = fields.Image("Case Image 256", max_width=256, max_height=256, store=True)
    case_image_url_128 = fields.Image("Case Image 128", max_width=128, max_height=128, store=True)

    rim_type = fields.Many2one('product.rim.type.spt','Rim Type')
    custom_message = fields.Text(string='Custom Message', default='', translate=True)
    country_of_origin = fields.Many2one('res.country','Country Of Origin')
    gender = fields.Selection([('male','Male'),('female','Female'),('m/f','Unisex')], string='Gender')
    bridge_size = fields.Many2one('product.bridge.size.spt','Bridge Size')
    temple_size = fields.Many2one('product.temple.size.spt','Temple Size')
    lense_color_name =  fields.Many2one('product.color.spt','Lense Color Name')
    aging = fields.Many2one('product.aging.spt','Aging')
    size = fields.Many2one('product.size.spt','Size')
    material = fields.Char('Material',compute="_compute_material",store=True)

    on_consignment = fields.Boolean('On Consignment')
    minimum_qty = fields.Integer('Minimum Qty')
    is_new_price = fields.Boolean('New Price (Flag)')
    actual_stock = fields.Float('Actual Stock',compute="calculate_actual_stock")
    # assign_qty = fields.Float('Assign Qty',default=0.0)
    # kits_ecom_categ_id = fields.Many2one('product.public.category','Website Public Product Category')
    variant_count = fields.Integer('Product Variant',compute="_get_product_variant")
    # is_image_missing = fields.Boolean(compute="_get_image",store=True)
    is_image_missing = fields.Boolean()
    is_forcefully_unpublished = fields.Boolean('Forcefully Unpublished')
    primary_image_url = fields.Char('Primary Url',compute='_compute_primary_image_url',store=True)
    sec_image_url = fields.Char('Secondary Url',compute='_compute_primary_image_url',store=True)
    last_qty_update = fields.Datetime('Last Qty Update')
    updated_on = fields.Datetime('Updated On')
    updated_by = fields.Many2one('res.users','Updated By')
    product_seo_keyword = fields.Char(string='Seo keyword')
    product_seo_url = fields.Char(compute='_compute_product_seo_url',string='Seo Url')
    reversed_qty_spt = fields.Integer('Reserved Qty', compute='_compute_reversed_qty_spt',store=True)
    available_qty_spt = fields.Integer('Available Qty', compute='_compute_reversed_qty_spt',store=True)
    stock_move_line_ids = fields.One2many('stock.move.line','product_id',string='Move Lines')
    stock_quant_ids = fields.One2many('stock.quant','product_id',string='Quant Lines')
    order_not_invoice = fields.Float('Open Orders ')
    not_open_order = fields.Boolean('Not Open Orders (Flag)',compute="_compute_order_not_invoice",default=False,store=True,compute_sudo=True)
    ideal_product  = fields.Boolean('Ideal Product (Flag)',compute="_compute_order_not_invoice",default=False,store=True,compute_sudo=True)
    sale_order_line_ids = fields.One2many('sale.order.line','product_id',string='Order Lines',store=True)
    update_open_order = fields.Boolean(compute="_compute_order_not_invoice",compute_sudo=True,string="Update Open Order (Flag)")
    product_brand_commission = fields.Float('Product Brand Commission')
    application_type = fields.Selection([('0','B2B'),('1','Both'),('2','Only B2C')],'Application Type')
    is_3d_model = fields.Boolean('Is 3D Model')
    meta_keyword = fields.Char('Meta Keyword')
    meta_title = fields.Char('Meta Title')
    meta_description = fields.Text('Meta Description')
    is_b2c_published = fields.Boolean('B2C1 Published')
    is_select_for_lenses = fields.Boolean('Is Allow For Add Lenses')
    hs_code = fields.Char(
        string="HS Code",
        help="Standardized code for international shipping and goods declaration. At the moment, only used for the FedEx shipping provider.",
    )
    product_pricelist_item_ids = fields.One2many('product.pricelist.item','product_id')
    is_pending_price = fields.Boolean( string='Is Pending Price (Flag)',compute='_compute_pending_price',store=True)
    
    @api.depends('lst_price','price_wholesale','price_msrp','product_pricelist_item_ids')
    def _compute_pending_price(self):
        for record in self:
            is_pending_price =False
            # if record.type== 'product' and (not record.lst_price or not (record.price_wholesale or record.is_case_product) or not (record.price_msrp or record.is_case_product) or  any(record.product_pricelist_item_ids.mapped(lambda pp : not pp.fixed_price))):
            if not record.is_case_product:
                if record.type== 'product' and (not record.lst_price or not record.price_wholesale or not record.price_msrp or  any(record.product_pricelist_item_ids.mapped(lambda pp : not pp.fixed_price))):
                    is_pending_price = True
            if is_pending_price:
                record.is_published_spt = False
            record.is_pending_price = is_pending_price
    
    @api.model
    def search(self, args, offset=0, limit=None, order=None, count=False):
        args.append(('is_pending_price','!=',True))
        if self._context.get('pending_price'):
            args.remove(('is_pending_price','!=',True))
        return super(ProductProduct, self).search(args,offset,limit,order,count)
            
    _sql_constraints = [
        ('seo_keyword', 'unique(product_seo_keyword)', 'Seo keyword already exists!')
    ]

    # Case
    case_product_id = fields.Many2one('product.product','Case')

    # @api.depends('clearance_usd','clearance_usd_in_percentage')
    # def _compute_clearance_price(self):
    #     ir_config_parameter_on_sale_usd = float(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.on_sale_cad_spt', default=0))
    #     for record in self:
    #         clearance_cad = 0.00
    #         clearance_cad_in_percentage = 0.00
    #         if record.clearance_usd:
    #             clearance_cad = round(record.clearance_usd * ir_config_parameter_on_sale_usd,2)
    #             if clearance_cad and record.lst_price:
    #                 clearance_cad_in_percentage = round((1 -(clearance_cad/record.lst_price))*100,2)

    #         record.clearance_cad = clearance_cad
    #         record.clearance_cad_in_percentage = clearance_cad_in_percentage

    @api.depends('image_url','image_secondary_url')
    def _compute_primary_image_url(self):
        for rec in self:
            rec.primary_image_url = False
            rec.sec_image_url = False
            if rec.image_url:
                rec.primary_image_url = rec.image_url
            if rec.image_secondary_url:
                rec.sec_image_url = rec.image_secondary_url

    # @api.depends('case_product_id')
    # def _compute_case_product(self):
    #     for rec in self._origin:
    #         rec.case_type = rec.case_product_id.case_type
    #         rec.case_image_url = rec.case_product_id.image_url

    @api.constrains("case_product_id")
    def _check_case_product(self):
        for s in self:
            s.case_type = s.case_product_id.case_type
            s.case_image_url = s.case_product_id.image_url

    @api.onchange('image_url','image_secondary_url','case_image_url')
    def get_image(self):
        for rec in self:
            is_image_missing = False
            try:
                img_primary = urllib.request.Request(rec.image_url, headers={"User-Agent": "Chrome"})
                urllib.request.urlopen(img_primary)

                img_secondary = urllib.request.Request(rec.image_secondary_url, headers={"User-Agent": "Chrome"})
                urllib.request.urlopen(img_secondary)

                case_image_url = urllib.request.Request(rec.case_image_url, headers={"User-Agent": "Chrome"})
                urllib.request.urlopen(case_image_url)

            except:
                is_image_missing = True

            rec.is_image_missing = is_image_missing

    def _get_product_variant(self):
        for rec in self:
            rec.variant_count = 0
            country_id = self.env.user.country_id.ids
            product_ids = self.env['product.product'].search([('geo_restriction','not in',country_id),('brand','=',rec.product_variant_ids.brand.name),('model','=',rec.product_variant_ids.model.name),('categ_id','=',rec.categ_id.id),('id','!=',rec.id) if type(rec.id) == int else ('id','!=',rec._origin.id)])
            if product_ids:
                rec.variant_count = len(product_ids)
    
    def _read_group_categ_id(self, categories, domain, order):
        category_ids = self.env.context.get('default_categ_id')
        if not category_ids and self.env.context.get('group_expand'):
            category_ids = categories._search([], order=order, access_rights_uid=SUPERUSER_ID)
        return categories.browse(category_ids)

    def calculate_actual_stock(self):
        for rec in self:
            rec.actual_stock = 0.0
            if rec.on_consignment:
                rec.actual_stock = rec.available_qty_spt - rec.minimum_qty

    @api.depends('material_id')
    def _compute_material(self):
        for rec in self:
            if rec.material_id:
                material_names = []
                for material in rec.material_id:
                    material_names.append(material.name.strip())
                rec.material = '/'.join(material_names)
            else:
                rec.material = ''


    @api.depends('new_arrivals')
    def _onchange_new_arrivals(self):
        self.sudo().write({'new_arrival_update' : datetime.now()})

    @api.onchange('clearance_usd')
    def onchange_clearance_usd(self):
        for record in self: 
            record.clearance_usd_in_percentage = round((1 -(record.clearance_usd/record.lst_price))*100,2)
   
    @api.onchange('clearance_usd_in_percentage')
    def onchange_clearance_usd_in_percentage(self):
        for record in self: 
            record.clearance_usd =  round(record.lst_price-((record.clearance_usd_in_percentage * record.lst_price)*0.01),2)

    # @api.depends('on_sale_usd','on_sale_usd_in_percentage')
    # def _compute_onsale_price(self):
    #     ir_config_parameter_on_sale_usd = float(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.on_sale_cad_spt', default=0))
    #     for record in self:
    #         on_sale_cad = 0.00
    #         on_sale_cad_in_percentage = 0.00
    #         if record.on_sale_usd:
    #             on_sale_cad = round(record.on_sale_usd * ir_config_parameter_on_sale_usd,2)
    #             if on_sale_cad and record.lst_price:
    #                 on_sale_cad_in_percentage = round((1 -(on_sale_cad/record.lst_price))*100,2)

    #         record.on_sale_cad = on_sale_cad
    #         record.on_sale_cad_in_percentage = on_sale_cad_in_percentage

    def calculate_clearance_price_for_product_import(self):
        for record in self:
            if record.clearance_usd and record.lst_price:
                record.onchange_clearance_usd()
            if record.clearance_usd_in_percentage:
                record.onchange_clearance_usd_in_percentage()
            # record._compute_clearance_price()

    def calculate_onsale_price_for_product_import(self):
        for record in self:
            if record.on_sale_usd and record.lst_price:
                record.onchange_on_sale_usd()
            if record.on_sale_usd_in_percentage:
                record.onchange_on_sale_usd_in_percentage()
            # record._compute_onsale_price()

    @api.depends('temple_size','temple_size.name')
    def _compute_temple_size(self):
        for record in self:
            record.temple_size_compute = int(record.temple_size.name)
   
    @api.onchange('on_sale_usd')
    def onchange_on_sale_usd(self):
        for record in self: 
            record.on_sale_usd_in_percentage = round((1 -(record.on_sale_usd/record.lst_price))*100,2)
   
    @api.onchange('on_sale_usd_in_percentage')
    def onchange_on_sale_usd_in_percentage(self):
        for record in self: 
            record.on_sale_usd =  round(record.lst_price-((record.on_sale_usd_in_percentage * record.lst_price)*0.01),2)


    @api.depends('bridge_size','bridge_size.name')
    def _compute_bridge_size(self):
        for record in self:
            record.bridge_size_compute = int(record.bridge_size.name)

    @api.depends('color_code','eye_size')
    # @api.depends('product_template_attribute_value_ids','color_code','eye_size')
    def _compute_eye_size(self):
        # attribute_obj = self.env['product.attribute']
        # product_template_attribute_value_obj = self.env['product.template.attribute.value']
        # attribute_id = attribute_obj.search([('name','=','Eye Size')])
        # color_attribute_id = attribute_obj.search([('name','=','Color')])
        for record in self:
            record.eye_size_compute = int(record.eye_size.name)
            record.manufacture_color_code = str(record.color_code.name)
            # product_template_attribute_line_id = product_template_attribute_value_obj.search([('attribute_id','=',attribute_id.id),('id','in',record.product_template_attribute_value_ids.ids)])
            # if product_template_attribute_line_id:
            #     record.eye_size_compute = int(product_template_attribute_line_id.name)
            # product_template_colour_attribute_line_id = product_template_attribute_value_obj.search([('attribute_id','=',color_attribute_id.id),('id','in',record.product_template_attribute_value_ids.ids)])
            # if product_template_colour_attribute_line_id:
            #     record.manufacture_color_code = str(product_template_colour_attribute_line_id.name)

    # @api.depends('product_tmpl_id','product_tmpl_id.website_id','product_tmpl_id.product_variant_count')
    # # @api.depends('product_tmpl_id','product_tmpl_id.website_id','product_tmpl_id.product_variant_count')
    # def _compute_fields_stored(self):
    #     for rec in self:
    #         rec.website_id = rec.product_tmpl_id.website_id.id

    @api.depends('product_seo_keyword')
    def _compute_product_seo_url(self):
        # base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        web_url = self.env['kits.b2b.website'].search([],limit=1).url
        for record in self:
            # if record.website_id and record.website_id.domain:
            #     base_url = 'http://'+ record.website_id.domain
            if record.product_seo_keyword:
                record.product_seo_url = web_url + '/product/%s' % (record.product_seo_keyword)
            else:
                try:
                    if record.website_url:
                        record.product_seo_url = web_url + record.website_url
                    else:
                        record.product_seo_url = web_url + '/shop/product/' + slug(record)
                except Exception as e:
                    record.product_seo_url = web_url + '/'
    
    @api.depends('stock_move_line_ids','stock_move_ids','sale_order_line_ids','sale_order_line_ids.write_date','sale_order_line_ids.state','stock_move_line_ids.write_date','stock_move_ids.write_date','stock_move_ids.state')
    def _compute_order_not_invoice(self):
        for record in self:
            orders = self.env['sale.order']
            not_open_order = False
            ideal_product = False
            for line in record.sale_order_line_ids:
                if line.order_id.state not in ['paid','merged','open_inv','draft_inv','cancel']:
                    orders |= line.order_id
                # if (line.order_id.source_spt == 'Website' and line.order_id.state == 'draft'):
                #     orders -= line.order_id

            if not orders and record.sale_order_line_ids:
                not_open_order = True
            if not orders and not record.sale_order_line_ids :
                ideal_product = True
            record.order_not_invoice = len(orders)
            record.update_open_order = False
            record.ideal_product = ideal_product
            record.not_open_order = not_open_order

    @api.depends('stock_quant_ids','stock_move_line_ids','stock_move_ids','stock_move_ids.product_qty','stock_move_ids.quantity_done', 'stock_move_ids.state')
    @api.depends_context('lot_id', 'owner_id','from_date', 'to_date','company_owned', 'force_company',)
    def _compute_reversed_qty_spt(self):
        move_obj = self.env['stock.move']
        for record in self:
            move_ids = move_obj.search([('state','not in',['done','cancel']),('product_id','=',record.id)])
            record.reversed_qty_spt =  sum(move_ids.mapped('quantity_done'))
            record.available_qty_spt =  record.qty_available - record.reversed_qty_spt 


    @api.onchange('name')
    def onchange_name(self):
        for record in self:
            if not record.variant_name :
                record.variant_name = record.name
    
    @api.onchange('variant_name')
    def onchange_case_variant_name(self):
        for rec in self:
            if rec.is_case_product:
                rec.name = rec.variant_name

    @api.onchange('case_image_url')
    def onchange_case_image_url(self):
        for record in self:
            case_image_url_1920 = False
            try:
                if not record.case_image_url:
                    raise UserError(_())
                res = requests.get(record.case_image_url.replace('\r',''))
                if res.ok:
                    image = base64.b64encode(res.content)
                    case_image_url_1920 = image
                else:
                    raise UserError(_())
            except:
                # img_path = get_module_resource('tzc_sales_customization_spt', 'static/src/img', 'default_product_img.png')
                # if img_path:
                #     with open(img_path, 'rb') as f:
                #         image = f.read()
                #     case_image_url_1920 = base64.b64encode(image)
                # case_image_url_1920 = False
                pass
            record.write({
                        'case_image_url_128': case_image_url_1920,
                        'case_image_url_256': case_image_url_1920,
                        'case_image_url_512': case_image_url_1920,
                        'case_image_url_1024': case_image_url_1920,
                        'case_image_url_1920': case_image_url_1920,
                        })


    @api.onchange('image_url')
    def onchange_image_url(self):
        for record in self:
            image_variant_1920 = False
            try:
                if not record.image_url:
                    raise UserError(_())
                res = requests.get(record.image_url.replace('\r',''))
                if res.ok:
                    image = base64.b64encode(res.content)
                    image_variant_1920 = image
                else:
                    raise UserError(_())
            except:
                # img_path = get_module_resource('tzc_sales_customization_spt', 'static/src/img', 'default_product_img.png')
                # if img_path:
                #     with open(img_path, 'rb') as f:
                #         image = f.read()
                #     image_variant_1920 = base64.b64encode(image)
                # image_variant_1920 = False
                pass
            record.write({'image_variant_1920': image_variant_1920})

            
    @api.onchange('image_secondary_url')
    def onchange_image_secondary(self):
        for record in self:
            image_secondary = False
            try:
                if not record.image_secondary_url:
                    raise UserError(_())
                res = requests.get(record.image_secondary_url.replace('\r',''))
                if res.ok:
                    image = base64.b64encode(res.content)
                    image_secondary = image
                else:
                    raise UserError(_())
            except:
                # img_path = get_module_resource('tzc_product_import_spt', 'static/src/img', 'default_product_img.png')
                # if img_path:
                #     with open(img_path, 'rb') as f:
                #         image = f.read()
                #         image_secondary = base64.b64encode(image)
                # image_secondary = False
                pass
            record.write({'image_secondary':image_secondary})

    def action_product_reserv_qty(self):
        sml= self.env['stock.move.line'].search([('state','not in',['done','cancel']),('qty_done','!=',0),('product_id','in',self.ids)])
        pick_id = sml.mapped("picking_id")
        order_ids = pick_id.mapped('order_id').ids
        order_ids = list(set(order_ids))
        return {
            'name': _('Reserved Products'),
            'view_mode': 'tree,form',
            'res_model': 'sale.order',
            'views': [
                (self.env.ref('sale.view_order_tree').id, 'tree'),
                (self.env.ref('sale.view_order_form').id, 'form'),],
            'type': 'ir.actions.act_window',
            'domain': [('id','in',order_ids)],
        }

    def actoin_get_sale_order(self):
        for record in self:
            orders = self.env['sale.order']
            for line in record.sale_order_line_ids:
                if line.order_id.state not in ['paid','merged','open_inv','draft_inv','cancel']:
                    orders |= line.order_id
                if (line.order_id.source_spt == 'Website' and line.order_id.state == 'draft'):
                        orders -= line.order_id
        return {
            'name': _('Orders'),
            'res_model': 'sale.order',
            'view_mode': 'tree,form',
            'views': [
                (self.env.ref('sale.view_order_tree').id, 'tree'),
                (self.env.ref('sale.view_order_form').id, 'form'),
                ],
            'type': 'ir.actions.act_window',
            'domain': [('id', 'in',orders.ids)],
            'context': {
                'search_default_my_quotation' : 0
            }
        }


    def website_is_publish_form(self):
        self.write({'is_published_spt':True})
    
    def website_is_unpublish_form(self):
        self.write({'is_published_spt':False})

    def open_product_variant_spt(self):
        tree_view_id = self.env.ref('product.product_product_tree_view').id
        form_view_id = self.env.ref('product.product_normal_form_view').id
        # product_ids = self.search([('product_tmpl_id','=',self.product_tmpl_id.id)])
        country_id = self.env.user.country_id.ids
        product_ids = self.env['product.product'].search([('geo_restriction','not in',country_id),('brand','=',self.product_variant_ids.brand.name),('model','=',self.product_variant_ids.model.name),('categ_id','=',self.product_variant_ids.categ_id.name),('id','!=',self.id)])
        return {
            'name': _('Product Variants'),
            'view_mode': 'tree,form',
            'view_type':'form',
            'views':[[tree_view_id,'tree'],[form_view_id,'form']],
            'domain': [('id','in',product_ids.ids)],
            'res_model': 'product.product',
            'type': 'ir.actions.act_window',
        }

    def action_view_sales(self):
        tree_view_id = self.env.ref('sale.view_quotation_tree_with_onboarding').id
        form_view_id = self.env.ref('sale.view_order_form').id

        order_ids = self.env['sale.order.line'].search([('product_id','=',self.id)]).mapped('order_id')

        return {
            'name': _('Orders'),
            'view_mode': 'tree,form',
            'view_type':'form',
            'views':[[tree_view_id,'tree'],[form_view_id,'form']],
            'domain': [('id','in',order_ids.ids)],
            'res_model': 'sale.order',
            'type': 'ir.actions.act_window'
        }
    
    def action_update_quantity_on_hand(self):
        return self.product_tmpl_id.with_context(default_product_id=self.id,reserve_calculated=True, create=True).action_update_quantity_on_hand()

    def refresh_product_image(self):
        for rec in self.search([]):
            rec.get_image()

    def action_create_bundal_product(self):
        unpublished = self.filtered(lambda product: not product.is_published_spt)
        self = self - unpublished
        out_of_stock = self.filtered(lambda x: x.available_qty_spt <= 0)
        self = self - out_of_stock
        return {
            'name':_('Create Package'),
            'type':"ir.actions.act_window",
            'res_model':'kits.warning.wizard',
            'view_mode':'form',
            'view_id':self.env.ref('tzc_sales_customization_spt.kits_warning_wizard_unpublished_product_form_view').id,
            'context':{
                'default_allowed_products':[(6,0,self.ids)],
                'default_out_of_stock_products':[(6,0,out_of_stock.ids)],
                'default_unpublished_products':[(6,0,unpublished.ids)],
                },
            'target':'new',
        }

    def action_open_clearance_wizard(self):
        return {
            'name': _('Set On Sale Price'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'on.sale.price.wizard.spt',
            'target': 'new',
            'context' : {
                'default_product_ids': [(6,0,self.ids)],'default_sale_type': 'clearance'}
        }
    
    def action_open_on_sale_wizard(self):
        return {
            'name': _('Set On Sale Price'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'on.sale.price.wizard.spt',
            'target': 'new',
            'context' : {
                'default_product_ids': [(6,0,self.ids)],'default_sale_type' :'on_sale'}
        }
    
    def action_remove_on_sale(self):
        for record in self:
            record.sale_type = False

    def action_product_published_spt(self):
        for record in self:
            record.is_published_spt = True
    
    def action_product_unpublished_spt(self):
        for record in self:
            record.is_published_spt = False

    def action_add_in_new_arrivals(self):
        for rec in self:
            if not rec.new_arrivals:
                rec.new_arrivals = True

    def action_remove_from_new_arrivals(self):
        for rec in self:
            if rec.new_arrivals:
                rec.new_arrivals = False

    def action_add_to_new_price(self):
        for rec in self:
            rec.is_new_price = True

    def action_remove_to_new_price(self):
        for rec in self:
            rec.is_new_price = False

    @api.model
    def create(self, vals):
        res = super(ProductProduct, self).create(vals)
        # no field as inventory_availability
        # if res.type == 'product':
        #     res.inventory_availability = 'always'
        return res

    def write(self, vals):
        field_list = ['variant_name','brand','default_code','image_url','lst_price','case_type','height','width','length','weight','volume']
        update = self.env['ir.model']._updated_data_validation(field_list,vals,self._name)
        if update:
            vals.update({'updated_by':self.env.user.id,'updated_on':datetime.today()})
        if 'active' in vals and (vals['active'] == False or vals['active']) and not self.env.context.get('from_product_import'):
            # if not self.env.user.has_group('base.group_system'):
            raise UserError(_('Due to security restrictions, you are not allowed to "%s" this record.'%('Unarchive' if vals['active'] else 'Archive')))
        # elif 'active' in vals and vals['active'] == True:
        #     if not self.env.user.has_group('base.group_system'):
        #         raise UserError(_('Due to security restrictions, you are not allowed to "Unarchive" this record \n Contact your administrator to request access if necessary.'))

        return super(ProductProduct,self).write(vals)

    # def action_open_eto_method_wizard(self):
    #     form_view_id = self.env.ref('tzc_sales_customization_spt.eto_sale_method_wizard_from_view_spt').id
    #     return {
    #         'name': _('Set Pricelist'),
    #         'type': 'ir.actions.act_window',
    #         'view_mode': 'form',
    #         'view_type':'form',
    #         'views':[[form_view_id,'form']],
    #         'res_model': 'eto.sale.method.wizard.spt',
    #         'target': 'new',
    #         'context' : {
    #             'default_product_ids': [(6,0,self.ids)],}
    #     }

    def name_get(self):
        res = super(ProductProduct, self).name_get()
        result = []
        for record in res:
            name = ''
            product_id = self.browse(record[0])
            if product_id:
                if product_id.variant_name:
                    name = product_id.variant_name
                else:
                    name = product_id.name
                    if product_id.color_code or product_id.eye_size:
                        eye_size = str(00)
                        # for attribute in product_id.product_template_attribute_value_ids:
                        #     if attribute.attribute_id.name == 'Color':
                        #         color_name = attribute.product_attribute_value_id.name.split('-')[0]
                        #     if attribute.attribute_id.name == 'Eye Size':
                        #         eye_size = attribute.product_attribute_value_id.name
                        color_name = product_id.color_code.name if product_id.color_code.name else '00'
                        eye_size = product_id.eye_size.name if product_id.eye_size.name else '00'
                        name = name +' ' + color_name + ' ' + eye_size + ' ' + str(product_id.bridge_size.name if product_id.bridge_size else '00') + ' ' + str(product_id.temple_size.name if product_id.temple_size else '00') + (' (' + str(product_id.categ_id.name) + ')' if product_id.categ_id else '')
                # if product_id.categ_id.name not in ['s', 'All', 'Deliveries', 'Expenses', 'Saleable', 'PoS', 'E', 'S', 'e'] and product_id.length:
                #     name += str(product_id.length)
            result.append((product_id.id, name))
        return result

    @tools.ormcache('self.id')
    def _get_first_possible_variant_id(self):
        self.ensure_one()
        return self.id

    def _get_attribute_exclusions(self, parent_combination=None, parent_name=None):
        """Return the list of attribute exclusions of a product.

        :param parent_combination: the combination from which
            `self` is an optional or accessory product. Indeed exclusions
            rules on one product can concern another product.
        :type parent_combination: recordset `product.template.attribute.value`
        :param parent_name: the name of the parent product combination.
        :type parent_name: str

        :return: dict of exclusions
            - exclusions: from this product itself
            - parent_combination: ids of the given parent_combination
            - parent_exclusions: from the parent_combination
           - parent_product_name: the name of the parent product if any, used in the interface
               to explain why some combinations are not available.
               (e.g: Not available with Customizable Desk (Legs: Steel))
           - mapped_attribute_names: the name of every attribute values based on their id,
               used to explain in the interface why that combination is not available
               (e.g: Not available with Color: Black)
        """
        self.ensure_one()
        parent_combination = parent_combination or self.env['product.template.attribute.value']
        return {
            'exclusions': self.product_tmpl_id._complete_inverse_exclusions(self.product_tmpl_id._get_own_attribute_exclusions()),
            'parent_exclusions': self.product_tmpl_id._get_parent_attribute_exclusions(parent_combination),
            'parent_combination': parent_combination.ids,
            'parent_product_name': parent_name,
            'mapped_attribute_names': self.product_tmpl_id._get_mapped_attribute_names(parent_combination),
        }

    def action_view_stock_move_lines(self):
        action = super(ProductProduct,self).action_view_stock_move_lines()
        action['context'] ="{'create': 0}"
        return action
    
    def unpublished_product_from_website_spt(self):

        self.env['product.product'].search([('is_published_spt','=',True)]).write({'is_published_spt':False})

        product_obj = self.env['product.product']
        color_obj = self.env['product.color.spt']
        material_obj = self.env['product.material.spt']
        rim_type_obj = self.env['product.rim.type.spt']
        shape_obj = self.env['product.shape.spt']

        # find other values
        material_ids = material_obj.search([('name','ilike','other')]).ids
        shape_ids = shape_obj.search([('name','ilike','other')]).ids
        rim_type_ids = rim_type_obj.search([('name','ilike','other')]).ids
        color_ids = color_obj.search([('name','ilike','other')]).ids

        # publish products
        # ignoring case products
        product_ids = product_obj.search([('is_case_product','=',False)])
        product_ids = product_ids.filtered(lambda x: x.eye_size_compute > 1 and x.available_qty_spt > 0 and not x.is_image_missing)
        product_ids = product_ids.filtered(lambda x: x.product_color_name.id and x.product_color_name.id not in color_ids)
        product_ids = product_ids.filtered(lambda x: x.rim_type.id and x.rim_type.id not in rim_type_ids)
        # product_ids = product_ids.filtered(lambda x: x.secondary_color_name.id and x.secondary_color_name.id not in color_ids)
        product_ids = product_ids.filtered(lambda x: x.lense_color_name.id and x.lense_color_name.id not in color_ids)
        product_ids = product_ids.filtered(lambda x: x.material_id.filtered(lambda y:y.id not in material_ids))
        product_ids = product_ids.filtered(lambda x: x.shape_id.filtered(lambda y:y.id not in shape_ids))
        product_ids.write({'is_published_spt':True})
        
        # self._cr.commit()

        # unpublish product
        # self.search([('is_published_spt','=',True),('is_forcefully_unpublished','=',True)]).write({'is_published_spt':False})
        # product_ids = self.search([('is_published_spt','=',True)])
        # product_obj |= product_ids.filtered(lambda x: x.available_qty_spt <= 0 or x.eye_size_compute < 1 and x.is_image_missing)
        # product_obj |= product_ids.filtered(lambda x: x.product_color_name.id in color_ids) 
        # product_obj |= product_ids.filtered(lambda x: x.rim_type.id in rim_type_ids)
        # product_obj |= product_ids.filtered(lambda x: x.secondary_color_name.id in color_ids)
        # product_obj |= product_ids.filtered(lambda x: x.lense_color_name.id in color_ids)
        # product_obj |= product_ids.filtered(lambda x: x.material_id.filtered(lambda y:y.id in material_ids))
        # product_obj |= product_ids.filtered(lambda x: x.shape_id.filtered(lambda y:y.id in shape_ids))
        # product_obj.write({'is_published_spt':False})
        # self._cr.commit()

        # Declare
        # eye_glass_cated_id = self.env['product.category'].search([('name','=','E')])
        # sun_glass_cated_id = self.env['product.category'].search([('name','=','S')])
        # eye_glass_website_slider_product_ids = self.search([('is_published_spt','=',True),('available_in_slider_spt','=',True),('categ_id','=',eye_glass_cated_id.id)])
        # sun_glass_website_slider_product_ids = self.search([('is_published_spt','=',True),('available_in_slider_spt','=',True),('categ_id','=',sun_glass_cated_id.id)])

        # check eye glass slider products
        # if len(eye_glass_website_slider_product_ids) < 10:
        #     eye_glass_product_ids = self.search([('is_published_spt','=',True),('categ_id','=',eye_glass_cated_id.id )])
        #     eye_glass_count = len(eye_glass_website_slider_product_ids)
        #     if eye_glass_product_ids :
        #         while(eye_glass_count < 10):
        #             product_id = random.choices(eye_glass_product_ids)[0]
        #             if product_id and product_id.available_qty_spt > 0.00 and not product_id.is_image_missing:
        #                 product_id.action_set_available_in_slider()
        #                 eye_glass_count += 1
        # self._cr.commit()

        # check sunglass slider products
        # if len(sun_glass_website_slider_product_ids) < 10:
        #     sun_glass_product_ids = self.search([('is_published_spt','=',True),('categ_id','=',sun_glass_cated_id.id )])
        #     sun_glass_count = len(sun_glass_website_slider_product_ids)
        #     if sun_glass_product_ids :
        #         while(sun_glass_count < 10):
        #             product_id = random.choices(sun_glass_product_ids)[0]
        #             if product_id and product_id.available_qty_spt > 0.00 and not product_id.is_image_missing:
        #                 product_id.action_set_available_in_slider()
        #                 sun_glass_count += 1
        # self._cr.commit()

        # find published products
        # product_ids = self.search([('is_published_spt','=',True)])

        # publish brands of published products
        # on_brand_ids = product_ids.mapped('brand')
        # brand_ids = brand_obj.search([('id','not in',on_brand_ids.ids)])
        # brand_ids.unpublish_brands_spt()
        # publish_brand_ids =  brand_obj.search([('id','in',on_brand_ids.ids)])
        # publish_brand_ids.publish_brands_spt()
        # self._cr.commit()
        
        # publish categories of published products
        # on_categ_ids = product_ids.mapped('categ_id')
        # categ_ids = categ_obj.search([('id','not in',on_categ_ids.ids)])
        # categ_ids.unpublish_public_category()
        # publish_categ_ids =  categ_obj.search([('id','in',on_categ_ids.ids)])
        # publish_categ_ids.publish_public_category()
        # self._cr.commit()
        
        # publish colors of published products
        # on_color_ids = product_ids.mapped('product_color_name')
        # on_color_ids_spt = product_ids.mapped('secondary_color_name')
        # on_color_ids |= on_color_ids_spt
        # color_ids = color_obj.search([('id','not in',on_color_ids.ids)])
        # color_ids.unpublish_colors_spt()
        # publish_color_ids =  color_obj.search([('id','in',on_color_ids.ids)])
        # publish_color_ids.publish_colors_spt()
        # self._cr.commit()

        # publish material filters of published products
        # on_material_ids = product_ids.mapped('material_id')
        # material_ids = material_obj.search([('id','not in',on_material_ids.ids),('is_published','=',True)])
        # material_ids.unpublish_material_spt()
        # publish_material_ids =  material_obj.search([('id','in',on_material_ids.ids),('is_published','=',False)])
        # publish_material_ids.publish_material_spt()
        # self._cr.commit()

        # publish rim filters of published products
        # on_rim_type_ids = product_ids.mapped('rim_type')
        # rim_type_ids = rim_type_obj.search([('id','not in',on_rim_type_ids.ids),('is_published','=',True)])
        # rim_type_ids.unpublish_rim_type_spt()
        # publish_rim_type_ids =  rim_type_obj.search([('id','in',on_rim_type_ids.ids),('is_published','=',False)])
        # publish_rim_type_ids.publish_rim_type_spt()
        # self._cr.commit()

        # publish shape filters of published products
        # on_shape_ids = product_ids.mapped('shape_id')
        # shape_ids = shape_obj.search([('id','not in',on_shape_ids.ids),('is_published','=',True)])
        # shape_ids.unpublish_shape_spt()
        # publish_shape_ids =  shape_obj.search([('id','in',on_shape_ids.ids),('is_published','=',False)])
        # publish_shape_ids.publish_shape_spt()
        # self._cr.commit()

        # product published for pages
        # without_clearance_products = product_obj.search([('is_published_spt','=',True),('sale_type','!=','clearance'),('is_forcefully_unpublished','=',False)])
        # clearance_products = product_obj.search([('is_published_spt','=',True),('sale_type','=','clearance'),('is_forcefully_unpublished','=',False)])
        # eyeglass_products = without_clearance_products.filtered(lambda pro : pro.categ_id.name in ['E','e'])
        # sunglass_products = without_clearance_products.filtered(lambda pro : pro.categ_id.name in ['s','S'])
        # new_arrivals_products =  product_obj.search([('is_published_spt','=',True),('new_arrivals','=',True),('is_forcefully_unpublished','=',False)])
        # sale_type_products =  product_obj.search([('is_published_spt','=',True),('sale_type','!=',False),('is_forcefully_unpublished','=',False)])
        
        # for brand
        # brand_obj.search([]).write({'website_published':False,'eyeglass_avl_brand':False,'sunglass_avl_brand':False,'new_arrival_avl_brand':False,'sale_avl_brand':False})
        
        # without_clearance_products.mapped('brand').write({'website_published': True}) 
        # clearance_products.mapped('brand').write({'website_published': True}) 
        # eyeglass_products.mapped('brand').write({'eyeglass_avl_brand': True})
        # sunglass_products.mapped('brand').write({'sunglass_avl_brand': True})
        # new_arrivals_products.mapped('brand').write({'new_arrival_avl_brand': True})
        # sale_type_products.mapped('brand').write({'sale_avl_brand': True})

        #for color
        # color_obj.search([]).write({'is_published':False,'eyeglass_avl_colour':False,'sunglass_avl_colour':False,'new_arrival_avl_colour':False,'sale_avl_colour':False})

        # without_clearance_products.mapped('product_color_name').write({'is_published': True})
        # eyeglass_products.mapped('product_color_name').write({'eyeglass_avl_colour': True})
        # sunglass_products.mapped('product_color_name').write({'sunglass_avl_colour': True})
        # new_arrivals_products.mapped('product_color_name').write({'new_arrival_avl_colour': True})
        # sale_type_products.mapped('product_color_name').write({'sale_avl_colour': True})
        
        # without_clearance_products.mapped('secondary_color_name').write({'is_published': True})
        # eyeglass_products.mapped('secondary_color_name').write({'sale_avl_colour': True})
        # sunglass_products.mapped('secondary_color_name').write({'sunglass_avl_colour': True})
        # new_arrivals_products.mapped('secondary_color_name').write({'new_arrival_avl_colour': True})
        # sale_type_products.mapped('secondary_color_name').write({'sale_avl_colour': True})
        
        # shape
        # shape_ids.search([]).write({'is_published':False,'eyeglass_avl_shape':False,'sunglass_avl_shape':False,'new_arrival_avl_shape':False,'sale_avl_shape':False})

        # without_clearance_products.mapped('shape_id').write({'is_published': True})
        # eyeglass_products.mapped('shape_id').write({'eyeglass_avl_shape': True})
        # sunglass_products.mapped('shape_id').write({'sunglass_avl_shape': True})
        # new_arrivals_products.mapped('shape_id').write({'new_arrival_avl_shape': True})
        # sale_type_products.mapped('shape_id').write({'sale_avl_shape': True})
        
        # rim_type
        # rim_type_obj.search([]).write({'is_published':False,'eyeglass_avl_rim_type':False,'sunglass_avl_rim_type':False,'new_arrival_avl_rim_type':False,'sale_avl_rim_type':False})

        # without_clearance_products.mapped('rim_type').write({'is_published': True})
        # eyeglass_products.mapped('rim_type').write({'eyeglass_avl_rim_type': True})
        # sunglass_products.mapped('rim_type').write({'sunglass_avl_rim_type': True})
        # new_arrivals_products.mapped('rim_type').write({'new_arrival_avl_rim_type': True})
        # sale_type_products.mapped('rim_type').write({'sale_avl_rim_type': True})
        
        # material
        # material_obj.search([]).write({'is_published':False,'eyeglass_avl_material':False,'sunglass_avl_material':False,'new_arrival_avl_material':False,'sale_avl_material':False})

        # without_clearance_products.mapped('material_id').write({'is_published': True})
        # eyeglass_products.mapped('material_id').write({'eyeglass_avl_material': True})
        # sunglass_products.mapped('material_id').write({'sunglass_avl_material': True})
        # new_arrivals_products.mapped('material_id').write({'new_arrival_avl_material': True})
        # sale_type_products.mapped('material_id').write({'sale_avl_material': True})

        # product_brand_ids = brand_obj.search([]).with_context(active_test=False).filtered(lambda x:not x.kits_product_ids).unlink()
        # product_color_ids = color_obj.search([]).with_context(active_test=False).filtered(lambda x:x.primary_color_products <= 0 and x.secondary_color_products <= 0).unlink()
        # product_rim_type_ids = rim_type_obj.search([]).with_context(active_test=False).filtered(lambda x:x.products_count <= 0).unlink()
        # product_material_ids =  material_obj.search([]).with_context(active_test=False).filtered(lambda x:not x.kits_product_ids).unlink()
        # product_shape_ids = shape_obj.search([]).with_context(active_test=False).filtered(lambda x:not x.kits_product_ids).unlink()

    def get_onsale_price_spt(self,min_price,max_price,pricelist):
        query = """
                    DROP TABLE IF EXISTS temp_price_table;
                    CREATE TEMP TABLE temp_price_table(
                        id int PRIMARY KEY,
                        price numeric,
                        discount numeric,
                        inflation_price numeric
                    );
                    INSERT INTO temp_price_table(
                        id,
                        price,
                        inflation_price,
                        discount
                    )
                    (
                    SELECT
                        pp.id,
                        pp.%(field)s,
                        (
                            select inflation_rate from kits_inflation_rule kir
                            left join inflation_country_rel icr on kir.id = icr.discount_id
                            left join inflation_brand_rel ibr on kir.id = ibr.model_id
                            left join kits_inflation ki on kir.inflation_id = ki.id
                            where icr.brand_id = %(country_id)s and ibr.brand_id = pp.brand and ki.is_active = true
                            and (
                                case when ki.from_date is not null and ki.to_date is not null then (ki.from_date <= NOW() and NOW() <= ki.to_date)
                                when ki.from_date is not null then ki.from_date <= NOW()
                                when ki.to_date is not null then ki.to_date >= NOW()
                                when (ki.from_date is null or ki.to_date is null ) then true
                                else false
                                end
                            )
                        ),
                        (
                        select discount from kits_special_discount ksd
                            left join special_discount_country_rel cids on ksd.id = cids.discount_id
                            left join special_discount_brand_rel bids on ksd.id = bids.model_id
                            left join tzc_fest_discount fd on ksd.tzc_fest_id = fd.id
                                where cids.brand_id = %(country_id)s and bids.brand_id = pp.brand and fd.is_active = true and (
                                            case when fd.from_date is not null and fd.to_date is not null then (fd.from_date <= NOW() and NOW() <= fd.to_date)
                                            when fd.from_date is not null then fd.from_date <= NOW()
                                            when fd.to_date is not null then fd.to_date >= NOW()
                                            when (fd.from_date is null or fd.to_date is null ) then true
                                            else false
                                            end
                                        ))
                        FROM product_pricelist_item ppi
                        LEFT JOIN product_product pp ON ppi.product_id = pp.id
                        WHERE pp.id in %(pids)s and ppi.pricelist_id = %(pricelist_id)s and pp.is_published_spt = true
                    );

                    select id from temp_price_table
                    where 
                    (
                        (case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)-
                        ((case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)*0.01*coalesce(discount,0.0))
                    ) between %(min_price)s and %(max_price)s
                    ;
        """%({
                        'field':'on_sale_usd',
                        'pids': str(self.ids).replace('[','(').replace(']',')'),
                        'country_id': self.env.user.country_id.id,
                        'min_price': min_price,
                        'max_price': max_price,
                        'pricelist_id': pricelist.id
                    })
        self._cr.execute(query)
        result = self._cr.fetchall()
        result = [i[0] for i in result]
        return result

    def get_clearance_price_spt(self,min_price,max_price,pricelist):
        query = """ 
                    DROP TABLE IF EXISTS temp_price_table;
                    CREATE TEMP TABLE temp_price_table(
                        id int PRIMARY KEY,
                        price numeric,
                        discount numeric,
                        inflation_price numeric
                    );
                    INSERT INTO temp_price_table(
                        id,
                        price,
                        inflation_price,
                        discount
                    )
                    (
                    SELECT
                        pp.id,
                        pp.%(field)s,
                        (
                            select inflation_rate from kits_inflation_rule kir
                            left join inflation_country_rel icr on kir.id = icr.discount_id
                            left join inflation_brand_rel ibr on kir.id = ibr.model_id
                            left join kits_inflation ki on kir.inflation_id = ki.id
                            where icr.brand_id = %(country_id)s and ibr.brand_id = pp.brand and ki.is_active = true
                            and (
                                case when ki.from_date is not null and ki.to_date is not null then (ki.from_date <= NOW() and NOW() <= ki.to_date)
                                when ki.from_date is not null then ki.from_date <= NOW()
                                when ki.to_date is not null then ki.to_date >= NOW()
                                when (ki.from_date is null or ki.to_date is null ) then true
                                else false
                                end
                            )
                        ),
                        (
                        select discount from kits_special_discount ksd
                            left join special_discount_country_rel cids on ksd.id = cids.discount_id
                            left join special_discount_brand_rel bids on ksd.id = bids.model_id
                            left join tzc_fest_discount fd on ksd.tzc_fest_id = fd.id
                                where cids.brand_id = %(country_id)s and bids.brand_id = pp.brand and fd.is_active = true and (
                                            case when fd.from_date is not null and fd.to_date is not null then (fd.from_date <= NOW() and NOW() <= fd.to_date)
                                            when fd.from_date is not null then fd.from_date <= NOW()
                                            when fd.to_date is not null then fd.to_date >= NOW()
                                            when (fd.from_date is null or fd.to_date is null ) then true
                                            else false
                                            end
                                        ))
                        FROM product_pricelist_item ppi
                        LEFT JOIN product_product pp ON ppi.product_id = pp.id
                        WHERE pp.id in %(pids)s and ppi.pricelist_id = %(pricelist_id)s and pp.is_published_spt = true
                    );

                    select id from temp_price_table
                    where 
                    (
                        (case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)-
                        ((case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)*0.01*coalesce(discount,0.0))
                    ) between %(min_price)s and %(max_price)s
                    ;"""%({
                            'field':'clearance_cad' if pricelist.currency_id.name == 'CAD' else 'clearance_usd',
                            'pids': str(self.ids).replace('[','(').replace(']',')'),
                            'country_id': self.env.user.country_id.id,
                            'min_price': min_price,
                            'max_price': max_price,
                            'pricelist_id': pricelist.id
                        })
        self._cr.execute(query)
        result = self._cr.fetchall()
        result = [i[0] for i in result]
        return result

    def get_pricelist_price_for_product(self,min_price,max_price,price_list):
        products,limit,offset = [],5000,0
        for i in range(0,math.ceil(len(self.ids)/limit)):
            prods = self.ids[offset:offset+limit]
            query = """
                        DROP TABLE IF EXISTS temp_price_table;
                        CREATE TEMP TABLE temp_price_table(
                            id int PRIMARY KEY,
                            price numeric,
                            discount numeric,
                            inflation_price numeric
                        );
                        INSERT INTO temp_price_table(
                            id,
                            price,
                            inflation_price,
                            discount
                        )
                        (
                        SELECT
                            pp.id,
                            ppi.%(field)s,
                            (
                                select inflation_rate from kits_inflation_rule kir
                                left join inflation_country_rel icr on kir.id = icr.discount_id
                                left join inflation_brand_rel ibr on kir.id = ibr.model_id
                                left join kits_inflation ki on kir.inflation_id = ki.id
                                where icr.brand_id = %(country_id)s and ibr.brand_id = pp.brand and ki.is_active = true
                                and (
                                    case when ki.from_date is not null and ki.to_date is not null then (ki.from_date <= NOW() and NOW() <= ki.to_date)
                                    when ki.from_date is not null then ki.from_date <= NOW()
                                    when ki.to_date is not null then ki.to_date >= NOW()
                                    when (ki.from_date is null or ki.to_date is null ) then true
                                    else false
                                    end
                                )
                            ),
                            (
                            select discount from kits_special_discount ksd
                                left join special_discount_country_rel cids on ksd.id = cids.discount_id
                                left join special_discount_brand_rel bids on ksd.id = bids.model_id
                                left join tzc_fest_discount fd on ksd.tzc_fest_id = fd.id
                                    where cids.brand_id = %(country_id)s and bids.brand_id = pp.brand and fd.is_active = true and (
                                                case when fd.from_date is not null and fd.to_date is not null then (fd.from_date <= NOW() and NOW() <= fd.to_date)
                                                when fd.from_date is not null then fd.from_date <= NOW()
                                                when fd.to_date is not null then fd.to_date >= NOW()
                                                when (fd.from_date is null or fd.to_date is null ) then true
                                                else false
                                                end
                                            ))
                            FROM product_pricelist_item ppi
                            LEFT JOIN product_product pp ON ppi.product_id = pp.id
                            WHERE pp.id in %(pids)s and ppi.pricelist_id = %(pricelist_id)s and pp.is_published_spt = true
                        );

                        select id from temp_price_table
                        where 
                        (
                            (case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)-
                            ((case when inflation_price is not null then price+(price*0.01*inflation_price) else price end)*0.01*coalesce(discount,0.0))
                        ) between %(min_price)s and %(max_price)s
                        ;
                        """%({
                                'field':'fixed_price',
                                'pids': str(self.ids).replace('[','(').replace(']',')'),
                                'country_id': self.env.user.country_id.id,
                                'min_price': min_price,
                                'max_price': max_price,
                                'pricelist_id': price_list.id
                            })
            self._cr.execute(query)
            product_price_dict_spt = self._cr.fetchall()
            products.extend([i[0] for i in product_price_dict_spt])
            offset+=limit
        return products
    
    def action_edit_products(self):
        wizard_list = []
        for record in self.sorted(lambda x: x.variant_name):
            wizard_list.append((0,0,{
                'product_id' : record.id,
                'product_color_name' : record.product_color_name.id,
                'secondary_color_name' : record.secondary_color_name.id,
                'shape' : [(6,0,record.shape_id.ids)],
                'lense_color_name' : record.lense_color_name.id,
                'rim_type' : record.rim_type.id,
                'material' : [(6,0,record.material_id.ids)],
                'gender' : record.gender,
                'image_1' : record.image_url,
                'image_2' : record.image_secondary_url,
                'onhand_qty':record.qty_available,
            }))
        wizard_id = self.env['product.edit.wizard.spt'].create({'line_ids': wizard_list})

        return {
            'name': _('Product Edit'),
            'view_mode': 'form',
            'res_model': 'product.edit.wizard.spt',
            'view_id': self.env.ref('tzc_sales_customization_spt.product_edit_wizard_form_view_spt').id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'res_id' : wizard_id.id
        }
    
    def get_random_product_image_fields(self,products):
        # image_fields = ['image_variant_512','image_secondary_512'] 
        image_fields = ['image_url','image_secondary_url']
        product_dict = {}
        for product in products:
            # product_dict[product.id] = random.choice(image_fields)
            url = getattr(product,'%s'%random.choice(image_fields))
            try:
                if url:
                    img = urlopen(url).read()
                    image_data = url
            except:
                image_data = "/web/static/src/img/placeholder.png"
            product_dict[product.id] = image_data
        return product_dict

    def unlink(self):
        try:
            if not self.env.user.has_group('base.group_system'):
                raise UserError(_('Due to security restrictions, you are not allowed to delete this record \n Contact your administrator to request access if necessary.'))
            else:
                return super(ProductProduct,self).unlink()
        except Exception as e:
            raise UserError('This product cannot be deleted since there might be some data attached to it. You may delete those data and try again.\n\n'+e.pgerror)


    def _prepare_variant_vals(self):
        self.ensure_one()
        res = {}
        partner_price_list = self.env.user.partner_id.property_product_pricelist.currency_id.name
        if self:
            product_name = self.display_name
            product_brand = self.brand.name
            product_model = self.model.name
            product_barcode = self.barcode
            product_length = self.length
            # if self.product_template_attribute_value_ids:
            #     for val_vart in self.product_template_attribute_value_ids:
            #         if val_vart.attribute_id.name == 'Eye Size':
            #             product_eye_size = val_vart.name.split('-')[0]
            #         if val_vart.attribute_id.name == 'Color':
            #             product_att_color = val_vart.name.split('-')[0]
            product_eye_size = self.eye_size.name or 'N/A'
            product_att_color = self.color_code.name or 'N/A'
            
            res.update({
              'product_name':product_name,
              'product_brand':product_brand or 'N/A',
              'product_model':product_model or 'N/A',
              'product_eye_size':product_eye_size or 'N/A',
              'product_length':product_length or 'N/A',
              'product_att_color':product_att_color or 'N/A',
              'partner_price_list':partner_price_list,
              'product_barcode':product_barcode or 'N/A',
              'product_category':self.categ_id.name,
              'is_in_wish':self._is_in_wishlist()
            })
        return res

    def _compute_sales_count(self):
        for record in self:
            order_ids = self.env['sale.order.line'].search([('product_id','=',record.id)]).mapped('order_id')
            record.sales_count = len(order_ids) if order_ids else 0.0

    def get_product_image(self,image_url):
        img_data = "/web/static/src/img/placeholder.png"
        try:
            if image_url:
                if self.env.context.get('cron'):
                    img_data = []
                    for url in image_url:
                        try:
                            img_url = urlopen(url)
                            img_data.append(True)
                        except:
                            img_data.append(False)
                else:
                    img = urlopen(image_url)
                    img_data = image_url
        except:
            pass

        return img_data
    
    def catalog_report_product_name(self):
        name = ''
        if self.brand:
            name += self.brand.name
        if self.model:
            name += '\n'+self.model.name
        if self.color_code:
            name += '\n'+self.color_code.name
        if self.eye_size:
            name += '\n'+self.eye_size.name
        if self.bridge_size:
            name += ' ' + self.bridge_size.name
        if self.temple_size:
            name += ' ' + self.temple_size.name
        if self.categ_id:
            name += ' (' + self.categ_id.name + ')'

        return name
    
    def get_black_special_friday_sale(self):
        data = self._get_black_special_friday_sale()
        if data.get('current_activate') and data.get('dynamic_label_icon'):
            data['icon'] = str(self.env['tzc.fest.discount'].browse(data.get('dynamic_label_icon')).dynamic_label_icon)[2:-1]
        
        return data

    def cron_notify_on_consignment_product(self):
        product_ids = self.env['product.product'].search([]).filtered(lambda x:x.on_consignment and x.minimum_qty and x.available_qty_spt < x.minimum_qty)
        product_list = []
       
        for product in product_ids:
            product_list.append(product)
        if product_list:
            template_id = self.env.ref('tzc_sales_customization_spt.mail_template_notify_admin_of_on_consigment_product_qty')
            template_id.with_context(product_name = product_list).send_mail(product.id,force_send=True,email_layout_xmlid="mail.mail_notification_light")

    def cron_remove_new_arrival_product(self):
        days = int(self.env['ir.config_parameter'].sudo().get_param('tzc_sales_customization_spt.new_arraival_remove_after'))
        if days:
            products = self.env['product.product'].search([('new_arrivals','=',True)])
            limit = datetime.now()-timedelta(days=days)
            products_to_remove = products.filtered(lambda x : x.new_arrival_update <= limit)
            products_to_remove.action_remove_from_new_arrivals()

    def action_product_data(self):
        wizard_obj = self.env['product.info.wizard.spt']
        query = '''select
            pp.default_code as "SKU",
            pp.variant_name as "Name",
            pbs.name as "Brand",
            pms.name as "Model",
            pcc.name as "Manufacturing Color Code",
            pp.eye_size_compute as "Eye Size",
            pc.name as "Category",
            (pp.available_qty_spt+pp.reversed_qty_spt) as "Total Qty",
            pp.available_qty_spt as "Available Qty",
            pp.reversed_qty_spt as "Reserved Qty",
            pp.lst_price_usd as "Price",
            pp.barcode as "Barcode",
            pp.image_url as "Image",
            pp.image_secondary_url as "Image 2",
            (select case when pp.is_published_spt = True then 'Yes' else 'No' end) as "Is Published",
            (SELECT CASE WHEN pp.is_image_missing = True then 'false' else 'true' end) as "Image Set",
            pcs.name as "Color Name",
            (select name from product_color_spt where id = pp.secondary_color_name) as "Secondary Color Name",
            COALESCE (pp.temporary_out_of_stock,false)::varchar as "Temporary Out Of Stock",
            pp.manufacture_color_code as "Manufacturer Color Code",
            pbss.name as "Bridge Size",
            COALESCE (ptss.name,'')::varchar as "Temple Size",
            (select name from product_color_spt where id = pp.lense_color_name) as "Lence Color Name",
            prts.name as "Rim Type",
            shape.name as "Shape",
            material.name as "Material",
            pp.flex_hinges as "Flex Hinges",
            COALESCE (pp.weight,0.0)::float as "Weight",
            (select case when pp.gender = 'male' then 'M' when pp.gender = 'female' then 'F' when pp.gender = 'm/f' then 'M/F' else '' end ) as "Gender",
            pp.create_date::Date as "Create Date",
            pp.write_Date::Date as "Modify Date",
            case when country.name is NULL then 'N/A' else country.name->>'en_US' end as "Country of Origin",
            pp.order_not_invoice as "#Open Orders",
            (select count(oder.id) from sale_order_line sol 
                        left join sale_order oder on sol.order_id = oder.id
                        where sol.product_id = pp.id and oder.state not in ('cancel','merged')) as "#Order",
            (SELECT CASE WHEN pp.is_forcefully_unpublished = True then 'Yes' else 'No' end) as "Is Forcefully Unpublished"
        from product_product pp
            left join product_template pt on pp.product_tmpl_id = pt.id
            left join product_brand_spt pbs on pp.brand=pbs.id
            left join product_model_spt pms on pp.model=pms.id
            left join product_category pc on pp.categ_id=pc.id
            left join product_color_spt pcs on pp.product_color_name=pcs.id
            left join product_size_spt pss on pp.size=pss.id 
            left join product_bridge_size_spt pbss on pp.bridge_size = pbss.id
            left join product_temple_size_spt ptss on pp.temple_size = ptss.id
            left join product_rim_type_spt prts on pp.rim_type = prts.id
            FULL OUTER join product_with_material_real material_real on pp.id = material_real.product_id
            left join product_material_spt material on material_real.material_id = material.id
            FULL OUTER join product_with_shape_real shape_real on pp.id = shape_real.product_id
            left join product_shape_spt shape on shape_real.shape_id = shape.id 
            left join res_country country on pp.country_of_origin = country.id
            left join kits_product_color_code pcc on pp.color_code = pcc.id
            where pp.active = 'True' order by pp.default_code;'''

        print('\n')
        print(query)
        print('\n')
        self.env.cr.execute(query)
        product_ids = self._cr.fetchall()
        columns = [desc[0] for desc in self.env.cr.description]
        df = pd.DataFrame(product_ids,columns=columns)
        writer = pd.ExcelWriter('/tmp/All_Prodcts_Export.xlsx')
        df.to_excel(writer,index=False,sheet_name="Products")
        writer.save()
        message=  f"From {self.env['product.product'].search_count([('active','=',True)])} products {product_ids.__len__()} products are exported."
        context = {"default_message":message,'all_product':True}
        return {
            "name": _("Exported Products"),
            "type":"ir.actions.act_window",
            "res_model":"warning.spt.wizard",
            "view_mode":"form",
            "view_id":self.env.ref('tzc_sales_customization_spt.warning_wizard_spt_form_view').id,
            "context":context,
            'target':"new",
        }

    def _get_black_special_friday_sale(self):
        active_fest_id = self.env['tzc.fest.discount'].search([('is_active','=',True)])
        if self._context.get('partner_id'):
            partner_id = self._context.get('partner_id')
        else:
            partner_id = self.env.user.partner_id
        records = self.env['kits.special.discount'].search([('country_id','in',partner_id.country_id.ids),('brand_ids','in',self.brand.ids),('tzc_fest_id','=',active_fest_id.id)])
        current_activate = False
        if active_fest_id.from_date and active_fest_id.to_date:
            current_activate = True if datetime.today().date() >= active_fest_id.from_date and datetime.today().date() <= active_fest_id.to_date else False
        elif active_fest_id.from_date and not active_fest_id.to_date:
            current_activate = True if datetime.today().date() >= active_fest_id.from_date else False
        elif not active_fest_id.from_date and active_fest_id.to_date:
            current_activate = True if datetime.today().date() <= active_fest_id.to_date else False
        elif not active_fest_id.from_date and not active_fest_id.to_date:
            current_activate = True
        if records:
            vals = {'discount':records[-1].discount,
                    'active_dynamic_label':active_fest_id.active_dynamic_label_name if active_fest_id.active_dynamic_label_name else False,
                    'dynamic_label_icon':active_fest_id.id if active_fest_id else False,
                    'special_disc_active':True if active_fest_id else False,
                    'is_set_discount_date':True if active_fest_id.from_date or active_fest_id.to_date else False,
                    'current_activate':current_activate}
            return vals
        else:
            return {}

    def get_black_special_friday_sale(self):
        data = self._get_black_special_friday_sale()
        if data.get('current_activate') and data.get('dynamic_label_icon'):
            data['icon'] = str(self.env['tzc.fest.discount'].browse(data.get('dynamic_label_icon')).dynamic_label_icon)[2:-1]
        
        return data

       
    def add_to_catolog(self):
        catalog_obj = self.env['sale.catalog']
        catalog_line_obj = self.env['sale.catalog.line']
        catalog_name = self.env['ir.sequence'].next_by_code('sale.catalog') or 'New'
        catalog_id = catalog_obj.create({'name':catalog_name,'state':'draft'})
        for record in self:
            product_price = record.lst_price_usd
            if record.sale_type:
                if record.sale_type == 'on_sale':
                    product_price = record.on_sale_usd
                if record.sale_type == 'clearance':
                    product_price = record.clearance_usd
                    
            catalog_line_id = catalog_line_obj.create({
                'catalog_id':catalog_id.id,
                'product_pro_id': record.id,
                'product_price_msrp': record.price_msrp_usd,
                'product_price': record.lst_price_usd,
                'product_price_wholesale': record.price_wholesale_usd,
                'product_qty': 1,
                'sale_type' : record.sale_type,
                'unit_discount_price' : product_price,
                })
            catalog_line_id._onchange_fix_discount_price()
            catalog_id._onchange_pricelist_id()
        
        return {
            'name': _('Sale Catalog'),
            'view_mode': 'form',
            'res_id': catalog_id.id,
            'res_model': 'sale.catalog',
            'type': 'ir.actions.act_window',
        }
    
    @api.model
    def get_product_name(self,barcode):
        if barcode:
            product = self.env['product.product'].search([('barcode','=',barcode)])
            if product:
                return {'product_name':product.variant_name ,'product_id':product.id}
            return False
        return False

    # Method for check Inflation & Special Discount.
    ''' :param : list ID's - Country of customer,
                 Boolean - Bypass Inflation & Special Discount
    
        return {
                    'Infaltion': True or False,
                    'Special Discount': True or False,
                    'Infaltion Rate': in %,
                    'Special Discount': in %
                } '''
    def inflation_special_discount(self,country_ids,bypass_flag=False):
        is_special_discount = False 
        is_inflation = False

        # Check Inflation.
        active_inflation = self.env['kits.inflation'].search([('is_active','=',True)])
        inflation_rule_ids = self.env['kits.inflation.rule'].search([('country_id','in',country_ids),('brand_ids','in',self.brand.ids),('inflation_id','=',active_inflation.id)])
        inflation_rule = inflation_rule_ids[-1] if inflation_rule_ids else False
            
        # Validation of Inflation Rule.
        if inflation_rule:
            if active_inflation.from_date and active_inflation.to_date :
                if active_inflation.from_date <= datetime.now().date() and active_inflation.to_date >= datetime.now().date():
                    is_inflation = True
            elif active_inflation.from_date:
                if active_inflation.from_date <= datetime.now().date():
                    is_inflation = True
            elif active_inflation.to_date:
                if active_inflation.to_date >= datetime.now().date():
                    is_inflation = True
            else:
                if not active_inflation.from_date:
                    is_inflation = True
                if not active_inflation.to_date:
                    is_inflation = True
            
        # Check Discount.
        active_fest_id = self.env['tzc.fest.discount'].search([('is_active','=',True)])
        special_disocunt_id = self.env['kits.special.discount'].search([('country_id','in',country_ids),('brand_ids','in',self.brand.ids),('tzc_fest_id','=',active_fest_id.id)])
        price_rule_id = special_disocunt_id[-1] if special_disocunt_id else False
        if price_rule_id:
            # Validation of Special Discount Rule.
            if active_fest_id.from_date and active_fest_id.to_date :
                if active_fest_id.from_date <= datetime.now().date() and active_fest_id.to_date >= datetime.now().date():
                    is_special_discount = True
            elif active_fest_id.from_date:
                if active_fest_id.from_date <= datetime.now().date():
                    is_special_discount = True
            elif active_fest_id.to_date:
                if active_fest_id.to_date >= datetime.now().date():
                    is_special_discount = True
            else:
                if not active_fest_id.from_date:
                    is_special_discount = True
                if not active_fest_id.to_date:
                    is_special_discount = True

        return {
            'is_inflation':is_inflation,
            'inflation_rate':inflation_rule.inflation_rate if inflation_rule else 0,
            'is_special_discount':is_special_discount,
            'special_disc_rate':price_rule_id.discount if price_rule_id else 0,
        }

    def action_open_case_products(self):
        product_categ_id = self.env.ref('tzc_sales_customization_spt.case_product_category')
        if product_categ_id:
            return {
                'name': _('Case Products'),
                'view_mode': 'tree,form',
                'res_model': 'product.product',
                'views': [(self.env.ref('tzc_sales_customization_spt.case_product_view_tree_spt').id,'tree'),
                            (self.env.ref('product.product_normal_form_view').id,'form')],
                'type': 'ir.actions.act_window',
                'domain': [('is_case_product','=',True)],
                # 'context' : {"search_default_filter_is_case_product":True,'case_product':True,'default_is_case_product':True,'default_purchase_ok':False,'default_default_code':"test"}
                'context' : {"search_default_filter_is_case_product":True,'case_product':True,'default_is_case_product':True,'default_purchase_ok':False,'default_categ_id':product_categ_id.id,'default_detailed_type':'product','pending_price':True}
            }
    @api.model
    def _get_view(self, view_id=None, view_type='search', **options):
        arch, view = super(ProductProduct,self)._get_view(view_id, view_type, **options)
        if self._context.get('case_product'):
            if view.type == 'search':
                doc = arch
                str_xml = etree.tostring(doc, encoding='unicode')
                soup = BeautifulSoup(str_xml, "html.parser")
                filters = soup.find_all('filter')
                ignore_filter_names = ['product_brand_group_by','categ_id','product_updated_by_group_by']
                for filter_search in filters:
                    if filter_search.attrs.get('name') not in ignore_filter_names:
                        filter_search['invisible']=1
                arch = etree.fromstring(str(soup))
            if view.type == 'form':
                doc = arch
                str_xml = etree.tostring(doc, encoding='unicode')
                soup = BeautifulSoup(str_xml, "html.parser")
                soup.form['edit']=1
                soup.form['create']=1
                arch = etree.fromstring(str(soup))
        return arch, view
    
    @api.onchange('brand')
    def onchange_brand_internal_ref(self):
        for rec in self:
            if rec.is_case_product:
                if rec.brand:
                    rec.default_code = '-'.join(rec.brand.name.lower().split(' '))+'-case'
                    rec.barcode = rec.default_code
                else:
                    rec.default_code = ''
                if not rec.variant_name and rec.brand:
                    rec.variant_name = rec.brand.name + ' Case'
                rec.barcode = rec.default_code

    @api.onchange('default_code')
    def onchange_default_code_barcode(self):
        for rec in self:
            if rec.is_case_product:
                rec.barcode = rec.default_code

    
    def print_pending_price_product(self):
        header = ['','','Id','Wholesale Price','MSRP Price']
        pricelist_name = self.env['product.pricelist'].search([]).mapped('name')
        header.extend(pricelist_name)
        product_list = []
        for record in self:
            record_list = [record.primary_image_url,record.sec_image_url,record.default_code,record.price_wholesale,record.price_msrp ]
            for pricelist in pricelist_name:
                record_list.append(record.product_pricelist_item_ids.filtered(lambda pl : pl.pricelist_id.name == pricelist).fixed_price)
            product_list.append(record_list)
        
        workbook = Workbook()
        sheet = workbook.create_sheet(title='Product', index=0)
        bd = Side(style='thin', color="000000")
        top_bottom_border = Border(top=bd, bottom=bd)
        heading_font = Font(name="Garamond", size="10", bold=True)
        image_style = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrapText=True)
        row_n = 1
        col = 1
        for heading_data in header:
            sheet.cell(row =row_n,column = col).value = heading_data
            col+=1
        row_n+=1
        for worksheet_line in product_list:
            col_n = 65 
            col = 1
            for worksheet_line_data in worksheet_line:
                sheet.column_dimensions[chr(col_n)].width = 15
                try:
                    if worksheet_line_data and worksheet_line.index(worksheet_line_data) in [0,1]:
                        sheet.row_dimensions[row_n].height = 60
                        img = BytesIO()
                        img.flush()
                        img.write(requests.get(worksheet_line_data).content)
                        image = openpyxl.drawing.image.Image(img)
                        image.width = 128
                        image.height = 65
                        sheet.add_image(image, chr(col_n)+str(row_n))
                        sheet[chr(col_n)+str(row_n)].alignment = image_style
                    else:
                        sheet.cell(row_n,col).value = worksheet_line_data
                    col_n += 1  
                    col += 1  
                except:
                    pass
            row_n +=1
        fp = BytesIO()
        workbook.save(fp)
        img.close()
        fp.seek(0)
        data = fp.read()
        fp.close()
        wiz_id = self.env['warning.spt.wizard'].create({'file':base64.b64encode(data)})

        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.xlsx' % (wiz_id.id, 'pending_price'),
            'target': 'self',
        }

    def wix_product_export_excel(self):
        rows = []
        for rec in self:
            row=[]
            row.append(rec.default_code)
            row.append("Product")
            row.append(f"{str(rec.brand.name)} {str(rec.model.name)} {str(rec.product_color_name.name)}{('/'+str(rec.secondary_color_name.name)) if rec.secondary_color_name else ''}")
            # declaring \n in variable as we cannot use directly in string formatting.
            new_line = '\n'
            description = f"""<p>Product Specifications:</p>
<ul>
<li>Colour - {str(rec.product_color_name.name)}</li>{(new_line+'<li>Secondary Colour - ' + str(rec.secondary_color_name.name) + '</li>') if rec.secondary_color_name.name else ''}
<li>Bridge Size - {(str(rec.bridge_size.name) if rec.bridge_size.name else '00')}</li>
<li>Eyesize - {(str(rec.eye_size.name) if rec.eye_size.name else '00')}</li>
<li>Temple Size - {(str(rec.temple_size.name) if rec.temple_size.name else '00')}</li>
<li>Lens Colour - {(str(rec.lense_color_name.name) if rec.lense_color_name.name else '')}</li>
<li>Rim Type - {(str(rec.rim_type.name) if rec.rim_type.name else '')}</li>
<li>Shape - {(str(rec.shape_id.name) if rec.shape_id.name else '')}</li>
<li>Material - {(str(rec.material_id.name) if rec.material_id.name else '')}</li>
<li>Flex Hinges - {rec.flex_hinges}</li>
<li>Gender - {dict(rec._fields['gender'].selection).get(rec.gender)}</li>
</ul>
<p><em>Actual case may vary from image</em></p>
                """
        
            row.append(description)
            row.append(str(rec.primary_image_url) + ";" + str(rec.sec_image_url) + ";" + str(rec.case_image_url))
            row.append(rec.brand.name)
            row.append(rec.default_code)
            row.append("")
            row.append(rec.lst_price)
            row.append("")
            row.append("TRUE")
            row.append("PERCENT")
            row.append("0")
            row.append("InStock")
            row.append(rec.weight)
            row.append("")
            row.append("Colour")
            row.append("COLOR")
            row.append(f"{str(rec.product_color_name.color)}:{str(rec.product_color_name.name)}")
            rows.append(row)

        # creating a csv file
        with open('wix_product_export.csv', mode='w') as file:
            writer = csv.writer(file, delimiter=',')
            writer.writerow(['handleId','fieldType','name','description','productImageUrl','collection','sku','ribbon','price','surcharge','visible','discountMode','discountValue','inventory','weight','cost','productOptionName1','productOptionType1','productOptionDescription1','productOptionName2','productOptionType2','productOptionDescription2','productOptionName3','productOptionType3','productOptionDescription3','productOptionName4','productOptionType4','productOptionDescription4','productOptionName5','productOptionType5','productOptionDescription5','productOptionName6','productOptionType6','productOptionDescription6','additionalInfoTitle1','additionalInfoDescription1','additionalInfoTitle2','additionalInfoDescription2','additionalInfoTitle3','additionalInfoDescription3','additionalInfoTitle4','additionalInfoDescription4','additionalInfoTitle5','additionalInfoDescription5','additionalInfoTitle6','additionalInfoDescription6','customTextField1','customTextCharLimit1','customTextMandatory1','customTextField2','customTextCharLimit2','customTextMandatory2','brand'])
            for row in rows:
                writer.writerow(row)
        with open('wix_product_export.csv', 'r', encoding="utf-8") as f2:
            data = str.encode(f2.read(), 'utf-8')
            wiz_id = self.env['warning.spt.wizard'].create({'file':base64.encodebytes(data)})
            os.remove('wix_product_export.csv')
            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.csv' % (wiz_id.id, 'wix_product_export'),
                'target': 'self',
            }

    def shopify_product_export_excel(self):
        rows = []
        for rec in self:
            row=[]
            row.append(rec.default_code)
            title = f"{str(rec.brand.name)} {str(rec.model.name)} {str(rec.product_color_name.name)}{('/'+str(rec.secondary_color_name.name)) if rec.secondary_color_name else ''}"
            row.append(title)
            # declaring \n in variable as we cannot use directly in string formatting.
            new_line = '\n'
            body_html = f"""<p>Product Specifications:</p>
<ul>
<li>Colour - {str(rec.product_color_name.name)}</li>{(new_line+'<li>Secondary Colour - ' + str(rec.secondary_color_name.name) + '</li>') if rec.secondary_color_name.name else ''}
<li>Bridge Size - {(str(rec.bridge_size.name) if rec.bridge_size.name else '00')}</li>
<li>Eyesize - {(str(rec.eye_size.name) if rec.eye_size.name else '00')}</li>
<li>Temple Size - {(str(rec.temple_size.name) if rec.temple_size.name else '00')}</li>
<li>Lens Colour - {(str(rec.lense_color_name.name) if rec.lense_color_name.name else '')}</li>
<li>Rim Type - {(str(rec.rim_type.name) if rec.rim_type.name else '')}</li>
<li>Shape - {(str(rec.shape_id.name) if rec.shape_id.name else '')}</li>
<li>Material - {(str(rec.material_id.name) if rec.material_id.name else '')}</li>
<li>Flex Hinges - {rec.flex_hinges}</li>
<li>Gender - {dict(rec._fields['gender'].selection).get(rec.gender)}</li>
</ul>
<p><em>Actual case may vary from image</em></p>
                """
            row.append(body_html)
            row.append(str(rec.brand.name))
            categ_dict = {
                "E" : "Health & Beauty > Personal Care > Vision Care > Eyeglasses",
                "S" : 'Apparel & Accessories > Clothing Accessories > Sunglasses'
            }
            row.append(categ_dict.get(rec.categ_id.name))
            row.append("")
            row.append(f"{rec.rim_type.name}, {dict(rec._fields['gender'].selection).get(rec.gender)}, {rec.shape_id.name}")
            row.append("true")
            row.append("Color")
            row.append(str(rec.product_color_name.name))
            row.append("Eye Size")
            row.append(str(rec.eye_size.name))
            row.append("Material")
            row.append(str(rec.material_id.name))
            row.append(str(rec.default_code))
            row.append(str(rec.weight))
            row.append("shopify")
            row.append(rec.available_qty_spt)
            row.append('deny')
            row.append('manual')
            usd_price_id = self.env['product.pricelist'].search([('name','=','USD Price List')],limit=1).id
            eto_dubai_price_id = self.env['product.pricelist'].search([('name','=','ETO Dubai Price')],limit=1).id
            eto_other_price_id = self.env['product.pricelist'].search([('name','=','Other ETO Branch Price')],limit=1).id
            row.append(rec.product_pricelist_item_ids.filtered(lambda x:x.pricelist_id.id==usd_price_id).fixed_price)
            row.append(rec.product_pricelist_item_ids.filtered(lambda x:x.pricelist_id.id==eto_dubai_price_id).fixed_price)
            row.append(rec.product_pricelist_item_ids.filtered(lambda x:x.pricelist_id.id==eto_other_price_id).fixed_price)
            row.append(rec.price_wholesale)
            row.append('true')
            row.append('false' if rec.categ_id.name=='E' else 'true')
            row.append(rec.barcode)
            row.append(rec.primary_image_url)
            row.append('1')
            row.append('')
            row.append('false')
            row.extend(['','','','','','','','','','','','','','','','',rec.weight_uom_name,'','','true','true','','','true','','','active'])
            # Adding first row with details.
            rows.append(row)
            # Adding second row with secondary image.
            s_row = [rec.default_code,'','','','','','','','','','','','','','','','','','','','','','','','','','',rec.sec_image_url,'2']
            t_row = [rec.default_code,'','','','','','','','','','','','','','','','','','','','','','','','','','',rec.case_image_url,'3']
            rows.append(s_row)
            rows.append(t_row)
        # creating a csv file
        with open('shopify_product_export.csv', mode='w') as file:
            writer = csv.writer(file, delimiter=',')
            writer.writerow(['Handle', 'Title', 'Body (HTML)', 'Vendor', 'Product Category', 'Type', 'Tags', 'Published', 'Option1 Name', 'Option1 Value', 'Option2 Name', 'Option2 Value', 'Option3 Name', 'Option3 Value', 'Variant SKU', 'Variant Grams', 'Variant Inventory Tracker', 'Variant Inventory Qty', 'Variant Inventory Policy', 'Variant Fulfillment Service', 'Variant Price', 'Price 2', 'Price 3','Variant Compare At Price','Variant Requires Shipping', 'Variant Taxable', 'Variant Barcode', 'Image Src', 'Image Position', 'Image Alt Text', 'Gift Card', 'SEO Title', 'SEO Description', 'Google Shopping / Google Product Category', 'Google Shopping / Gender', 'Google Shopping / Age Group', 'Google Shopping / MPN', 'Google Shopping / AdWords Grouping', 'Google Shopping / AdWords Labels', 'Google Shopping / Condition', 'Google Shopping / Custom Product', 'Google Shopping / Custom Label 0', 'Google Shopping / Custom Label 1', 'Google Shopping / Custom Label 2', 'Google Shopping / Custom Label 3', 'Google Shopping / Custom Label 4', 'Variant Image', 'Variant Weight Unit', 'Variant Tax Code', 'Cost per item', 'Included / Canada', 'Included / International', 'Price / International', 'Compare At Price / International', 'Included / United States', 'Price / United States', 'Compare At Price / United States', 'Status'])
            for row in rows:
                writer.writerow(row)
        with open('shopify_product_export.csv', 'r', encoding="utf-8") as f2:
            data = str.encode(f2.read(), 'utf-8')
            # base64.encodestring = base64.encodebytes
            wiz_id = self.env['warning.spt.wizard'].create({'file':base64.encodebytes(data)})
            os.remove('shopify_product_export.csv')
            return {
                'type': 'ir.actions.act_url',
                'url': 'web/content/?model=warning.spt.wizard&download=true&field=file&id=%s&filename=%s.csv' % (wiz_id.id, 'shopify_product_export'),
                'target': 'self',
            }
